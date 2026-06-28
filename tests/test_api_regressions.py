import os
import json
import sqlite3
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch
from zipfile import ZipFile

from fastapi import HTTPException
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from tests.test_support import database, main
import email_service
import license_runtime
import update_service
from db_adapter import CompatRow, PostgresCursorAdapter, _is_postgres_integrity_error, _rewrite_sql_for_postgres


class ApiRegressionTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.client = TestClient(main.app)

    def setUp(self):
        self.connection = database.get_connection()
        self._reset_database()
        with main.AUTH_LOGIN_ATTEMPTS_LOCK:
            main.AUTH_LOGIN_ATTEMPTS.clear()
        with main.FEEDBACK_ATTEMPTS_LOCK:
            main.FEEDBACK_ATTEMPTS.clear()

    def tearDown(self):
        self.connection.close()

    def _reset_database(self):
        cursor = self.connection.cursor()
        for table in (
            "auth_email_verification_tokens",
            "auth_password_reset_tokens",
            "auth_sessions",
            "auth_audit_events",
            "feedback_reports",
            "license_activation_attempts",
            "license_events",
            "licenses",
            "desktop_sync_outbox",
            "organization_invitations",
            "user_department_access",
            "organization_memberships",
            "users",
            "schedule_entries",
            "app_settings",
            "employee_day_statuses",
            "employee_recurring_preferences",
            "employee_week_preference_requests",
            "employee_week_preferences",
            "employee_preferences",
            "coverage_requirements",
            "shift_requirements",
            "employee_positions",
            "shift_templates",
            "positions",
            "employees",
        ):
            cursor.execute(f"DELETE FROM {table}")
        cursor.execute(
            """
            INSERT OR IGNORE INTO departments (
                id, organization_id, public_id, name, description, display_order, is_active
            )
            VALUES (1, 1, 'dep_00000000000000000000000000000001', 'Main department', NULL, 0, 1)
            """
        )
        cursor.execute("DELETE FROM departments WHERE id <> 1")
        cursor.execute(
            """
            UPDATE departments
            SET organization_id = 1,
                public_id = 'dep_00000000000000000000000000000001',
                name = 'Main department',
                description = NULL,
                display_order = 0,
                is_active = 1
            WHERE id = 1
            """
        )
        cursor.execute(
            "UPDATE organizations SET name = 'Local Organization', status = 'active', created_at = CURRENT_TIMESTAMP, updated_at = CURRENT_TIMESTAMP WHERE id = 1"
        )
        self.connection.commit()
        backup_dir = database.get_backup_dir()
        for backup_path in backup_dir.glob("*.db"):
            backup_path.unlink()

    def _employee_payload(self, **overrides):
        payload = {
            "full_name": "Employee A",
            "sex": "female",
            "min_shifts_per_week": 1,
            "target_shifts_per_week": 3,
            "max_shifts_per_week": 5,
            "can_work_night": True,
            "can_work_weekends": True,
            "can_work_evenings_after_night": True,
            "can_work_mornings_and_evenings": True,
        }
        payload.update(overrides)
        return payload

    def _position_payload(self, **overrides):
        payload = {
            "name": "Nurse",
            "requires_continuous_coverage": False,
            "minimum_staff_presence": 0,
        }
        payload.update(overrides)
        return payload

    def _template_payload(self, **overrides):
        position_id = overrides.pop("position_id", None)
        if position_id is None:
            position_id = self._create_position(name="Template Position")
        payload = {
            "position_id": position_id,
            "name": "Morning",
            "category": "morning",
            "start_time": "06:00",
            "end_time": "14:00",
            "is_overnight": False,
            "is_active": True,
            "is_split_only": False,
        }
        payload.update(overrides)
        return payload

    def _create_employee(self, headers=None, **overrides):
        response = self.client.post("/api/employees", json=self._employee_payload(**overrides), headers=headers)
        self.assertEqual(response.status_code, 200)
        return response.json()["employee"]["id"]

    def _create_position(self, headers=None, **overrides):
        response = self.client.post("/api/positions", json=self._position_payload(**overrides), headers=headers)
        self.assertEqual(response.status_code, 200)
        return response.json()["position"]["id"]

    def _create_shift_template(self, headers=None, **overrides):
        response = self.client.post("/api/shift-templates", json=self._template_payload(**overrides), headers=headers)
        self.assertEqual(response.status_code, 200)
        return response.json()["shift_template"]["id"]

    def test_authorization_schema_is_initialized(self):
        cursor = self.connection.cursor()
        for table_name in (
            "organizations",
            "users",
            "organization_memberships",
            "organization_invitations",
            "auth_audit_events",
            "auth_sessions",
            "auth_password_reset_tokens",
            "auth_email_verification_tokens",
            "desktop_sync_outbox",
            "licenses",
            "license_events",
            "license_activation_attempts",
            "feedback_reports",
            "schema_metadata",
            "schema_migrations",
            "departments",
            "user_department_access",
        ):
            cursor.execute(
                "SELECT name FROM sqlite_master WHERE type = 'table' AND name = ?",
                (table_name,),
            )
            self.assertIsNotNone(cursor.fetchone(), table_name)

        cursor.execute("SELECT public_id, name, status FROM organizations WHERE id = 1")
        default_organization = cursor.fetchone()
        self.assertIsNotNone(default_organization)
        self.assertEqual(default_organization["public_id"], "local-default")
        self.assertEqual(default_organization["status"], "active")

        for table_name in (
            "employees",
            "departments",
            "positions",
            "shift_templates",
            "schedule_entries",
            "shift_requirements",
            "employee_preferences",
            "employee_week_preferences",
            "employee_week_preference_requests",
            "employee_recurring_preferences",
            "employee_day_statuses",
            "coverage_requirements",
        ):
            cursor.execute(f"PRAGMA table_info({table_name})")
            columns = {row["name"] for row in cursor.fetchall()}
            self.assertIn("organization_id", columns, table_name)
            self.assertIn("public_id", columns, table_name)
            self.assertIn("updated_by", columns, table_name)

            cursor.execute(
                "SELECT name FROM sqlite_master WHERE type = 'index' AND name = ?",
                (f"idx_{table_name}_public_id",),
            )
            self.assertIsNotNone(cursor.fetchone(), table_name)

        cursor.execute("PRAGMA table_info(app_settings)")
        app_settings_columns = {row["name"] for row in cursor.fetchall()}
        self.assertIn("organization_id", app_settings_columns)

        cursor.execute("SELECT value FROM schema_metadata WHERE key = 'schema_version'")
        schema_version_row = cursor.fetchone()
        self.assertIsNotNone(schema_version_row)
        self.assertEqual(int(schema_version_row["value"]), database.CURRENT_SCHEMA_VERSION)

        cursor.execute("SELECT to_version FROM schema_migrations ORDER BY id DESC LIMIT 1")
        migration_row = cursor.fetchone()
        self.assertIsNotNone(migration_row)
        self.assertEqual(migration_row["to_version"], database.CURRENT_SCHEMA_VERSION)

    def test_shift_template_rebuild_recovers_from_stale_old_table(self):
        connection = sqlite3.connect(":memory:")
        connection.row_factory = sqlite3.Row
        cursor = connection.cursor()
        cursor.execute("PRAGMA foreign_keys = ON")
        cursor.execute(
            """
            CREATE TABLE positions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL
            )
            """
        )
        cursor.execute(
            """
            CREATE TABLE shift_templates (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                position_id INTEGER,
                category TEXT NOT NULL CHECK (category IN ('morning', 'evening', 'night')),
                name TEXT NOT NULL,
                start_time TEXT NOT NULL,
                end_time TEXT NOT NULL,
                is_overnight INTEGER NOT NULL DEFAULT 0,
                is_active INTEGER NOT NULL DEFAULT 1,
                is_split_only INTEGER NOT NULL DEFAULT 0,
                organization_id INTEGER NOT NULL DEFAULT 1,
                public_id TEXT,
                created_at TEXT,
                updated_at TEXT,
                updated_by INTEGER,
                UNIQUE(position_id, name)
            )
            """
        )
        cursor.execute("INSERT INTO positions (id, name) VALUES (1, 'Nurse')")
        cursor.execute(
            """
            INSERT INTO shift_templates (
                id, position_id, category, name, start_time, end_time, is_overnight, is_active,
                is_split_only, organization_id, public_id, created_at, updated_at, updated_by
            )
            VALUES (1, 1, 'morning', 'Morning', '08:00', '16:00', 0, 1, 0, 1, 'tpl_existing', NULL, NULL, NULL)
            """
        )
        cursor.execute(
            """
            CREATE TABLE shift_templates_old (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                category TEXT NOT NULL,
                start_time TEXT NOT NULL,
                end_time TEXT NOT NULL,
                is_overnight INTEGER NOT NULL DEFAULT 0,
                is_active INTEGER NOT NULL DEFAULT 1,
                is_split_only INTEGER NOT NULL DEFAULT 0
            )
            """
        )
        cursor.execute(
            """
            INSERT INTO shift_templates_old (id, name, category, start_time, end_time)
            VALUES (99, 'Stale', 'night', '22:00', '06:00')
            """
        )

        database._rebuild_shift_templates_table(cursor)

        cursor.execute("SELECT COUNT(*) AS count FROM shift_templates")
        self.assertEqual(cursor.fetchone()["count"], 1)
        cursor.execute("SELECT name, public_id FROM shift_templates WHERE id = 1")
        row = cursor.fetchone()
        self.assertEqual(row["name"], "Morning")
        self.assertEqual(row["public_id"], "tpl_existing")
        cursor.execute("SELECT name FROM sqlite_master WHERE type = 'table' AND name = 'shift_templates_old'")
        self.assertIsNone(cursor.fetchone())
        connection.close()

    def test_openapi_exposes_bearer_auth_for_protected_routes(self):
        main.app.openapi_schema = None
        schema = main.app.openapi()

        security_schemes = schema["components"]["securitySchemes"]
        self.assertEqual(security_schemes["BearerAuth"]["type"], "http")
        self.assertEqual(security_schemes["BearerAuth"]["scheme"], "bearer")

        employees_get = schema["paths"]["/api/employees"]["get"]
        employees_post = schema["paths"]["/api/employees"]["post"]
        self.assertIn({"BearerAuth": []}, employees_get["security"])
        self.assertIn({"BearerAuth": []}, employees_post["security"])
        self.assertNotIn(
            ("authorization", "header"),
            [(parameter["name"], parameter["in"]) for parameter in employees_get.get("parameters", [])],
        )
        self.assertNotIn("security", schema["paths"]["/api/auth/status"]["get"])

    def test_docs_page_uses_current_browser_session_token(self):
        response = self.client.get("/docs")

        self.assertEqual(response.status_code, 200)
        self.assertIn("schedule_app_auth_token", response.text)
        self.assertIn("requestInterceptor", response.text)
        self.assertIn("BearerAuth", response.text)
        self.assertIn("Bearer \" + token", response.text)

    def test_license_status_defaults_to_trial_runtime(self):
        response = self.client.get("/api/license/status")
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["status"], "trial")
        self.assertEqual(payload["source"], "trial")
        self.assertEqual(payload["plan_code"], "trial")
        self.assertEqual(payload["employee_limit"], 15)
        self.assertTrue(payload["enforcement"]["can_generate_schedule"])

    def test_license_status_uses_local_developer_bypass_only_with_developer_mode(self):
        with patch.dict(
            os.environ,
            {
                "APP_ENV": "development",
                "K_SERVICE": "",
                "SCHEDULE_APP_DEVELOPER_MODE": "1",
                "SCHEDULE_APP_LICENSE_BYPASS": "1",
            },
            clear=False,
        ):
            response = self.client.get("/api/license/status")
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["status"], "active")
        self.assertEqual(payload["source"], "developer_bypass")
        self.assertEqual(payload["plan_code"], "developer")
        self.assertEqual(payload["employee_limit"], 9999)
        self.assertTrue(payload["developer_bypass"])
        self.assertEqual(payload["message"], "Developer license bypass enabled")
        self.assertTrue(payload["enforcement"]["can_generate_schedule"])

    def test_license_bypass_is_ignored_in_deployed_runtime(self):
        with patch.dict(
            os.environ,
            {
                "APP_ENV": "staging",
                "K_SERVICE": "",
                "SCHEDULE_APP_DEVELOPER_MODE": "1",
                "SCHEDULE_APP_LICENSE_BYPASS": "1",
            },
            clear=False,
        ):
            response = self.client.get("/api/license/status")
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["status"], "trial")
        self.assertEqual(payload["source"], "trial")
        self.assertNotIn("developer_bypass", payload)

    def test_license_import_stores_signed_certificate_and_updates_status(self):
        certificate = {
            "license_id": "lic_test_001",
            "organization_public_id": "local-default",
            "customer_legal_name": "Beta Clinic",
            "branch_id": "main",
            "plan_code": "team_35",
            "employee_limit": 35,
            "features": ["desktop", "employee_portal"],
            "issued_at": "2026-05-01T00:00:00",
            "support_cloud_expires_at": "2027-05-01",
            "grace_ends_at": "2027-05-15",
            "key_id": "dev",
            "signature_scheme": "unsigned-dev-v1",
            "signature": "development",
        }
        with patch.dict(os.environ, {"SCHEDULE_APP_DEVELOPER_MODE": "1"}):
            response = self.client.post("/api/license/import-file", json={"certificate": certificate})
        self.assertEqual(response.status_code, 200)
        payload = response.json()["license"]
        self.assertEqual(payload["status"], "active")
        self.assertEqual(payload["source"], "license")
        self.assertEqual(payload["license_id"], "lic_test_001")
        self.assertEqual(payload["employee_limit"], 35)

        cursor = self.connection.cursor()
        cursor.execute("SELECT event_type FROM license_events WHERE license_id = ?", ("lic_test_001",))
        self.assertEqual(cursor.fetchone()["event_type"], "license_imported")

    def test_license_import_rejects_wrong_organization(self):
        certificate = {
            "license_id": "lic_test_wrong_org",
            "organization_public_id": "other-org",
            "plan_code": "team_35",
            "employee_limit": 35,
            "issued_at": "2026-05-01T00:00:00",
            "signature_scheme": "unsigned-dev-v1",
            "signature": "development",
        }
        with patch.dict(os.environ, {"SCHEDULE_APP_DEVELOPER_MODE": "1"}):
            response = self.client.post("/api/license/import-file", json={"certificate": certificate})
        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()["detail"], "License certificate belongs to a different organization")

    def test_license_activation_code_imports_signed_certificate(self):
        certificate = {
            "license_id": "lic_activation_001",
            "organization_public_id": "local-default",
            "plan_code": "team_75",
            "employee_limit": 75,
            "features": ["desktop", "employee_portal"],
            "issued_at": "2026-05-01T00:00:00Z",
            "support_cloud_expires_at": "2027-05-01",
            "grace_ends_at": "2027-05-15",
            "key_id": "test",
            "signature_scheme": "hmac-sha256-v1",
        }
        certificate["signature"] = license_runtime.hmac_signature(certificate, "test-secret")
        activation_code = license_runtime.encode_activation_code(certificate)

        with patch.dict(os.environ, {"SCHEDULE_APP_LICENSE_SIGNING_SECRET": "test-secret"}):
            response = self.client.post("/api/license/activate-code", json={"activation_code": activation_code})

        self.assertEqual(response.status_code, 200)
        payload = response.json()["license"]
        self.assertEqual(payload["license_id"], "lic_activation_001")
        self.assertEqual(payload["employee_limit"], 75)
        self.assertEqual(payload["source"], "license")

        cursor = self.connection.cursor()
        cursor.execute("SELECT status FROM license_activation_attempts ORDER BY id DESC LIMIT 1")
        self.assertEqual(cursor.fetchone()["status"], "success")

    def test_runtime_database_seeds_matching_license_from_bundled_database(self):
        certificate = {
            "license_id": "lic_bundled_seed",
            "organization_public_id": "local-default",
            "plan_code": "full_access",
            "employee_limit": 9999,
            "issued_at": "2026-05-06T00:00:00Z",
            "support_cloud_expires_at": "2036-05-06",
            "grace_ends_at": "2036-05-20",
            "status": "active",
            "signature": "bundled-support-signature",
            "signature_scheme": "support-v1",
        }
        with tempfile.TemporaryDirectory() as temp_dir:
            bundled_path = Path(temp_dir) / "schedule_app.db"
            bundled_connection = sqlite3.connect(bundled_path)
            try:
                bundled_cursor = bundled_connection.cursor()
                bundled_cursor.execute(
                    """
                    CREATE TABLE licenses (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        organization_id INTEGER NOT NULL DEFAULT 1,
                        license_id TEXT NOT NULL UNIQUE,
                        status TEXT NOT NULL,
                        plan_code TEXT NOT NULL,
                        employee_limit INTEGER NOT NULL,
                        support_cloud_expires_at TEXT,
                        grace_ends_at TEXT,
                        certificate_json TEXT NOT NULL,
                        signature TEXT NOT NULL,
                        key_id TEXT,
                        source TEXT NOT NULL,
                        imported_at TEXT NOT NULL,
                        last_verified_at TEXT,
                        revoked_at TEXT
                    )
                    """
                )
                bundled_cursor.execute(
                    """
                    INSERT INTO licenses (
                        organization_id, license_id, status, plan_code, employee_limit,
                        support_cloud_expires_at, grace_ends_at, certificate_json, signature,
                        key_id, source, imported_at, last_verified_at, revoked_at
                    )
                    VALUES (1, ?, 'active', 'full_access', 9999, '2036-05-06', '2036-05-20', ?, ?, 'support', 'support', '2026-05-06T00:00:00', '2026-05-06T00:00:00', NULL)
                    """,
                    (
                        certificate["license_id"],
                        json.dumps(certificate, sort_keys=True),
                        certificate["signature"],
                    ),
                )
                bundled_connection.commit()
            finally:
                bundled_connection.close()

            cursor = self.connection.cursor()
            cursor.execute("DELETE FROM licenses")
            self.connection.commit()

            with patch.object(database, "get_bundled_database_path", return_value=bundled_path):
                database._seed_licenses_from_bundled_database(cursor)
            self.connection.commit()

            cursor.execute("SELECT license_id, plan_code, employee_limit FROM licenses")
            license_row = cursor.fetchone()
            self.assertIsNotNone(license_row)
            self.assertEqual(license_row["license_id"], "lic_bundled_seed")
            self.assertEqual(license_row["plan_code"], "full_access")
            self.assertEqual(license_row["employee_limit"], 9999)

    def test_expired_trial_blocks_employee_creation_manual_schedule_and_generation(self):
        employee_id = self._create_employee(full_name="Licensed Employee")
        position_id = self._create_position(name="Licensed Position")
        template_id = self._create_shift_template(position_id=position_id, name="Licensed Morning")

        cursor = self.connection.cursor()
        cursor.execute("UPDATE organizations SET created_at = '2026-01-01T00:00:00' WHERE id = 1")
        self.connection.commit()

        employee_response = self.client.post("/api/employees", json=self._employee_payload(full_name="Blocked Employee"))
        self.assertEqual(employee_response.status_code, 402)
        self.assertEqual(employee_response.json()["detail"]["license_status"], "expired")

        schedule_response = self.client.post(
            "/api/schedule",
            json={
                "employee_id": employee_id,
                "position_id": position_id,
                "date": "2026-05-03",
                "shift_template_id": template_id,
            },
        )
        self.assertEqual(schedule_response.status_code, 402)
        self.assertEqual(schedule_response.json()["detail"]["capability"], "can_create_shift")

        generate_response = self.client.post(
            "/api/schedule/auto-generate",
            json={"position_id": position_id, "week_start_date": "2026-05-03"},
        )
        self.assertEqual(generate_response.status_code, 402)
        self.assertEqual(generate_response.json()["detail"]["capability"], "can_generate_schedule")

    def test_trial_employee_limit_blocks_new_employee_and_developer_bypass_allows_it(self):
        for index in range(15):
            self._create_employee(full_name=f"Trial Employee {index}")

        blocked_response = self.client.post("/api/employees", json=self._employee_payload(full_name="Blocked Limit"))
        self.assertEqual(blocked_response.status_code, 402)
        detail = blocked_response.json()["detail"]
        self.assertEqual(detail["capability"], "can_add_employee")
        self.assertTrue(detail["employee_limit_reached"])

        with patch.dict(
            os.environ,
            {
                "APP_ENV": "development",
                "K_SERVICE": "",
                "SCHEDULE_APP_DEVELOPER_MODE": "1",
                "SCHEDULE_APP_LICENSE_BYPASS": "1",
            },
            clear=False,
        ):
            bypass_response = self.client.post("/api/employees", json=self._employee_payload(full_name="Bypass Employee"))
        self.assertEqual(bypass_response.status_code, 200)

    def test_health_endpoints_report_runtime_readiness(self):
        live_response = self.client.get("/api/health/live")
        self.assertEqual(live_response.status_code, 200)
        self.assertEqual(live_response.json()["status"], "ok")

        ready_response = self.client.get("/api/health/ready")
        self.assertEqual(ready_response.status_code, 200)
        ready_payload = ready_response.json()
        self.assertEqual(ready_payload["status"], "ok")
        self.assertEqual(ready_payload["database"]["status"], "ok")
        self.assertEqual(ready_payload["runtime"]["database_engine"], "sqlite")

    def test_client_config_exposes_public_employee_portal_url(self):
        with patch.dict(os.environ, {"PUBLIC_APP_BASE_URL": "https://shiftcare.example.com/"}, clear=False):
            response = self.client.get("/api/client-config")
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["public_app_base_url"], "https://shiftcare.example.com")
        self.assertEqual(payload["employee_portal_url"], "https://shiftcare.example.com/login")
        self.assertEqual(payload["employee_invitation_url_base"], "https://shiftcare.example.com/accept-invitation")

    def test_auth_status_reports_bootstrap_availability(self):
        empty_response = self.client.get("/api/auth/status")
        self.assertEqual(empty_response.status_code, 200)
        self.assertTrue(empty_response.json()["bootstrap_available"])
        self.assertEqual(empty_response.json()["active_user_count"], 0)

        bootstrap_response = self.client.post(
            "/api/auth/bootstrap",
            json={
                "organization_name": "Beta Clinic",
                "full_name": "Owner User",
                "email": "owner@example.com",
                "password": "CorrectHorse123",
            },
        )
        self.assertEqual(bootstrap_response.status_code, 200)

        populated_response = self.client.get("/api/auth/status")
        self.assertEqual(populated_response.status_code, 200)
        self.assertFalse(populated_response.json()["bootstrap_available"])
        self.assertEqual(populated_response.json()["active_user_count"], 1)

    def test_health_readiness_reports_postgres_configuration_errors(self):
        with patch.dict(os.environ, {"DATABASE_ENGINE": "postgresql"}):
            response = self.client.get("/api/health/ready")
        self.assertEqual(response.status_code, 503)
        payload = response.json()["detail"]
        self.assertEqual(payload["status"], "blocked")
        self.assertTrue(payload["runtime"]["issues"])
        self.assertNotIn("data layer is not yet switched", " ".join(payload["runtime"]["issues"]))

    def test_postgres_adapter_rewrites_sqlite_placeholders_and_rows(self):
        self.assertEqual(
            _rewrite_sql_for_postgres("SELECT * FROM users WHERE id = ? AND email = ?"),
            "SELECT * FROM users WHERE id = %s AND email = %s",
        )
        self.assertEqual(
            _rewrite_sql_for_postgres("ON CONFLICT(key) DO UPDATE SET value = excluded.value"),
            "ON CONFLICT (key) DO UPDATE SET value = excluded.value",
        )
        row = CompatRow(["id", "email"], (7, "owner@example.com"))
        self.assertEqual(row[0], 7)
        self.assertEqual(row["email"], "owner@example.com")
        self.assertEqual(dict(row), {"id": 7, "email": "owner@example.com"})

    def test_postgres_schema_cursor_does_not_request_lastrowid(self):
        class FailingLastvalConnection:
            def execute(self, sql):
                raise AssertionError(f"unexpected connection execute: {sql}")

        class FakeCursor:
            rowcount = 1
            description = None

            def __init__(self):
                self.connection = FailingLastvalConnection()
                self.executed = []

            def execute(self, sql, params=None):
                self.executed.append((sql, params))

        fake_cursor = FakeCursor()
        adapter = PostgresCursorAdapter(fake_cursor, track_lastrowid=False)
        adapter.execute("INSERT INTO app_settings (key, value) VALUES (?, ?)", ("a", "b"))
        self.assertEqual(fake_cursor.executed, [("INSERT INTO app_settings (key, value) VALUES (%s, %s)", ("a", "b"))])
        self.assertIsNone(adapter.lastrowid)

    def test_postgres_restrict_violation_maps_to_sqlite_integrity_error(self):
        RestrictViolation = type(
            "RestrictViolation",
            (Exception,),
            {"__module__": "psycopg.errors", "sqlstate": "23001"},
        )
        self.assertTrue(_is_postgres_integrity_error(RestrictViolation("restricted")))

    def test_auth_bootstrap_login_me_and_logout_flow(self):
        payload = {
            "organization_name": "Beta Clinic",
            "full_name": "Owner User",
            "email": "Owner@Example.COM",
            "password": "CorrectHorse123",
        }
        bootstrap_response = self.client.post("/api/auth/bootstrap", json=payload)
        self.assertEqual(bootstrap_response.status_code, 200)
        bootstrap_body = bootstrap_response.json()
        self.assertEqual(bootstrap_body["token_type"], "bearer")
        self.assertTrue(bootstrap_body["access_token"])
        self.assertEqual(bootstrap_body["user"]["email"], "owner@example.com")
        self.assertEqual(bootstrap_body["user"]["memberships"][0]["role"], "owner")
        self.assertEqual(bootstrap_body["user"]["memberships"][0]["organization_name"], "Beta Clinic")

        duplicate_response = self.client.post("/api/auth/bootstrap", json=payload)
        self.assertEqual(duplicate_response.status_code, 409)

        login_response = self.client.post(
            "/api/auth/login",
            json={"email": "owner@example.com", "password": "CorrectHorse123"},
        )
        self.assertEqual(login_response.status_code, 200)
        token = login_response.json()["access_token"]

        me_response = self.client.get("/api/auth/me", headers={"Authorization": f"Bearer {token}"})
        self.assertEqual(me_response.status_code, 200)
        self.assertEqual(me_response.json()["user"]["email"], "owner@example.com")

        logout_response = self.client.post("/api/auth/logout", headers={"Authorization": f"Bearer {token}"})
        self.assertEqual(logout_response.status_code, 200)

        revoked_response = self.client.get("/api/auth/me", headers={"Authorization": f"Bearer {token}"})
        self.assertEqual(revoked_response.status_code, 401)

    def test_feedback_report_endpoint_saves_report_and_sends_notification(self):
        owner_response = self.client.post(
            "/api/auth/bootstrap",
            json={
                "organization_name": "Feedback Clinic",
                "full_name": "Owner User",
                "email": "owner@example.com",
                "password": "password123",
            },
        )
        self.assertEqual(owner_response.status_code, 200)
        token = owner_response.json()["access_token"]
        organization_id = owner_response.json()["user"]["memberships"][0]["organization_id"]
        headers = {"Authorization": f"Bearer {token}"}

        with patch("main.send_feedback_report_email", return_value=email_service.EmailSendResult("sent")) as send_email:
            response = self.client.post(
                "/api/feedback/reports",
                headers=headers,
                json={
                    "report_type": "bug",
                    "severity": "major",
                    "area": "schedule",
                    "title": "Schedule does not save",
                    "description": "Saving the schedule returns an error from the API.",
                    "steps_to_reproduce": "Open schedule\nClick save",
                    "actual_result": "Error is shown",
                    "expected_result": "Schedule is saved",
                    "organization_id": organization_id,
                    "page_url": "http://testserver/schedule",
                    "client_context": {"frontend_errors": [{"message": "boom"}]},
                },
            )

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertEqual(body["notification"]["status"], "sent")
        self.assertTrue(body["report"]["public_id"].startswith("fbr_"))
        send_email.assert_called_once()

        cursor = self.connection.cursor()
        cursor.execute("SELECT * FROM feedback_reports WHERE public_id = ?", (body["report"]["public_id"],))
        report = cursor.fetchone()
        self.assertIsNotNone(report)
        self.assertEqual(report["report_type"], "bug")
        self.assertEqual(report["severity"], "major")
        self.assertEqual(report["area"], "schedule")
        self.assertEqual(report["notification_status"], "sent")
        self.assertIn("boom", report["client_context_json"])

        cursor.execute(
            "SELECT event_type FROM auth_audit_events WHERE event_type = 'feedback_report_created'"
        )
        self.assertIsNotNone(cursor.fetchone())

    def test_employee_can_login_with_linked_id_card(self):
        owner_response = self.client.post(
            "/api/auth/create-organization",
            json={
                "organization_name": "Beta Clinic",
                "full_name": "Owner User",
                "email": "owner@example.com",
                "password": "CorrectHorse123",
            },
        )
        self.assertEqual(owner_response.status_code, 200)
        owner_headers = {"Authorization": f"Bearer {owner_response.json()['access_token']}"}
        organization_id = owner_response.json()["user"]["memberships"][0]["organization_id"]
        employee_id = self._create_employee(headers=owner_headers, full_name="Employee User", id_card="123-456-789")

        invitation_response = self.client.post(
            f"/api/organizations/{organization_id}/invitations",
            headers=owner_headers,
            json={"email": "employee@example.com", "employee_id": employee_id, "role": "employee", "expires_in_days": 7},
        )
        self.assertEqual(invitation_response.status_code, 200)
        accept_response = self.client.post(
            "/api/auth/accept-invitation",
            json={
                "token": invitation_response.json()["invitation_token"],
                "password": "EmployeePass123",
                "confirm_password": "EmployeePass123",
            },
        )
        self.assertEqual(accept_response.status_code, 200)

        id_login_response = self.client.post(
            "/api/auth/login",
            json={"email": "123456789", "password": "EmployeePass123"},
        )
        self.assertEqual(id_login_response.status_code, 200)
        membership = id_login_response.json()["user"]["memberships"][0]
        self.assertEqual(membership["role"], "employee")
        self.assertEqual(membership["employee_id"], employee_id)

    def test_unlinked_employee_member_is_repaired_for_members_list_and_id_login(self):
        owner_response = self.client.post(
            "/api/auth/create-organization",
            json={
                "organization_name": "Beta Clinic",
                "full_name": "Owner User",
                "email": "owner@example.com",
                "password": "CorrectHorse123",
            },
        )
        self.assertEqual(owner_response.status_code, 200)
        owner_headers = {"Authorization": f"Bearer {owner_response.json()['access_token']}"}
        organization_id = owner_response.json()["user"]["memberships"][0]["organization_id"]
        employee_id = self._create_employee(headers=owner_headers, full_name="Employee User", id_card="123456789")

        invitation_response = self.client.post(
            f"/api/organizations/{organization_id}/invitations",
            headers=owner_headers,
            json={"email": "employee@example.com", "employee_id": employee_id, "role": "employee", "expires_in_days": 7},
        )
        self.assertEqual(invitation_response.status_code, 200)
        accept_response = self.client.post(
            "/api/auth/accept-invitation",
            json={
                "token": invitation_response.json()["invitation_token"],
                "password": "EmployeePass123",
                "confirm_password": "EmployeePass123",
            },
        )
        self.assertEqual(accept_response.status_code, 200)

        with database.get_connection() as connection:
            cursor = connection.cursor()
            cursor.execute(
                """
                UPDATE organization_memberships
                SET employee_id = NULL
                WHERE organization_id = ? AND role = 'employee'
                """,
                (organization_id,),
            )
            connection.commit()

        members_response = self.client.get(f"/api/organizations/{organization_id}/members", headers=owner_headers)
        self.assertEqual(members_response.status_code, 200)
        employee_members = [
            member for member in members_response.json()["members"]
            if member["email"] == "employee@example.com"
        ]
        self.assertEqual(employee_members[0]["employee_id"], employee_id)
        self.assertEqual(employee_members[0]["employee_name"], "Employee User")

        id_login_response = self.client.post(
            "/api/auth/login",
            json={"email": "123-456-789", "password": "EmployeePass123"},
        )
        self.assertEqual(id_login_response.status_code, 200)
        self.assertEqual(id_login_response.json()["user"]["memberships"][0]["employee_id"], employee_id)

    def test_id_card_login_repairs_legacy_unlinked_employee_member_by_unique_name(self):
        owner_response = self.client.post(
            "/api/auth/create-organization",
            json={
                "organization_name": "Beta Clinic",
                "full_name": "Owner User",
                "email": "owner@example.com",
                "password": "CorrectHorse123",
            },
        )
        self.assertEqual(owner_response.status_code, 200)
        owner_headers = {"Authorization": f"Bearer {owner_response.json()['access_token']}"}
        organization_id = owner_response.json()["user"]["memberships"][0]["organization_id"]
        employee_id = self._create_employee(headers=owner_headers, full_name="Legacy Employee", id_card="987654321")
        now = "2026-04-30T12:00:00"
        with database.get_connection() as connection:
            cursor = connection.cursor()
            cursor.execute(
                """
                INSERT INTO users (email, full_name, password_hash, status, email_verified, created_at, updated_at)
                VALUES (?, ?, ?, 'active', 0, ?, ?)
                """,
                ("legacy@example.com", "Legacy Employee", main.hash_password("EmployeePass123"), now, now),
            )
            user_id = cursor.lastrowid
            cursor.execute(
                """
                INSERT INTO organization_memberships (organization_id, user_id, role, status, employee_id, created_at, updated_at)
                VALUES (?, ?, 'employee', 'active', NULL, ?, ?)
                """,
                (organization_id, user_id, now, now),
            )
            connection.commit()

        id_login_response = self.client.post(
            "/api/auth/login",
            json={"email": "987654321", "password": "EmployeePass123"},
        )
        self.assertEqual(id_login_response.status_code, 200)
        self.assertEqual(id_login_response.json()["user"]["memberships"][0]["employee_id"], employee_id)

    def test_desktop_cloud_login_imports_cloud_organization_into_local_sqlite(self):
        cloud_user = {
            "id": 44,
            "email": "owner@example.com",
            "full_name": "Cloud Owner",
            "status": "active",
            "email_verified": True,
            "memberships": [
                {
                    "organization_id": 42,
                    "organization_public_id": "org_cloud",
                    "organization_name": "Cloud Clinic",
                    "role": "owner",
                    "status": "active",
                    "employee_id": None,
                }
            ],
        }
        cloud_bundle = {
            "format": "shiftcare.organization.v1",
            "app_version": "0.15.1_beta",
            "organization": {"id": 42, "public_id": "org_cloud", "name": "Cloud Clinic", "status": "active"},
            "records": {
                "employees": [
                    {
                        "id": 7,
                        "organization_id": 42,
                        "public_id": "emp_cloud",
                        "full_name": "Imported Employee",
                        "sex": "female",
                        "min_shifts_per_week": 1,
                        "target_shifts_per_week": 3,
                        "max_shifts_per_week": 5,
                        "can_work_night": 1,
                        "can_work_weekends": 1,
                        "can_work_evenings_after_night": 1,
                        "can_work_mornings_and_evenings": 1,
                    }
                ],
                "positions": [
                    {
                        "id": 9,
                        "organization_id": 42,
                        "public_id": "pos_cloud",
                        "name": "Caregiver",
                        "color": "#eff6ff",
                        "requires_continuous_coverage": 0,
                        "minimum_staff_presence": 0,
                    }
                ],
                "shift_templates": [],
                "shift_requirements": [],
                "coverage_requirements": [],
                "employee_preferences": [],
                "employee_week_preferences": [],
                "employee_day_statuses": [],
                "schedule_entries": [],
                "licenses": [
                    {
                        "id": 3,
                        "organization_id": 42,
                        "license_id": "lic_cloud_full_access",
                        "status": "active",
                        "plan_code": "full_access",
                        "employee_limit": 9999,
                        "support_cloud_expires_at": "2036-05-06",
                        "grace_ends_at": "2036-05-20",
                        "certificate_json": '{"license_id":"lic_cloud_full_access","organization_public_id":"org_cloud","plan_code":"full_access","employee_limit":9999,"issued_at":"2026-05-06T00:00:00Z","status":"active","signature":"cloud-admin-grant","signature_scheme":"cloud-admin-v1"}',
                        "signature": "cloud-admin-grant",
                        "key_id": "cloud-admin-grant",
                        "source": "support",
                        "imported_at": "2026-05-06T00:00:00Z",
                        "last_verified_at": "2026-05-06T00:00:00Z",
                        "revoked_at": None,
                    }
                ],
                "app_settings": [],
                "employee_positions": [],
            },
        }

        def fake_cloud_request(base_url, path, **kwargs):
            if path == "/api/auth/login":
                self.assertEqual(kwargs["method"], "POST")
                return {"access_token": "cloud-token", "user": cloud_user}
            if path == "/api/organizations/42/cloud-export":
                self.assertEqual(kwargs["token"], "cloud-token")
                return cloud_bundle
            if path == "/api/organizations/42/members":
                self.assertEqual(kwargs["token"], "cloud-token")
                return {
                    "members": [
                        {
                            "user_id": 77,
                            "email": "employee@example.com",
                            "full_name": "Imported Employee",
                            "role": "employee",
                            "membership_status": "active",
                            "employee_id": 7,
                            "employee_public_id": "emp_cloud",
                            "employee_name": "Imported Employee",
                        },
                        {
                            "user_id": 78,
                            "email": "former@example.com",
                            "full_name": "Former Employee",
                            "role": "employee",
                            "membership_status": "disabled",
                            "employee_id": None,
                            "employee_public_id": None,
                            "employee_name": None,
                        }
                    ]
                }
            if path == "/api/organizations/42/invitations":
                self.assertEqual(kwargs["token"], "cloud-token")
                return {
                    "invitations": [
                        {
                            "id": 88,
                            "email": "pending@example.com",
                            "employee_id": 7,
                            "employee_public_id": "emp_cloud",
                            "employee_name": "Imported Employee",
                            "role": "employee",
                            "status": "pending",
                        }
                    ]
                }
            raise AssertionError(path)

        with patch.object(main, "request_cloud_json", side_effect=fake_cloud_request):
            response = self.client.post(
                "/api/desktop/cloud-login",
                json={"email": "owner@example.com", "password": "CorrectHorse123"},
            )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["user"]["email"], "owner@example.com")
        self.assertEqual(payload["user"]["memberships"][0]["organization_name"], "Cloud Clinic")
        self.assertEqual(payload["desktop_sync"]["cloud_organization_id"], 42)
        self.assertEqual(payload["desktop_sync"]["imported"]["employees"], 1)

        cursor = self.connection.cursor()
        cursor.execute("SELECT full_name FROM employees WHERE public_id = 'emp_cloud'")
        self.assertEqual(cursor.fetchone()["full_name"], "Imported Employee")
        cursor.execute("SELECT value FROM app_settings WHERE key = 'cloud_organization_id'")
        self.assertEqual(cursor.fetchone()["value"], "42")
        cursor.execute("SELECT plan_code, employee_limit FROM licenses WHERE license_id = 'lic_cloud_full_access'")
        license_row = cursor.fetchone()
        self.assertIsNotNone(license_row)
        self.assertEqual(license_row["plan_code"], "full_access")
        self.assertEqual(license_row["employee_limit"], 9999)

    def test_desktop_linked_organization_reads_cloud_members_and_invitations(self):
        owner_response = self.client.post(
            "/api/auth/bootstrap",
            json={
                "organization_name": "Local Linked",
                "full_name": "Owner User",
                "email": "owner@example.com",
                "password": "CorrectHorse123",
            },
        )
        self.assertEqual(owner_response.status_code, 200)
        headers = {"Authorization": f"Bearer {owner_response.json()['access_token']}"}
        cursor = self.connection.cursor()
        for key, value in {
            "cloud_api_base_url": "https://schedule-app-beta.web.app",
            "cloud_organization_id": "42",
            "cloud_organization_public_id": "org_cloud",
            "desktop_cloud_access_token": "cloud-token",
        }.items():
            cursor.execute(
                """
                INSERT INTO app_settings (organization_id, key, value)
                VALUES (1, ?, ?)
                ON CONFLICT(key) DO UPDATE SET value = excluded.value
                """,
                (key, value),
            )
        self.connection.commit()

        def fake_cloud_request(base_url, path, **kwargs):
            self.assertEqual(base_url, "https://schedule-app-beta.web.app")
            self.assertEqual(kwargs["token"], "cloud-token")
            if path == "/api/organizations/42/members":
                return {
                    "members": [
                        {
                            "user_id": 77,
                            "email": "employee@example.com",
                            "full_name": "Imported Employee",
                            "role": "employee",
                            "membership_status": "active",
                            "employee_id": 7,
                            "employee_public_id": "emp_cloud",
                            "employee_name": "Imported Employee",
                        }
                    ]
                }
            if path == "/api/organizations/42/invitations":
                return {
                    "invitations": [
                        {
                            "id": 88,
                            "email": "pending@example.com",
                            "employee_id": 7,
                            "employee_public_id": "emp_cloud",
                            "employee_name": "Imported Employee",
                            "role": "employee",
                            "status": "pending",
                        }
                    ]
                }
            if path == "/api/organizations/42/members/77":
                self.assertEqual(kwargs["method"], "DELETE")
                return {"message": "Organization member access removed"}
            if path == "/api/organizations/42/invitations/88/regenerate-token":
                self.assertEqual(kwargs["method"], "POST")
                return {
                    "invitation": {
                        "id": 88,
                        "organization_id": 42,
                        "email": "pending@example.com",
                        "employee_id": 7,
                        "role": "employee",
                        "status": "pending",
                        "expires_at": "2026-05-07T00:00:00",
                    },
                    "invitation_token": "new-cloud-token",
                    "invitation_url": "https://schedule-app-beta.web.app/accept-invitation?token=new-cloud-token",
                }
            if path == "/api/organizations/42/invitations/88":
                self.assertEqual(kwargs["method"], "DELETE")
                return {"message": "Invitation revoked"}
            raise AssertionError(path)

        with (
            patch.object(main, "request_cloud_json", side_effect=fake_cloud_request),
            patch.object(main, "is_desktop_invitation_request", return_value=True),
        ):
            members_response = self.client.get("/api/organizations/1/members", headers=headers)
            invitations_response = self.client.get("/api/organizations/1/invitations", headers=headers)
            remove_member_response = self.client.delete("/api/organizations/1/members/77", headers=headers)
            regenerate_response = self.client.post(
                "/api/organizations/1/invitations/88/regenerate-token",
                headers=headers,
            )
            revoke_response = self.client.delete("/api/organizations/1/invitations/88", headers=headers)

        self.assertEqual(members_response.status_code, 200)
        self.assertEqual(len(members_response.json()["members"]), 1)
        self.assertEqual(members_response.json()["members"][0]["employee_public_id"], "emp_cloud")
        self.assertNotIn("former@example.com", [member["email"] for member in members_response.json()["members"]])
        self.assertEqual(invitations_response.status_code, 200)
        self.assertEqual(invitations_response.json()["invitations"][0]["employee_public_id"], "emp_cloud")
        self.assertEqual(remove_member_response.status_code, 200)
        self.assertEqual(regenerate_response.status_code, 200)
        self.assertEqual(regenerate_response.json()["invitation_token"], "new-cloud-token")
        self.assertEqual(revoke_response.status_code, 200)

    def test_cloud_export_import_round_trip_preserves_setup_data(self):
        bootstrap_response = self.client.post(
            "/api/auth/bootstrap",
            json={
                "organization_name": "Round Trip Clinic",
                "full_name": "Owner User",
                "email": "owner@example.com",
                "password": "CorrectHorse123",
            },
        )
        self.assertEqual(bootstrap_response.status_code, 200)
        token = bootstrap_response.json()["access_token"]
        headers = {"Authorization": f"Bearer {token}"}
        organization_id = bootstrap_response.json()["user"]["memberships"][0]["organization_id"]

        employee_id = self._create_employee(headers=headers, full_name="Nurse One")
        position_id = self._create_position(headers=headers, name="Ward Nurse")
        template_id = self._create_shift_template(headers=headers, position_id=position_id, name="Day")
        assignment_response = self.client.post(
            "/api/employee-positions",
            json={
                "employee_id": employee_id,
                "position_id": position_id,
                "is_primary": True,
                "priority_score": 80,
                "is_fallback_only": False,
            },
            headers=headers,
        )
        self.assertEqual(assignment_response.status_code, 200)
        schedule_response = self.client.post(
            "/api/schedule",
            json={
                "employee_id": employee_id,
                "position_id": position_id,
                "date": "2026-05-04",
                "shift_template_id": template_id,
            },
            headers=headers,
        )
        self.assertEqual(schedule_response.status_code, 200)
        cursor = self.connection.cursor()
        cursor.execute(
            """
            INSERT INTO licenses (
                organization_id, license_id, status, plan_code, employee_limit,
                support_cloud_expires_at, grace_ends_at, certificate_json, signature,
                key_id, source, imported_at, last_verified_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
            """,
            (
                organization_id,
                "lic_round_trip_full_access",
                "active",
                "full_access",
                9999,
                "2036-05-06",
                "2036-05-20",
                '{"license_id":"lic_round_trip_full_access","organization_public_id":"local-default","plan_code":"full_access","employee_limit":9999,"issued_at":"2026-05-06T00:00:00Z","status":"active","signature":"local-admin-grant","signature_scheme":"local-admin-v1"}',
                "local-admin-grant",
                "local-admin-grant",
                "support",
            ),
        )
        self.connection.commit()

        export_response = self.client.get(f"/api/organizations/{organization_id}/cloud-export", headers=headers)
        self.assertEqual(export_response.status_code, 200)
        bundle = export_response.json()
        self.assertEqual(bundle["format"], "shiftcare.organization.v1")
        self.assertEqual(len(bundle["records"]["employees"]), 1)
        self.assertEqual(len(bundle["records"]["employee_positions"]), 1)
        self.assertEqual(len(bundle["records"]["licenses"]), 1)

        import_response = self.client.post(
            f"/api/organizations/{organization_id}/cloud-import",
            json={"bundle": bundle, "replace_existing": True},
            headers=headers,
        )
        self.assertEqual(import_response.status_code, 200)
        imported = import_response.json()["imported"]
        self.assertEqual(imported["employees"], 1)
        self.assertEqual(imported["positions"], 1)
        self.assertEqual(imported["shift_templates"], 1)
        self.assertEqual(imported["schedule_entries"], 1)
        self.assertEqual(imported["licenses"], 1)

        employees_response = self.client.get("/api/employees", headers=headers)
        self.assertEqual(employees_response.status_code, 200)
        self.assertEqual([employee["full_name"] for employee in employees_response.json()], ["Nurse One"])

        assignments_response = self.client.get("/api/employee-positions", headers=headers)
        self.assertEqual(assignments_response.status_code, 200)
        self.assertEqual(len(assignments_response.json()), 1)
        cursor.execute("SELECT plan_code, employee_limit FROM licenses WHERE license_id = 'lic_round_trip_full_access'")
        license_row = cursor.fetchone()
        self.assertIsNotNone(license_row)
        self.assertEqual(license_row["plan_code"], "full_access")
        self.assertEqual(license_row["employee_limit"], 9999)

        empty_link_response = self.client.get(f"/api/organizations/{organization_id}/cloud-link", headers=headers)
        self.assertEqual(empty_link_response.status_code, 200)
        self.assertFalse(empty_link_response.json()["linked"])

        save_link_response = self.client.post(
            f"/api/organizations/{organization_id}/cloud-link",
            json={
                "cloud_api_base_url": "https://portal.shiftcare.co.il/",
                "cloud_organization_id": 42,
                "cloud_organization_public_id": "org_cloud_public",
                "linked_at": "2026-04-29T03:00:00.000Z",
            },
            headers=headers,
        )
        self.assertEqual(save_link_response.status_code, 200)

        linked_response = self.client.get(f"/api/organizations/{organization_id}/cloud-link", headers=headers)
        self.assertEqual(linked_response.status_code, 200)
        linked_payload = linked_response.json()
        self.assertTrue(linked_payload["linked"])
        self.assertEqual(linked_payload["cloud_api_base_url"], "https://portal.shiftcare.co.il")
        self.assertEqual(linked_payload["cloud_organization_id"], 42)
        self.assertEqual(linked_payload["cloud_organization_public_id"], "org_cloud_public")

        unlink_response = self.client.delete(f"/api/organizations/{organization_id}/cloud-link", headers=headers)
        self.assertEqual(unlink_response.status_code, 200)
        unlinked_response = self.client.get(f"/api/organizations/{organization_id}/cloud-link", headers=headers)
        self.assertEqual(unlinked_response.status_code, 200)
        self.assertFalse(unlinked_response.json()["linked"])

    def test_new_organization_owned_records_receive_public_ids(self):
        employee_id = self._create_employee()
        position_id = self._create_position()
        template_id = self._create_shift_template(position_id=position_id)

        self.client.post(
            "/api/employee-positions",
            json={
                "employee_id": employee_id,
                "position_id": position_id,
                "is_primary": True,
                "priority_score": 100,
                "is_fallback_only": False,
            },
        )
        schedule_response = self.client.post(
            "/api/schedule",
            json={
                "employee_id": employee_id,
                "position_id": position_id,
                "date": "2026-04-20",
                "shift_template_id": template_id,
            },
        )
        self.assertEqual(schedule_response.status_code, 200)
        recurring_response = self.client.post(
            "/api/employee-recurring-preferences",
            json={
                "employee_id": employee_id,
                "rules": [
                    {
                        "preference_kind": "strict",
                        "day_of_week": 1,
                        "preference_type": "only_morning",
                    }
                ],
            },
        )
        self.assertEqual(recurring_response.status_code, 200)

        cursor = self.connection.cursor()
        cursor.execute(
            """
            INSERT INTO employee_week_preference_requests (
                employee_id, week_start_date, preference_date, preference_type,
                request_type, target_category, status
            )
            VALUES (?, '2026-04-20', '2026-04-22', 'only_morning', 'request_shift', 'morning', 'pending')
            """,
            (employee_id,),
        )
        self.connection.commit()
        for table_name, prefix in database.PUBLIC_ID_TABLE_PREFIXES.items():
            cursor.execute(f"SELECT public_id FROM {table_name} WHERE public_id IS NOT NULL LIMIT 1")
            public_id_row = cursor.fetchone()
            if table_name in {"shift_requirements", "employee_preferences", "employee_week_preferences", "employee_day_statuses", "coverage_requirements"}:
                continue
            self.assertIsNotNone(public_id_row, table_name)
            self.assertRegex(public_id_row["public_id"], rf"^{prefix}_[0-9a-f]{{32}}$")

    def test_auth_login_rejects_invalid_password(self):
        self.client.post(
            "/api/auth/bootstrap",
            json={
                "organization_name": "Beta Clinic",
                "full_name": "Owner User",
                "email": "owner@example.com",
                "password": "CorrectHorse123",
            },
        )
        response = self.client.post(
            "/api/auth/login",
            json={"email": "owner@example.com", "password": "wrong-password"},
        )
        self.assertEqual(response.status_code, 401)

    def test_auth_login_rate_limits_repeated_failures(self):
        self.client.post(
            "/api/auth/bootstrap",
            json={
                "organization_name": "Beta Clinic",
                "full_name": "Owner User",
                "email": "owner@example.com",
                "password": "CorrectHorse123",
            },
        )

        with patch.object(main, "AUTH_LOGIN_RATE_LIMIT_ATTEMPTS", 3):
            first_response = self.client.post(
                "/api/auth/login",
                json={"email": "owner@example.com", "password": "wrong-password"},
            )
            second_response = self.client.post(
                "/api/auth/login",
                json={"email": "owner@example.com", "password": "wrong-password"},
            )
            limited_response = self.client.post(
                "/api/auth/login",
                json={"email": "owner@example.com", "password": "wrong-password"},
            )
            correct_password_response = self.client.post(
                "/api/auth/login",
                json={"email": "owner@example.com", "password": "CorrectHorse123"},
            )

        self.assertEqual(first_response.status_code, 401)
        self.assertEqual(second_response.status_code, 401)
        self.assertEqual(limited_response.status_code, 429)
        self.assertEqual(correct_password_response.status_code, 429)

        cursor = self.connection.cursor()
        cursor.execute(
            """
            SELECT event_type
            FROM auth_audit_events
            WHERE event_type IN ('login_failed', 'login_rate_limited')
            ORDER BY id
            """
        )
        event_types = [row["event_type"] for row in cursor.fetchall()]
        self.assertEqual(event_types.count("login_failed"), 3)
        self.assertGreaterEqual(event_types.count("login_rate_limited"), 1)

    def test_auth_profile_update_and_password_change(self):
        bootstrap_response = self.client.post(
            "/api/auth/bootstrap",
            json={
                "organization_name": "Beta Clinic",
                "full_name": "Owner User",
                "email": "owner@example.com",
                "password": "CorrectHorse123",
            },
        )
        self.assertEqual(bootstrap_response.status_code, 200)
        original_token = bootstrap_response.json()["access_token"]
        headers = {"Authorization": f"Bearer {original_token}"}

        profile_response = self.client.put(
            "/api/auth/profile",
            headers=headers,
            json={"full_name": "Updated Owner"},
        )
        self.assertEqual(profile_response.status_code, 200)
        self.assertEqual(profile_response.json()["user"]["full_name"], "Updated Owner")

        wrong_password_response = self.client.post(
            "/api/auth/change-password",
            headers=headers,
            json={"current_password": "wrong-password", "new_password": "NewPassword123"},
        )
        self.assertEqual(wrong_password_response.status_code, 401)

        second_login_response = self.client.post(
            "/api/auth/login",
            json={"email": "owner@example.com", "password": "CorrectHorse123"},
        )
        self.assertEqual(second_login_response.status_code, 200)
        second_token = second_login_response.json()["access_token"]

        change_response = self.client.post(
            "/api/auth/change-password",
            headers=headers,
            json={"current_password": "CorrectHorse123", "new_password": "NewPassword123"},
        )
        self.assertEqual(change_response.status_code, 200)

        old_login_response = self.client.post(
            "/api/auth/login",
            json={"email": "owner@example.com", "password": "CorrectHorse123"},
        )
        self.assertEqual(old_login_response.status_code, 401)

        new_login_response = self.client.post(
            "/api/auth/login",
            json={"email": "owner@example.com", "password": "NewPassword123"},
        )
        self.assertEqual(new_login_response.status_code, 200)

        kept_session_response = self.client.get("/api/auth/me", headers=headers)
        self.assertEqual(kept_session_response.status_code, 200)

        revoked_session_response = self.client.get(
            "/api/auth/me",
            headers={"Authorization": f"Bearer {second_token}"},
        )
        self.assertEqual(revoked_session_response.status_code, 401)

    def test_password_reset_and_email_verification_flows(self):
        bootstrap_response = self.client.post(
            "/api/auth/bootstrap",
            json={
                "organization_name": "Beta Clinic",
                "full_name": "Owner User",
                "email": "owner@example.com",
                "password": "CorrectHorse123",
            },
        )
        self.assertEqual(bootstrap_response.status_code, 200)
        token = bootstrap_response.json()["access_token"]

        reset_request = self.client.post(
            "/api/auth/request-password-reset",
            json={"email": "owner@example.com"},
        )
        self.assertEqual(reset_request.status_code, 200)
        self.assertEqual(reset_request.json()["email_status"]["status"], "disabled")
        reset_token = reset_request.json()["reset_token"]
        self.assertTrue(reset_token)

        reset_response = self.client.post(
            "/api/auth/reset-password",
            json={"token": reset_token, "new_password": "NewPassword123"},
        )
        self.assertEqual(reset_response.status_code, 200)

        old_login = self.client.post(
            "/api/auth/login",
            json={"email": "owner@example.com", "password": "CorrectHorse123"},
        )
        self.assertEqual(old_login.status_code, 401)
        new_login = self.client.post(
            "/api/auth/login",
            json={"email": "owner@example.com", "password": "NewPassword123"},
        )
        self.assertEqual(new_login.status_code, 200)

        verification_request = self.client.post(
            "/api/auth/request-email-verification",
            headers={"Authorization": f"Bearer {new_login.json()['access_token']}"},
        )
        self.assertEqual(verification_request.status_code, 200)
        self.assertEqual(verification_request.json()["email_status"]["status"], "disabled")
        verification_token = verification_request.json()["verification_token"]

        verify_response = self.client.post(
            "/api/auth/verify-email",
            json={"token": verification_token},
        )
        self.assertEqual(verify_response.status_code, 200)

        cursor = self.connection.cursor()
        cursor.execute(
            """
            SELECT event_type
            FROM auth_audit_events
            WHERE event_type IN (
                'password_reset_requested',
                'password_reset_completed',
                'email_verification_requested',
                'email_verified'
            )
            ORDER BY id
            """
        )
        self.assertEqual(
            [row["event_type"] for row in cursor.fetchall()],
            [
                "password_reset_requested",
                "password_reset_completed",
                "email_verification_requested",
                "email_verified",
            ],
        )

    def test_email_enabled_sends_invitation_and_hides_auth_debug_tokens(self):
        owner_response = self.client.post(
            "/api/auth/bootstrap",
            json={
                "organization_name": "Beta Clinic",
                "full_name": "Owner User",
                "email": "owner@example.com",
                "password": "CorrectHorse123",
            },
        )
        self.assertEqual(owner_response.status_code, 200)
        owner_token = owner_response.json()["access_token"]
        owner_headers = {"Authorization": f"Bearer {owner_token}"}
        employee_record_id = self._create_employee(headers=owner_headers, full_name="Employee User")

        with (
            patch.dict(os.environ, {"PUBLIC_APP_BASE_URL": "https://shiftcare.example.com"}, clear=False),
            patch.object(main, "send_invitation_email", return_value=email_service.EmailSendResult("sent")) as send_invitation,
        ):
            invitation_response = self.client.post(
                "/api/organizations/1/invitations",
                headers=owner_headers,
                json={
                    "email": "employee@example.com",
                    "employee_id": employee_record_id,
                    "role": "employee",
                    "expires_in_days": 7,
                },
            )

        self.assertEqual(invitation_response.status_code, 200)
        payload = invitation_response.json()
        self.assertEqual(payload["email_status"]["status"], "sent")
        send_invitation.assert_called_once()
        self.assertEqual(send_invitation.call_args.kwargs["to_email"], "employee@example.com")
        self.assertEqual(send_invitation.call_args.kwargs["organization_name"], "Beta Clinic")
        self.assertIn(payload["invitation_token"], send_invitation.call_args.kwargs["invitation_url"])

        with (
            patch.object(main, "email_delivery_is_enabled", return_value=True),
            patch.object(main, "send_password_reset_email", return_value=email_service.EmailSendResult("sent")) as send_reset,
        ):
            reset_request = self.client.post(
                "/api/auth/request-password-reset",
                json={"email": "owner@example.com"},
            )
        self.assertEqual(reset_request.status_code, 200)
        self.assertIsNone(reset_request.json()["reset_token"])
        self.assertEqual(reset_request.json()["email_status"]["status"], "sent")
        send_reset.assert_called_once()

        login_response = self.client.post(
            "/api/auth/login",
            json={"email": "owner@example.com", "password": "CorrectHorse123"},
        )
        self.assertEqual(login_response.status_code, 200)
        with (
            patch.object(main, "email_delivery_is_enabled", return_value=True),
            patch.object(main, "send_email_verification_email", return_value=email_service.EmailSendResult("sent")) as send_verify,
        ):
            verification_request = self.client.post(
                "/api/auth/request-email-verification",
                headers={"Authorization": f"Bearer {login_response.json()['access_token']}"},
            )
        self.assertEqual(verification_request.status_code, 200)
        self.assertIsNone(verification_request.json()["verification_token"])
        self.assertEqual(verification_request.json()["email_status"]["status"], "sent")
        send_verify.assert_called_once()

    def test_invitation_flow_creates_employee_membership_and_enforces_roles(self):
        owner_response = self.client.post(
            "/api/auth/bootstrap",
            json={
                "organization_name": "Beta Clinic",
                "full_name": "Owner User",
                "email": "owner@example.com",
                "password": "CorrectHorse123",
            },
        )
        self.assertEqual(owner_response.status_code, 200)
        owner_user_id = owner_response.json()["user"]["id"]
        owner_token = owner_response.json()["access_token"]

        members_response = self.client.get(
            "/api/organizations/1/members",
            headers={"Authorization": f"Bearer {owner_token}"},
        )
        self.assertEqual(members_response.status_code, 200)
        self.assertEqual(len(members_response.json()["members"]), 1)

        owner_headers = {"Authorization": f"Bearer {owner_token}"}
        employee_record_id = self._create_employee(headers=owner_headers, full_name="Employee User")
        with patch.dict(os.environ, {"PUBLIC_APP_BASE_URL": "https://shiftcare.example.com"}, clear=False):
            invitation_response = self.client.post(
                "/api/organizations/1/invitations",
                headers={"Authorization": f"Bearer {owner_token}"},
                json={
                    "email": "employee@example.com",
                    "employee_id": employee_record_id,
                    "role": "employee",
                    "expires_in_days": 7,
                },
            )
        self.assertEqual(invitation_response.status_code, 200)
        self.assertEqual(invitation_response.json()["invitation"]["employee_id"], employee_record_id)
        self.assertEqual(invitation_response.json()["email_status"]["status"], "disabled")
        invitation_token = invitation_response.json()["invitation_token"]
        self.assertEqual(
            invitation_response.json()["invitation_url"],
            f"https://shiftcare.example.com/accept-invitation?token={invitation_token}",
        )

        preview_response = self.client.get(f"/api/auth/invitation-preview?token={invitation_token}")
        self.assertEqual(preview_response.status_code, 200)
        self.assertEqual(preview_response.json()["employee_name"], "Employee User")
        self.assertFalse(preview_response.json()["requires_name"])

        accept_response = self.client.post(
            "/api/auth/accept-invitation",
            json={
                "token": invitation_token,
                "password": "EmployeePass123",
                "confirm_password": "EmployeePass123",
            },
        )
        self.assertEqual(accept_response.status_code, 200)
        employee_token = accept_response.json()["access_token"]
        employee_user_id = accept_response.json()["user"]["id"]
        employee_memberships = accept_response.json()["user"]["memberships"]
        self.assertEqual(employee_memberships[0]["role"], "employee")
        self.assertEqual(employee_memberships[0]["organization_id"], 1)
        self.assertEqual(employee_memberships[0]["employee_id"], employee_record_id)

        forbidden_members_response = self.client.get(
            "/api/organizations/1/members",
            headers={"Authorization": f"Bearer {employee_token}"},
        )
        self.assertEqual(forbidden_members_response.status_code, 403)

        forbidden_invitation_response = self.client.post(
            "/api/organizations/1/invitations",
            headers={"Authorization": f"Bearer {employee_token}"},
            json={"email": "second@example.com", "role": "employee"},
        )
        self.assertEqual(forbidden_invitation_response.status_code, 403)

        invitations_response = self.client.get(
            "/api/organizations/1/invitations",
            headers={"Authorization": f"Bearer {owner_token}"},
        )
        self.assertEqual(invitations_response.status_code, 200)
        self.assertEqual(invitations_response.json()["invitations"][0]["status"], "accepted")
        self.assertEqual(invitations_response.json()["invitations"][0]["employee_name"], "Employee User")

        remove_self_response = self.client.delete(
            f"/api/organizations/1/members/{owner_user_id}",
            headers={"Authorization": f"Bearer {owner_token}"},
        )
        self.assertEqual(remove_self_response.status_code, 400)

        remove_employee_response = self.client.delete(
            f"/api/organizations/1/members/{employee_user_id}",
            headers={"Authorization": f"Bearer {owner_token}"},
        )
        self.assertEqual(remove_employee_response.status_code, 200)

        revoked_session_response = self.client.get(
            "/api/auth/me",
            headers={"Authorization": f"Bearer {employee_token}"},
        )
        self.assertEqual(revoked_session_response.status_code, 401)

        members_after_remove_response = self.client.get(
            "/api/organizations/1/members",
            headers={"Authorization": f"Bearer {owner_token}"},
        )
        self.assertEqual(members_after_remove_response.status_code, 200)
        self.assertNotIn(
            "employee@example.com",
            [
                member["email"]
                for member in members_after_remove_response.json()["members"]
            ],
        )
        connection = database.get_connection()
        try:
            cursor = connection.cursor()
            cursor.execute(
                """
                SELECT COUNT(*) AS account_count
                FROM users
                WHERE lower(email) = 'employee@example.com'
                """,
            )
            self.assertEqual(cursor.fetchone()["account_count"], 0)
        finally:
            connection.close()

        reinvite_employee_response = self.client.post(
            "/api/organizations/1/invitations",
            headers={"Authorization": f"Bearer {owner_token}"},
            json={
                "email": "employee@example.com",
                "employee_id": employee_record_id,
                "role": "employee",
                "expires_in_days": 7,
            },
        )
        self.assertEqual(reinvite_employee_response.status_code, 200)
        reaccept_response = self.client.post(
            "/api/auth/accept-invitation",
            json={
                "token": reinvite_employee_response.json()["invitation_token"],
                "password": "EmployeePass123",
                "confirm_password": "EmployeePass123",
            },
        )
        self.assertEqual(reaccept_response.status_code, 200)
        self.assertNotEqual(reaccept_response.json()["user"]["id"], employee_user_id)
        reaccepted_membership = reaccept_response.json()["user"]["memberships"][0]
        self.assertEqual(reaccepted_membership["status"], "active")
        self.assertEqual(reaccepted_membership["employee_id"], employee_record_id)

        pending_invitation_response = self.client.post(
            "/api/organizations/1/invitations",
            headers={"Authorization": f"Bearer {owner_token}"},
            json={"email": "pending@example.com", "role": "read_only", "expires_in_days": 7},
        )
        self.assertEqual(pending_invitation_response.status_code, 200)
        pending_invitation_id = pending_invitation_response.json()["invitation"]["id"]
        original_pending_token = pending_invitation_response.json()["invitation_token"]

        regenerated_response = self.client.post(
            f"/api/organizations/1/invitations/{pending_invitation_id}/regenerate-token",
            headers={"Authorization": f"Bearer {owner_token}"},
        )
        self.assertEqual(regenerated_response.status_code, 200)
        regenerated_token = regenerated_response.json()["invitation_token"]
        self.assertNotEqual(regenerated_token, original_pending_token)
        old_token_accept_response = self.client.post(
            "/api/auth/accept-invitation",
            json={
                "token": original_pending_token,
                "full_name": "Pending User",
                "password": "PendingPass123",
            },
        )
        self.assertEqual(old_token_accept_response.status_code, 404)

        revoke_invitation_response = self.client.delete(
            f"/api/organizations/1/invitations/{pending_invitation_id}",
            headers={"Authorization": f"Bearer {owner_token}"},
        )
        self.assertEqual(revoke_invitation_response.status_code, 200)

        invitations_after_revoke_response = self.client.get(
            "/api/organizations/1/invitations",
            headers={"Authorization": f"Bearer {owner_token}"},
        )
        self.assertEqual(invitations_after_revoke_response.status_code, 200)
        revoked_invitation = next(
            invitation
            for invitation in invitations_after_revoke_response.json()["invitations"]
            if invitation["id"] == pending_invitation_id
        )
        self.assertEqual(revoked_invitation["status"], "revoked")
        regenerate_revoked_response = self.client.post(
            f"/api/organizations/1/invitations/{pending_invitation_id}/regenerate-token",
            headers={"Authorization": f"Bearer {owner_token}"},
        )
        self.assertEqual(regenerate_revoked_response.status_code, 409)

    def test_admin_department_access_limits_schedule_and_directories(self):
        owner_response = self.client.post(
            "/api/auth/bootstrap",
            json={
                "organization_name": "Beta Clinic",
                "full_name": "Owner User",
                "email": "owner@example.com",
                "password": "CorrectHorse123",
            },
        )
        self.assertEqual(owner_response.status_code, 200)
        owner_headers = {"Authorization": f"Bearer {owner_response.json()['access_token']}"}

        care_department = self.client.post(
            "/api/departments",
            headers=owner_headers,
            json={"name": "Care teams", "description": "", "display_order": 1, "is_active": True},
        )
        self.assertEqual(care_department.status_code, 200)
        cleaning_department = self.client.post(
            "/api/departments",
            headers=owner_headers,
            json={"name": "Cleaning", "description": "", "display_order": 2, "is_active": True},
        )
        self.assertEqual(cleaning_department.status_code, 200)
        care_department_id = care_department.json()["department"]["id"]
        cleaning_department_id = cleaning_department.json()["department"]["id"]

        care_position_id = self._create_position(headers=owner_headers, department_id=care_department_id, name="Nurse")
        cleaning_position_id = self._create_position(headers=owner_headers, department_id=cleaning_department_id, name="Cleaner")
        employee_id = self._create_employee(headers=owner_headers, full_name="Shared Employee")
        care_template_id = self._create_shift_template(headers=owner_headers, position_id=care_position_id, name="Care Morning")
        cleaning_template_id = self._create_shift_template(headers=owner_headers, position_id=cleaning_position_id, name="Clean Morning")

        cursor = self.connection.cursor()
        now = main.current_utc_timestamp()
        cursor.execute(
            """
            INSERT INTO users (email, full_name, password_hash, status, email_verified, created_at, updated_at)
            VALUES ('scheduler@example.com', 'Senior Nurse', ?, 'active', 1, ?, ?)
            """,
            (main.hash_password("SchedulerPass123"), now, now),
        )
        scheduler_user_id = cursor.lastrowid
        cursor.execute(
            """
            INSERT INTO organization_memberships (organization_id, user_id, role, status, created_at, updated_at)
            VALUES (1, ?, 'scheduler', 'active', ?, ?)
            """,
            (scheduler_user_id, now, now),
        )
        scheduler_session = main.build_auth_response(self.connection, scheduler_user_id)
        self.connection.commit()
        scheduler_headers = {"Authorization": f"Bearer {scheduler_session['access_token']}"}

        access_response = self.client.put(
            f"/api/organizations/1/members/{scheduler_user_id}/department-access",
            headers=owner_headers,
            json={"department_ids": [care_department_id]},
        )
        self.assertEqual(access_response.status_code, 200)
        self.assertEqual([item["id"] for item in access_response.json()["department_access"]], [care_department_id])

        departments_response = self.client.get("/api/departments", headers=scheduler_headers)
        self.assertEqual(departments_response.status_code, 200)
        self.assertEqual([item["id"] for item in departments_response.json()], [care_department_id])

        positions_response = self.client.get("/api/positions", headers=scheduler_headers)
        self.assertEqual(positions_response.status_code, 200)
        self.assertEqual([item["id"] for item in positions_response.json()], [care_position_id])

        forbidden_schedule_response = self.client.post(
            "/api/schedule",
            headers=scheduler_headers,
            json={
                "employee_id": employee_id,
                "position_id": cleaning_position_id,
                "date": "2026-06-23",
                "shift_template_id": cleaning_template_id,
            },
        )
        self.assertEqual(forbidden_schedule_response.status_code, 403)

        care_schedule_response = self.client.post(
            "/api/schedule",
            headers=scheduler_headers,
            json={
                "employee_id": employee_id,
                "position_id": care_position_id,
                "date": "2026-06-23",
                "shift_template_id": care_template_id,
            },
        )
        self.assertEqual(care_schedule_response.status_code, 200)

        owner_cleaning_schedule_response = self.client.post(
            "/api/schedule",
            headers=owner_headers,
            json={
                "employee_id": employee_id,
                "position_id": cleaning_position_id,
                "date": "2026-06-23",
                "shift_template_id": cleaning_template_id,
            },
        )
        self.assertEqual(owner_cleaning_schedule_response.status_code, 200)

        schedule_response = self.client.get("/api/schedule", headers=scheduler_headers)
        self.assertEqual(schedule_response.status_code, 200)
        self.assertEqual({entry["position_id"] for entry in schedule_response.json()}, {care_position_id})

        members_response = self.client.get("/api/organizations/1/members", headers=owner_headers)
        self.assertEqual(members_response.status_code, 200)
        scheduler_member = next(member for member in members_response.json()["members"] if member["user_id"] == scheduler_user_id)
        self.assertEqual([department["id"] for department in scheduler_member["department_access"]], [care_department_id])

    def test_reinvite_purges_legacy_disabled_account_email(self):
        owner_response = self.client.post(
            "/api/auth/bootstrap",
            json={
                "organization_name": "Beta Clinic",
                "full_name": "Owner User",
                "email": "owner@example.com",
                "password": "CorrectHorse123",
            },
        )
        self.assertEqual(owner_response.status_code, 200)
        owner_user_id = owner_response.json()["user"]["id"]
        owner_headers = {"Authorization": f"Bearer {owner_response.json()['access_token']}"}
        employee_record_id = self._create_employee(headers=owner_headers, full_name="Rehired Employee")

        cursor = self.connection.cursor()
        now = main.current_utc_timestamp()
        cursor.execute(
            """
            INSERT INTO users (email, full_name, password_hash, status, email_verified, created_at, updated_at)
            VALUES ('rehire@example.com', 'Former Employee', ?, 'active', 0, ?, ?)
            """,
            (main.hash_password("OldEmployeePass123"), now, now),
        )
        stale_user_id = cursor.lastrowid
        cursor.execute(
            """
            INSERT INTO organization_memberships (organization_id, user_id, role, status, employee_id, created_at, updated_at)
            VALUES (1, ?, 'employee', 'disabled', NULL, ?, ?)
            """,
            (stale_user_id, now, now),
        )
        cursor.execute(
            """
            INSERT INTO organization_invitations (
                organization_id, email, employee_id, role, token_hash, status,
                expires_at, accepted_at, created_by_user_id, created_at
            )
            VALUES (1, 'rehire@example.com', ?, 'employee', ?, 'accepted', ?, ?, ?, ?)
            """,
            (
                employee_record_id,
                main.hash_session_token("legacy-rehire-token"),
                now,
                now,
                owner_user_id,
                now,
            ),
        )
        self.connection.commit()

        invitation_response = self.client.post(
            "/api/organizations/1/invitations",
            headers=owner_headers,
            json={
                "email": "rehire@example.com",
                "employee_id": employee_record_id,
                "role": "employee",
                "expires_in_days": 7,
            },
        )
        self.assertEqual(invitation_response.status_code, 200)
        cursor.execute("SELECT COUNT(*) AS account_count FROM users WHERE id = ?", (stale_user_id,))
        self.assertEqual(cursor.fetchone()["account_count"], 0)
        cursor.execute(
            """
            SELECT COUNT(*) AS historical_invitation_count
            FROM organization_invitations
            WHERE lower(email) = 'rehire@example.com'
              AND status != 'pending'
            """
        )
        self.assertEqual(cursor.fetchone()["historical_invitation_count"], 0)

        accept_response = self.client.post(
            "/api/auth/accept-invitation",
            json={
                "token": invitation_response.json()["invitation_token"],
                "password": "NewEmployeePass123",
                "confirm_password": "NewEmployeePass123",
            },
        )
        self.assertEqual(accept_response.status_code, 200)
        self.assertNotEqual(accept_response.json()["user"]["id"], stale_user_id)
        membership = accept_response.json()["user"]["memberships"][0]
        self.assertEqual(membership["status"], "active")
        self.assertEqual(membership["employee_id"], employee_record_id)

    def test_recurring_preferences_require_admin_or_owner_after_auth_bootstrap(self):
        owner_response = self.client.post(
            "/api/auth/bootstrap",
            json={
                "organization_name": "Beta Clinic",
                "full_name": "Owner User",
                "email": "owner@example.com",
                "password": "CorrectHorse123",
            },
        )
        self.assertEqual(owner_response.status_code, 200)
        owner_headers = {"Authorization": f"Bearer {owner_response.json()['access_token']}"}
        employee_id = self._create_employee(headers=owner_headers, full_name="Employee User")

        invitation_response = self.client.post(
            "/api/organizations/1/invitations",
            headers=owner_headers,
            json={
                "email": "employee@example.com",
                "employee_id": employee_id,
                "role": "employee",
                "expires_in_days": 7,
            },
        )
        self.assertEqual(invitation_response.status_code, 200)
        accept_response = self.client.post(
            "/api/auth/accept-invitation",
            json={
                "token": invitation_response.json()["invitation_token"],
                "password": "EmployeePass123",
                "confirm_password": "EmployeePass123",
            },
        )
        self.assertEqual(accept_response.status_code, 200)
        employee_headers = {"Authorization": f"Bearer {accept_response.json()['access_token']}"}

        payload = {
            "employee_id": employee_id,
            "rules": [
                {
                    "preference_kind": "strict",
                    "day_of_week": 0,
                    "request_type": "request_shift",
                    "target_category": "night",
                },
                {
                    "preference_kind": "strict",
                    "day_of_week": 0,
                    "request_type": "exclude_shift",
                    "target_category": "morning",
                },
                {
                    "preference_kind": "strict",
                    "day_of_week": 0,
                    "request_type": "exclude_shift",
                    "target_category": "evening",
                },
                {
                    "preference_kind": "soft",
                    "day_of_week": 1,
                    "request_type": "day_off",
                },
            ],
        }
        forbidden_save = self.client.post("/api/employee-recurring-preferences", headers=employee_headers, json=payload)
        self.assertEqual(forbidden_save.status_code, 403)
        forbidden_read = self.client.get(
            "/api/employee-recurring-preferences",
            headers=employee_headers,
            params={"employee_id": employee_id},
        )
        self.assertEqual(forbidden_read.status_code, 403)

        owner_save = self.client.post("/api/employee-recurring-preferences", headers=owner_headers, json=payload)
        self.assertEqual(owner_save.status_code, 200)
        owner_read = self.client.get(
            "/api/employee-recurring-preferences",
            headers=owner_headers,
            params={"employee_id": employee_id},
        )
        self.assertEqual(owner_read.status_code, 200)
        self.assertEqual(
            {
                (rule["preference_kind"], rule["day_of_week"], rule["request_type"], rule["target_category"])
                for rule in owner_read.json()
            },
            {
                ("strict", 0, "request_shift", "night"),
                ("strict", 0, "exclude_shift", "morning"),
                ("strict", 0, "exclude_shift", "evening"),
                ("soft", 1, "day_off", None),
            },
        )

    def test_employee_schedule_scope_returns_primary_position_team_entries(self):
        owner_response = self.client.post(
            "/api/auth/create-organization",
            json={
                "organization_name": "Beta Clinic",
                "full_name": "Owner User",
                "email": "owner@example.com",
                "password": "CorrectHorse123",
            },
        )
        self.assertEqual(owner_response.status_code, 200)
        owner_headers = {"Authorization": f"Bearer {owner_response.json()['access_token']}"}
        organization_id = owner_response.json()["user"]["memberships"][0]["organization_id"]

        employee_a = self._create_employee(headers=owner_headers, full_name="Employee A", id_card="111111111")
        employee_b = self._create_employee(headers=owner_headers, full_name="Employee B", id_card="222222222")
        employee_c = self._create_employee(headers=owner_headers, full_name="Employee C", id_card="333333333")
        nurse_position = self._create_position(headers=owner_headers, name="Nurse")
        caregiver_position = self._create_position(headers=owner_headers, name="Caregiver")
        other_position = self._create_position(headers=owner_headers, name="Admin")
        nurse_template = self._create_shift_template(headers=owner_headers, position_id=nurse_position, name="Nurse Morning")
        caregiver_template = self._create_shift_template(headers=owner_headers, position_id=caregiver_position, name="Caregiver Morning")
        other_template = self._create_shift_template(headers=owner_headers, position_id=other_position, name="Admin Morning")

        for payload in (
            {"employee_id": employee_a, "position_id": caregiver_position, "is_primary": False, "priority_score": 30},
            {"employee_id": employee_a, "position_id": nurse_position, "is_primary": True, "priority_score": 90},
            {"employee_id": employee_b, "position_id": other_position, "is_primary": True, "priority_score": 90},
            {"employee_id": employee_c, "position_id": nurse_position, "is_primary": True, "priority_score": 90},
        ):
            response = self.client.post("/api/employee-positions", headers=owner_headers, json=payload)
            self.assertEqual(response.status_code, 200)

        for payload in (
            {"employee_id": employee_a, "position_id": nurse_position, "date": "2026-04-20", "shift_template_id": nurse_template},
            {"employee_id": employee_a, "position_id": caregiver_position, "date": "2026-04-21", "shift_template_id": caregiver_template},
            {"employee_id": employee_b, "position_id": other_position, "date": "2026-04-20", "shift_template_id": other_template},
            {"employee_id": employee_c, "position_id": nurse_position, "date": "2026-04-22", "shift_template_id": nurse_template},
        ):
            response = self.client.post("/api/schedule", headers=owner_headers, json=payload)
            self.assertEqual(response.status_code, 200)

        invitation_response = self.client.post(
            f"/api/organizations/{organization_id}/invitations",
            headers=owner_headers,
            json={"email": "employee-a@example.com", "employee_id": employee_a, "role": "employee", "expires_in_days": 7},
        )
        self.assertEqual(invitation_response.status_code, 200)
        accept_response = self.client.post(
            "/api/auth/accept-invitation",
            json={"token": invitation_response.json()["invitation_token"], "password": "EmployeePass123"},
        )
        self.assertEqual(accept_response.status_code, 200)
        employee_headers = {"Authorization": f"Bearer {accept_response.json()['access_token']}"}

        positions_response = self.client.get("/api/positions", headers=employee_headers)
        self.assertEqual(positions_response.status_code, 200)
        positions = positions_response.json()
        self.assertEqual([position["id"] for position in positions], [nurse_position, caregiver_position])
        self.assertTrue(positions[0]["is_primary"])

        assignments_response = self.client.get("/api/employee-positions", headers=employee_headers)
        self.assertEqual(assignments_response.status_code, 200)
        self.assertEqual({item["employee_id"] for item in assignments_response.json()}, {employee_a})

        employees_for_position_response = self.client.get(
            "/api/employees",
            headers=employee_headers,
            params={"position_id": nurse_position},
        )
        self.assertEqual(employees_for_position_response.status_code, 200)
        self.assertEqual({item["id"] for item in employees_for_position_response.json()}, {employee_a, employee_c})

        assignments_for_position_response = self.client.get(
            "/api/employee-positions",
            headers=employee_headers,
            params={"position_id": nurse_position},
        )
        self.assertEqual(assignments_for_position_response.status_code, 200)
        self.assertEqual({item["employee_id"] for item in assignments_for_position_response.json()}, {employee_a, employee_c})

        schedule_response = self.client.get(
            "/api/schedule",
            headers=employee_headers,
            params={"position_id": nurse_position},
        )
        self.assertEqual(schedule_response.status_code, 200)
        schedule_entries = schedule_response.json()
        self.assertEqual({entry["employee_id"] for entry in schedule_entries}, {employee_a, employee_c})
        self.assertEqual({entry["position_id"] for entry in schedule_entries}, {nurse_position})

        unauthorized_position_response = self.client.get(
            "/api/schedule",
            headers=employee_headers,
            params={"position_id": other_position},
        )
        self.assertEqual(unauthorized_position_response.status_code, 403)

        own_schedule_response = self.client.get("/api/schedule", headers=employee_headers)
        self.assertEqual(own_schedule_response.status_code, 200)
        self.assertEqual({entry["employee_id"] for entry in own_schedule_response.json()}, {employee_a})

    def test_cloud_import_preserves_employee_portal_schedule_links(self):
        owner_response = self.client.post(
            "/api/auth/create-organization",
            json={
                "organization_name": "Beta Clinic",
                "full_name": "Owner User",
                "email": "owner@example.com",
                "password": "CorrectHorse123",
            },
        )
        self.assertEqual(owner_response.status_code, 200)
        owner_headers = {"Authorization": f"Bearer {owner_response.json()['access_token']}"}
        owner_user_id = owner_response.json()["user"]["id"]
        organization_id = owner_response.json()["user"]["memberships"][0]["organization_id"]

        employee_id = self._create_employee(headers=owner_headers, full_name="Care Worker A", id_card="111111111")
        position_id = self._create_position(headers=owner_headers, name="Nurse")
        template_id = self._create_shift_template(headers=owner_headers, position_id=position_id, name="Nurse Morning")
        assignment_response = self.client.post(
            "/api/employee-positions",
            headers=owner_headers,
            json={"employee_id": employee_id, "position_id": position_id, "is_primary": True, "priority_score": 90},
        )
        self.assertEqual(assignment_response.status_code, 200)
        schedule_response = self.client.post(
            "/api/schedule",
            headers=owner_headers,
            json={
                "employee_id": employee_id,
                "position_id": position_id,
                "date": "2026-04-20",
                "shift_template_id": template_id,
            },
        )
        self.assertEqual(schedule_response.status_code, 200)

        invitation_response = self.client.post(
            f"/api/organizations/{organization_id}/invitations",
            headers=owner_headers,
            json={"email": "employee-a@example.com", "employee_id": employee_id, "role": "employee", "expires_in_days": 7},
        )
        self.assertEqual(invitation_response.status_code, 200)
        accept_response = self.client.post(
            "/api/auth/accept-invitation",
            json={
                "token": invitation_response.json()["invitation_token"],
                "full_name": "Portal Login Name",
                "password": "EmployeePass123",
            },
        )
        self.assertEqual(accept_response.status_code, 200)
        employee_headers = {"Authorization": f"Bearer {accept_response.json()['access_token']}"}

        bundle = main.build_organization_export_bundle(self.connection, organization_id, exported_by_user_id=owner_user_id)
        import_response = self.client.post(
            f"/api/organizations/{organization_id}/cloud-import",
            headers=owner_headers,
            json={"bundle": bundle, "replace_existing": True},
        )
        self.assertEqual(import_response.status_code, 200)
        self.assertEqual(import_response.json()["restored_employee_links"]["memberships"], 1)

        imported_schedule_response = self.client.get("/api/schedule", headers=employee_headers)
        self.assertEqual(imported_schedule_response.status_code, 200)
        imported_schedule = imported_schedule_response.json()
        self.assertEqual(len(imported_schedule), 1)
        self.assertEqual(imported_schedule[0]["date"], "2026-04-20")
        self.assertEqual(imported_schedule[0]["shift_template_name"], "Nurse Morning")

    def test_invitation_url_uses_public_app_base_without_manual_cloud_link(self):
        owner_response = self.client.post(
            "/api/auth/bootstrap",
            json={
                "organization_name": "Portal Clinic",
                "full_name": "Owner User",
                "email": "owner@example.com",
                "password": "CorrectHorse123",
            },
        )
        self.assertEqual(owner_response.status_code, 200)
        headers = {"Authorization": f"Bearer {owner_response.json()['access_token']}"}
        organization_id = owner_response.json()["user"]["memberships"][0]["organization_id"]

        unlinked_employee_id = self._create_employee(headers=headers, full_name="Unlinked Employee")
        with patch.dict(os.environ, {"PUBLIC_APP_BASE_URL": "https://portal.shiftcare.co.il"}, clear=False):
            unlinked_invitation = self.client.post(
                f"/api/organizations/{organization_id}/invitations",
                headers=headers,
                json={
                    "email": "unlinked@example.com",
                    "employee_id": unlinked_employee_id,
                    "role": "employee",
                    "expires_in_days": 7,
                },
            )
        self.assertEqual(unlinked_invitation.status_code, 200)
        linked_token = unlinked_invitation.json()["invitation_token"]
        self.assertEqual(
            unlinked_invitation.json()["invitation_url"],
            f"https://portal.shiftcare.co.il/accept-invitation?token={linked_token}",
        )

    def test_login_page_returns_auth_shell(self):
        response = self.client.get("/login")
        self.assertEqual(response.status_code, 200)
        self.assertIn("Authorize user", response.text)
        self.assertIn("Add organization", response.text)
        self.assertIn("Local scheduling workspace", response.text)
        self.assertIn("Load organization data to this computer", response.text)
        self.assertIn("data-login-method=\"email\"", response.text)
        self.assertIn("data-login-method=\"id_card\"", response.text)
        self.assertIn("class=\"lang-switcher\"", response.text)
        self.assertIn("href=\"/static/icons/app-icon.svg\"", response.text)
        self.assertIn("/static/js/i18n.js", response.text)
        self.assertNotIn("Cloud portal and migration", response.text)
        self.assertNotIn("Cloud is the primary workspace", response.text)
        self.assertIn("/static/js/auth.js", response.text)

    def test_favicon_route_serves_app_icon(self):
        response = self.client.get("/favicon.ico")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.headers["content-type"], "image/x-icon")
        self.assertGreater(len(response.content), 0)

    def test_shared_i18n_contains_auth_page_keys(self):
        i18n_js = Path("static/js/i18n.js").read_text(encoding="utf-8")
        for key in [
            "auth_employee_portal",
            "auth_employee_login",
            "auth_employee_login_action_text",
            "auth_msg_employee_login_ready",
        ]:
            self.assertEqual(i18n_js.count(f"{key}:"), 3)

    def test_login_frontend_uses_desktop_cloud_login_without_workspace_switch(self):
        auth_js = Path("static/js/auth.js").read_text(encoding="utf-8")
        auth_client_js = Path("static/js/auth_client.js").read_text(encoding="utf-8")
        auth_i18n_js = Path("static/js/auth_i18n.js").read_text(encoding="utf-8")
        self.assertIn("initializeDefaultApiMode", auth_js)
        self.assertIn("window.scheduleAuth.useLocalApi();", auth_js)
        self.assertIn("/api/desktop/cloud-login", auth_js)
        self.assertIn("/api/desktop/cloud-create-organization", auth_js)
        self.assertIn("&& !isCloudEmployeePortalMode()", auth_js)
        self.assertIn("isDesktopLocalOrigin", auth_client_js)
        self.assertIn("isEmployeePortalMode", auth_i18n_js)
        self.assertNotIn("apiCloudButton", auth_js)
        self.assertIn("CLOUD_API_FALLBACK_BASE_URL", auth_client_js)
        self.assertIn("https://schedule-app-beta.web.app", auth_client_js)
        self.assertNotIn("initializeCloudFirstMode", auth_js)

    def test_schedule_frontend_auto_loads_employee_primary_position(self):
        schedule_js = Path("static/js/schedule.js").read_text(encoding="utf-8")
        self.assertIn("function isEmployeeUser()", schedule_js)
        self.assertIn("allPositions.find(position => position.is_primary) || allPositions[0]", schedule_js)
        self.assertIn("await loadSchedulePageData({ showLoadedMessage: false });", schedule_js)

    def test_hosted_web_schedule_hides_coverage_display_without_desktop_removal(self):
        schedule_html = Path("templates/schedule.html").read_text(encoding="utf-8")
        schedule_js = Path("static/js/schedule.js").read_text(encoding="utf-8")
        self.assertIn('data-web-hide="coverage"', schedule_html)
        self.assertIn("function shouldShowCoverageInSchedule()", schedule_js)
        self.assertIn("window.scheduleAuth?.isHostedCloudOrigin?.()", schedule_js)
        self.assertIn("hosted-employee-schedule", schedule_js)
        self.assertIn("coverageHeaderRow", schedule_js)

    def test_employee_portal_frontend_uses_server_portal_mode(self):
        auth_client_js = Path("static/js/auth_client.js").read_text(encoding="utf-8")
        access_control_js = Path("static/js/access_control.js").read_text(encoding="utf-8")
        schedule_html = Path("templates/schedule.html").read_text(encoding="utf-8")
        weekly_html = Path("templates/weekly_preferences.html").read_text(encoding="utf-8")
        shared_css = Path("static/css/style.css").read_text(encoding="utf-8")

        self.assertIn("function isEmployeePortalMode()", auth_client_js)
        self.assertIn('document.body?.dataset?.employeePortalMode === "1"', auth_client_js)
        self.assertIn("window.scheduleAuth?.isEmployeePortalMode?.()", access_control_js)
        self.assertIn('data-employee-portal-mode="{{ \'1\' if cloud_employee_portal_mode() else \'0\' }}"', schedule_html)
        self.assertIn('data-employee-portal-mode="{{ \'1\' if cloud_employee_portal_mode() else \'0\' }}"', weekly_html)
        self.assertIn("employee-portal-preferences", weekly_html)
        self.assertIn("data-employee-portal-hidden", weekly_html)
        self.assertIn("window.scheduleAccessControl?.apply?.();", weekly_html)
        self.assertIn("await loadWeekPreferences();", weekly_html)
        self.assertIn("body.employee-portal-preferences", shared_css)

    def test_service_worker_caches_current_employee_portal_assets(self):
        service_worker_js = Path("static/service-worker.js").read_text(encoding="utf-8")
        pwa_js = Path("static/js/pwa.js").read_text(encoding="utf-8")

        self.assertIn("20260624", service_worker_js)
        self.assertIn("/login", service_worker_js)
        self.assertIn("/departments", service_worker_js)
        self.assertIn('requestUrl.searchParams.get("embedded") === "1"', service_worker_js)
        self.assertIn("/static/css/auth.css?v=0.20.9_beta-desktop-1080p-readability", service_worker_js)
        self.assertIn("/static/js/auth.js?v=0.20.9_beta-portal-entry-employee-mode", service_worker_js)
        self.assertIn("/static/js/schedule.js?v=0.20.9_beta-schedule-sync-manual-time", service_worker_js)
        self.assertIn("/static/js/update_notifier.js?v=0.20.9_beta-startup-updates", service_worker_js)
        self.assertNotIn("/static/css/style.css?v=0.20.1_beta-generation-modes-rtl", service_worker_js)
        self.assertNotIn("/static/css/schedule.css?v=0.20.1_beta-generation-modes", service_worker_js)
        self.assertIn("registration.update()", pwa_js)

    def test_update_notifier_script_is_loaded_on_startup_pages(self):
        for template_name in [
            "index.html",
            "login.html",
            "schedule.html",
            "settings.html",
            "employees.html",
            "weekly_preferences.html",
            "organization.html",
            "feedback.html",
            "departments.html",
            "positions.html",
            "employee_positions.html",
            "coverage_requirements.html",
            "shift_templates.html",
            "guide.html",
        ]:
            with self.subTest(template_name=template_name):
                html = Path("templates") / template_name
                self.assertIn("/static/js/update_notifier.js", html.read_text(encoding="utf-8"))

        notifier_js = Path("static/js/update_notifier.js").read_text(encoding="utf-8")
        self.assertIn('document.body.dataset.demoMode === "1"', notifier_js)
        self.assertIn('document.body.dataset.employeePortalMode === "1"', notifier_js)

        i18n_js = Path("static/js/i18n.js").read_text(encoding="utf-8")
        for key in [
            "common_ok",
            "updates_modal_title",
            "updates_modal_install",
            "updates_changelog_title",
            "updates_changelog_empty",
        ]:
            self.assertEqual(i18n_js.count(f"{key}:"), 3)

    def test_departments_page_is_embedded_directory_without_global_sidebar_link(self):
        settings_response = self.client.get("/settings")
        self.assertEqual(settings_response.status_code, 200)
        self.assertNotIn('href="/departments"', settings_response.text)
        self.assertIn('data-directory-src="/departments?embedded=1"', settings_response.text)

        departments_response = self.client.get("/departments?embedded=1")
        self.assertEqual(departments_response.status_code, 200)
        self.assertIn("departments_page_title", departments_response.text)
        self.assertNotIn("settings_page_title", departments_response.text)

        shared_css = Path("static/css/style.css").read_text(encoding="utf-8")
        self.assertIn("body.embedded-admin .department-form-panel", shared_css)

        access_control_js = Path("static/js/access_control.js").read_text(encoding="utf-8")
        self.assertIn('"/departments"', access_control_js)
        self.assertNotIn('["/departments", "D", "nav_departments", "Departments"]', access_control_js)
        self.assertNotIn('"/departments": ["nav_departments", "Departments"]', access_control_js)

    def test_read_only_role_does_not_get_weekly_preferences_navigation(self):
        access_control_js = Path("static/js/access_control.js").read_text(encoding="utf-8")
        self.assertIn('read_only: {\n            pages: new Set(["/", "/schedule", "/organization", "/feedback", "/guide", "/docs"])', access_control_js)
        self.assertIn('nav: new Set(["/", "/schedule", "/organization", "/feedback"])', access_control_js)

    def test_organization_pages_return_auth_shells(self):
        organization_response = self.client.get("/organization")
        self.assertEqual(organization_response.status_code, 200)
        self.assertIn("/static/js/organization.js", organization_response.text)
        self.assertIn("Invite member", organization_response.text)
        self.assertIn("Public page for employee wishes", organization_response.text)
        self.assertIn('id="invite-role"', organization_response.text)
        self.assertIn('value="read_only"', organization_response.text)
        self.assertIn('id="invite-employee-field"', organization_response.text)
        self.assertNotIn("No employee link", organization_response.text)
        self.assertNotIn("Optional beta add-on", organization_response.text)
        self.assertNotIn("Upload and link cloud organization", organization_response.text)

        organization_alias_response = self.client.get("/organizations")
        self.assertEqual(organization_alias_response.status_code, 200)
        self.assertIn("/static/js/organization.js", organization_alias_response.text)

        invitation_response = self.client.get("/accept-invitation")
        self.assertEqual(invitation_response.status_code, 200)
        self.assertIn("/static/js/accept_invitation.js", invitation_response.text)
        self.assertIn("Accept invitation", invitation_response.text)

    def test_organization_frontend_uses_invitation_employee_link_without_manual_member_link(self):
        organization_js = Path("static/js/organization.js").read_text(encoding="utf-8")
        self.assertIn("role: selectedRole", organization_js)
        self.assertIn("payload.employee_id = Number(elements.inviteEmployee.value)", organization_js)
        self.assertIn('data-organization-action="member-role"', organization_js)
        self.assertIn("/members/${userId}/role", organization_js)
        self.assertNotIn("link-member-employee", organization_js)
        self.assertNotIn("data-member-employee-select", organization_js)

    def test_cloud_web_is_limited_to_employee_portal_surfaces(self):
        with patch.dict(os.environ, {"APP_ENV": "staging"}, clear=False):
            login_response = self.client.get("/login")
            self.assertEqual(login_response.status_code, 200)
            self.assertIn("Employee login", login_response.text)
            self.assertNotIn("Add organization", login_response.text)
            self.assertNotIn("bootstrap-form", login_response.text)

            blocked_create_response = self.client.post(
                "/api/auth/create-organization",
                json={
                    "organization_name": "Web Org",
                    "full_name": "Web Owner",
                    "email": "web-owner@example.com",
                    "password": "CorrectHorse123",
                },
            )
            self.assertEqual(blocked_create_response.status_code, 404)

            trusted_create_response = self.client.post(
                "/api/auth/create-organization",
                headers={"X-ShiftCare-Desktop-Client": "1"},
                json={
                    "organization_name": "Desktop Org",
                    "full_name": "Desktop Owner",
                    "email": "desktop-owner@example.com",
                    "password": "CorrectHorse123",
                },
            )
            self.assertEqual(trusted_create_response.status_code, 200)

            self.assertEqual(self.client.get("/employees").status_code, 404)
            self.assertEqual(self.client.get("/settings").status_code, 404)
            self.assertEqual(self.client.get("/departments").status_code, 404)
            self.assertEqual(self.client.get("/positions").status_code, 404)
            self.assertEqual(self.client.get("/weekly-preferences").status_code, 200)
            self.assertEqual(self.client.get("/schedule").status_code, 200)
            self.assertEqual(self.client.get("/organization").status_code, 200)

    def test_developer_support_dashboard_requires_feature_flag_and_admin(self):
        with tempfile.TemporaryDirectory() as disabled_app_data:
            with patch.dict(os.environ, {"LOCALAPPDATA": disabled_app_data, "SCHEDULE_APP_DEVELOPER_MODE": "0"}, clear=False):
                page_disabled = self.client.get("/support")
                self.assertEqual(page_disabled.status_code, 404)

        owner_response = self.client.post(
            "/api/auth/bootstrap",
            json={
                "organization_name": "Support Org",
                "full_name": "Owner User",
                "email": "owner@example.com",
                "password": "CorrectHorse123",
            },
        )
        self.assertEqual(owner_response.status_code, 200)
        owner_token = owner_response.json()["access_token"]
        owner_headers = {"Authorization": f"Bearer {owner_token}"}

        with tempfile.TemporaryDirectory() as disabled_app_data:
            with patch.dict(os.environ, {"LOCALAPPDATA": disabled_app_data, "SCHEDULE_APP_DEVELOPER_MODE": "0"}, clear=False):
                disabled_api = self.client.get("/api/support/accounts", headers=owner_headers)
                self.assertEqual(disabled_api.status_code, 404)

        employee_id = self._create_employee(headers=owner_headers, full_name="Linked Employee")
        invitation_response = self.client.post(
            "/api/organizations/1/invitations",
            headers=owner_headers,
            json={
                "email": "employee@example.com",
                "role": "employee",
                "employee_id": employee_id,
                "expires_in_days": 7,
            },
        )
        self.assertEqual(invitation_response.status_code, 200)
        employee_response = self.client.post(
            "/api/auth/accept-invitation",
            json={
                "token": invitation_response.json()["invitation_token"],
                "full_name": "Employee User",
                "password": "CorrectHorse123",
            },
        )
        self.assertEqual(employee_response.status_code, 200)
        employee_headers = {"Authorization": f"Bearer {employee_response.json()['access_token']}"}

        with patch.dict(os.environ, {"SCHEDULE_APP_DEVELOPER_MODE": "1"}, clear=False):
            page_enabled = self.client.get("/support")
            self.assertEqual(page_enabled.status_code, 200)
            self.assertIn("/static/js/support.js", page_enabled.text)
            support_js = Path("static/js/support.js").read_text(encoding="utf-8")
            self.assertIn("nativeFetch", support_js)
            self.assertNotIn("scheduleAuth.request(\"/api/support/accounts\")", support_js)

            employee_forbidden = self.client.get("/api/support/accounts", headers=employee_headers)
            self.assertEqual(employee_forbidden.status_code, 403)

            support_response = self.client.get("/api/support/accounts", headers=owner_headers)
            self.assertEqual(support_response.status_code, 200)
            payload = support_response.json()
            self.assertTrue(payload["developer_mode"])
            self.assertEqual(payload["organizations"][0]["name"], "Support Org")
            account_emails = {account["email"] for account in payload["accounts"]}
            self.assertIn("owner@example.com", account_emails)
            self.assertIn("employee@example.com", account_emails)
            self.assertEqual(payload["employees"][0]["full_name"], "Linked Employee")

        with tempfile.TemporaryDirectory() as temp_app_data:
            flag_path = Path(temp_app_data) / "Schedule App" / "developer_mode.flag"
            flag_path.parent.mkdir(parents=True)
            flag_path.write_text("enabled", encoding="utf-8")
            with patch.dict(os.environ, {"LOCALAPPDATA": temp_app_data, "SCHEDULE_APP_DEVELOPER_MODE": "0"}, clear=False):
                flag_enabled = self.client.get("/api/support/accounts", headers=owner_headers)
            self.assertEqual(flag_enabled.status_code, 200)

    def _save_shift_requirement(self, **overrides):
        payload = {
            "position_id": overrides.pop("position_id"),
            "shift_category": "morning",
            "required_total": 1,
            "required_female_min": 0,
            "required_male_min": 0,
        }
        payload.update(overrides)
        response = self.client.post("/api/shift-requirements", json=payload)
        self.assertEqual(response.status_code, 200)
        return response.json()["requirement"]

    def test_employee_crud_flow_round_trips_flags_and_values(self):
        create_response = self.client.post("/api/employees", json=self._employee_payload())
        self.assertEqual(create_response.status_code, 200)
        employee = create_response.json()["employee"]
        employee_id = employee["id"]
        self.assertEqual(employee["full_name"], "Employee A")

        list_response = self.client.get("/api/employees")
        self.assertEqual(list_response.status_code, 200)
        employees = list_response.json()
        self.assertEqual(len(employees), 1)
        self.assertTrue(employees[0]["can_work_night"])

        update_response = self.client.put(
            f"/api/employees/{employee_id}",
            json=self._employee_payload(
                full_name="Employee A Updated",
                target_shifts_per_week=4,
                can_work_night=False,
            ),
        )
        self.assertEqual(update_response.status_code, 200)
        self.assertEqual(update_response.json()["employee"]["full_name"], "Employee A Updated")

        list_response = self.client.get("/api/employees")
        updated_employee = list_response.json()[0]
        self.assertEqual(updated_employee["target_shifts_per_week"], 4)
        self.assertFalse(updated_employee["can_work_night"])

        delete_response = self.client.delete(f"/api/employees/{employee_id}")
        self.assertEqual(delete_response.status_code, 200)
        self.assertEqual(self.client.get("/api/employees").json(), [])

    def test_position_crud_blocks_duplicate_names(self):
        position_id = self._create_position()
        duplicate_response = self.client.post("/api/positions", json=self._position_payload())
        self.assertEqual(duplicate_response.status_code, 400)
        self.assertEqual(duplicate_response.json()["detail"], "Position already exists in this department")

        update_response = self.client.put(
            f"/api/positions/{position_id}",
            json=self._position_payload(
                name="Charge Nurse",
                color="#dbeafe",
                requires_continuous_coverage=True,
                minimum_staff_presence=2,
                allow_same_day_other_positions=True,
                max_consecutive_nights=1,
                emergency_max_consecutive_nights=2,
                max_consecutive_split_days=3,
                emergency_max_consecutive_split_days=4,
            ),
        )
        self.assertEqual(update_response.status_code, 200)

        positions = self.client.get("/api/positions").json()
        self.assertEqual(positions[0]["name"], "Charge Nurse")
        self.assertEqual(positions[0]["color"], "#dbeafe")
        self.assertTrue(positions[0]["requires_continuous_coverage"])
        self.assertEqual(positions[0]["minimum_staff_presence"], 2)
        self.assertTrue(positions[0]["allow_same_day_other_positions"])
        self.assertEqual(positions[0]["max_consecutive_nights"], 1)
        self.assertEqual(positions[0]["emergency_max_consecutive_nights"], 2)
        self.assertEqual(positions[0]["max_consecutive_split_days"], 3)
        self.assertEqual(positions[0]["emergency_max_consecutive_split_days"], 4)
        effective_settings = main.get_position_app_settings(self.connection, position_id)
        self.assertEqual(effective_settings["max_consecutive_nights"], 1)
        self.assertEqual(effective_settings["emergency_max_consecutive_nights"], 2)
        self.assertEqual(effective_settings["max_consecutive_split_days"], 3)
        self.assertEqual(effective_settings["emergency_max_consecutive_split_days"], 4)

    def test_same_position_name_is_allowed_in_different_departments(self):
        default_department = self.client.get("/api/departments").json()[0]
        second_department_response = self.client.post(
            "/api/departments",
            json={
                "name": "Ward 2",
                "description": None,
                "display_order": 2,
                "is_active": True,
            },
        )
        self.assertEqual(second_department_response.status_code, 200)
        second_department_id = second_department_response.json()["department"]["id"]

        first_position = self.client.post(
            "/api/positions",
            json=self._position_payload(department_id=default_department["id"], name="Nurse"),
        )
        self.assertEqual(first_position.status_code, 200)
        second_position = self.client.post(
            "/api/positions",
            json=self._position_payload(department_id=second_department_id, name="Nurse"),
        )
        self.assertEqual(second_position.status_code, 200)

        duplicate_in_same_department = self.client.post(
            "/api/positions",
            json=self._position_payload(department_id=second_department_id, name="Nurse"),
        )
        self.assertEqual(duplicate_in_same_department.status_code, 400)

        positions = self.client.get("/api/positions").json()
        self.assertEqual(
            {(position["department_name"], position["name"]) for position in positions},
            {("Main department", "Nurse"), ("Ward 2", "Nurse")},
        )

    def test_validation_rejects_conflicting_assignment_flags_and_invalid_time_windows(self):
        employee_id = self._create_employee()
        position_id = self._create_position()

        assignment_response = self.client.post(
            "/api/employee-positions",
            json={
                "employee_id": employee_id,
                "position_id": position_id,
                "is_primary": True,
                "priority_score": 100,
                "is_fallback_only": True,
            },
        )
        self.assertEqual(assignment_response.status_code, 422)

        template_response = self.client.post(
            "/api/shift-templates",
            json=self._template_payload(start_time="14:00", end_time="06:00", is_overnight=False),
        )
        self.assertEqual(template_response.status_code, 422)

        coverage_response = self.client.post(
            "/api/coverage-requirements",
            json={
                "position_id": position_id,
                "start_time": "14:00",
                "end_time": "06:00",
                "required_total": 1,
                "required_female_min": 0,
                "required_male_min": 0,
                "is_overnight": False,
            },
        )
        self.assertEqual(coverage_response.status_code, 422)

    def test_shift_template_crud_and_active_filter(self):
        position_id = self._create_position(name="Ward A")
        template_id = self._create_shift_template(position_id=position_id)

        update_response = self.client.put(
            f"/api/shift-templates/{template_id}",
            json=self._template_payload(
                position_id=position_id,
                name="Morning Inactive",
                is_active=False,
            ),
        )
        self.assertEqual(update_response.status_code, 200)

        all_templates = self.client.get("/api/shift-templates").json()
        self.assertEqual(len(all_templates), 1)
        self.assertFalse(all_templates[0]["is_active"])

        active_templates = self.client.get("/api/shift-templates", params={"active_only": True}).json()
        self.assertEqual(active_templates, [])

        delete_response = self.client.delete(f"/api/shift-templates/{template_id}")
        self.assertEqual(delete_response.status_code, 200)
        self.assertEqual(self.client.get("/api/shift-templates").json(), [])

    def test_weekly_preference_upsert_and_delete(self):
        employee_id = self._create_employee()

        create_response = self.client.post(
            "/api/employee-week-preferences",
            json={
                "employee_id": employee_id,
                "week_start_date": "2026-04-20",
                "preference_date": "2026-04-21",
                "preference_type": "off_day",
            },
        )
        self.assertEqual(create_response.status_code, 200)

        update_response = self.client.post(
            "/api/employee-week-preferences",
            json={
                "employee_id": employee_id,
                "week_start_date": "2026-04-20",
                "preference_date": "2026-04-21",
                "preference_type": "vacation",
            },
        )
        self.assertEqual(update_response.status_code, 200)

        list_response = self.client.get(
            "/api/employee-week-preferences",
            params={"week_start_date": "2026-04-20"},
        )
        self.assertEqual(list_response.status_code, 200)
        preferences = list_response.json()
        self.assertEqual(len(preferences), 1)
        self.assertEqual(preferences[0]["preference_type"], "vacation")

        delete_response = self.client.delete(
            "/api/employee-week-preferences",
            params={"employee_id": employee_id, "preference_date": "2026-04-21"},
        )
        self.assertEqual(delete_response.status_code, 200)
        self.assertEqual(delete_response.json()["deleted_count"], 1)

    def test_weekly_preferences_allow_multiple_shift_requests_same_day(self):
        employee_id = self._create_employee()
        for request_type, target_category in (
            ("request_shift", "morning"),
            ("request_shift", "night"),
            ("exclude_shift", "evening"),
        ):
            response = self.client.post(
                "/api/employee-week-preferences",
                json={
                    "employee_id": employee_id,
                    "week_start_date": "2026-04-20",
                    "preference_date": "2026-04-21",
                    "request_type": request_type,
                    "target_category": target_category,
                },
            )
            self.assertEqual(response.status_code, 200)

        preferences = self.client.get(
            "/api/employee-week-preferences",
            params={"week_start_date": "2026-04-20"},
        ).json()
        self.assertEqual(len(preferences), 3)
        self.assertEqual(
            {(item["request_type"], item["target_category"]) for item in preferences},
            {("request_shift", "morning"), ("request_shift", "night"), ("exclude_shift", "evening")},
        )

        delete_response = self.client.delete(
            "/api/employee-week-preferences",
            params={
                "employee_id": employee_id,
                "preference_date": "2026-04-21",
                "preference_id": preferences[0]["id"],
            },
        )
        self.assertEqual(delete_response.status_code, 200)
        self.assertEqual(delete_response.json()["deleted_count"], 1)
        remaining = self.client.get(
            "/api/employee-week-preferences",
            params={"week_start_date": "2026-04-20"},
        ).json()
        self.assertEqual(len(remaining), 2)

    def test_employee_weekly_preferences_after_two_days_require_approval(self):
        owner_response = self.client.post(
            "/api/auth/bootstrap",
            json={
                "organization_name": "Beta Clinic",
                "full_name": "Owner User",
                "email": "owner@example.com",
                "password": "CorrectHorse123",
            },
        )
        self.assertEqual(owner_response.status_code, 200)
        owner_user_id = owner_response.json()["user"]["id"]
        owner_headers = {"Authorization": f"Bearer {owner_response.json()['access_token']}"}
        employee_id = self._create_employee(
            headers=owner_headers,
            full_name="Employee User",
            id_card="123456789",
        )
        position_id = self._create_position(headers=owner_headers, name="Ward A")
        assignment_response = self.client.post(
            "/api/employee-positions",
            headers=owner_headers,
            json={
                "employee_id": employee_id,
                "position_id": position_id,
                "is_primary": True,
                "priority_score": 100,
                "is_fallback_only": False,
            },
        )
        self.assertEqual(assignment_response.status_code, 200)
        invitation_response = self.client.post(
            "/api/organizations/1/invitations",
            headers=owner_headers,
            json={"email": "employee@example.com", "employee_id": employee_id, "role": "employee", "expires_in_days": 7},
        )
        self.assertEqual(invitation_response.status_code, 200)
        employee_response = self.client.post(
            "/api/auth/accept-invitation",
            json={
                "token": invitation_response.json()["invitation_token"],
                "full_name": "Employee User",
                "password": "EmployeePass123",
            },
        )
        self.assertEqual(employee_response.status_code, 200)
        employee_headers = {"Authorization": f"Bearer {employee_response.json()['access_token']}"}

        cursor = self.connection.cursor()
        cursor.execute(
            """
            INSERT INTO users (email, full_name, status, email_verified, created_at, updated_at)
            VALUES ('ward-admin@example.com', 'Ward Admin', 'active', 1, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
            """
        )
        department_admin_user_id = cursor.lastrowid
        cursor.execute(
            """
            INSERT INTO organization_memberships (organization_id, user_id, role, status, created_at, updated_at)
            VALUES (1, ?, 'admin', 'active', CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
            """,
            (department_admin_user_id,),
        )
        cursor.execute(
            """
            INSERT INTO user_department_access (
                organization_id, user_id, department_id, created_at, updated_at, updated_by
            )
            VALUES (1, ?, 1, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP, ?)
            """,
            (department_admin_user_id, owner_user_id),
        )
        department_admin_session = main.build_auth_response(self.connection, department_admin_user_id)
        self.connection.commit()
        department_admin_headers = {"Authorization": f"Bearer {department_admin_session['access_token']}"}

        for preference_date in ("2026-04-20", "2026-04-21"):
            response = self.client.post(
                "/api/employee-week-preferences",
                headers=employee_headers,
                json={
                    "employee_id": employee_id,
                    "week_start_date": "2026-04-20",
                    "preference_date": preference_date,
                    "request_type": "request_shift",
                    "target_category": "morning",
                },
            )
            self.assertEqual(response.status_code, 200)
            self.assertEqual(response.json()["status"], "saved")

        third_day_response = self.client.post(
            "/api/employee-week-preferences",
            headers=employee_headers,
            json={
                "employee_id": employee_id,
                "week_start_date": "2026-04-20",
                "preference_date": "2026-04-22",
                "request_type": "exclude_shift",
                "target_category": "night",
            },
        )
        self.assertEqual(third_day_response.status_code, 200)
        self.assertEqual(third_day_response.json()["status"], "pending_approval")

        fourth_day_response = self.client.post(
            "/api/employee-week-preferences",
            headers=employee_headers,
            json={
                "employee_id": employee_id,
                "week_start_date": "2026-04-20",
                "preference_date": "2026-04-23",
                "request_type": "request_shift",
                "target_category": "evening",
            },
        )
        self.assertEqual(fourth_day_response.status_code, 200)
        self.assertEqual(fourth_day_response.json()["status"], "pending_approval")

        preferences = self.client.get(
            "/api/employee-week-preferences",
            headers=owner_headers,
            params={"week_start_date": "2026-04-20"},
        ).json()
        self.assertEqual({item["preference_date"] for item in preferences}, {"2026-04-20", "2026-04-21"})

        pending_requests = self.client.get(
            "/api/employee-week-preference-requests",
            headers=department_admin_headers,
            params={"week_start_date": "2026-04-20", "status": "pending"},
        ).json()
        pending_by_date = {request["preference_date"]: request for request in pending_requests}
        self.assertEqual(set(pending_by_date), {"2026-04-22", "2026-04-23"})
        self.assertEqual(pending_by_date["2026-04-22"]["employee_id"], employee_id)

        delete_response = self.client.delete(
            f"/api/employee-week-preference-requests/{pending_by_date['2026-04-23']['id']}",
            headers=department_admin_headers,
        )
        self.assertEqual(delete_response.status_code, 200)
        self.assertEqual(delete_response.json()["deleted_count"], 1)

        pending_requests = self.client.get(
            "/api/employee-week-preference-requests",
            headers=department_admin_headers,
            params={"week_start_date": "2026-04-20", "status": "pending"},
        ).json()
        self.assertEqual(len(pending_requests), 1)
        self.assertEqual(pending_requests[0]["preference_date"], "2026-04-22")

        approve_response = self.client.patch(
            f"/api/employee-week-preference-requests/{pending_requests[0]['id']}",
            headers=department_admin_headers,
            json={"status": "approved"},
        )
        self.assertEqual(approve_response.status_code, 200)
        self.assertEqual(approve_response.json()["status"], "approved")
        self.assertIsNotNone(approve_response.json()["approved_preference_id"])

        preferences_after_approval = self.client.get(
            "/api/employee-week-preferences",
            headers=owner_headers,
            params={"week_start_date": "2026-04-20"},
        ).json()
        self.assertEqual(
            {item["preference_date"] for item in preferences_after_approval},
            {"2026-04-20", "2026-04-21", "2026-04-22"},
        )
        reviewed_requests = self.client.get(
            "/api/employee-week-preference-requests",
            headers=employee_headers,
            params={"week_start_date": "2026-04-20"},
        ).json()
        self.assertEqual(reviewed_requests[0]["status"], "approved")

        employee_delete_candidate = self.client.post(
            "/api/employee-week-preferences",
            headers=employee_headers,
            json={
                "employee_id": employee_id,
                "week_start_date": "2026-04-20",
                "preference_date": "2026-04-23",
                "request_type": "request_shift",
                "target_category": "evening",
            },
        )
        self.assertEqual(employee_delete_candidate.status_code, 200)
        self.assertEqual(employee_delete_candidate.json()["status"], "pending_approval")

        employee_delete_response = self.client.delete(
            f"/api/employee-week-preference-requests/{employee_delete_candidate.json()['request']['id']}",
            headers=employee_headers,
        )
        self.assertEqual(employee_delete_response.status_code, 200)
        self.assertEqual(employee_delete_response.json()["deleted_count"], 1)

        employee_requests_after_delete = self.client.get(
            "/api/employee-week-preference-requests",
            headers=employee_headers,
            params={"week_start_date": "2026-04-20"},
        ).json()
        self.assertEqual(
            {request["preference_date"] for request in employee_requests_after_delete},
            {"2026-04-22"},
        )

    def test_vacation_day_status_blocks_schedule_cell(self):
        employee_id = self._create_employee()
        response = self.client.post(
            "/api/employee-day-statuses",
            json={"employee_id": employee_id, "date": "2026-04-21", "status_type": "vacation"},
        )
        self.assertEqual(response.status_code, 200)
        statuses = self.client.get("/api/employee-day-statuses").json()
        self.assertEqual(statuses[0]["status_type"], "vacation")

    def test_weekly_preference_date_must_belong_to_selected_week(self):
        employee_id = self._create_employee()

        response = self.client.post(
            "/api/employee-week-preferences",
            json={
                "employee_id": employee_id,
                "week_start_date": "2026-05-03",
                "preference_date": "2026-05-10",
                "preference_type": "off_day",
            },
        )

        self.assertEqual(response.status_code, 422)
        self.assertIn("preference_date must belong to the selected week", response.text)

    def test_pending_local_preferences_prevent_cloud_pull_before_generation(self):
        employee_id = self._create_employee()
        response = self.client.post(
            "/api/employee-week-preferences",
            json={
                "employee_id": employee_id,
                "week_start_date": "2026-05-03",
                "preference_date": "2026-05-04",
                "preference_type": "off_day",
            },
        )
        self.assertEqual(response.status_code, 200)

        cursor = self.connection.cursor()
        for key, value in {
            "cloud_api_base_url": "https://schedule-app-beta.web.app",
            "cloud_organization_id": "42",
            "desktop_cloud_access_token": "cloud-token",
        }.items():
            cursor.execute(
                """
                INSERT INTO app_settings (organization_id, key, value)
                VALUES (1, ?, ?)
                ON CONFLICT(key) DO UPDATE SET value = excluded.value
                """,
                (key, value),
            )
        self.connection.commit()

        with patch.object(main, "request_cloud_json", side_effect=AssertionError("cloud pull should be skipped")):
            main.pull_cloud_preferences_for_desktop_generation(self.connection)

        cursor.execute(
            """
            SELECT preference_type
            FROM employee_week_preferences
            WHERE employee_id = ? AND preference_date = '2026-05-04'
            """,
            (employee_id,),
        )
        self.assertEqual(cursor.fetchone()["preference_type"], "off_day")

    def test_weekly_preferences_list_pulls_cloud_preferences_when_no_local_pending_changes(self):
        employee_id = self._create_employee()
        cursor = self.connection.cursor()
        cursor.execute("UPDATE employees SET public_id = 'emp_cloud' WHERE id = ?", (employee_id,))
        for key, value in {
            "cloud_api_base_url": "https://schedule-app-beta.web.app",
            "cloud_organization_id": "42",
            "desktop_cloud_access_token": "cloud-token",
        }.items():
            cursor.execute(
                """
                INSERT INTO app_settings (organization_id, key, value)
                VALUES (1, ?, ?)
                ON CONFLICT(key) DO UPDATE SET value = excluded.value
                """,
                (key, value),
            )
        cursor.execute("DELETE FROM desktop_sync_outbox")
        self.connection.commit()

        cloud_bundle = {
            "format": "shiftcare.organization.v1",
            "app_version": "0.15.15_beta",
            "records": {
                "employees": [{"id": 7, "public_id": "emp_cloud"}],
                "employee_preferences": [],
                "employee_week_preferences": [
                    {
                        "id": 11,
                        "public_id": "wpr_cloud",
                        "employee_id": 7,
                        "week_start_date": "2026-05-03",
                        "preference_date": "2026-05-04",
                        "preference_type": "off_day",
                        "created_at": "2026-05-06T01:00:00",
                        "updated_at": "2026-05-06T01:00:00",
                    }
                ],
                "employee_recurring_preferences": [],
            },
        }

        def fake_cloud_request(base_url, path, **kwargs):
            self.assertEqual(path, "/api/organizations/42/cloud-export")
            self.assertEqual(kwargs["token"], "cloud-token")
            return cloud_bundle

        with patch.object(main, "request_cloud_json", side_effect=fake_cloud_request):
            response = self.client.get(
                "/api/employee-week-preferences",
                params={"week_start_date": "2026-05-03"},
            )

        self.assertEqual(response.status_code, 200)
        preferences = response.json()
        self.assertEqual(len(preferences), 1)
        self.assertEqual(preferences[0]["employee_id"], employee_id)
        self.assertEqual(preferences[0]["preference_type"], "off_day")

    def test_cloud_preference_pull_does_not_queue_imported_rows_for_push(self):
        employee_id = self._create_employee()
        cursor = self.connection.cursor()
        cursor.execute("UPDATE employees SET public_id = 'emp_cloud' WHERE id = ?", (employee_id,))
        cursor.execute("DELETE FROM desktop_sync_outbox")
        self.connection.commit()

        cloud_bundle = {
            "format": "shiftcare.organization.v1",
            "app_version": "0.20.3_beta",
            "records": {
                "employees": [{"id": 7, "public_id": "emp_cloud"}],
                "employee_preferences": [],
                "employee_week_preferences": [
                    {
                        "id": 11,
                        "public_id": "wpr_cloud",
                        "employee_id": 7,
                        "week_start_date": "2026-05-03",
                        "preference_date": "2026-05-04",
                        "preference_type": "off_day",
                        "request_type": "day_off",
                        "target_category": None,
                        "created_at": "2026-05-06T01:00:00",
                        "updated_at": "2026-05-06T01:00:00",
                    }
                ],
                "employee_recurring_preferences": [
                    {
                        "id": 12,
                        "public_id": "rpr_cloud",
                        "employee_id": 7,
                        "preference_kind": "strict",
                        "day_of_week": 6,
                        "preference_type": "only_night",
                        "request_type": "request_shift",
                        "target_category": "night",
                        "created_at": "2026-05-06T01:00:00",
                        "updated_at": "2026-05-06T01:00:00",
                    }
                ],
            },
        }

        def fake_cloud_request(base_url, path, **kwargs):
            self.assertEqual(path, "/api/organizations/42/cloud-export")
            return cloud_bundle

        with patch.object(main, "request_cloud_json", side_effect=fake_cloud_request):
            self.assertTrue(main.sync_cloud_preferences_to_desktop(self.connection, {
                "cloud_api_base_url": "https://schedule-app-beta.web.app",
                "cloud_organization_id": "42",
                "desktop_cloud_access_token": "cloud-token",
            }))
        self.connection.commit()

        cursor.execute("SELECT COUNT(*) AS count FROM desktop_sync_outbox")
        self.assertEqual(cursor.fetchone()["count"], 0)
        cursor.execute(
            """
            SELECT request_type, target_category
            FROM employee_recurring_preferences
            WHERE employee_id = ? AND day_of_week = 6
            """,
            (employee_id,),
        )
        recurring_row = cursor.fetchone()
        self.assertIsNotNone(recurring_row)
        self.assertEqual(recurring_row["request_type"], "request_shift")
        self.assertEqual(recurring_row["target_category"], "night")

    def test_auto_generate_uses_local_preferences_without_cloud_pull(self):
        position_id = self._create_position(name="Caregiver")
        self._create_shift_template(position_id=position_id)
        employee_id = self._create_employee(full_name="Local Employee")
        response = self.client.post(
            "/api/employee-positions",
            json={
                "employee_id": employee_id,
                "position_id": position_id,
                "is_primary": True,
                "priority_score": 100,
                "is_fallback_only": False,
            },
        )
        self.assertEqual(response.status_code, 200)
        recurring_response = self.client.post(
            "/api/employee-recurring-preferences",
            json={
                "employee_id": employee_id,
                "rules": [
                    {
                        "preference_kind": "strict",
                        "day_of_week": 1,
                        "request_type": "request_shift",
                        "target_category": "morning",
                    }
                ],
            },
        )
        self.assertEqual(recurring_response.status_code, 200)
        self._save_shift_requirement(position_id=position_id)

        cursor = self.connection.cursor()
        for key, value in {
            "cloud_api_base_url": "https://schedule-app-beta.web.app",
            "cloud_organization_id": "42",
            "desktop_cloud_access_token": "cloud-token",
        }.items():
            cursor.execute(
                """
                INSERT INTO app_settings (organization_id, key, value)
                VALUES (1, ?, ?)
                ON CONFLICT(key) DO UPDATE SET value = excluded.value
                """,
                (key, value),
            )
        cursor.execute("DELETE FROM desktop_sync_outbox")
        self.connection.commit()

        with patch.object(main, "request_cloud_json", side_effect=AssertionError("generation must not pull cloud preferences")):
            generate_response = self.client.post(
                "/api/schedule/auto-generate",
                json={"position_id": position_id, "week_start_date": "2026-04-20"},
            )

        self.assertEqual(generate_response.status_code, 200)
        cursor.execute(
            """
            SELECT request_type, target_category
            FROM employee_recurring_preferences
            WHERE employee_id = ?
            """,
            (employee_id,),
        )
        recurring_row = cursor.fetchone()
        self.assertIsNotNone(recurring_row)
        self.assertEqual(recurring_row["request_type"], "request_shift")
        self.assertEqual(recurring_row["target_category"], "morning")

    def test_request_cloud_json_retries_transient_network_error(self):
        class FakeResponse:
            def __enter__(self):
                return self

            def __exit__(self, exc_type, exc, traceback):
                return False

            def read(self):
                return b'{"ok": true}'

        with (
            patch.object(
                main.urllib.request,
                "urlopen",
                side_effect=[main.urllib.error.URLError("temporary dns failure"), FakeResponse()],
            ) as urlopen_mock,
            patch.object(main, "sleep") as sleep_mock,
        ):
            response = main.request_cloud_json("https://cloud.example.com", "/api/ping")

        self.assertEqual(response, {"ok": True})
        self.assertEqual(urlopen_mock.call_count, 2)
        self.assertEqual(urlopen_mock.call_args.kwargs["timeout"], 45.0)
        sleep_mock.assert_called_once_with(1.0)

    def test_request_cloud_json_retries_retryable_cloud_status(self):
        class FakeResponse:
            def __enter__(self):
                return self

            def __exit__(self, exc_type, exc, traceback):
                return False

            def read(self):
                return b'{"ok": true}'

        http_error = main.urllib.error.HTTPError(
            "https://cloud.example.com/api/ping",
            503,
            "Service Unavailable",
            {"Retry-After": "2"},
            main.BytesIO(b'{"detail": "temporary outage"}'),
        )

        with (
            patch.object(main.urllib.request, "urlopen", side_effect=[http_error, FakeResponse()]) as urlopen_mock,
            patch.object(main, "sleep") as sleep_mock,
        ):
            response = main.request_cloud_json("https://cloud.example.com", "/api/ping")

        self.assertEqual(response, {"ok": True})
        self.assertEqual(urlopen_mock.call_count, 2)
        sleep_mock.assert_called_once_with(2.0)

    def test_desktop_sync_pushes_pending_preferences_without_pre_pull_delete(self):
        employee_id = self._create_employee()
        cursor = self.connection.cursor()
        cursor.execute("DELETE FROM desktop_sync_outbox")
        for key, value in {
            "cloud_api_base_url": "https://schedule-app-beta.web.app",
            "cloud_organization_id": "42",
            "desktop_cloud_access_token": "cloud-token",
        }.items():
            cursor.execute(
                """
                INSERT INTO app_settings (organization_id, key, value)
                VALUES (1, ?, ?)
                ON CONFLICT(key) DO UPDATE SET value = excluded.value
                """,
                (key, value),
            )
        self.connection.commit()

        response = self.client.post(
            "/api/employee-week-preferences",
            json={
                "employee_id": employee_id,
                "week_start_date": "2026-05-03",
                "preference_date": "2026-05-04",
                "preference_type": "off_day",
            },
        )
        self.assertEqual(response.status_code, 200)

        cloud_import_payloads = []

        def fake_cloud_request(base_url, path, **kwargs):
            if path == "/api/organizations/42/cloud-export":
                raise AssertionError("pending local preferences must not be replaced by a cloud pull before push")
            if path == "/api/organizations/42/cloud-import":
                cloud_import_payloads.append(kwargs["payload"])
                return {"message": "imported"}
            raise AssertionError(path)

        with patch.object(main, "request_cloud_json", side_effect=fake_cloud_request):
            self.assertTrue(main.run_desktop_sync_once())

        cursor.execute(
            """
            SELECT preference_type
            FROM employee_week_preferences
            WHERE employee_id = ? AND preference_date = '2026-05-04'
            """,
            (employee_id,),
        )
        self.assertEqual(cursor.fetchone()["preference_type"], "off_day")
        synced_preferences = cloud_import_payloads[0]["bundle"]["records"]["employee_week_preferences"]
        self.assertEqual(synced_preferences[0]["preference_type"], "off_day")

    def test_weekly_preference_permissions_after_auth_bootstrap(self):
        employee_a = self._create_employee(full_name="Employee A")
        employee_b = self._create_employee(full_name="Employee B")
        owner_response = self.client.post(
            "/api/auth/bootstrap",
            json={
                "organization_name": "Beta Clinic",
                "full_name": "Owner User",
                "email": "owner@example.com",
                "password": "CorrectHorse123",
            },
        )
        self.assertEqual(owner_response.status_code, 200)
        owner_token = owner_response.json()["access_token"]
        owner_headers = {"Authorization": f"Bearer {owner_token}"}

        invitation_response = self.client.post(
            "/api/organizations/1/invitations",
            headers=owner_headers,
            json={"email": "employee@example.com", "employee_id": employee_a, "role": "employee", "expires_in_days": 7},
        )
        employee_response = self.client.post(
            "/api/auth/accept-invitation",
            json={
                "token": invitation_response.json()["invitation_token"],
                "full_name": "Employee User",
                "password": "EmployeePass123",
            },
        )
        self.assertEqual(employee_response.status_code, 200)
        employee_user_id = employee_response.json()["user"]["id"]
        employee_headers = {"Authorization": f"Bearer {employee_response.json()['access_token']}"}

        own_response = self.client.post(
            "/api/employee-week-preferences",
            headers=employee_headers,
            json={
                "employee_id": employee_a,
                "week_start_date": "2026-04-20",
                "preference_date": "2026-04-21",
                "preference_type": "off_day",
            },
        )
        self.assertEqual(own_response.status_code, 200)

        other_response = self.client.post(
            "/api/employee-week-preferences",
            headers=employee_headers,
            json={
                "employee_id": employee_b,
                "week_start_date": "2026-04-20",
                "preference_date": "2026-04-22",
                "preference_type": "off_day",
            },
        )
        self.assertEqual(other_response.status_code, 403)

        list_response = self.client.get(
            "/api/employee-week-preferences",
            headers=employee_headers,
            params={"week_start_date": "2026-04-20"},
        )
        self.assertEqual(list_response.status_code, 200)
        employee_preferences = list_response.json()
        self.assertEqual([item["employee_id"] for item in employee_preferences], [employee_a])
        self.assertEqual(employee_preferences[0]["preference_type"], "off_day")
        self.assertEqual(employee_preferences[0]["request_type"], "day_off")

        employee_list_response = self.client.get("/api/employees", headers=employee_headers)
        self.assertEqual(employee_list_response.status_code, 200)
        self.assertEqual([item["id"] for item in employee_list_response.json()], [employee_a])

        employee_schedule_response = self.client.get("/api/schedule", headers=employee_headers)
        self.assertEqual(employee_schedule_response.status_code, 200)

        denied_day_status = self.client.post(
            "/api/employee-day-statuses",
            headers=employee_headers,
            json={"employee_id": employee_a, "date": "2026-04-20", "status_type": "sick"},
        )
        self.assertEqual(denied_day_status.status_code, 403)

        owner_employee_list_response = self.client.get("/api/employees", headers=owner_headers)
        self.assertEqual(owner_employee_list_response.status_code, 200)
        self.assertEqual(
            {item["id"] for item in owner_employee_list_response.json()},
            {employee_a, employee_b},
        )

    def test_employee_invitation_requires_employee_link(self):
        self._create_employee(full_name="Employee A")
        owner_response = self.client.post(
            "/api/auth/bootstrap",
            json={
                "organization_name": "Beta Clinic",
                "full_name": "Owner User",
                "email": "owner@example.com",
                "password": "CorrectHorse123",
            },
        )
        owner_headers = {"Authorization": f"Bearer {owner_response.json()['access_token']}"}

        invitation_response = self.client.post(
            "/api/organizations/1/invitations",
            headers=owner_headers,
            json={"email": "employee@example.com", "role": "employee", "expires_in_days": 7},
        )
        self.assertEqual(invitation_response.status_code, 400)

    def test_employee_preference_access_repairs_accepted_invitation_link(self):
        employee_id = self._create_employee(full_name="Employee A")
        owner_response = self.client.post(
            "/api/auth/bootstrap",
            json={
                "organization_name": "Beta Clinic",
                "full_name": "Owner User",
                "email": "owner@example.com",
                "password": "CorrectHorse123",
            },
        )
        owner_headers = {"Authorization": f"Bearer {owner_response.json()['access_token']}"}
        invitation_response = self.client.post(
            "/api/organizations/1/invitations",
            headers=owner_headers,
            json={"email": "employee@example.com", "employee_id": employee_id, "role": "employee", "expires_in_days": 7},
        )
        employee_response = self.client.post(
            "/api/auth/accept-invitation",
            json={
                "token": invitation_response.json()["invitation_token"],
                "password": "EmployeePass123",
            },
        )
        employee_user_id = employee_response.json()["user"]["id"]
        employee_headers = {"Authorization": f"Bearer {employee_response.json()['access_token']}"}
        cursor = self.connection.cursor()
        cursor.execute("UPDATE organization_memberships SET employee_id = NULL WHERE user_id = ?", (employee_user_id,))
        self.connection.commit()

        employee_list_response = self.client.get("/api/employees", headers=employee_headers)
        self.assertEqual(employee_list_response.status_code, 200)
        self.assertEqual([item["id"] for item in employee_list_response.json()], [employee_id])

        cursor.execute("SELECT employee_id FROM organization_memberships WHERE user_id = ?", (employee_user_id,))
        self.assertEqual(cursor.fetchone()["employee_id"], employee_id)

    def test_unlinked_employee_preference_access_returns_specific_error(self):
        owner_response = self.client.post(
            "/api/auth/bootstrap",
            json={
                "organization_name": "Beta Clinic",
                "full_name": "Owner User",
                "email": "owner@example.com",
                "password": "CorrectHorse123",
            },
        )
        owner_headers = {"Authorization": f"Bearer {owner_response.json()['access_token']}"}
        invitation_response = self.client.post(
            "/api/organizations/1/invitations",
            headers=owner_headers,
            json={"email": "employee@example.com", "role": "read_only", "expires_in_days": 7},
        )
        self.assertEqual(invitation_response.status_code, 200)
        employee_response = self.client.post(
            "/api/auth/accept-invitation",
            json={
                "token": invitation_response.json()["invitation_token"],
                "full_name": "Employee User",
                "password": "EmployeePass123",
            },
        )
        self.assertEqual(employee_response.status_code, 200)
        employee_headers = {"Authorization": f"Bearer {employee_response.json()['access_token']}"}

        employee_list_response = self.client.get("/api/employees", headers=employee_headers)
        self.assertEqual(employee_list_response.status_code, 403)
        self.assertEqual(employee_list_response.json()["detail"], "Preference permissions are required")

    def test_owner_can_update_employee_member_link(self):
        employee_id = self._create_employee(full_name="Employee A")
        cursor = self.connection.cursor()
        cursor.execute("SELECT public_id FROM employees WHERE id = ?", (employee_id,))
        employee_public_id = cursor.fetchone()["public_id"]
        owner_response = self.client.post(
            "/api/auth/bootstrap",
            json={
                "organization_name": "Beta Clinic",
                "full_name": "Owner User",
                "email": "owner@example.com",
                "password": "CorrectHorse123",
            },
        )
        owner_headers = {"Authorization": f"Bearer {owner_response.json()['access_token']}"}
        invitation_response = self.client.post(
            "/api/organizations/1/invitations",
            headers=owner_headers,
            json={"email": "employee@example.com", "employee_id": employee_id, "role": "employee", "expires_in_days": 7},
        )
        employee_response = self.client.post(
            "/api/auth/accept-invitation",
            json={"token": invitation_response.json()["invitation_token"], "password": "EmployeePass123"},
        )
        employee_user_id = employee_response.json()["user"]["id"]
        cursor.execute("UPDATE organization_memberships SET employee_id = NULL WHERE user_id = ?", (employee_user_id,))
        self.connection.commit()

        link_response = self.client.put(
            f"/api/organizations/1/members/{employee_user_id}/employee-link",
            headers=owner_headers,
            json={"employee_id": employee_id},
        )
        self.assertEqual(link_response.status_code, 200)

        cursor.execute("SELECT employee_id FROM organization_memberships WHERE user_id = ?", (employee_user_id,))
        self.assertEqual(cursor.fetchone()["employee_id"], employee_id)

        cursor.execute("UPDATE organization_memberships SET employee_id = NULL WHERE user_id = ?", (employee_user_id,))
        self.connection.commit()
        public_link_response = self.client.put(
            f"/api/organizations/1/members/{employee_user_id}/employee-link",
            headers=owner_headers,
            json={"employee_public_id": employee_public_id},
        )
        self.assertEqual(public_link_response.status_code, 200)

        cursor.execute("SELECT employee_id FROM organization_memberships WHERE user_id = ?", (employee_user_id,))
        self.assertEqual(cursor.fetchone()["employee_id"], employee_id)

    def test_owner_can_update_member_role_and_employee_link_together(self):
        employee_id = self._create_employee(full_name="Employee B")
        owner_response = self.client.post(
            "/api/auth/bootstrap",
            json={
                "organization_name": "Beta Clinic",
                "full_name": "Owner User",
                "email": "owner@example.com",
                "password": "CorrectHorse123",
            },
        )
        owner_headers = {"Authorization": f"Bearer {owner_response.json()['access_token']}"}
        invitation_response = self.client.post(
            "/api/organizations/1/invitations",
            headers=owner_headers,
            json={"email": "viewer@example.com", "role": "read_only", "expires_in_days": 7},
        )
        self.assertEqual(invitation_response.status_code, 200)
        viewer_response = self.client.post(
            "/api/auth/accept-invitation",
            json={
                "token": invitation_response.json()["invitation_token"],
                "full_name": "Viewer User",
                "password": "ViewerPass123",
            },
        )
        self.assertEqual(viewer_response.status_code, 200)
        viewer_user_id = viewer_response.json()["user"]["id"]

        scheduler_response = self.client.put(
            f"/api/organizations/1/members/{viewer_user_id}/role",
            headers=owner_headers,
            json={"role": "scheduler"},
        )
        self.assertEqual(scheduler_response.status_code, 200)
        cursor = self.connection.cursor()
        cursor.execute("SELECT role, employee_id FROM organization_memberships WHERE user_id = ?", (viewer_user_id,))
        updated_role = cursor.fetchone()
        self.assertEqual(updated_role["role"], "scheduler")
        self.assertIsNone(updated_role["employee_id"])

        employee_role_response = self.client.put(
            f"/api/organizations/1/members/{viewer_user_id}/role",
            headers=owner_headers,
            json={"role": "employee", "employee_id": employee_id},
        )
        self.assertEqual(employee_role_response.status_code, 200)
        cursor.execute("SELECT role, employee_id FROM organization_memberships WHERE user_id = ?", (viewer_user_id,))
        employee_role = cursor.fetchone()
        self.assertEqual(employee_role["role"], "employee")
        self.assertEqual(employee_role["employee_id"], employee_id)

        self_update_response = self.client.put(
            f"/api/organizations/1/members/{owner_response.json()['user']['id']}/role",
            headers=owner_headers,
            json={"role": "read_only"},
        )
        self.assertEqual(self_update_response.status_code, 400)

    def test_role_management_protects_owner_only_roles(self):
        owner_response = self.client.post(
            "/api/auth/bootstrap",
            json={
                "organization_name": "Beta Clinic",
                "full_name": "Owner User",
                "email": "owner@example.com",
                "password": "CorrectHorse123",
            },
        )
        owner_headers = {"Authorization": f"Bearer {owner_response.json()['access_token']}"}
        admin_invitation = self.client.post(
            "/api/organizations/1/invitations",
            headers=owner_headers,
            json={"email": "admin@example.com", "role": "admin", "expires_in_days": 7},
        )
        self.assertEqual(admin_invitation.status_code, 200)
        admin_response = self.client.post(
            "/api/auth/accept-invitation",
            json={
                "token": admin_invitation.json()["invitation_token"],
                "full_name": "Admin User",
                "password": "AdminPass123",
            },
        )
        self.assertEqual(admin_response.status_code, 200)
        admin_headers = {"Authorization": f"Bearer {admin_response.json()['access_token']}"}

        owner_invitation = self.client.post(
            "/api/organizations/1/invitations",
            headers=admin_headers,
            json={"email": "next-owner@example.com", "role": "owner", "expires_in_days": 7},
        )
        self.assertEqual(owner_invitation.status_code, 403)

    def test_schedule_edit_permissions_after_auth_bootstrap(self):
        owner_response = self.client.post(
            "/api/auth/bootstrap",
            json={
                "organization_name": "Beta Clinic",
                "full_name": "Owner User",
                "email": "owner@example.com",
                "password": "CorrectHorse123",
            },
        )
        self.assertEqual(owner_response.status_code, 200)
        owner_headers = {"Authorization": f"Bearer {owner_response.json()['access_token']}"}

        read_only_invitation = self.client.post(
            "/api/organizations/1/invitations",
            headers=owner_headers,
            json={"email": "viewer@example.com", "role": "read_only", "expires_in_days": 7},
        )
        self.assertEqual(read_only_invitation.status_code, 200)
        read_only_response = self.client.post(
            "/api/auth/accept-invitation",
            json={
                "token": read_only_invitation.json()["invitation_token"],
                "full_name": "Viewer User",
                "password": "ViewerPass123",
            },
        )
        self.assertEqual(read_only_response.status_code, 200)
        read_only_headers = {"Authorization": f"Bearer {read_only_response.json()['access_token']}"}

        self.assertEqual(self.client.get("/api/schedule", headers=read_only_headers).status_code, 200)
        denied_create = self.client.post(
            "/api/schedule",
            headers=read_only_headers,
            json={"employee_id": 1, "position_id": 1, "date": "2026-04-20", "shift_template_id": 1},
        )
        self.assertEqual(denied_create.status_code, 403)
        denied_day_status = self.client.post(
            "/api/employee-day-statuses",
            headers=read_only_headers,
            json={"employee_id": 1, "date": "2026-04-20", "status_type": "sick"},
        )
        self.assertEqual(denied_day_status.status_code, 403)

    def test_schedule_entry_status_and_clear_week_flow(self):
        employee_id = self._create_employee()
        position_id = self._create_position()
        template_id = self._create_shift_template(position_id=position_id)

        assignment_response = self.client.post(
            "/api/employee-positions",
            json={
                "employee_id": employee_id,
                "position_id": position_id,
                "is_primary": True,
                "priority_score": 100,
                "is_fallback_only": False,
            },
        )
        self.assertEqual(assignment_response.status_code, 200)

        day_status_response = self.client.post(
            "/api/employee-day-statuses",
            json={
                "employee_id": employee_id,
                "date": "2026-04-20",
                "status_type": "day_off",
            },
        )
        self.assertEqual(day_status_response.status_code, 200)

        create_response = self.client.post(
            "/api/schedule",
            json={
                "employee_id": employee_id,
                "position_id": position_id,
                "date": "2026-04-20",
                "shift_template_id": template_id,
            },
        )
        self.assertEqual(create_response.status_code, 200)
        schedule_entry_id = create_response.json()["schedule_entry"]["id"]

        status_response = self.client.patch(
            f"/api/schedule/{schedule_entry_id}/status",
            json={"no_show": True},
        )
        self.assertEqual(status_response.status_code, 200)
        self.assertTrue(status_response.json()["no_show"])

        schedule_entries = self.client.get("/api/schedule").json()
        self.assertEqual(len(schedule_entries), 1)
        self.assertTrue(schedule_entries[0]["no_show"])

        clear_response = self.client.post(
            "/api/schedule/clear-week",
            json={
                "position_id": position_id,
                "week_start_date": "2026-04-20",
            },
        )
        self.assertEqual(clear_response.status_code, 200)
        self.assertEqual(clear_response.json()["deleted_count"], 1)
        self.assertEqual(self.client.get("/api/schedule").json(), [])
        self.assertEqual(self.client.get("/api/employee-day-statuses").json(), [])

    def test_schedule_entry_time_override_affects_returned_times_and_coverage(self):
        employee_id = self._create_employee(sex="female")
        position_id = self._create_position()
        template_id = self._create_shift_template(
            position_id=position_id,
            start_time="06:30",
            end_time="15:00",
        )

        create_response = self.client.post(
            "/api/schedule",
            json={
                "employee_id": employee_id,
                "position_id": position_id,
                "date": "2026-04-20",
                "shift_template_id": template_id,
            },
        )
        self.assertEqual(create_response.status_code, 200)
        schedule_entry_id = create_response.json()["schedule_entry"]["id"]

        update_response = self.client.patch(
            f"/api/schedule/{schedule_entry_id}/time",
            json={"start_time": "06:30", "end_time": "12:00"},
        )
        self.assertEqual(update_response.status_code, 200)
        updated_entry = update_response.json()["schedule_entry"]
        self.assertEqual(updated_entry["start_time"], "06:30")
        self.assertEqual(updated_entry["end_time"], "12:00")
        self.assertEqual(updated_entry["template_end_time"], "15:00")
        self.assertTrue(updated_entry["time_overridden"])

        entries = self.client.get("/api/schedule").json()
        before_departure_slot = {"start": 7 * 60, "end": 12 * 60, "required_total": 1, "required_female_min": 0, "required_male_min": 0}
        after_departure_slot = {"start": 12 * 60, "end": 15 * 60, "required_total": 1, "required_female_min": 0, "required_male_min": 0}
        self.assertEqual(main.count_slot_coverage(entries, before_departure_slot), (1, 1, 0))
        self.assertEqual(main.count_slot_coverage(entries, after_departure_slot), (0, 0, 0))

        extend_response = self.client.patch(
            f"/api/schedule/{schedule_entry_id}/time",
            json={"start_time": "06:30", "end_time": "20:00"},
        )
        self.assertEqual(extend_response.status_code, 200)
        entries = self.client.get("/api/schedule").json()
        evening_slot = {"start": 15 * 60, "end": 20 * 60, "required_total": 1, "required_female_min": 0, "required_male_min": 0}
        self.assertEqual(main.count_slot_coverage(entries, evening_slot), (1, 1, 0))

        reset_response = self.client.patch(
            f"/api/schedule/{schedule_entry_id}/time",
            json={"reset": True},
        )
        self.assertEqual(reset_response.status_code, 200)
        reset_entry = reset_response.json()["schedule_entry"]
        self.assertEqual(reset_entry["start_time"], "06:30")
        self.assertEqual(reset_entry["end_time"], "15:00")
        self.assertFalse(reset_entry["time_overridden"])

    def test_manual_schedule_edits_keep_day_off_status_in_sync(self):
        employee_id = self._create_employee()
        position_id = self._create_position()
        template_id = self._create_shift_template(position_id=position_id)

        response = self.client.post(
            "/api/employee-positions",
            json={
                "employee_id": employee_id,
                "position_id": position_id,
                "is_primary": True,
                "priority_score": 100,
                "is_fallback_only": False,
            },
        )
        self.assertEqual(response.status_code, 200)

        response = self.client.post(
            "/api/employee-day-statuses",
            json={
                "employee_id": employee_id,
                "date": "2026-04-20",
                "status_type": "day_off",
            },
        )
        self.assertEqual(response.status_code, 200)

        create_response = self.client.post(
            "/api/schedule",
            json={
                "employee_id": employee_id,
                "position_id": position_id,
                "date": "2026-04-20",
                "shift_template_id": template_id,
            },
        )
        self.assertEqual(create_response.status_code, 200)
        schedule_entry_id = create_response.json()["schedule_entry"]["id"]

        day_statuses = self.client.get("/api/employee-day-statuses").json()
        self.assertEqual(day_statuses, [])

        delete_response = self.client.delete(f"/api/schedule/{schedule_entry_id}")
        self.assertEqual(delete_response.status_code, 200)

        day_statuses = self.client.get("/api/employee-day-statuses").json()
        self.assertEqual(len(day_statuses), 1)
        self.assertEqual(day_statuses[0]["employee_id"], employee_id)
        self.assertEqual(day_statuses[0]["date"], "2026-04-20")
        self.assertEqual(day_statuses[0]["status_type"], "day_off")

    def test_manual_edits_after_auto_generation_restore_day_off_for_removed_shift(self):
        employee_id = self._create_employee(full_name="Employee A")
        position_id = self._create_position()
        template_id = self._create_shift_template(position_id=position_id)

        response = self.client.post(
            "/api/employee-positions",
            json={
                "employee_id": employee_id,
                "position_id": position_id,
                "is_primary": True,
                "priority_score": 100,
                "is_fallback_only": False,
            },
        )
        self.assertEqual(response.status_code, 200)

        self._save_shift_requirement(position_id=position_id)

        generate_response = self.client.post(
            "/api/schedule/auto-generate",
            json={"position_id": position_id, "week_start_date": "2026-04-20"},
        )
        self.assertEqual(generate_response.status_code, 200)
        self.assertEqual(generate_response.json()["created_count"], 5)

        generated_entries = self.client.get("/api/schedule").json()
        monday_entry = next(entry for entry in generated_entries if entry["date"] == "2026-04-20")

        delete_response = self.client.delete(f"/api/schedule/{monday_entry['id']}")
        self.assertEqual(delete_response.status_code, 200)

        day_statuses = self.client.get("/api/employee-day-statuses").json()
        monday_status = next(
            status
            for status in day_statuses
            if status["employee_id"] == employee_id and status["date"] == "2026-04-20"
        )
        self.assertEqual(monday_status["status_type"], "day_off")

    def test_app_settings_persist_generation_weights_and_display_mode(self):
        initial_response = self.client.get("/api/app-settings")
        self.assertEqual(initial_response.status_code, 200)
        initial_settings = initial_response.json()
        self.assertEqual(initial_settings["schedule_coverage_display_mode"], "interval")
        self.assertEqual(initial_settings["schedule_morning_color"], "#ecfeff")
        self.assertEqual(initial_settings["schedule_evening_color"], "#fff7ed")
        self.assertEqual(initial_settings["schedule_night_color"], "#eef2ff")
        self.assertEqual(initial_settings["schedule_status_color"], "#f5f3ff")
        self.assertFalse(initial_settings["allow_multiple_positions_per_day"])
        self.assertEqual(initial_settings["max_daily_work_minutes"], 720)
        self.assertEqual(initial_settings["coverage_shortage_gain_weight"], 100)
        self.assertEqual(initial_settings["balance_target_distance_weight"], 70)

        update_response = self.client.put(
            "/api/app-settings",
            json={
                "schedule_coverage_display_mode": "category",
                "schedule_morning_color": "#d1fae5",
                "schedule_evening_color": "#ffedd5",
                "schedule_night_color": "#e0e7ff",
                "schedule_status_color": "#ede9fe",
                "allow_multiple_positions_per_day": True,
                "max_daily_work_minutes": 960,
                "coverage_shortage_gain_weight": 180,
                "balance_target_distance_weight": 95,
                "after_night_evening_penalty": 1600,
            },
        )
        self.assertEqual(update_response.status_code, 200)

        stored_settings = update_response.json()["settings"]
        self.assertEqual(stored_settings["schedule_coverage_display_mode"], "category")
        self.assertEqual(stored_settings["schedule_morning_color"], "#d1fae5")
        self.assertEqual(stored_settings["schedule_evening_color"], "#ffedd5")
        self.assertEqual(stored_settings["schedule_night_color"], "#e0e7ff")
        self.assertEqual(stored_settings["schedule_status_color"], "#ede9fe")
        self.assertTrue(stored_settings["allow_multiple_positions_per_day"])
        self.assertEqual(stored_settings["max_daily_work_minutes"], 960)
        self.assertEqual(stored_settings["coverage_shortage_gain_weight"], 180)
        self.assertEqual(stored_settings["balance_target_distance_weight"], 95)
        self.assertEqual(stored_settings["after_night_evening_penalty"], 1600)

        direct_read = main.get_app_settings(self.connection)
        self.assertEqual(direct_read["schedule_coverage_display_mode"], "category")
        self.assertEqual(direct_read["schedule_morning_color"], "#d1fae5")
        self.assertEqual(direct_read["schedule_evening_color"], "#ffedd5")
        self.assertEqual(direct_read["schedule_night_color"], "#e0e7ff")
        self.assertEqual(direct_read["schedule_status_color"], "#ede9fe")
        self.assertTrue(direct_read["allow_multiple_positions_per_day"])
        self.assertEqual(direct_read["max_daily_work_minutes"], 960)
        self.assertEqual(direct_read["coverage_shortage_gain_weight"], 180)
        self.assertEqual(direct_read["balance_target_distance_weight"], 95)
        self.assertEqual(direct_read["after_night_evening_penalty"], 1600)

        cursor = self.connection.cursor()
        cursor.execute(
            """
            SELECT organization_id, value
            FROM app_settings
            WHERE key = 'schedule_coverage_display_mode'
            """
        )
        coverage_display_row = cursor.fetchone()
        self.assertEqual(coverage_display_row["organization_id"], 1)
        self.assertEqual(coverage_display_row["value"], "category")

        color_position_id = self._create_position(name="Color Test", color="#fde68a")
        reset_response = self.client.post("/api/app-settings/reset-colors")
        self.assertEqual(reset_response.status_code, 200)
        reset_payload = reset_response.json()
        self.assertEqual(reset_payload["settings"]["schedule_morning_color"], "#ecfeff")
        self.assertEqual(reset_payload["settings"]["schedule_evening_color"], "#fff7ed")
        self.assertEqual(reset_payload["settings"]["schedule_night_color"], "#eef2ff")
        self.assertEqual(reset_payload["settings"]["schedule_status_color"], "#f5f3ff")
        self.assertEqual(reset_payload["default_position_color"], "#eff6ff")

        positions = self.client.get("/api/positions").json()
        color_position = next(position for position in positions if position["id"] == color_position_id)
        self.assertEqual(color_position["color"], "#eff6ff")

    def test_schedule_rejects_template_from_another_position(self):
        employee_id = self._create_employee()
        ward_a_id = self._create_position(name="Ward A")
        ward_b_id = self._create_position(name="Ward B")
        template_id = self._create_shift_template(position_id=ward_a_id, name="Ward A Morning")

        response = self.client.post(
            "/api/employee-positions",
            json={
                "employee_id": employee_id,
                "position_id": ward_b_id,
                "is_primary": True,
                "priority_score": 100,
                "is_fallback_only": False,
            },
        )
        self.assertEqual(response.status_code, 200)

        create_response = self.client.post(
            "/api/schedule",
            json={
                "employee_id": employee_id,
                "position_id": ward_b_id,
                "date": "2026-04-20",
                "shift_template_id": template_id,
            },
        )
        self.assertEqual(create_response.status_code, 404)
        self.assertEqual(create_response.json()["detail"], "Shift template not found for this position")

    def test_manual_schedule_allows_cross_position_same_day_shifts(self):
        reset_response = self.client.put(
            "/api/app-settings",
            json={"allow_multiple_positions_per_day": False},
        )
        self.assertEqual(reset_response.status_code, 200)

        employee_id = self._create_employee(full_name="Employee A", max_shifts_per_week=7)
        ward_a_id = self._create_position(name="Ward A")
        ward_b_id = self._create_position(name="Ward B")
        morning_id = self._create_shift_template(position_id=ward_a_id, name="Ward A Morning")
        evening_id = self._create_shift_template(
            position_id=ward_b_id,
            name="Ward B Evening",
            category="evening",
            start_time="14:00",
            end_time="20:00",
        )

        for position_id in (ward_a_id, ward_b_id):
            response = self.client.post(
                "/api/employee-positions",
                json={
                    "employee_id": employee_id,
                    "position_id": position_id,
                    "is_primary": True,
                    "priority_score": 100,
                    "is_fallback_only": False,
                },
            )
            self.assertEqual(response.status_code, 200)

        first_shift = self.client.post(
            "/api/schedule",
            json={
                "employee_id": employee_id,
                "position_id": ward_a_id,
                "date": "2026-04-20",
                "shift_template_id": morning_id,
            },
        )
        self.assertEqual(first_shift.status_code, 200)

        allowed_second_shift = self.client.post(
            "/api/schedule",
            json={
                "employee_id": employee_id,
                "position_id": ward_b_id,
                "date": "2026-04-20",
                "shift_template_id": evening_id,
            },
        )
        self.assertEqual(allowed_second_shift.status_code, 200)

    def test_auto_generate_end_to_end_for_two_positions_fills_full_week(self):
        ward_a_id = self._create_position(name="Ward A")
        ward_b_id = self._create_position(name="Ward B")
        self.assertIsInstance(self._create_shift_template(position_id=ward_a_id), int)
        self.assertIsInstance(self._create_shift_template(position_id=ward_b_id), int)

        employee_ids = [
            self._create_employee(full_name="Employee A"),
            self._create_employee(full_name="Employee B"),
            self._create_employee(full_name="Employee C"),
            self._create_employee(full_name="Employee D"),
        ]

        for employee_id in employee_ids[:2]:
            response = self.client.post(
                "/api/employee-positions",
                json={
                    "employee_id": employee_id,
                    "position_id": ward_a_id,
                    "is_primary": True,
                    "priority_score": 100,
                    "is_fallback_only": False,
                },
            )
            self.assertEqual(response.status_code, 200)

        for employee_id in employee_ids[2:]:
            response = self.client.post(
                "/api/employee-positions",
                json={
                    "employee_id": employee_id,
                    "position_id": ward_b_id,
                    "is_primary": True,
                    "priority_score": 100,
                    "is_fallback_only": False,
                },
            )
            self.assertEqual(response.status_code, 200)

        self._save_shift_requirement(position_id=ward_a_id)
        self._save_shift_requirement(position_id=ward_b_id)

        generate_a = self.client.post(
            "/api/schedule/auto-generate",
            json={"position_id": ward_a_id, "week_start_date": "2026-04-20"},
        )
        self.assertEqual(generate_a.status_code, 200)

        generate_b = self.client.post(
            "/api/schedule/auto-generate",
            json={"position_id": ward_b_id, "week_start_date": "2026-04-20"},
        )
        self.assertEqual(generate_b.status_code, 200)

        for response in (generate_a, generate_b):
            payload = response.json()
            self.assertEqual(payload["created_count"], 7)
            self.assertEqual(payload["feasibility_report"]["status"], "ok")
            self.assertEqual(payload["unfilled_reports"], [])

        schedule_entries = self.client.get("/api/schedule").json()
        self.assertEqual(len(schedule_entries), 14)
        self.assertEqual({entry["position_id"] for entry in schedule_entries}, {ward_a_id, ward_b_id})
        self.assertEqual({entry["date"] for entry in schedule_entries}, {
            "2026-04-20",
            "2026-04-21",
            "2026-04-22",
            "2026-04-23",
            "2026-04-24",
            "2026-04-25",
            "2026-04-26",
        })

    def test_single_position_auto_generate_does_not_mark_empty_days_as_day_off(self):
        generated_position_id = self._create_position(name="Generated Ward")
        secondary_position_id = self._create_position(name="Secondary Ward")
        self._create_shift_template(position_id=generated_position_id, name="Generated Morning")

        employee_id = self._create_employee(
            full_name="Shared Employee",
            min_shifts_per_week=0,
            target_shifts_per_week=1,
            max_shifts_per_week=1,
        )
        for position_id, is_primary, priority_score in (
            (generated_position_id, True, 100),
            (secondary_position_id, False, 50),
        ):
            response = self.client.post(
                "/api/employee-positions",
                json={
                    "employee_id": employee_id,
                    "position_id": position_id,
                    "is_primary": is_primary,
                    "priority_score": priority_score,
                    "is_fallback_only": False,
                },
            )
            self.assertEqual(response.status_code, 200)

        self._save_shift_requirement(position_id=generated_position_id)

        generate_response = self.client.post(
            "/api/schedule/auto-generate",
            json={"position_id": generated_position_id, "week_start_date": "2026-04-20"},
        )
        self.assertEqual(generate_response.status_code, 200)
        payload = generate_response.json()
        self.assertEqual(payload["created_count"], 1)
        self.assertEqual(payload["day_off_count"], 0)

        day_statuses = self.client.get("/api/employee-day-statuses").json()
        self.assertEqual(day_statuses, [])

    def test_auto_generate_preserves_existing_manual_week_schedule(self):
        position_id = self._create_position(name="Replace Ward")
        template_id = self._create_shift_template(position_id=position_id, name="Morning")
        employee_id = self._create_employee(full_name="Employee A")
        response = self.client.post(
            "/api/employee-positions",
            json={
                "employee_id": employee_id,
                "position_id": position_id,
                "is_primary": True,
                "priority_score": 100,
                "is_fallback_only": False,
            },
        )
        self.assertEqual(response.status_code, 200)
        self._save_shift_requirement(position_id=position_id)

        create_response = self.client.post(
            "/api/schedule",
            json={
                "employee_id": employee_id,
                "position_id": position_id,
                "date": "2026-04-20",
                "shift_template_id": template_id,
            },
        )
        self.assertEqual(create_response.status_code, 200)
        existing_entry_id = create_response.json()["schedule_entry"]["id"]

        generate_response = self.client.post(
            "/api/schedule/auto-generate",
            json={"position_id": position_id, "week_start_date": "2026-04-20"},
        )
        self.assertEqual(generate_response.status_code, 200)
        payload = generate_response.json()
        self.assertEqual(payload["created_count"], 4)

        schedule_entries = self.client.get("/api/schedule").json()
        self.assertIn(existing_entry_id, {entry["id"] for entry in schedule_entries})
        self.assertEqual(len(schedule_entries), 5)
        manual_entry = next(entry for entry in schedule_entries if entry["id"] == existing_entry_id)
        self.assertEqual(manual_entry["date"], "2026-04-20")
        self.assertEqual(manual_entry["employee_id"], employee_id)

    def test_auto_generate_respects_strict_employee_card_preferences(self):
        position_id = self._create_position(name="Strict Preference Ward")
        self._create_shift_template(position_id=position_id, name="Morning")

        strict_employee_id = self._create_employee(
            full_name="Strict Employee",
            min_shifts_per_week=0,
            target_shifts_per_week=3,
            max_shifts_per_week=5,
        )
        backup_employee_id = self._create_employee(
            full_name="Backup Employee",
            min_shifts_per_week=0,
            target_shifts_per_week=3,
            max_shifts_per_week=5,
        )

        for assignment in (
            {
                "employee_id": strict_employee_id,
                "position_id": position_id,
                "is_primary": True,
                "priority_score": 100,
                "is_fallback_only": False,
            },
            {
                "employee_id": backup_employee_id,
                "position_id": position_id,
                "is_primary": False,
                "priority_score": 10,
                "is_fallback_only": False,
            },
        ):
            response = self.client.post("/api/employee-positions", json=assignment)
            self.assertEqual(response.status_code, 200)

        recurring_response = self.client.post(
            "/api/employee-recurring-preferences",
            json={
                "employee_id": strict_employee_id,
                "rules": [
                    {
                        "preference_kind": "strict",
                        "day_of_week": 1,
                        "request_type": "request_shift",
                        "target_category": "night",
                    },
                    {
                        "preference_kind": "strict",
                        "day_of_week": 1,
                        "request_type": "exclude_shift",
                        "target_category": "morning",
                    }
                ],
            },
        )
        self.assertEqual(recurring_response.status_code, 200)
        self._save_shift_requirement(position_id=position_id)

        generate_response = self.client.post(
            "/api/schedule/auto-generate",
            json={"position_id": position_id, "week_start_date": "2026-04-19"},
        )
        self.assertEqual(generate_response.status_code, 200)

        monday_entries = [
            entry
            for entry in self.client.get("/api/schedule").json()
            if entry["date"] == "2026-04-20"
        ]
        self.assertEqual(len(monday_entries), 1)
        self.assertEqual(monday_entries[0]["employee_id"], backup_employee_id)

    def test_auto_generate_all_positions_returns_successes_and_failures(self):
        ward_a_id = self._create_position(name="Ward A")
        ward_b_id = self._create_position(name="Ward B")
        ward_c_id = self._create_position(name="Ward C")

        self._create_shift_template(position_id=ward_a_id, name="Ward A Morning")
        self._create_shift_template(position_id=ward_b_id, name="Ward B Morning")

        employee_a = self._create_employee(full_name="Employee A")
        employee_b = self._create_employee(full_name="Employee B")

        for employee_id, position_id in ((employee_a, ward_a_id), (employee_b, ward_b_id), (employee_a, ward_c_id)):
            response = self.client.post(
                "/api/employee-positions",
                json={
                    "employee_id": employee_id,
                    "position_id": position_id,
                    "is_primary": True,
                    "priority_score": 100,
                    "is_fallback_only": False,
                },
            )
            self.assertEqual(response.status_code, 200)

        self._save_shift_requirement(position_id=ward_a_id)
        self._save_shift_requirement(position_id=ward_b_id)
        self._save_shift_requirement(position_id=ward_c_id)

        response = self.client.post(
            "/api/schedule/auto-generate-all",
            json={"week_start_date": "2026-04-20"},
        )
        self.assertEqual(response.status_code, 200)

        payload = response.json()
        self.assertEqual(payload["generated_positions"], 2)
        self.assertEqual(payload["failed_positions"], 1)
        self.assertEqual(len(payload["results"]), 2)
        self.assertEqual(len(payload["failures"]), 1)
        self.assertEqual(payload["failures"][0]["position_id"], ward_c_id)
        self.assertEqual(payload["failures"][0]["detail"], "No active shift templates found")

        schedule_entries = self.client.get("/api/schedule").json()
        self.assertEqual({entry["position_id"] for entry in schedule_entries}, {ward_a_id, ward_b_id})

    def test_auto_generate_all_prioritizes_primary_position_before_secondary_assignment(self):
        settings_response = self.client.put("/api/app-settings", json={"max_work_days_per_week": 7})
        self.assertEqual(settings_response.status_code, 200)

        secondary_position_id = self._create_position(name="Secondary Ward")
        primary_position_id = self._create_position(name="Primary Ward")
        self._create_shift_template(
            position_id=secondary_position_id,
            name="Secondary Evening",
            category="evening",
            start_time="14:00",
            end_time="20:00",
        )
        self._create_shift_template(position_id=primary_position_id, name="Primary Morning")

        shared_employee_id = self._create_employee(
            full_name="Shared Employee",
            min_shifts_per_week=0,
            target_shifts_per_week=7,
            max_shifts_per_week=7,
        )
        backup_employee_id = self._create_employee(
            full_name="Backup Employee",
            min_shifts_per_week=0,
            target_shifts_per_week=7,
            max_shifts_per_week=7,
        )

        for assignment in (
            {
                "employee_id": shared_employee_id,
                "position_id": secondary_position_id,
                "is_primary": False,
                "priority_score": 10,
                "is_fallback_only": False,
            },
            {
                "employee_id": shared_employee_id,
                "position_id": primary_position_id,
                "is_primary": True,
                "priority_score": 100,
                "is_fallback_only": False,
            },
            {
                "employee_id": backup_employee_id,
                "position_id": primary_position_id,
                "is_primary": False,
                "priority_score": 50,
                "is_fallback_only": False,
            },
        ):
            response = self.client.post("/api/employee-positions", json=assignment)
            self.assertEqual(response.status_code, 200)

        self._save_shift_requirement(
            position_id=secondary_position_id,
            shift_category="evening",
        )
        self._save_shift_requirement(position_id=primary_position_id)

        response = self.client.post(
            "/api/schedule/auto-generate-all",
            json={"week_start_date": "2026-04-20"},
        )
        self.assertEqual(response.status_code, 200)

        schedule_entries = self.client.get("/api/schedule").json()
        primary_entries = [entry for entry in schedule_entries if entry["position_id"] == primary_position_id]
        secondary_entries = [entry for entry in schedule_entries if entry["position_id"] == secondary_position_id]

        self.assertEqual(len(primary_entries), 7)
        self.assertEqual({entry["employee_name"] for entry in primary_entries}, {"Shared Employee"})
        self.assertEqual(secondary_entries, [])

    def test_auto_generate_spreads_night_shifts_evenly_between_night_staff(self):
        position_id = self._create_position(name="Night Ward")
        night_template_id = self._create_shift_template(
            position_id=position_id,
            name="Night",
            category="night",
            start_time="23:00",
            end_time="07:00",
            is_overnight=True,
        )
        self.assertIsInstance(night_template_id, int)
        employee_ids = [
            self._create_employee(full_name="Vaheed", sex="male", target_shifts_per_week=3, max_shifts_per_week=7),
            self._create_employee(full_name="Employee B", sex="female", target_shifts_per_week=3, max_shifts_per_week=7),
            self._create_employee(full_name="Employee C", sex="male", target_shifts_per_week=3, max_shifts_per_week=7),
        ]

        for employee_id in employee_ids:
            response = self.client.post(
                "/api/employee-positions",
                json={
                    "employee_id": employee_id,
                    "position_id": position_id,
                    "is_primary": True,
                    "priority_score": 100,
                    "is_fallback_only": False,
                },
            )
            self.assertEqual(response.status_code, 200)

        self._save_shift_requirement(
            position_id=position_id,
            shift_category="night",
            required_total=1,
            required_female_min=0,
            required_male_min=0,
        )

        generate_response = self.client.post(
            "/api/schedule/auto-generate",
            json={"position_id": position_id, "week_start_date": "2026-04-20"},
        )
        self.assertEqual(generate_response.status_code, 200)
        self.assertEqual(generate_response.json()["unfilled_reports"], [])

        schedule_entries = self.client.get(
            "/api/schedule",
            params={"position_id": position_id, "week_start_date": "2026-04-20"},
        ).json()

        self.assertEqual(len(schedule_entries), 7)
        self.assertTrue(all(entry["shift_category"] == "night" for entry in schedule_entries))

        counts: dict[str, int] = {}
        for entry in schedule_entries:
            counts[entry["employee_name"]] = counts.get(entry["employee_name"], 0) + 1

        self.assertEqual(set(counts.keys()), {"Vaheed", "Employee B", "Employee C"})
        self.assertLessEqual(max(counts.values()) - min(counts.values()), 1)

    def test_auto_generate_avoids_locking_same_shift_type_to_one_employee(self):
        position_id = self._create_position(name="Mixed Ward")
        self._create_shift_template(position_id=position_id, name="Morning", category="morning", start_time="06:00", end_time="14:00")
        self._create_shift_template(position_id=position_id, name="Evening", category="evening", start_time="14:00", end_time="20:00")
        employee_ids = [
            self._create_employee(full_name="Employee A", target_shifts_per_week=4, max_shifts_per_week=7),
            self._create_employee(full_name="Employee B", target_shifts_per_week=4, max_shifts_per_week=7),
        ]

        for employee_id in employee_ids:
            response = self.client.post(
                "/api/employee-positions",
                json={
                    "employee_id": employee_id,
                    "position_id": position_id,
                    "is_primary": True,
                    "priority_score": 100,
                    "is_fallback_only": False,
                },
            )
            self.assertEqual(response.status_code, 200)

        self._save_shift_requirement(
            position_id=position_id,
            shift_category="morning",
            required_total=1,
            required_female_min=0,
            required_male_min=0,
        )
        self._save_shift_requirement(
            position_id=position_id,
            shift_category="evening",
            required_total=1,
            required_female_min=0,
            required_male_min=0,
        )

        generate_response = self.client.post(
            "/api/schedule/auto-generate",
            json={"position_id": position_id, "week_start_date": "2026-04-20"},
        )
        self.assertEqual(generate_response.status_code, 200)
        self.assertEqual(len(generate_response.json()["unfilled_reports"]), 2)

        schedule_entries = self.client.get(
            "/api/schedule",
            params={"position_id": position_id, "week_start_date": "2026-04-20"},
        ).json()

        self.assertEqual(len(schedule_entries), 12)

        counts: dict[str, dict[str, int]] = {}
        for entry in schedule_entries:
            counts.setdefault(entry["employee_name"], {"morning": 0, "evening": 0, "night": 0})
            counts[entry["employee_name"]][entry["shift_category"]] += 1

        for employee_name in ("Employee A", "Employee B"):
            self.assertGreater(counts[employee_name]["morning"], 0)
            self.assertGreater(counts[employee_name]["evening"], 0)

    def test_export_excel_for_empty_week_has_headers_and_zero_total(self):
        employee_id = self._create_employee(full_name="Employee A")
        position_id = self._create_position()
        response = self.client.post(
            "/api/employee-positions",
            json={
                "employee_id": employee_id,
                "position_id": position_id,
                "is_primary": True,
                "priority_score": 100,
                "is_fallback_only": False,
            },
        )
        self.assertEqual(response.status_code, 200)

        export_response = self.client.get(
            "/api/schedule/export-excel",
            params={"week_start_date": "2026-04-20", "position_id": position_id, "lang": "en"},
        )
        self.assertEqual(export_response.status_code, 200)

        workbook = load_workbook(filename=main.BytesIO(export_response.content))
        worksheet = workbook.active
        self.assertEqual(worksheet["A1"].value, "Schedule export - Nurse - week starting 2026-04-20")
        self.assertEqual(worksheet["A3"].value, "Employee")
        self.assertEqual(worksheet["I5"].value, "0 / 0h")

    def test_schedule_export_writes_audit_event_when_authenticated(self):
        owner_response = self.client.post(
            "/api/auth/bootstrap",
            json={
                "organization_name": "Beta Clinic",
                "full_name": "Owner User",
                "email": "owner@example.com",
                "password": "CorrectHorse123",
            },
        )
        self.assertEqual(owner_response.status_code, 200)
        owner_token = owner_response.json()["access_token"]
        position_id = self._create_position(headers={"Authorization": f"Bearer {owner_token}"})

        export_response = self.client.get(
            "/api/schedule/export-excel",
            headers={"Authorization": f"Bearer {owner_token}"},
            params={"week_start_date": "2026-04-20", "position_id": position_id, "lang": "en"},
        )
        self.assertEqual(export_response.status_code, 200)

        cursor = self.connection.cursor()
        cursor.execute(
            """
            SELECT user_id, organization_id, metadata_json
            FROM auth_audit_events
            WHERE event_type = 'schedule_exported'
            """
        )
        event = cursor.fetchone()
        self.assertIsNotNone(event)
        self.assertEqual(event["user_id"], owner_response.json()["user"]["id"])
        self.assertEqual(event["organization_id"], 1)
        metadata = main.json.loads(event["metadata_json"])
        self.assertEqual(metadata["format"], "excel")
        self.assertEqual(metadata["scope"], "position")
        self.assertEqual(metadata["position_id"], position_id)

    def test_export_word_for_empty_week_has_headers_and_zero_total(self):
        employee_id = self._create_employee(full_name="Employee A")
        position_id = self._create_position()
        response = self.client.post(
            "/api/employee-positions",
            json={
                "employee_id": employee_id,
                "position_id": position_id,
                "is_primary": True,
                "priority_score": 100,
                "is_fallback_only": False,
            },
        )
        self.assertEqual(response.status_code, 200)

        export_response = self.client.get(
            "/api/schedule/export-word",
            params={"week_start_date": "2026-04-20", "position_id": position_id, "lang": "en"},
        )
        self.assertEqual(export_response.status_code, 200)
        self.assertEqual(
            export_response.headers["content-type"],
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        )

        with ZipFile(main.BytesIO(export_response.content)) as document:
            document_xml = document.read("word/document.xml").decode("utf-8")
        self.assertIn("<w:tblPr>", document_xml)
        self.assertIn("<w:tblGrid>", document_xml)
        self.assertIn("<w:tblBorders>", document_xml)
        self.assertIn("<w:tblHeader/>", document_xml)
        self.assertIn("<w:tblLayout w:type=\"fixed\"/>", document_xml)
        self.assertIn("Schedule export - Nurse - week starting 2026-04-20", document_xml)
        self.assertIn("Employee A", document_xml)
        self.assertIn("0 / 0h", document_xml)

    def test_export_word_keeps_shift_no_show_and_day_off_inside_table(self):
        employee_id = self._create_employee(full_name="Employee A")
        position_id = self._create_position()
        morning_id = self._create_shift_template(position_id=position_id, name="Morning")
        response = self.client.post(
            "/api/employee-positions",
            json={
                "employee_id": employee_id,
                "position_id": position_id,
                "is_primary": True,
                "priority_score": 100,
                "is_fallback_only": False,
            },
        )
        self.assertEqual(response.status_code, 200)

        response = self.client.post(
            "/api/schedule",
            json={
                "employee_id": employee_id,
                "position_id": position_id,
                "date": "2026-04-20",
                "shift_template_id": morning_id,
            },
        )
        self.assertEqual(response.status_code, 200)
        entry_id = response.json()["schedule_entry"]["id"]
        response = self.client.patch(f"/api/schedule/{entry_id}/status", json={"no_show": True})
        self.assertEqual(response.status_code, 200)

        response = self.client.post(
            "/api/employee-day-statuses",
            json={
                "employee_id": employee_id,
                "date": "2026-04-21",
                "status_type": "day_off",
            },
        )
        self.assertEqual(response.status_code, 200)

        export_response = self.client.get(
            "/api/schedule/export-word",
            params={"week_start_date": "2026-04-20", "position_id": position_id, "lang": "en"},
        )
        self.assertEqual(export_response.status_code, 200)

        with ZipFile(main.BytesIO(export_response.content)) as document:
            document_xml = document.read("word/document.xml").decode("utf-8")
        first_table = document_xml[document_xml.index("<w:tbl>"):document_xml.index("</w:tbl>")]
        self.assertIn("Morning - No-show", first_table)
        self.assertIn("Day off", first_table)
        self.assertIn("<w:tcW w:w=\"2200\" w:type=\"dxa\"/>", first_table)
        self.assertIn("<w:tcW w:w=\"1600\" w:type=\"dxa\"/>", first_table)
        self.assertIn("<w:tcW w:w=\"1998\" w:type=\"dxa\"/>", first_table)

    def test_export_excel_includes_shift_no_show_and_day_off_labels(self):
        employee_id = self._create_employee(full_name="Employee A")
        position_id = self._create_position()
        morning_id = self._create_shift_template(position_id=position_id, name="Morning")
        evening_id = self._create_shift_template(
            position_id=position_id,
            name="Evening",
            category="evening",
            start_time="14:00",
            end_time="20:00",
        )
        night_id = self._create_shift_template(
            position_id=position_id,
            name="Night",
            category="night",
            start_time="20:00",
            end_time="07:00",
            is_overnight=True,
        )

        response = self.client.put("/api/app-settings", json={"allow_multiple_positions_per_day": True})
        self.assertEqual(response.status_code, 200)

        response = self.client.post(
            "/api/employee-positions",
            json={
                "employee_id": employee_id,
                "position_id": position_id,
                "is_primary": True,
                "priority_score": 100,
                "is_fallback_only": False,
            },
        )
        self.assertEqual(response.status_code, 200)

        response = self.client.post(
            "/api/schedule",
            json={
                "employee_id": employee_id,
                "position_id": position_id,
                "date": "2026-04-20",
                "shift_template_id": morning_id,
            },
        )
        self.assertEqual(response.status_code, 200)
        entry_id = response.json()["schedule_entry"]["id"]

        response = self.client.post(
            "/api/schedule",
            json={
                "employee_id": employee_id,
                "position_id": position_id,
                "date": "2026-04-21",
                "shift_template_id": evening_id,
            },
        )
        self.assertEqual(response.status_code, 200)
        no_show_entry_id = response.json()["schedule_entry"]["id"]

        response = self.client.patch(f"/api/schedule/{no_show_entry_id}/status", json={"no_show": True})
        self.assertEqual(response.status_code, 200)

        response = self.client.post(
            "/api/employee-day-statuses",
            json={
                "employee_id": employee_id,
                "date": "2026-04-22",
                "status_type": "day_off",
            },
        )
        self.assertEqual(response.status_code, 200)

        other_position_id = self._create_position(name="Reception", color="#fde68a")
        other_template_id = self._create_shift_template(position_id=other_position_id, name="Other Morning")
        response = self.client.post(
            "/api/employee-positions",
            json={
                "employee_id": employee_id,
                "position_id": other_position_id,
                "is_primary": False,
                "priority_score": 50,
                "is_fallback_only": False,
            },
        )
        self.assertEqual(response.status_code, 200)
        response = self.client.post(
            "/api/schedule",
            json={
                "employee_id": employee_id,
                "position_id": other_position_id,
                "date": "2026-04-23",
                "shift_template_id": other_template_id,
            },
        )
        self.assertEqual(response.status_code, 200)
        response = self.client.post(
            "/api/schedule",
            json={
                "employee_id": employee_id,
                "position_id": position_id,
                "date": "2026-04-23",
                "shift_template_id": night_id,
            },
        )
        self.assertEqual(response.status_code, 200)

        response = self.client.post(
            "/api/schedule",
            json={
                "employee_id": employee_id,
                "position_id": position_id,
                "date": "2026-04-24",
                "shift_template_id": morning_id,
            },
        )
        self.assertEqual(response.status_code, 200)
        response = self.client.post(
            "/api/employee-day-statuses",
            json={
                "employee_id": employee_id,
                "date": "2026-04-24",
                "status_type": "sick",
            },
        )
        self.assertEqual(response.status_code, 200)

        export_response = self.client.get(
            "/api/schedule/export-excel",
            params={"week_start_date": "2026-04-20", "position_id": position_id, "lang": "ru"},
        )
        self.assertEqual(export_response.status_code, 200)

        workbook = load_workbook(filename=main.BytesIO(export_response.content))
        worksheet = workbook.active
        self.assertEqual(worksheet["B5"].value, "Morning")
        self.assertIn("B5:B6", {str(cell_range) for cell_range in worksheet.merged_cells.ranges})
        self.assertEqual(worksheet["C5"].value, "Evening - Неявка")
        self.assertIn("C5:C6", {str(cell_range) for cell_range in worksheet.merged_cells.ranges})
        self.assertEqual(worksheet["D5"].value, "Выходной")
        self.assertIn("D5:D6", {str(cell_range) for cell_range in worksheet.merged_cells.ranges})
        self.assertEqual(worksheet["E5"].value, "Reception: Other Morning")
        self.assertEqual(worksheet["E6"].value, "Night")
        self.assertEqual(worksheet["E5"].font.color.rgb, "FF000000")
        self.assertEqual(worksheet["E5"].fill.fgColor.rgb, "FFFDE68A")
        self.assertEqual(worksheet["E6"].fill.fill_type, None)
        self.assertEqual(worksheet["F5"].value, "Morning")
        self.assertEqual(worksheet["F6"].value, "Больничный")
        self.assertEqual(worksheet["I5"].value, "3 / 27ч")
        self.assertIn("A5:A6", {str(cell_range) for cell_range in worksheet.merged_cells.ranges})
        self.assertIn("I5:I6", {str(cell_range) for cell_range in worksheet.merged_cells.ranges})
        self.assertEqual(worksheet["A5"].border.left.style, "medium")
        self.assertEqual(worksheet["A5"].border.top.style, "medium")
        self.assertEqual(worksheet["I5"].border.right.style, "medium")
        self.assertEqual(worksheet["E6"].border.bottom.style, "medium")
        summary_sheet = workbook["Сводка координатора"]
        self.assertEqual(summary_sheet["A1"].value, "Сводка координатора")
        self.assertEqual(summary_sheet["B3"].value, "Nurse")
        self.assertEqual(summary_sheet["B6"].value, 5)
        self.assertEqual(summary_sheet["B8"].value, 1)
        self.assertEqual(summary_sheet["B9"].value, 1)
        self.assertEqual(summary_sheet["B10"].value, 1)

    def test_delete_impact_endpoints_report_related_records(self):
        employee_id = self._create_employee(full_name="Employee A")
        position_id = self._create_position()
        template_id = self._create_shift_template(position_id=position_id, name="Morning")

        response = self.client.post(
            "/api/employee-positions",
            json={
                "employee_id": employee_id,
                "position_id": position_id,
                "is_primary": True,
                "priority_score": 100,
                "is_fallback_only": False,
            },
        )
        self.assertEqual(response.status_code, 200)

        response = self.client.post(
            "/api/employee-preferences",
            json={
                "employee_id": employee_id,
                "allow_morning": True,
                "allow_evening": True,
                "allow_night": True,
                "allow_morning_evening_combo": True,
            },
        )
        self.assertEqual(response.status_code, 200)

        response = self.client.post(
            "/api/employee-week-preferences",
            json={
                "employee_id": employee_id,
                "week_start_date": "2026-04-20",
                "preference_date": "2026-04-21",
                "preference_type": "vacation",
            },
        )
        self.assertEqual(response.status_code, 200)

        response = self.client.post(
            "/api/employee-recurring-preferences",
            json={
                "employee_id": employee_id,
                "rules": [
                    {
                        "preference_kind": "strict",
                        "day_of_week": 0,
                        "preference_type": "only_night",
                    }
                ],
            },
        )
        self.assertEqual(response.status_code, 200)

        response = self.client.post(
            "/api/employee-day-statuses",
            json={
                "employee_id": employee_id,
                "date": "2026-04-22",
                "status_type": "day_off",
            },
        )
        self.assertEqual(response.status_code, 200)

        response = self.client.post(
            "/api/shift-requirements",
            json={
                "position_id": position_id,
                "shift_category": "morning",
                "required_total": 1,
                "required_female_min": 0,
                "required_male_min": 0,
            },
        )
        self.assertEqual(response.status_code, 200)

        response = self.client.post(
            "/api/coverage-requirements",
            json={
                "position_id": position_id,
                "start_time": "06:00",
                "end_time": "14:00",
                "required_total": 1,
                "required_female_min": 0,
                "required_male_min": 0,
                "is_overnight": False,
            },
        )
        self.assertEqual(response.status_code, 200)

        response = self.client.post(
            "/api/schedule",
            json={
                "employee_id": employee_id,
                "position_id": position_id,
                "date": "2026-04-20",
                "shift_template_id": template_id,
            },
        )
        self.assertEqual(response.status_code, 200)

        employee_impact = self.client.get(f"/api/employees/{employee_id}/delete-impact")
        self.assertEqual(employee_impact.status_code, 200)
        self.assertEqual(
            employee_impact.json(),
            {
                "employee_id": employee_id,
                "employee_name": "Employee A",
                "assignments": 1,
                "schedule_entries": 1,
                "general_preferences": 1,
                "weekly_preferences": 1,
                "recurring_preferences": 1,
                "day_statuses": 1,
            },
        )

        position_impact = self.client.get(f"/api/positions/{position_id}/delete-impact")
        self.assertEqual(position_impact.status_code, 200)
        self.assertEqual(position_impact.json()["position_name"], "Nurse")
        self.assertEqual(position_impact.json()["assignments"], 1)
        self.assertEqual(position_impact.json()["schedule_entries"], 1)
        self.assertEqual(position_impact.json()["shift_requirements"], 1)
        self.assertEqual(position_impact.json()["coverage_requirements"], 1)

        assignment_impact = self.client.get(
            "/api/employee-positions/delete-impact",
            params={"employee_id": employee_id, "position_id": position_id},
        )
        self.assertEqual(assignment_impact.status_code, 200)
        self.assertEqual(assignment_impact.json()["employee_name"], "Employee A")
        self.assertEqual(assignment_impact.json()["position_name"], "Nurse")
        self.assertEqual(assignment_impact.json()["schedule_entries"], 1)

        template_impact = self.client.get(f"/api/shift-templates/{template_id}/delete-impact")
        self.assertEqual(template_impact.status_code, 200)
        self.assertEqual(template_impact.json()["template_name"], "Morning")
        self.assertEqual(template_impact.json()["position_name"], "Nurse")
        self.assertEqual(template_impact.json()["schedule_entries"], 1)

    def test_clear_week_preview_reports_entries_and_day_off_statuses(self):
        employee_id = self._create_employee(full_name="Employee A")
        position_id = self._create_position(name="Ward A")
        template_id = self._create_shift_template(position_id=position_id)

        response = self.client.post(
            "/api/employee-positions",
            json={
                "employee_id": employee_id,
                "position_id": position_id,
                "is_primary": True,
                "priority_score": 100,
                "is_fallback_only": False,
            },
        )
        self.assertEqual(response.status_code, 200)

        response = self.client.post(
            "/api/schedule",
            json={
                "employee_id": employee_id,
                "position_id": position_id,
                "date": "2026-04-20",
                "shift_template_id": template_id,
            },
        )
        self.assertEqual(response.status_code, 200)

        response = self.client.post(
            "/api/employee-day-statuses",
            json={
                "employee_id": employee_id,
                "date": "2026-04-21",
                "status_type": "day_off",
            },
        )
        self.assertEqual(response.status_code, 200)

        preview_response = self.client.get(
            "/api/schedule/clear-week-preview",
            params={"position_id": position_id, "week_start_date": "2026-04-20"},
        )
        self.assertEqual(preview_response.status_code, 200)
        self.assertEqual(
            preview_response.json(),
            {
                "position_id": position_id,
                "position_name": "Ward A",
                "week_start_date": "2026-04-20",
                "assigned_employees": 1,
                "schedule_entries": 1,
                "day_off_statuses": 1,
            },
        )

    def test_clear_week_actions_return_recovery_backup_names(self):
        employee_id = self._create_employee(full_name="Employee A")
        position_id = self._create_position(name="Ward A")
        template_id = self._create_shift_template(position_id=position_id)
        response = self.client.post(
            "/api/employee-positions",
            json={
                "employee_id": employee_id,
                "position_id": position_id,
                "is_primary": True,
                "priority_score": 100,
                "is_fallback_only": False,
            },
        )
        self.assertEqual(response.status_code, 200)

        response = self.client.post(
            "/api/schedule",
            json={
                "employee_id": employee_id,
                "position_id": position_id,
                "date": "2026-04-20",
                "shift_template_id": template_id,
            },
        )
        self.assertEqual(response.status_code, 200)
        response = self.client.post(
            "/api/employee-day-statuses",
            json={
                "employee_id": employee_id,
                "date": "2026-04-21",
                "status_type": "day_off",
            },
        )
        self.assertEqual(response.status_code, 200)

        clear_response = self.client.post(
            "/api/schedule/clear-week",
            json={"position_id": position_id, "week_start_date": "2026-04-20"},
        )
        self.assertEqual(clear_response.status_code, 200)
        self.assertEqual(clear_response.json()["deleted_count"], 1)
        self.assertTrue(clear_response.json()["backup_name"])

        response = self.client.post(
            "/api/schedule",
            json={
                "employee_id": employee_id,
                "position_id": position_id,
                "date": "2026-04-22",
                "shift_template_id": template_id,
            },
        )
        self.assertEqual(response.status_code, 200)
        response = self.client.post(
            "/api/employee-day-statuses",
            json={
                "employee_id": employee_id,
                "date": "2026-04-23",
                "status_type": "day_off",
            },
        )
        self.assertEqual(response.status_code, 200)

        clear_all_response = self.client.post(
            "/api/schedule/clear-week-all",
            json={"week_start_date": "2026-04-20"},
        )
        self.assertEqual(clear_all_response.status_code, 200)
        self.assertEqual(clear_all_response.json()["deleted_count"], 1)
        self.assertEqual(clear_all_response.json()["day_off_deleted_count"], 1)
        self.assertTrue(clear_all_response.json()["backup_name"])

    def test_get_base_path_uses_meipass_in_frozen_mode(self):
        with patch.object(main.sys, "frozen", True, create=True), patch.object(
            main.sys,
            "_MEIPASS",
            str(Path("D:/fake_bundle")),
            create=True,
        ):
            self.assertEqual(main.get_base_path(), Path("D:/fake_bundle"))

    def test_frozen_database_path_uses_local_app_data(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            local_app_data = temp_path / "LocalAppData"
            bundled_dir = temp_path / "bundle"
            install_dir = temp_path / "Program Files" / "Schedule App"
            bundled_dir.mkdir(parents=True)
            install_dir.mkdir(parents=True)

            bundled_database = bundled_dir / "schedule_app.db"
            connection = sqlite3.connect(bundled_database)
            connection.execute("CREATE TABLE marker (id INTEGER PRIMARY KEY)")
            connection.close()

            with (
                patch.dict(os.environ, {"LOCALAPPDATA": str(local_app_data)}),
                patch.object(sys, "frozen", True, create=True),
                patch.object(sys, "_MEIPASS", str(bundled_dir), create=True),
                patch.object(sys, "executable", str(install_dir / "ScheduleApp.exe")),
            ):
                runtime_path = database.get_database_path()

            expected_path = local_app_data / "Schedule App" / "schedule_app.db"
            self.assertEqual(runtime_path, expected_path)
            self.assertTrue(runtime_path.exists())
            self.assertFalse((install_dir / "schedule_app.db").exists())

    def test_update_check_reports_newer_github_release_asset(self):
        releases = [
            {
                "tag_name": "v0.13.6-beta",
                "name": "Schedule App 0.13.6-beta",
                "draft": False,
                "prerelease": True,
                "body": "## What Changed\n- Startup update modal\n- Post-install changelog",
                "published_at": "2026-04-26T00:00:00Z",
                "html_url": "https://github.com/LittleDespairs/Schedule_app/releases/tag/v0.13.6-beta",
                "assets": [
                    {
                        "name": "ScheduleApp_Setup_0.13.6-beta.exe",
                        "browser_download_url": "https://github.com/LittleDespairs/Schedule_app/releases/download/v0.13.6-beta/ScheduleApp_Setup_0.13.6-beta.exe",
                        "size": 123,
                    }
                ],
            }
        ]

        with patch.object(main, "APP_VERSION", "0.13.5_beta"), patch.object(main, "request_github_releases", return_value=releases):
            payload = main.get_update_status()

        self.assertTrue(payload["update_available"])
        self.assertEqual(payload["latest"]["version"], "0.13.6-beta")
        self.assertEqual(payload["latest"]["asset_name"], "ScheduleApp_Setup_0.13.6-beta.exe")
        self.assertEqual(payload["latest"]["changelog_summary"], ["Startup update modal", "Post-install changelog"])

    def test_release_notes_summary_prefers_main_change_bullets(self):
        body = """
## Verification
- Windows installers
- SHA256: abc123

## What Changed
- Preserves manually assigned shifts during generation
- Shows an update modal when a newer installer is available
- Shows the main changelog after the updated app starts

## Details
- Internal build metadata
"""

        summary = update_service.summarize_release_notes(body)

        self.assertEqual(
            summary,
            [
                "Preserves manually assigned shifts during generation",
                "Shows an update modal when a newer installer is available",
                "Shows the main changelog after the updated app starts",
            ],
        )

    def test_update_check_ignores_non_installer_assets(self):
        releases = [
            {
                "tag_name": "v0.13.7-beta",
                "draft": False,
                "assets": [
                    {
                        "name": "source.zip",
                        "browser_download_url": "https://github.com/LittleDespairs/Schedule_app/releases/download/v0.13.7-beta/source.zip",
                        "size": 123,
                    }
                ],
            }
        ]

        with patch.object(main, "request_github_releases", return_value=releases):
            payload = main.get_update_status()

        self.assertFalse(payload["update_available"])
        self.assertIn("No installable", payload["message"])

    def test_update_install_verifies_downloaded_installer_signature_before_launch(self):
        latest = {
            "version": "0.15.18-beta",
            "asset_name": "ShiftCare_Setup_0.15.18-beta.exe",
            "download_url": "https://github.com/LittleDespairs/Schedule_app_releases/releases/download/v0.15.18-beta/ShiftCare_Setup_0.15.18-beta.exe",
        }
        installer_path = Path(tempfile.gettempdir()) / "ShiftCare_Setup_0.15.18-beta.exe"

        with (
            patch.object(
                main,
                "get_update_status",
                return_value={"current_version": "0.15.17-beta", "update_available": True, "latest": latest},
            ),
            patch.object(main, "download_update_installer", return_value=installer_path),
            patch.object(main, "verify_windows_installer_signature") as verify_signature,
            patch.object(main.subprocess, "Popen") as popen,
            patch.object(main, "schedule_desktop_shutdown") as schedule_shutdown,
        ):
            response = self.client.post(
                "/api/updates/install",
                json={"download_url": latest["download_url"], "asset_name": latest["asset_name"]},
            )

        self.assertEqual(response.status_code, 200)
        verify_signature.assert_called_once_with(installer_path)
        popen.assert_called_once_with([str(installer_path), "/CLOSEAPPLICATIONS"], close_fds=True)
        schedule_shutdown.assert_called_once()

    def test_update_install_records_changelog_for_next_startup(self):
        latest = {
            "version": "0.15.18-beta",
            "release_name": "ShiftCare 0.15.18 Beta",
            "asset_name": "ShiftCare_Setup_0.15.18-beta.exe",
            "download_url": "https://github.com/LittleDespairs/Schedule_app_releases/releases/download/v0.15.18-beta/ShiftCare_Setup_0.15.18-beta.exe",
            "body": "## Changes\n- Startup update modal\n- Post-install changelog",
            "changelog_summary": ["Startup update modal", "Post-install changelog"],
        }
        installer_path = Path(tempfile.gettempdir()) / "ShiftCare_Setup_0.15.18-beta.exe"

        with (
            patch.object(main, "update_notifications_are_enabled", return_value=True),
            patch.object(
                main,
                "get_update_status",
                return_value={"current_version": "0.15.17-beta", "update_available": True, "latest": latest},
            ),
            patch.object(main, "download_update_installer", return_value=installer_path),
            patch.object(main, "verify_windows_installer_signature"),
            patch.object(main.subprocess, "Popen"),
            patch.object(main, "schedule_desktop_shutdown"),
        ):
            install_response = self.client.post(
                "/api/updates/install",
                json={"download_url": latest["download_url"], "asset_name": latest["asset_name"]},
            )

        self.assertEqual(install_response.status_code, 200)

        with (
            patch.object(main, "APP_VERSION", "0.15.18_beta"),
            patch.object(main, "update_notifications_are_enabled", return_value=True),
            patch.object(
                main,
                "get_update_status",
                return_value={"current_version": "0.15.18_beta", "update_available": False},
            ),
        ):
            startup_response = self.client.get("/api/updates/startup")
            self.assertEqual(startup_response.status_code, 200)
            changelog = startup_response.json()["post_update_changelog"]
            self.assertEqual(changelog["version"], "0.15.18_beta")
            self.assertEqual(changelog["release_name"], "ShiftCare 0.15.18 Beta")
            self.assertEqual(changelog["summary"], ["Startup update modal", "Post-install changelog"])

            ack_response = self.client.post("/api/updates/post-install-changelog/ack")
            self.assertEqual(ack_response.status_code, 200)

            startup_after_ack = self.client.get("/api/updates/startup")
            self.assertIsNone(startup_after_ack.json()["post_update_changelog"])

    def test_windows_installer_signature_rejects_unexpected_publisher(self):
        completed = subprocess.CompletedProcess(
            args=[],
            returncode=0,
            stdout=json.dumps(
                {
                    "Status": "Valid",
                    "StatusMessage": "Signature verified.",
                    "Subject": "CN=Other Publisher",
                    "Thumbprint": "ABC123",
                }
            ),
            stderr="",
        )

        with tempfile.NamedTemporaryFile(suffix=".exe") as installer:
            with (
                patch.object(update_service.os, "name", "nt"),
                patch.object(update_service.subprocess, "run", return_value=completed),
                patch.dict(os.environ, {"SHIFTCARE_WINDOWS_SIGNER_SUBJECT": "ShiftCare"}),
            ):
                with self.assertRaises(HTTPException) as context:
                    update_service.verify_windows_installer_signature(Path(installer.name))

        self.assertEqual(context.exception.status_code, 400)
        self.assertIn("unexpected publisher", context.exception.detail)

    def test_download_page_uses_latest_installable_release_only(self):
        releases = [
            {
                "tag_name": "v0.15.11-beta",
                "name": "ShiftCare 0.15.11 Beta",
                "draft": False,
                "assets": [
                    {
                        "name": "ShiftCare_Setup_0.15.11-beta.exe",
                        "browser_download_url": "https://github.com/LittleDespairs/Schedule_app_releases/releases/download/v0.15.11-beta/ShiftCare_Setup_0.15.11-beta.exe",
                        "size": 100,
                    }
                ],
            },
            {
                "tag_name": "v0.15.18-beta",
                "name": "ShiftCare 0.15.18 Beta",
                "draft": False,
                "published_at": "2026-05-06T12:28:38Z",
                "html_url": "https://github.com/LittleDespairs/Schedule_app_releases/releases/tag/v0.15.18-beta",
                "assets": [
                    {
                        "name": "ShiftCare_Setup_0.15.18-beta.exe",
                        "browser_download_url": "https://github.com/LittleDespairs/Schedule_app_releases/releases/download/v0.15.18-beta/ShiftCare_Setup_0.15.18-beta.exe",
                        "size": 24878593,
                    }
                ],
            },
        ]

        with patch.object(main, "request_github_releases", return_value=releases):
            response = self.client.get("/download")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.headers["cache-control"], "no-store, no-cache, must-revalidate, max-age=0")
        self.assertIn("href=\"/static/manifest.webmanifest\"", response.text)
        self.assertIn("href=\"/static/icons/app-icon.svg\"", response.text)
        self.assertIn("href=\"/favicon.ico\"", response.text)
        self.assertIn("0.15.18-beta", response.text)
        self.assertIn("ShiftCare_Setup_0.15.18-beta.exe", response.text)
        self.assertNotIn("0.15.11-beta", response.text)

    def test_download_latest_redirects_to_latest_installer_asset(self):
        releases = [
            {
                "tag_name": "v0.15.11-beta",
                "draft": False,
                "assets": [
                    {
                        "name": "ShiftCare_Setup_0.15.11-beta.exe",
                        "browser_download_url": "https://github.com/LittleDespairs/Schedule_app_releases/releases/download/v0.15.11-beta/ShiftCare_Setup_0.15.11-beta.exe",
                    }
                ],
            },
            {
                "tag_name": "v0.15.18-beta",
                "draft": False,
                "assets": [
                    {
                        "name": "ShiftCare_Setup_0.15.18-beta.exe",
                        "browser_download_url": "https://github.com/LittleDespairs/Schedule_app_releases/releases/download/v0.15.18-beta/ShiftCare_Setup_0.15.18-beta.exe",
                    }
                ],
            },
        ]

        with patch.object(main, "request_github_releases", return_value=releases):
            response = self.client.get("/download/latest", follow_redirects=False)

        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            response.headers["location"],
            "https://github.com/LittleDespairs/Schedule_app_releases/releases/download/v0.15.18-beta/ShiftCare_Setup_0.15.18-beta.exe",
        )
        self.assertEqual(response.headers["cache-control"], "no-store, no-cache, must-revalidate, max-age=0")

    def test_download_latest_api_reports_latest_release(self):
        releases = [
            {
                "tag_name": "v0.15.18-beta",
                "draft": False,
                "assets": [
                    {
                        "name": "ShiftCare_Setup_0.15.18-beta.exe",
                        "browser_download_url": "https://github.com/LittleDespairs/Schedule_app_releases/releases/download/v0.15.18-beta/ShiftCare_Setup_0.15.18-beta.exe",
                        "size": 24878593,
                    }
                ],
            }
        ]

        with patch.object(main, "request_github_releases", return_value=releases):
            response = self.client.get("/api/download/latest")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["product"], "ShiftCare")
        self.assertEqual(payload["latest"]["version"], "0.15.18-beta")
        self.assertEqual(payload["latest"]["asset_name"], "ShiftCare_Setup_0.15.18-beta.exe")

    def test_database_backup_create_list_restore_round_trip(self):
        employee_id = self._create_employee(full_name="Employee A")
        self.assertIsInstance(employee_id, int)

        create_backup = self.client.post("/api/database/backups", json={"label": "manual"})
        self.assertEqual(create_backup.status_code, 200)
        backup_name = create_backup.json()["backup_name"]
        self.assertTrue(backup_name.endswith(".schedulebackup"))
        backup_path = database.get_backup_dir() / backup_name
        with ZipFile(backup_path) as archive:
            metadata = main.json.loads(archive.read("metadata.json").decode("utf-8"))
            self.assertEqual(metadata["app_version"], main.APP_VERSION)
            self.assertEqual(metadata["schema_version"], database.CURRENT_SCHEMA_VERSION)
            self.assertEqual(metadata["organization_id"], 1)
            self.assertIn("created_at", metadata)
            self.assertFalse(metadata["contains_access_tokens"])
            self.assertFalse(metadata["contains_server_secrets"])

        backup_list = self.client.get("/api/database/backups")
        self.assertEqual(backup_list.status_code, 200)
        self.assertTrue(any(item["name"] == backup_name for item in backup_list.json()["backups"]))

        delete_response = self.client.delete(f"/api/employees/{employee_id}")
        self.assertEqual(delete_response.status_code, 200)
        self.assertEqual(self.client.get("/api/employees").json(), [])

        restore_response = self.client.post("/api/database/restore", json={"backup_name": backup_name})
        self.assertEqual(restore_response.status_code, 200)

        employees = self.client.get("/api/employees").json()
        self.assertEqual(len(employees), 1)
        self.assertEqual(employees[0]["full_name"], "Employee A")

    def test_database_backup_restore_requires_admin_after_auth_bootstrap(self):
        owner_response = self.client.post(
            "/api/auth/bootstrap",
            json={
                "organization_name": "Beta Clinic",
                "full_name": "Owner User",
                "email": "owner@example.com",
                "password": "CorrectHorse123",
            },
        )
        self.assertEqual(owner_response.status_code, 200)
        owner_token = owner_response.json()["access_token"]
        owner_headers = {"Authorization": f"Bearer {owner_token}"}

        create_backup = self.client.post(
            "/api/database/backups",
            headers=owner_headers,
            json={"label": "manual"},
        )
        self.assertEqual(create_backup.status_code, 200)
        backup_name = create_backup.json()["backup_name"]
        cursor = self.connection.cursor()
        cursor.execute("SELECT COUNT(*) FROM auth_audit_events WHERE event_type = 'database_backup_created'")
        self.assertEqual(cursor.fetchone()[0], 1)

        unauthenticated_list = self.client.get("/api/database/backups")
        self.assertEqual(unauthenticated_list.status_code, 401)
        employee_record_id = self._create_employee(headers=owner_headers, full_name="Employee User")

        invitation_response = self.client.post(
            "/api/organizations/1/invitations",
            headers=owner_headers,
            json={
                "email": "employee@example.com",
                "employee_id": employee_record_id,
                "role": "employee",
                "expires_in_days": 7,
            },
        )
        self.assertEqual(invitation_response.status_code, 200)
        employee_response = self.client.post(
            "/api/auth/accept-invitation",
            json={
                "token": invitation_response.json()["invitation_token"],
                "full_name": "Employee User",
                "password": "EmployeePass123",
            },
        )
        self.assertEqual(employee_response.status_code, 200)
        employee_headers = {"Authorization": f"Bearer {employee_response.json()['access_token']}"}

        employee_restore = self.client.post(
            "/api/database/restore",
            headers=employee_headers,
            json={"backup_name": backup_name},
        )
        self.assertEqual(employee_restore.status_code, 403)

        restore_response = self.client.post(
            "/api/database/restore",
            headers=owner_headers,
            json={"backup_name": backup_name},
        )
        self.assertEqual(restore_response.status_code, 200)

        restored_connection = database.get_connection()
        try:
            restored_cursor = restored_connection.cursor()
            restored_cursor.execute(
                """
                SELECT event_type
                FROM auth_audit_events
                WHERE event_type = 'database_backup_restored'
                ORDER BY id
                """
            )
            self.assertEqual(
                [row["event_type"] for row in restored_cursor.fetchall()],
                ["database_backup_restored"],
            )
        finally:
            restored_connection.close()

    def test_delete_employee_cascades_related_records_and_returns_backup_name(self):
        employee_id = self._create_employee(full_name="Employee A")
        position_id = self._create_position()
        template_id = self._create_shift_template(position_id=position_id)

        response = self.client.post(
            "/api/employee-positions",
            json={
                "employee_id": employee_id,
                "position_id": position_id,
                "is_primary": True,
                "priority_score": 100,
                "is_fallback_only": False,
            },
        )
        self.assertEqual(response.status_code, 200)

        self.client.post(
            "/api/employee-preferences",
            json={
                "employee_id": employee_id,
                "allow_morning": True,
                "allow_evening": True,
                "allow_night": True,
                "allow_morning_evening_combo": True,
            },
        )
        self.client.post(
            "/api/employee-week-preferences",
            json={
                "employee_id": employee_id,
                "week_start_date": "2026-04-20",
                "preference_date": "2026-04-21",
                "preference_type": "vacation",
            },
        )
        self.client.post(
            "/api/employee-recurring-preferences",
            json={
                "employee_id": employee_id,
                "rules": [
                    {
                        "preference_kind": "soft",
                        "day_of_week": 2,
                        "preference_type": "off_day",
                    }
                ],
            },
        )
        self.client.post(
            "/api/employee-day-statuses",
            json={"employee_id": employee_id, "date": "2026-04-22", "status_type": "day_off"},
        )
        self.client.post(
            "/api/schedule",
            json={
                "employee_id": employee_id,
                "position_id": position_id,
                "date": "2026-04-20",
                "shift_template_id": template_id,
            },
        )

        delete_response = self.client.delete(f"/api/employees/{employee_id}")
        self.assertEqual(delete_response.status_code, 200)
        self.assertTrue(delete_response.json()["backup_name"])
        self.assertEqual(self.client.get("/api/employees").json(), [])
        self.assertEqual(self.client.get("/api/employee-positions").json(), [])
        self.assertEqual(self.client.get("/api/schedule").json(), [])
        self.assertEqual(self.client.get("/api/employee-day-statuses").json(), [])
        self.assertEqual(self.client.get("/api/employee-recurring-preferences").json(), [])
        preferences = self.client.get("/api/employee-week-preferences", params={"week_start_date": "2026-04-20"}).json()
        self.assertEqual(preferences, [])

    def test_delete_position_cascades_related_records_and_returns_backup_name(self):
        employee_id = self._create_employee(full_name="Employee A")
        position_id = self._create_position(name="Ward A")
        template_id = self._create_shift_template(position_id=position_id)

        self.client.post(
            "/api/employee-positions",
            json={
                "employee_id": employee_id,
                "position_id": position_id,
                "is_primary": True,
                "priority_score": 100,
                "is_fallback_only": False,
            },
        )
        self.client.post(
            "/api/shift-requirements",
            json={
                "position_id": position_id,
                "shift_category": "morning",
                "required_total": 1,
                "required_female_min": 0,
                "required_male_min": 0,
            },
        )
        self.client.post(
            "/api/coverage-requirements",
            json={
                "position_id": position_id,
                "start_time": "06:00",
                "end_time": "14:00",
                "required_total": 1,
                "required_female_min": 0,
                "required_male_min": 0,
                "is_overnight": False,
            },
        )
        self.client.post(
            "/api/schedule",
            json={
                "employee_id": employee_id,
                "position_id": position_id,
                "date": "2026-04-20",
                "shift_template_id": template_id,
            },
        )

        delete_response = self.client.delete(f"/api/positions/{position_id}")
        self.assertEqual(delete_response.status_code, 200)
        self.assertTrue(delete_response.json()["backup_name"])
        self.assertEqual(self.client.get("/api/positions").json(), [])
        self.assertEqual(self.client.get("/api/employee-positions").json(), [])
        self.assertEqual(self.client.get("/api/shift-requirements").json(), [])
        self.assertEqual(self.client.get("/api/coverage-requirements").json(), [])
        self.assertEqual(self.client.get("/api/schedule").json(), [])

    def test_delete_shift_template_in_use_is_blocked_and_plain_delete_returns_backup_name(self):
        position_id = self._create_position()
        template_id = self._create_shift_template(position_id=position_id, name="Morning")
        employee_id = self._create_employee()
        self.client.post(
            "/api/employee-positions",
            json={
                "employee_id": employee_id,
                "position_id": position_id,
                "is_primary": True,
                "priority_score": 100,
                "is_fallback_only": False,
            },
        )
        self.client.post(
            "/api/schedule",
            json={
                "employee_id": employee_id,
                "position_id": position_id,
                "date": "2026-04-20",
                "shift_template_id": template_id,
            },
        )

        in_use_delete = self.client.delete(f"/api/shift-templates/{template_id}")
        self.assertEqual(in_use_delete.status_code, 400)

        self.client.post(
            "/api/schedule/clear-week",
            json={"position_id": position_id, "week_start_date": "2026-04-20"},
        )
        delete_response = self.client.delete(f"/api/shift-templates/{template_id}")
        self.assertEqual(delete_response.status_code, 200)
        self.assertTrue(delete_response.json()["backup_name"])


if __name__ == "__main__":
    unittest.main()
