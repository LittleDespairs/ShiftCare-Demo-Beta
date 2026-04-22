import sqlite3
from datetime import datetime, timedelta
from typing import Literal, Optional

from fastapi import FastAPI, HTTPException, Request
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel, Field, model_validator

import sys
from pathlib import Path



from database import get_connection, init_db

import sys
from pathlib import Path

from io import BytesIO
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Tags metadata / Описание тегов для документации
tags_metadata = [
    {
        "name": "Pages",
        "description": "Frontend pages (HTML templates) / HTML страницы интерфейса",
    },
    {
        "name": "Employees",
        "description": "Operations with employees / Работа с сотрудниками",
    },
    {
        "name": "Positions",
        "description": "Operations with positions / Работа с должностями",
    },
    {
        "name": "Assignments",
        "description": "Employee-to-position assignments / Привязка сотрудников к должностям",
    },
    {
        "name": "Shift Templates",
        "description": "Shift template management / Управление шаблонами смен",
    },
    {
        "name": "Preferences",
        "description": "Employee preferences / Общие пожелания сотрудников",
    },
    {
        "name": "Weekly Preferences",
        "description": "Weekly employee preferences / Недельные пожелания сотрудников",
    },
    {
        "name": "Requirements",
        "description": "Shift requirements / Требования к сменам",
    },
    {
        "name": "Schedule",
        "description": "Schedule management / Управление расписанием",
    },
]

# Create FastAPI application / Создаём приложение FastAPI
app = FastAPI(
    title="Schedule App - Nursing Staff Scheduling 0.5.1_alpha",
    description="Web application for nursing staff scheduling",
    version="0.5.1_alpha",
    openapi_tags=tags_metadata, # Use tags metadata in OpenAPI docs / Используем описание тегов в документации OpenAPI
)

def get_base_path() -> Path: # Get base path for static and template files, compatible with PyInstaller / Получаем базовый путь для статических файлов и шаблонов, совместимый с PyInstaller
    if getattr(sys, "frozen", False):
        return Path(sys._MEIPASS)
    return Path(__file__).resolve().parent


BASE_PATH = get_base_path()

# Initialize database on startup / Инициализируем базу данных при запуске
init_db()

# Mount static files / Подключаем статические файлы
app.mount("/static", StaticFiles(directory=str(BASE_PATH / "static")), name="static")
print("BASE_PATH =", BASE_PATH)
print("STATIC PATH =", BASE_PATH / "static")
print("TEMPLATES PATH =", BASE_PATH / "templates")
templates = Jinja2Templates(directory=str(BASE_PATH / "templates"))


# =========================
# Pydantic models / Модели
# =========================

class EmployeeCreate(BaseModel): # Employee data for creation / Данные сотрудника для создания
    full_name: str = Field(min_length=2, max_length=100)
    sex: Literal["male", "female"]

    min_shifts_per_week: int = Field(ge=0, le=14)
    target_shifts_per_week: int = Field(ge=0, le=14)
    max_shifts_per_week: int = Field(ge=0, le=14)

    can_work_night: bool
    can_work_weekends: bool
    can_work_evenings_after_night: bool
    can_work_mornings_and_evenings: bool

    @model_validator(mode="after") # Validate that shift counts are consistent / Проверяем, что количество смен согласовано
    def validate_shift_range(self):
        if self.min_shifts_per_week > self.max_shifts_per_week:
            raise ValueError("min_shifts_per_week cannot be greater than max_shifts_per_week")

        if self.target_shifts_per_week < self.min_shifts_per_week:
            raise ValueError("target_shifts_per_week cannot be less than min_shifts_per_week")

        if self.target_shifts_per_week > self.max_shifts_per_week:
            raise ValueError("target_shifts_per_week cannot be greater than max_shifts_per_week")

        return self


class PositionCreate(BaseModel):
    # Position data / Данные должности
    name: str = Field(min_length=2, max_length=100)
    requires_continuous_coverage: bool = False
    minimum_staff_presence: int = Field(ge=0, le=20, default=0)

    @model_validator(mode="after")
    def validate_presence(self):
        if not self.requires_continuous_coverage and self.minimum_staff_presence != 0:
            raise ValueError("minimum_staff_presence must be 0 if continuous coverage is disabled")
        return self


class EmployeePositionCreate(BaseModel):
    # Employee-position assignment data / Данные для привязки сотрудника к должности
    employee_id: int
    position_id: int
    is_primary: bool = False
    priority_score: int = Field(ge=0, le=100, default=50)
    is_fallback_only: bool = False

class ClearWeekScheduleRequest(BaseModel):
    # Request data for clearing schedule of a week for a position / Данные запроса для очистки расписания на неделю для должности
    position_id: int
    week_start_date: str

class ShiftTemplateCreate(BaseModel):
    # Shift template / Шаблон смены
    name: str = Field(min_length=2, max_length=100)
    category: Literal["morning", "evening", "night"]
    start_time: str
    end_time: str
    is_overnight: bool = False
    is_active: bool = True
    is_split_only: bool = False


class ScheduleEntryCreate(BaseModel):
    # Scheduled shift / Назначенная смена
    employee_id: int
    position_id: int
    date: str
    shift_template_id: int


class AutoGenerateScheduleRequest(BaseModel):
    # Request data for auto-generating schedule for a week and position / Данные запроса для авто-генерации расписания на неделю для должности
    position_id: int
    week_start_date: str


class ShiftRequirementCreate(BaseModel):
    # Shift requirement data / Данные требования смены
    position_id: int
    shift_category: Literal["morning", "evening", "night"]
    required_total: int = Field(ge=1, le=20)
    required_female_min: int = Field(ge=0, le=20)

    @model_validator(mode="after")
    def validate_female_min(self):
        # Female minimum cannot exceed total / Минимум женщин не может быть больше общего количества
        if self.required_female_min > self.required_total:
            raise ValueError("required_female_min cannot be greater than required_total")
        return self


class EmployeePreferenceCreate(BaseModel):
    # Employee preference data / Данные общих пожеланий сотрудника
    employee_id: int
    allow_morning: bool
    allow_evening: bool
    allow_night: bool
    allow_morning_evening_combo: bool


class EmployeeWeekPreferenceCreate(BaseModel):
    # Weekly employee preference / Недельное пожелание сотрудника
    employee_id: int
    week_start_date: str
    preference_date: str
    preference_type: Literal[
        "no_preference",
        "off_day",
        "vacation",
        "only_morning",
        "only_evening",
        "only_night",
        "not_morning",
        "not_evening",
        "not_night",
        "no_morning_evening_combo",
    ]

class EmployeeDayStatusCreate(BaseModel):
    # Special status for employee on a specific date, like sick leave or no-show / Особый статус сотрудника на конкретную дату, например больничный или неявка
    employee_id: int
    date: str
    status_type: Literal["sick", "no_show"]


# =====================================
# Helper functions / Вспомогательные функции
# =====================================

def build_schedule_cell_text(entries: list[dict]) -> str:
    # Build compact text for one Excel cell / Строим компактный текст для ячейки Excel
    if not entries:
        return ""

    sorted_entries = sorted(entries, key=lambda item: (item["start_time"], item["end_time"]))

    parts = []
    for entry in sorted_entries:
        parts.append(f'{entry["shift_template_name"]}')

    return "\n".join(parts)

def get_employee_day_status(connection, employee_id: int, date: str):
    # Get special status of employee on the date, like sick leave or no-show / Получаем особый статус сотрудника на дату, например больничный или неявка
    cursor = connection.cursor()

    cursor.execute("""
        SELECT status_type
        FROM employee_day_statuses
        WHERE employee_id = ? AND date = ?
    """, (employee_id, date))

    return cursor.fetchone()

def row_to_employee_dict(row: sqlite3.Row) -> dict:
    # Convert SQLite row to employee dict / Преобразуем строку SQLite в словарь сотрудника
    return {
        "id": row["id"],
        "full_name": row["full_name"],
        "sex": row["sex"],
        "min_shifts_per_week": row["min_shifts_per_week"],
        "target_shifts_per_week": row["target_shifts_per_week"],
        "max_shifts_per_week": row["max_shifts_per_week"],
        "can_work_night": bool(row["can_work_night"]),
        "can_work_weekends": bool(row["can_work_weekends"]),
        "can_work_evenings_after_night": bool(row["can_work_evenings_after_night"]),
        "can_work_mornings_and_evenings": bool(row["can_work_mornings_and_evenings"]),
    }


def row_to_shift_template_dict(row: sqlite3.Row) -> dict:
    # Convert SQLite row to shift template dict / Преобразуем строку SQLite в словарь шаблона смены
    return {
        "id": row["id"],
        "name": row["name"],
        "category": row["category"],
        "start_time": row["start_time"],
        "end_time": row["end_time"],
        "is_overnight": bool(row["is_overnight"]),
        "is_active": bool(row["is_active"]),
        "is_split_only": bool(row["is_split_only"]),
    }

def row_to_position_dict(row: sqlite3.Row) -> dict:
    # Convert SQLite row to position dict / Преобразуем строку SQLite в словарь должности
    return {
        "id": row["id"],
        "name": row["name"],
        "requires_continuous_coverage": bool(row["requires_continuous_coverage"]),
        "minimum_staff_presence": row["minimum_staff_presence"],
    }


def parse_time_string(time_string: str) -> datetime.time:
    # Parse HH:MM string / Преобразуем строку HH:MM во время
    return datetime.strptime(time_string, "%H:%M").time()


def build_shift_datetimes(date_string: str, shift_template: dict) -> tuple[datetime, datetime]:
    # Build real start/end datetimes for a shift / Строим реальные datetime начала и конца смены
    shift_date = datetime.strptime(date_string, "%Y-%m-%d").date()

    start_dt = datetime.combine(shift_date, parse_time_string(shift_template["start_time"]))
    end_dt = datetime.combine(shift_date, parse_time_string(shift_template["end_time"]))

    if shift_template["is_overnight"]:
        end_dt += timedelta(days=1)

    return start_dt, end_dt

def intervals_overlap(start_a: datetime, end_a: datetime, start_b: datetime, end_b: datetime) -> bool:
    # Check if two time intervals overlap / Проверяем, пересекаются ли два временных интервала
    return start_a < end_b and start_b < end_a

def get_break_minutes_between_shifts(
    first_start: datetime,
    first_end: datetime,
    second_start: datetime,
    second_end: datetime
) -> float:
    # Return break in minutes between two non-overlapping shifts / Возвращаем перерыв в минутах между двумя непересекающимися сменами
    if first_end <= second_start:
        return (second_start - first_end).total_seconds() / 60

    if second_end <= first_start:
        return (first_start - second_end).total_seconds() / 60

    return -1  # Overlap / Пересечение


def previous_date(date_string: str) -> str:
    # Get previous date string / Получаем предыдущую дату
    current_date = datetime.strptime(date_string, "%Y-%m-%d").date()
    previous_day = current_date - timedelta(days=1)
    return previous_day.isoformat()


def build_week_dates(week_start_date: str) -> list[str]:
    # Build week dates from Sunday to Saturday / Строим неделю с воскресенья по субботу
    start_date = datetime.strptime(week_start_date, "%Y-%m-%d").date()
    return [(start_date + timedelta(days=i)).isoformat() for i in range(7)]


def get_week_end_date(week_start_date: str) -> str:
    # Get week end date / Получаем последний день недели
    start_date = datetime.strptime(week_start_date, "%Y-%m-%d").date()
    return (start_date + timedelta(days=6)).isoformat()


def get_employee_week_shift_count(connection, employee_id: int, week_start_date: str) -> int:
    # Count employee shifts in the selected week / Считаем смены сотрудника за выбранную неделю
    cursor = connection.cursor()
    week_end_date = get_week_end_date(week_start_date)

    cursor.execute("""
        SELECT COUNT(*) AS shift_count
        FROM schedule_entries
        WHERE employee_id = ?
          AND date >= ?
          AND date <= ?
    """, (employee_id, week_start_date, week_end_date))

    row = cursor.fetchone()
    return row["shift_count"] if row else 0

def get_employee_week_worked_days(connection, employee_id: int, week_start_date: str) -> int:
    # Count how many distinct days employee works in selected week
    cursor = connection.cursor()
    week_end_date = get_week_end_date(week_start_date)

    cursor.execute("""
        SELECT COUNT(DISTINCT date) AS worked_days
        FROM schedule_entries
        WHERE employee_id = ?
          AND date >= ?
          AND date <= ?
    """, (employee_id, week_start_date, week_end_date))

    row = cursor.fetchone()
    return row["worked_days"] if row else 0


def employee_has_shift_on_date(connection, employee_id: int, date: str) -> bool:
    cursor = connection.cursor()

    cursor.execute("""
        SELECT 1
        FROM schedule_entries
        WHERE employee_id = ?
          AND date = ?
        LIMIT 1
    """, (employee_id, date))

    return cursor.fetchone() is not None


def get_employee_consecutive_work_days_around_date(connection, employee_id: int, date_string: str) -> int:
    # Count consecutive worked days around target date, including current date if shift is added
    target_date = datetime.strptime(date_string, "%Y-%m-%d").date()
    streak = 1  # assume we are about to assign this day

    # Check backward
    current = target_date - timedelta(days=1)
    while employee_has_shift_on_date(connection, employee_id, current.isoformat()):
        streak += 1
        current -= timedelta(days=1)

    # Check forward
    current = target_date + timedelta(days=1)
    while employee_has_shift_on_date(connection, employee_id, current.isoformat()):
        streak += 1
        current += timedelta(days=1)

    return streak


def get_employee_day_preference(connection, employee_id: int, date: str):
    # Get employee preference for a specific date / Получаем пожелание сотрудника на дату
    cursor = connection.cursor()

    cursor.execute("""
        SELECT preference_type
        FROM employee_week_preferences
        WHERE employee_id = ? AND preference_date = ?
    """, (employee_id, date))

    return cursor.fetchone()


def get_shift_templates_by_category(connection, category: str) -> list[dict]:
    # Get active shift templates for one category / Получаем активные шаблоны смен по категории
    cursor = connection.cursor()

    cursor.execute("""
        SELECT *
        FROM shift_templates
        WHERE category = ? AND is_active = 1
        ORDER BY start_time, end_time
    """, (category,))

    return [row_to_shift_template_dict(row) for row in cursor.fetchall()]

def build_position_schedule_intervals_for_date(connection, position_id: int, date_string: str) -> list[tuple[datetime, datetime]]:
    # Build list of real time intervals when staff is scheduled for the position on the date / Строим список реальных временных интервалов, когда запланирован персонал на должности в дату
    cursor = connection.cursor()

    cursor.execute("""
        SELECT
            st.start_time,
            st.end_time,
            st.is_overnight
        FROM schedule_entries se
        JOIN shift_templates st ON se.shift_template_id = st.id
        WHERE se.position_id = ?
          AND se.date = ?
    """, (position_id, date_string))

    intervals = []
    for row in cursor.fetchall():
        template = {
            "start_time": row["start_time"],
            "end_time": row["end_time"],
            "is_overnight": bool(row["is_overnight"]),
        }
        intervals.append(build_shift_datetimes(date_string, template))

    return intervals


def count_staff_present_at_time(intervals: list[tuple[datetime, datetime]], moment: datetime) -> int:
    # Count how many intervals include the moment / Считаем, сколько интервалов включают момент
    count = 0
    for start_dt, end_dt in intervals:
        if start_dt <= moment < end_dt:
            count += 1
    return count

def cleanup_orphan_split_entries_for_position_week(connection, position_id: int, week_dates: list[str]) -> int:
# This function checks for "split-only" shift templates (like morning/evening pairs) and deletes any schedule entries that don't have their required counterpart on the same day. It returns the number of deleted entries.
# Эта функция проверяет шаблоны смен, которые должны идти парами (например, утренняя и вечерняя смены) и удаляет записи расписания, у которых нет своей пары в тот же день. Она возвращает количество удалённых записей.
    cursor = connection.cursor()

    deleted_ids = []

    for date in week_dates:
        cursor.execute("""
            SELECT
                se.id,
                se.employee_id,
                se.position_id,
                se.date,
                st.category
            FROM schedule_entries se
            JOIN shift_templates st ON se.shift_template_id = st.id
            WHERE se.position_id = ?
              AND se.date = ?
              AND st.is_split_only = 1
              AND st.category IN ('morning', 'evening')
        """, (position_id, date))

        entries = [dict(row) for row in cursor.fetchall()]

        for entry in entries:
            opposite_category = "evening" if entry["category"] == "morning" else "morning"

            cursor.execute("""
                SELECT 1
                FROM schedule_entries se
                JOIN shift_templates st ON se.shift_template_id = st.id
                WHERE se.employee_id = ?
                  AND se.position_id = ?
                  AND se.date = ?
                  AND st.is_split_only = 1
                  AND st.category = ?
                LIMIT 1
            """, (
                entry["employee_id"],
                entry["position_id"],
                entry["date"],
                opposite_category
            ))

            has_pair = cursor.fetchone() is not None

            if not has_pair:
                deleted_ids.append(entry["id"])

    if deleted_ids:
        cursor.execute(f"""
            DELETE FROM schedule_entries
            WHERE id IN ({",".join(["?"] * len(deleted_ids))})
        """, deleted_ids)

    return len(deleted_ids)


def validate_position_continuous_coverage_for_date(
    # Validate that for positions with continuous coverage requirement, there are enough staff scheduled at every moment of the day / Проверяем, что для должностей с требованием непрерывного покрытия достаточно персонала в каждый момент дня
    connection,
    position_id: int,
    date_string: str,
    required_staff: int
) -> None:
    if required_staff <= 0:
        return

    intervals = build_position_schedule_intervals_for_date(connection, position_id, date_string)

    if not intervals:
        raise HTTPException(
            status_code=400,
            detail="This position requires continuous coverage, but no staff is scheduled on this date"
        )

    day_start = datetime.combine(datetime.strptime(date_string, "%Y-%m-%d").date(), datetime.min.time())
    next_day_start = day_start + timedelta(days=1)

    critical_points = {day_start, next_day_start}

    for start_dt, end_dt in intervals:
        # clip interval to current calendar day
        clipped_start = max(start_dt, day_start)
        clipped_end = min(end_dt, next_day_start)

        if clipped_start < clipped_end:
            critical_points.add(clipped_start)
            critical_points.add(clipped_end)

    ordered_points = sorted(critical_points)

    for point in ordered_points[:-1]:
        present = count_staff_present_at_time(intervals, point)
        if present < required_staff:
            raise HTTPException(
                status_code=400,
                detail=f"Continuous coverage is broken on {date_string} for this position"
            )
        
def employee_has_shift_category_on_date(connection, employee_id: int, date_string: str, category: str) -> bool:
    # Check if employee has a shift of the category on the date / Проверяем, есть ли у сотрудника смена категории в дату
    cursor = connection.cursor()
    cursor.execute("""
        SELECT 1
        FROM schedule_entries se
        JOIN shift_templates st ON se.shift_template_id = st.id
        WHERE se.employee_id = ?
          AND se.date = ?
          AND st.category = ?
        LIMIT 1
    """, (employee_id, date_string, category))
    return cursor.fetchone() is not None

def get_employee_existing_shift_on_date_by_category(connection, employee_id: int, date_string: str, category: str):
    # Return existing shift template of employee for the date and category / Возвращаем существующую смену сотрудника по дате и категории
    cursor = connection.cursor()
    cursor.execute("""
        SELECT
            st.id,
            st.name,
            st.category,
            st.start_time,
            st.end_time,
            st.is_overnight,
            st.is_active,
            st.is_split_only
        FROM schedule_entries se
        JOIN shift_templates st ON se.shift_template_id = st.id
        WHERE se.employee_id = ?
          AND se.date = ?
          AND st.category = ?
        LIMIT 1
    """, (employee_id, date_string, category))

    row = cursor.fetchone()
    if not row:
        return None

    return row_to_shift_template_dict(row)

def get_sorted_templates_for_employee(connection, employee: dict, date_string: str, shift_category: str) -> list[dict]:
    # Get shift templates sorted by suitability for the employee on the date
    templates = get_shift_templates_by_category(connection, shift_category)

    paired_shift = None
    if shift_category == "morning":
        paired_shift = get_employee_existing_shift_on_date_by_category(
            connection,
            employee["id"],
            date_string,
            "evening"
        )
    elif shift_category == "evening":
        paired_shift = get_employee_existing_shift_on_date_by_category(
            connection,
            employee["id"],
            date_string,
            "morning"
        )

    def template_priority(template: dict):
        start_dt, end_dt = build_shift_datetimes(date_string, template)
        duration_minutes = int((end_dt - start_dt).total_seconds() // 60)

        split_penalty = 0
        pair_bonus = 0
        break_bonus = 0

        if paired_shift is not None:
            paired_start, paired_end = build_shift_datetimes(date_string, paired_shift)

            break_minutes = get_break_minutes_between_shifts(
                paired_start,
                paired_end,
                start_dt,
                end_dt
            )

            paired_is_split = paired_shift["is_split_only"]
            current_is_split = template["is_split_only"]

            # Оба шаблона должны быть одного типа:
            # либо оба split, либо оба обычные
            if paired_is_split != current_is_split:
                split_penalty += 300000

            # Если это split-пара и перерыв валидный — самый высокий приоритет
            elif current_is_split and break_minutes >= 60:
                pair_bonus -= 100000
                break_bonus -= int(break_minutes)

            # Если это split-шаблон, но перерыв слишком маленький — плохо
            elif current_is_split and break_minutes < 60:
                split_penalty += 200000

            # Если это обычная пара, тоже разрешаем, но менее приоритетно,
            # чем валидную split-пару
            elif not current_is_split:
                pair_bonus -= 1000

        else:
            # Если второй половины дня ещё нет, split-only шаблоны стараемся не брать
            # как одиночную смену
            if template["is_split_only"]:
                split_penalty += 10000

        return (
            split_penalty,
            pair_bonus,
            break_bonus,
            -duration_minutes,
            template["start_time"],
            template["end_time"]
        )

    return sorted(templates, key=template_priority)

def try_assign_split_pair_for_employee(
    connection,
    employee: dict,
    position_id: int,
    date: str
) -> list[dict] | None:
    """
    Try to assign one valid split morning + split evening pair
    for the same employee on the same date.
    Returns created entries list if success, otherwise None.
    """
    cursor = connection.cursor()

    morning_templates = [
        template
        for template in get_shift_templates_by_category(connection, "morning")
        if template["is_split_only"]
    ]

    evening_templates = [
        template
        for template in get_shift_templates_by_category(connection, "evening")
        if template["is_split_only"]
    ]

    if not morning_templates or not evening_templates:
        return None

    for morning_template in morning_templates:
        for evening_template in evening_templates:
            morning_start, morning_end = build_shift_datetimes(date, morning_template)
            evening_start, evening_end = build_shift_datetimes(date, evening_template)

            break_minutes = get_break_minutes_between_shifts(
                morning_start,
                morning_end,
                evening_start,
                evening_end
            )

            if break_minutes < 60:
                continue

            morning_entry = ScheduleEntryCreate(
                employee_id=employee["id"],
                position_id=position_id,
                date=date,
                shift_template_id=morning_template["id"]
            )

            evening_entry = ScheduleEntryCreate(
                employee_id=employee["id"],
                position_id=position_id,
                date=date,
                shift_template_id=evening_template["id"]
            )

            try:
                # Сначала вставляем утро без split-only запрета,
                # потому что мы уже заранее знаем, что сейчас же добавим пару.
                employee_data, position_data, _ = validate_schedule_entry_basic(connection, morning_entry, allow_split_without_pair=True)

                cursor.execute("""
                    INSERT INTO schedule_entries (
                        employee_id,
                        position_id,
                        date,
                        shift_template_id
                    )
                    VALUES (?, ?, ?, ?)
                """, (
                    morning_entry.employee_id,
                    morning_entry.position_id,
                    morning_entry.date,
                    morning_entry.shift_template_id
                ))

                try:
                    validate_schedule_entry_strict(connection, evening_entry)

                    cursor.execute("""
                        INSERT INTO schedule_entries (
                            employee_id,
                            position_id,
                            date,
                            shift_template_id
                        )
                        VALUES (?, ?, ?, ?)
                    """, (
                        evening_entry.employee_id,
                        evening_entry.position_id,
                        evening_entry.date,
                        evening_entry.shift_template_id
                    ))

                    return [
                        {
                            "employee_id": employee["id"],
                            "employee_name": employee["full_name"],
                            "date": date,
                            "shift_template_id": morning_template["id"],
                            "shift_template_name": morning_template["name"],
                            "shift_category": morning_template["category"]
                        },
                        {
                            "employee_id": employee["id"],
                            "employee_name": employee["full_name"],
                            "date": date,
                            "shift_template_id": evening_template["id"],
                            "shift_template_name": evening_template["name"],
                            "shift_category": evening_template["category"]
                        }
                    ]

                except HTTPException:
                    cursor.execute("""
                        DELETE FROM schedule_entries
                        WHERE employee_id = ?
                          AND position_id = ?
                          AND date = ?
                          AND shift_template_id = ?
                    """, (
                        morning_entry.employee_id,
                        morning_entry.position_id,
                        morning_entry.date,
                        morning_entry.shift_template_id
                    ))
                    continue

            except HTTPException:
                continue

    return None

def validate_schedule_entry_basic(connection, entry: ScheduleEntryCreate,  allow_split_without_pair: bool = False) -> tuple[dict, dict, dict]:
    # Basic validation of schedule entry before insert/ Стандартная проверка записи расписания перед вставкой, без проверки пожеланий и последовательных дней
    cursor = connection.cursor()

    cursor.execute("SELECT * FROM employees WHERE id = ?", (entry.employee_id,))
    employee_row = cursor.fetchone()
    if not employee_row:
        raise HTTPException(status_code=404, detail="Employee not found")
    employee = row_to_employee_dict(employee_row)

    cursor.execute("SELECT * FROM positions WHERE id = ?", (entry.position_id,))
    position_row = cursor.fetchone()
    if not position_row:
        raise HTTPException(status_code=404, detail="Position not found")
    position = row_to_position_dict(position_row)

    cursor.execute("""
        SELECT *
        FROM employee_positions
        WHERE employee_id = ? AND position_id = ?
    """, (entry.employee_id, entry.position_id))
    if not cursor.fetchone():
        raise HTTPException(status_code=400, detail="Employee is not assigned to this position")

    cursor.execute("""
        SELECT *
        FROM shift_templates
        WHERE id = ? AND is_active = 1
    """, (entry.shift_template_id,))
    shift_template_row = cursor.fetchone()
    if not shift_template_row:
        raise HTTPException(status_code=404, detail="Shift template not found")

    shift_template = row_to_shift_template_dict(shift_template_row)


    # Split-only templates should not be assigned without their pair on the same day, unless allow_split_without_pair is True/ Шаблоны, помеченные как split-only, не должны назначаться без своей пары в тот же день, если allow_split_without_pair не True
    if shift_template["is_split_only"] and not allow_split_without_pair:
        opposite_category = None
        if shift_template["category"] == "morning":
            opposite_category = "evening"
        elif shift_template["category"] == "evening":
            opposite_category = "morning"

        has_pair = False
        if opposite_category is not None:
            existing_pair = get_employee_existing_shift_on_date_by_category(
                connection,
                entry.employee_id,
                entry.date,
                opposite_category
            )
            has_pair = bool(existing_pair)

        if not has_pair:
            raise HTTPException(
                status_code=400,
                detail="This split-only shift cannot be assigned without its paired half on the same day"
            )

    return employee, position, shift_template

def validate_schedule_entry_strict(connection, entry: ScheduleEntryCreate) -> tuple[dict, dict]:
    # Strict validation of schedule entry before insert, with checking preferences and consecutive days / Строгая проверка записи расписания перед вставкой, с проверкой пожеланий и последовательных дней
    cursor = connection.cursor()

    # 1. Check employee exists
    cursor.execute("SELECT * FROM employees WHERE id = ?", (entry.employee_id,))
    employee_row = cursor.fetchone()
    if not employee_row:
        raise HTTPException(status_code=404, detail="Employee not found")

    employee = row_to_employee_dict(employee_row)

    # 2. Check position exists/ Проверяем, что должность существует
    cursor.execute("SELECT * FROM positions WHERE id = ?", (entry.position_id,))
    position_row = cursor.fetchone()
    if not position_row:
        raise HTTPException(status_code=404, detail="Position not found")

    position = row_to_position_dict(position_row)

    # 3. Check assignment exists/ Проверяем, что сотрудник привязан к должности
    cursor.execute("""
        SELECT *
        FROM employee_positions
        WHERE employee_id = ? AND position_id = ?
    """, (entry.employee_id, entry.position_id))

    if not cursor.fetchone():
        raise HTTPException(
            status_code=400,
            detail="Employee is not assigned to this position"
        )

    # 4. Check shift template exists/ Проверяем, что шаблон смены существует
    cursor.execute("""
        SELECT *
        FROM shift_templates
        WHERE id = ? AND is_active = 1
    """, (entry.shift_template_id,))
    shift_template_row = cursor.fetchone()

    if not shift_template_row:
        raise HTTPException(status_code=404, detail="Shift template not found")

    shift_template = row_to_shift_template_dict(shift_template_row)
        # Split-only templates should not be used as a single standalone shift
    if shift_template["is_split_only"]:
        opposite_category = "evening" if shift_template["category"] == "morning" else "morning" if shift_template["category"] == "evening" else None

        has_same_day_split_pair = False

        if opposite_category is not None:
            cursor.execute("""
                SELECT 1
                FROM schedule_entries se
                JOIN shift_templates st ON se.shift_template_id = st.id
                WHERE se.employee_id = ?
                  AND se.date = ?
                  AND st.category = ?
                  AND st.is_split_only = 1
                LIMIT 1
            """, (entry.employee_id, entry.date, opposite_category))

            has_same_day_split_pair = cursor.fetchone() is not None

        if not has_same_day_split_pair:
            raise HTTPException(
                status_code=400,
                detail="This shift template is marked as split-only and cannot be assigned as a standalone shift"
            )

        # 4.5. Prevent employee from working all 7 days of the week/ Предотвращаем ситуацию, когда сотрудник работает все 7 дней недели
    current_date_obj = datetime.strptime(entry.date, "%Y-%m-%d").date()
    sunday_offset = (current_date_obj.weekday() + 1) % 7
    week_start_date = (current_date_obj - timedelta(days=sunday_offset)).isoformat()

    worked_days = get_employee_week_worked_days(connection, entry.employee_id, week_start_date)
    already_has_shift_today = employee_has_shift_on_date(connection, entry.employee_id, entry.date)

    # If today is not yet a worked day, this assignment would add one more worked day/ Если сегодня ещё не считается отработанным днём, это назначение добавит ещё один отработанный день
    projected_worked_days = worked_days if already_has_shift_today else worked_days + 1

    if projected_worked_days > 6:
        raise HTTPException(
            status_code=400,
            detail="Employee must have at least 1 day off per week"
        )

    consecutive_days = get_employee_consecutive_work_days_around_date(
        connection,
        entry.employee_id,
        entry.date
    )

    if consecutive_days > 6:
        raise HTTPException(
            status_code=400,
            detail="Employee cannot work more than 6 consecutive days"
        )

    # 5. Check night permission
    if shift_template["category"] == "night" and not employee["can_work_night"]:
        raise HTTPException(
            status_code=400,
            detail="Employee cannot work night shifts"
        )

    # 6. Check weekly preference: off day
    day_preference = get_employee_day_preference(connection, entry.employee_id, entry.date)
    if day_preference and day_preference["preference_type"] in ("off_day", "vacation"):
        raise HTTPException(
            status_code=400,
            detail="Employee requested an off day for this date"
        )

    # 7. Respect category preferences
    if day_preference:
        preference_type = day_preference["preference_type"]

        if preference_type == "only_morning" and shift_template["category"] != "morning":
            raise HTTPException(status_code=400, detail="Employee requested only morning shifts for this date")

        if preference_type == "only_evening" and shift_template["category"] != "evening":
            raise HTTPException(status_code=400, detail="Employee requested only evening shifts for this date")

        if preference_type == "only_night" and shift_template["category"] != "night":
            raise HTTPException(status_code=400, detail="Employee requested only night shifts for this date")

        if preference_type == "not_morning" and shift_template["category"] == "morning":
            raise HTTPException(status_code=400, detail="Employee requested no morning shift for this date")

        if preference_type == "not_evening" and shift_template["category"] == "evening":
            raise HTTPException(status_code=400, detail="Employee requested no evening shift for this date")

        if preference_type == "not_night" and shift_template["category"] == "night":
            raise HTTPException(status_code=400, detail="Employee requested no night shift for this date")

    # 8. Check previous day night restrictions, including previous week Saturday
    prev_date = previous_date(entry.date)

    cursor.execute("""
        SELECT
            se.id,
            st.category
        FROM schedule_entries se
        JOIN shift_templates st ON se.shift_template_id = st.id
        WHERE se.employee_id = ?
          AND se.date = ?
          AND st.category = 'night'
    """, (entry.employee_id, prev_date))

    previous_night_shift = cursor.fetchone()

    if previous_night_shift:
        if shift_template["category"] == "morning":
            raise HTTPException(
                status_code=400,
                detail="Employee cannot work morning after a night shift"
            )

        if shift_template["category"] == "evening" and not employee["can_work_evenings_after_night"]:
            raise HTTPException(
                status_code=400,
                detail="Employee cannot work evening after a night shift"
            )

    # 9. Check time conflicts across all positions for same date
    cursor.execute("""
        SELECT
            se.id,
            se.date,
            se.shift_template_id,
            st.name,
            st.category,
            st.start_time,
            st.end_time,
            st.is_overnight
        FROM schedule_entries se
        JOIN shift_templates st ON se.shift_template_id = st.id
        WHERE se.employee_id = ? AND se.date = ?
    """, (entry.employee_id, entry.date))

    existing_shifts = [dict(row) for row in cursor.fetchall()]
    new_start, new_end = build_shift_datetimes(entry.date, shift_template)

    for existing in existing_shifts:
        existing["is_overnight"] = bool(existing["is_overnight"])
        existing_start, existing_end = build_shift_datetimes(entry.date, existing)

        if intervals_overlap(existing_start, existing_end, new_start, new_end):
            raise HTTPException(
                status_code=400,
                detail="Employee already has an overlapping shift on this date"
            )
    
        # 10. Default rule: evening + night on same day is forbidden
    for existing in existing_shifts:
        if {existing["category"], shift_template["category"]} == {"evening", "night"}:
            raise HTTPException(
                status_code=400,
                detail="Employee cannot work both evening and night on the same day"
            )

    #11. Default rule: morning + evening on same day is forbidden unless employee can work both and there is at least 1 hour break between them / Правило по умолчанию: утро + вечер в один день запрещены, если сотрудник не может работать оба и между ними нет перерыва минимум 1 час
    for existing in existing_shifts:
        if {existing["category"], shift_template["category"]} == {"morning", "evening"}:
            if not employee["can_work_mornings_and_evenings"]:
                raise HTTPException(
                    status_code=400,
                    detail="Employee cannot work both morning and evening on the same day"
                )

            existing["is_overnight"] = bool(existing["is_overnight"])

            cursor.execute("""
                SELECT is_split_only
                FROM shift_templates
                WHERE id = ?
            """, (existing["shift_template_id"],))
            existing_template_row = cursor.fetchone()

            existing_is_split_only = bool(existing_template_row["is_split_only"]) if existing_template_row else False
            new_is_split_only = bool(shift_template["is_split_only"])

            # Нельзя смешивать обычную и split-смену в одной паре
            if existing_is_split_only != new_is_split_only:
                raise HTTPException(
                    status_code=400,
                    detail="Morning + evening pair must use either two split shifts or two regular shifts"
                )

            existing_start, existing_end = build_shift_datetimes(entry.date, existing)
            new_start, new_end = build_shift_datetimes(entry.date, shift_template)

            break_minutes = get_break_minutes_between_shifts(
                existing_start,
                existing_end,
                new_start,
                new_end
            )

            if break_minutes < 60:
                raise HTTPException(
                    status_code=400,
                    detail="Employee must have at least 1 hour break between morning and evening shifts"
                )

def get_employee_shift_categories_on_date(connection, employee_id: int, date_string: str) -> list[str]:
    # Get list of shift categories employee has on the date / Получаем список категорий смен, которые есть у сотрудника в дату
    cursor = connection.cursor()
    cursor.execute("""
        SELECT st.category
        FROM schedule_entries se
        JOIN shift_templates st ON se.shift_template_id = st.id
        WHERE se.employee_id = ?
          AND se.date = ?
    """, (employee_id, date_string))

    return [row["category"] for row in cursor.fetchall()]


def employee_has_split_pair_on_date(connection, employee_id: int, date_string: str) -> bool:
    # Check if employee has both morning and evening split-only shifts on the date / Проверяем, есть ли у сотрудника и утренняя, и вечерняя split-only смена в дату
    cursor = connection.cursor()
    cursor.execute("""
        SELECT
            st.category,
            st.is_split_only
        FROM schedule_entries se
        JOIN shift_templates st ON se.shift_template_id = st.id
        WHERE se.employee_id = ?
          AND se.date = ?
    """, (employee_id, date_string))

    rows = cursor.fetchall()
    has_split_morning = any(
        row["category"] == "morning" and bool(row["is_split_only"])
        for row in rows
    )
    has_split_evening = any(
        row["category"] == "evening" and bool(row["is_split_only"])
        for row in rows
    )

    return has_split_morning and has_split_evening


def would_be_split_assignment(connection, employee_id: int, date_string: str, shift_template: dict) -> bool:
    # Check if assigning the shift template would create a split pair with another shift of the employee on the same date / Проверяем, создаст ли назначение шаблона пару split с другой сменой сотрудника в тот же день
    if shift_template["category"] not in ("morning", "evening"):
        return False

    if not shift_template["is_split_only"]:
        return False

    opposite_category = "evening" if shift_template["category"] == "morning" else "morning"

    existing_other_half = get_employee_existing_shift_on_date_by_category(
        connection,
        employee_id,
        date_string,
        opposite_category
    )

    return bool(existing_other_half and existing_other_half["is_split_only"])


def would_be_heavy_assignment_for_category(connection, employee_id: int, date_string: str, shift_category: str) -> bool:
    # Check if assigning a shift of the category would create a heavy day for the employee, either by being a night shift or by creating a split pair / Проверяем, создаст ли назначение смены категории тяжёлый день для сотрудника, либо будучи ночной сменой, либо создавая пару
    if shift_category == "night":
        return True

    if shift_category not in ("morning", "evening"):
        return False

    opposite_category = "evening" if shift_category == "morning" else "morning"
    paired_shift = get_employee_existing_shift_on_date_by_category(
        connection,
        employee_id,
        date_string,
        opposite_category
    )

    if paired_shift and paired_shift["is_split_only"]:
        return True

    return False


def get_employee_recent_heavy_stats(connection, employee_id: int, date_string: str) -> dict:
    # Get info about employee's recent heavy days around the date, including night shifts and split pairs in previous 3 days, and count of consecutive heavy days streak / Получаем информацию о недавних тяжёлых днях сотрудника вокруг даты, включая ночные смены и split-пары в предыдущие 3 дня, и считаем количество дней в текущей тяжёлой серии
    target_date = datetime.strptime(date_string, "%Y-%m-%d").date()

    previous_1 = (target_date - timedelta(days=1)).isoformat()
    previous_2 = (target_date - timedelta(days=2)).isoformat()
    previous_3 = (target_date - timedelta(days=3)).isoformat()

    prev1_night = employee_has_shift_category_on_date(connection, employee_id, previous_1, "night")
    prev2_night = employee_has_shift_category_on_date(connection, employee_id, previous_2, "night")
    prev3_night = employee_has_shift_category_on_date(connection, employee_id, previous_3, "night")

    prev1_split = employee_has_split_pair_on_date(connection, employee_id, previous_1)
    prev2_split = employee_has_split_pair_on_date(connection, employee_id, previous_2)
    prev3_split = employee_has_split_pair_on_date(connection, employee_id, previous_3)

    heavy_streak = 0
    for day in [previous_1, previous_2, previous_3]:
        if (
            employee_has_shift_category_on_date(connection, employee_id, day, "night")
            or employee_has_split_pair_on_date(connection, employee_id, day)
        ):
            heavy_streak += 1
        else:
            break

    return {
        "prev1_night": prev1_night,
        "prev2_night": prev2_night,
        "prev3_night": prev3_night,
        "prev1_split": prev1_split,
        "prev2_split": prev2_split,
        "prev3_split": prev3_split,
        "heavy_streak": heavy_streak,
    }


def can_assign_heavy_shift_today(connection, employee_id: int, date_string: str) -> bool:
    # Check if we can assign a heavy shift (night or split) to the employee on the date, based on recent heavy days and current day assignment / Проверяем, можно ли назначить тяжёлую смену (ночную или split) сотруднику в дату, исходя из недавних тяжёлых дней и текущих назначений
    recent = get_employee_recent_heavy_stats(connection, employee_id, date_string)

    # Не даём тяжёлые смены 3-й день подряд
    if recent["heavy_streak"] >= 2:
        return False

    return True


# =========================
# Pages / HTML pages / Страницы / HTML страницы
# =========================

@app.get("/shift-templates", tags=["Pages"])
def shift_templates_page(request: Request):
    # Render shift templates page / Отдаём страницу шаблонов смен
    return templates.TemplateResponse(
        request=request,
        name="shift_templates.html",
        context={},
    )

@app.get("/positions", tags=["Pages"])
def positions_page(request: Request):
    return templates.TemplateResponse(
        request=request,
        name="positions.html",
        context={},
    )


@app.get("/employee-positions", tags=["Pages"])
def employee_positions_page(request: Request):
    return templates.TemplateResponse(
        request=request,
        name="employee_positions.html",
        context={},
    )


@app.get("/", tags=["Pages"])
def home_page(request: Request):
    # Render home page / Отдаём главную HTML страницу
    return templates.TemplateResponse(
        request=request,
        name="index.html",
        context={},
    )


@app.get("/employees", tags=["Pages"])
def employees_page(request: Request):
    # Render employees page / Отдаём страницу сотрудников
    return templates.TemplateResponse(
        request=request,
        name="employees.html",
        context={},
    )


@app.get("/weekly-preferences", tags=["Pages"])
def weekly_preferences_page(request: Request):
    # Render weekly preferences page / Отдаём страницу недельных пожеланий
    return templates.TemplateResponse(
        request=request,
        name="weekly_preferences.html",
        context={},
    )

@app.get("/api/employee-day-statuses", tags=["Schedule"])
def get_employee_day_statuses():
    connection = get_connection()
    cursor = connection.cursor()

    cursor.execute("""
        SELECT
            eds.id,
            eds.employee_id,
            e.full_name AS employee_name,
            eds.date,
            eds.status_type
        FROM employee_day_statuses eds
        JOIN employees e ON eds.employee_id = e.id
        ORDER BY eds.date, eds.employee_id
    """)

    rows = [dict(row) for row in cursor.fetchall()]
    connection.close()
    return rows


@app.post("/api/employee-day-statuses", tags=["Schedule"])
def save_employee_day_status(status: EmployeeDayStatusCreate):
    connection = get_connection()
    cursor = connection.cursor()

    cursor.execute("SELECT id FROM employees WHERE id = ?", (status.employee_id,))
    if not cursor.fetchone():
        connection.close()
        raise HTTPException(status_code=404, detail="Employee not found")

    cursor.execute("""
        SELECT id
        FROM employee_day_statuses
        WHERE employee_id = ? AND date = ?
    """, (status.employee_id, status.date))

    existing = cursor.fetchone()

    if existing:
        cursor.execute("""
            UPDATE employee_day_statuses
            SET status_type = ?
            WHERE employee_id = ? AND date = ?
        """, (status.status_type, status.employee_id, status.date))
        saved_id = existing["id"]
    else:
        cursor.execute("""
            INSERT INTO employee_day_statuses (employee_id, date, status_type)
            VALUES (?, ?, ?)
        """, (status.employee_id, status.date, status.status_type))
        saved_id = cursor.lastrowid

    connection.commit()
    connection.close()

    return {
        "message": "Employee day status saved successfully",
        "status": {
            "id": saved_id,
            **status.model_dump()
        }
    }


@app.delete("/api/employee-day-statuses", tags=["Schedule"])
def delete_employee_day_status(employee_id: int, date: str):
    connection = get_connection()
    cursor = connection.cursor()

    cursor.execute("""
        DELETE FROM employee_day_statuses
        WHERE employee_id = ? AND date = ?
    """, (employee_id, date))

    deleted_count = cursor.rowcount
    connection.commit()
    connection.close()

    return {
        "message": "Employee day status deleted successfully",
        "deleted_count": deleted_count
    }

@app.get("/schedule", tags=["Pages"])
def schedule_page(request: Request):
    # Render schedule page / Отдаём страницу расписания
    return templates.TemplateResponse(
        request=request,
        name="schedule.html",
        context={},
    )

@app.get("/settings", tags=["Pages"])
def settings_page(request: Request):
    return templates.TemplateResponse(
        request=request,
        name="settings.html",
        context={},
    )

# =========================
# Employees API
# =========================

@app.get("/api/employees", tags=["Employees"])
def get_employees():
    # Return all employees from database / Возвращаем всех сотрудников из базы
    connection = get_connection()
    cursor = connection.cursor()

    cursor.execute("""
        SELECT *
        FROM employees
        ORDER BY id
    """)

    employees = [row_to_employee_dict(row) for row in cursor.fetchall()]
    connection.close()

    return employees


@app.post("/api/employees", tags=["Employees"])
def add_employee(employee: EmployeeCreate):
    # Save employee to database / Сохраняем сотрудника в базу
    connection = get_connection()
    cursor = connection.cursor()

    cursor.execute("""
    INSERT INTO employees (
        full_name,
        sex,
        min_shifts_per_week,
        target_shifts_per_week,
        max_shifts_per_week,
        can_work_night,
        can_work_weekends,
        can_work_evenings_after_night,
        can_work_mornings_and_evenings
    )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        employee.full_name,
        employee.sex,
        employee.min_shifts_per_week,
        employee.target_shifts_per_week,
        employee.max_shifts_per_week,
        int(employee.can_work_night),
        int(employee.can_work_weekends),
        int(employee.can_work_evenings_after_night),
        int(employee.can_work_mornings_and_evenings),
    ))

    connection.commit()
    new_employee_id = cursor.lastrowid
    connection.close()

    return {
        "message": "Employee added successfully",
        "employee": {
            "id": new_employee_id,
            **employee.model_dump(),
        },
    }


@app.put("/api/employees/{employee_id}", tags=["Employees"]) # Update employee / Обновляем сотрудника
def update_employee(employee_id: int, updated_employee: EmployeeCreate): # Employee data to update / Данные для обновления сотрудника
    # Update employee in database / Обновляем сотрудника в базе
    connection = get_connection() # Получаем соединение с базой данных
    cursor = connection.cursor() # Получаем курсор для выполнения SQL-запросов

    cursor.execute("SELECT * FROM employees WHERE id = ?", (employee_id,)) # Проверяем, что сотрудник существует
    existing_employee = cursor.fetchone() # Получаем существующего сотрудника по ID

    if not existing_employee:
        connection.close()
        raise HTTPException(status_code=404, detail="Employee not found")

    cursor.execute("""
    UPDATE employees
    SET
        full_name = ?,
        sex = ?,
        min_shifts_per_week = ?,
        target_shifts_per_week = ?,
        max_shifts_per_week = ?,
        can_work_night = ?,
        can_work_weekends = ?,
        can_work_evenings_after_night = ?,
        can_work_mornings_and_evenings = ?
    WHERE id = ?
""", (
    updated_employee.full_name,
    updated_employee.sex,
    updated_employee.min_shifts_per_week,
    updated_employee.target_shifts_per_week,
    updated_employee.max_shifts_per_week,
    int(updated_employee.can_work_night),
    int(updated_employee.can_work_weekends),
    int(updated_employee.can_work_evenings_after_night),
    int(updated_employee.can_work_mornings_and_evenings),
    employee_id,
))

    connection.commit()
    connection.close()

    return {
        "message": "Employee updated successfully",
        "employee": {
            "id": employee_id,
            **updated_employee.model_dump(),
        },
    }


@app.delete("/api/employees/{employee_id}", tags=["Employees"])
def delete_employee(employee_id: int):
    # Delete employee from database / Удаляем сотрудника из базы
    connection = get_connection()
    cursor = connection.cursor()

    cursor.execute("SELECT id FROM employees WHERE id = ?", (employee_id,))
    existing_employee = cursor.fetchone()

    if not existing_employee:
        connection.close()
        raise HTTPException(status_code=404, detail="Employee not found")

    # Delete related preferences / Удаляем связанные пожелания
    cursor.execute("DELETE FROM employee_preferences WHERE employee_id = ?", (employee_id,))
    cursor.execute("DELETE FROM employee_week_preferences WHERE employee_id = ?", (employee_id,))

    # Delete assignments first / Сначала удаляем связи
    cursor.execute("DELETE FROM employee_positions WHERE employee_id = ?", (employee_id,))

    # Delete schedule entries first / Сначала удаляем записи расписания
    cursor.execute("DELETE FROM schedule_entries WHERE employee_id = ?", (employee_id,))

    # Delete employee / Удаляем сотрудника
    cursor.execute("DELETE FROM employees WHERE id = ?", (employee_id,))

    connection.commit()
    connection.close()

    return {
        "message": "Employee deleted successfully"
    }


# =========================
# Positions API
# =========================

@app.get("/api/positions", tags=["Positions"])
def get_positions():
    # Return all positions from database / Возвращаем все должности из базы
    connection = get_connection()
    cursor = connection.cursor()

    cursor.execute("""
        SELECT *
        FROM positions
        ORDER BY id
    """)

    positions = [row_to_position_dict(row) for row in cursor.fetchall()]
    connection.close()

    return positions


@app.post("/api/positions", tags=["Positions"])
def add_position(position: PositionCreate):
    # Add position to database / Добавляем должность в базу
    connection = get_connection()
    cursor = connection.cursor()

    try:
        cursor.execute("""
            INSERT INTO positions (
                name,
                requires_continuous_coverage,
                minimum_staff_presence
            )
            VALUES (?, ?, ?)
        """, (
            position.name,
            int(position.requires_continuous_coverage),
            position.minimum_staff_presence
        ))

        connection.commit()
        new_position_id = cursor.lastrowid

    except sqlite3.IntegrityError:
        connection.close()
        raise HTTPException(status_code=400, detail="Position already exists")

    connection.close()

    return {
        "message": "Position added successfully",
        "position": {
            "id": new_position_id,
            "name": position.name,
            "requires_continuous_coverage": position.requires_continuous_coverage,
            "minimum_staff_presence": position.minimum_staff_presence,
        },
    }


# =========================
# Assignments API
# =========================

@app.get("/api/employee-positions", tags=["Assignments"])
def get_employee_positions():
    connection = get_connection()
    cursor = connection.cursor()

    cursor.execute("""
        SELECT
            ep.employee_id,
            ep.position_id,
            ep.is_primary,
            ep.priority_score,
            ep.is_fallback_only,
            e.full_name AS employee_name,
            p.name AS position_name
        FROM employee_positions ep
        JOIN employees e ON ep.employee_id = e.id
        JOIN positions p ON ep.position_id = p.id
        ORDER BY ep.employee_id, ep.position_id
    """)

    assignments = [dict(row) for row in cursor.fetchall()]
    connection.close()

    for item in assignments:
        item["is_primary"] = bool(item["is_primary"])
        item["is_fallback_only"] = bool(item["is_fallback_only"])

    return assignments


@app.post("/api/employee-positions", tags=["Assignments"])
def assign_employee_to_position(assignment: EmployeePositionCreate):
    connection = get_connection()
    cursor = connection.cursor()

    cursor.execute("SELECT id FROM employees WHERE id = ?", (assignment.employee_id,))
    employee = cursor.fetchone()
    if not employee:
        connection.close()
        raise HTTPException(status_code=404, detail="Employee not found")

    cursor.execute("SELECT id FROM positions WHERE id = ?", (assignment.position_id,))
    position = cursor.fetchone()
    if not position:
        connection.close()
        raise HTTPException(status_code=404, detail="Position not found")

    cursor.execute("""
        SELECT employee_id, position_id
        FROM employee_positions
        WHERE employee_id = ? AND position_id = ?
    """, (assignment.employee_id, assignment.position_id))

    existing_assignment = cursor.fetchone()
    if existing_assignment:
        connection.close()
        raise HTTPException(
            status_code=400,
            detail="Employee is already assigned to this position"
        )

    cursor.execute("""
        INSERT INTO employee_positions (
            employee_id,
            position_id,
            is_primary,
            priority_score,
            is_fallback_only
        )
        VALUES (?, ?, ?, ?, ?)
    """, (
        assignment.employee_id,
        assignment.position_id,
        int(assignment.is_primary),
        assignment.priority_score,
        int(assignment.is_fallback_only)
    ))

    connection.commit()
    connection.close()

    return {
        "message": "Employee assigned to position successfully",
        "assignment": assignment.model_dump(),
    }

@app.delete("/api/employee-positions", tags=["Assignments"])
def delete_employee_position(employee_id: int, position_id: int):
    # Delete employee-position assignment / Удаляем связь сотрудник-должность
    connection = get_connection()
    cursor = connection.cursor()

    cursor.execute("""
        SELECT employee_id, position_id
        FROM employee_positions
        WHERE employee_id = ? AND position_id = ?
    """, (employee_id, position_id))

    existing_assignment = cursor.fetchone()
    if not existing_assignment:
        connection.close()
        raise HTTPException(status_code=404, detail="Assignment not found")

    # Delete related schedule entries first / Сначала удаляем связанные записи расписания
    cursor.execute("""
        DELETE FROM schedule_entries
        WHERE employee_id = ? AND position_id = ?
    """, (employee_id, position_id))

    # Delete assignment / Удаляем связь
    cursor.execute("""
        DELETE FROM employee_positions
        WHERE employee_id = ? AND position_id = ?
    """, (employee_id, position_id))

    connection.commit()
    connection.close()

    return {
        "message": "Employee assignment deleted successfully"
    }


# =========================
# Shift Templates API
# =========================

@app.get("/api/shift-templates", tags=["Shift Templates"])
def get_shift_templates(active_only: bool = False):
    # Return all shift templates / Возвращаем все шаблоны смен
    connection = get_connection()
    cursor = connection.cursor()

    if active_only:
        cursor.execute("""
            SELECT *
            FROM shift_templates
            WHERE is_active = 1
            ORDER BY category, start_time, end_time
        """)
    else:
        cursor.execute("""
            SELECT *
            FROM shift_templates
            ORDER BY category, start_time, end_time
        """)

    templates_list = [row_to_shift_template_dict(row) for row in cursor.fetchall()]
    connection.close()

    return templates_list


@app.post("/api/shift-templates", tags=["Shift Templates"])
def add_shift_template(template: ShiftTemplateCreate):
    # Add shift template / Добавляем шаблон смены
    connection = get_connection()
    cursor = connection.cursor()

    try:
        cursor.execute("""
            INSERT INTO shift_templates (
                name,
                category,
                start_time,
                end_time,
                is_overnight,
                is_active,
                is_split_only
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            template.name,
            template.category,
            template.start_time,
            template.end_time,
            int(template.is_overnight),
            int(template.is_active),
            int(template.is_split_only),
        ))

        connection.commit()
        new_id = cursor.lastrowid

    except sqlite3.IntegrityError:
        connection.close()
        raise HTTPException(status_code=400, detail="Shift template already exists")

    connection.close()

    return {
        "message": "Shift template added successfully",
        "shift_template": {
            "id": new_id,
            **template.model_dump(),
        }
    }

@app.put("/api/shift-templates/{template_id}", tags=["Shift Templates"])
def update_shift_template(template_id: int, updated_template: ShiftTemplateCreate):
    # Update shift template / Обновляем шаблон смены
    connection = get_connection()
    cursor = connection.cursor()

    cursor.execute("SELECT * FROM shift_templates WHERE id = ?", (template_id,))
    existing_template = cursor.fetchone()

    if not existing_template:
        connection.close()
        raise HTTPException(status_code=404, detail="Shift template not found")

    try:
        cursor.execute("""
            UPDATE shift_templates
            SET
                name = ?,
                category = ?,
                start_time = ?,
                end_time = ?,
                is_overnight = ?,
                is_active = ?,
                is_split_only = ?
            WHERE id = ?
        """, (
            updated_template.name,
            updated_template.category,
            updated_template.start_time,
            updated_template.end_time,
            int(updated_template.is_overnight),
            int(updated_template.is_active),
            int(updated_template.is_split_only),
            template_id,
        ))

        connection.commit()

    except sqlite3.IntegrityError:
        connection.close()
        raise HTTPException(status_code=400, detail="Shift template with this name already exists")

    connection.close()

    return {
        "message": "Shift template updated successfully",
        "shift_template": {
            "id": template_id,
            **updated_template.model_dump(),
        }
    }

@app.delete("/api/shift-templates/{template_id}", tags=["Shift Templates"])
def delete_shift_template(template_id: int):
    # Delete shift template / Удаляем шаблон смены
    connection = get_connection()
    cursor = connection.cursor()

    cursor.execute("SELECT id FROM shift_templates WHERE id = ?", (template_id,)) # Check if the shift template exists / Проверяем, что шаблон смены существует
    existing_template = cursor.fetchone()

    if not existing_template:
        connection.close()
        raise HTTPException(status_code=404, detail="Shift template not found")

    # Check related schedule entries / Проверяем связанные записи расписания
    cursor.execute("""
        SELECT COUNT(*) AS count
        FROM schedule_entries
        WHERE shift_template_id = ?
    """, (template_id,))

    related_count = cursor.fetchone()["count"]

    if related_count > 0:
        connection.close()
        raise HTTPException(
            status_code=400,
            detail=f"Cannot delete shift template because it is used in {related_count} schedule entries"
        )

    # Delete shift template / Удаляем шаблон смены
    cursor.execute("DELETE FROM shift_templates WHERE id = ?", (template_id,))

    connection.commit()
    connection.close()

    return {
        "message": "Shift template deleted successfully"
    }

# =========================
# General Preferences API
# =========================

@app.get("/api/employee-preferences", tags=["Preferences"])
def get_employee_preferences():
    # Return all employee preferences / Возвращаем все общие пожелания сотрудников
    connection = get_connection()
    cursor = connection.cursor()

    cursor.execute("""
        SELECT
            ep.id,
            ep.employee_id,
            e.full_name AS employee_name,
            ep.allow_morning,
            ep.allow_evening,
            ep.allow_night,
            ep.allow_morning_evening_combo
        FROM employee_preferences ep
        JOIN employees e ON ep.employee_id = e.id
        ORDER BY ep.employee_id
    """)

    preferences = [dict(row) for row in cursor.fetchall()]
    connection.close()

    for preference in preferences:
        preference["allow_morning"] = bool(preference["allow_morning"])
        preference["allow_evening"] = bool(preference["allow_evening"])
        preference["allow_night"] = bool(preference["allow_night"])
        preference["allow_morning_evening_combo"] = bool(preference["allow_morning_evening_combo"])

    return preferences


@app.post("/api/employee-preferences", tags=["Preferences"])
def add_employee_preference(preference: EmployeePreferenceCreate):
    # Add employee preference / Добавляем общее пожелание сотрудника
    connection = get_connection()
    cursor = connection.cursor()

    # Check employee exists / Проверяем, что сотрудник существует
    cursor.execute("SELECT id FROM employees WHERE id = ?", (preference.employee_id,))
    employee = cursor.fetchone()

    if not employee:
        connection.close()
        raise HTTPException(status_code=404, detail="Employee not found")

    try:
        cursor.execute("""
            INSERT INTO employee_preferences (
                employee_id,
                allow_morning,
                allow_evening,
                allow_night,
                allow_morning_evening_combo
            )
            VALUES (?, ?, ?, ?, ?)
        """, (
            preference.employee_id,
            int(preference.allow_morning),
            int(preference.allow_evening),
            int(preference.allow_night),
            int(preference.allow_morning_evening_combo),
        ))

        connection.commit()
        new_id = cursor.lastrowid

    except sqlite3.IntegrityError:
        connection.close()
        raise HTTPException(
            status_code=400,
            detail="Preferences for this employee already exist"
        )

    connection.close()

    return {
        "message": "Employee preference added successfully",
        "preference": {
            "id": new_id,
            **preference.model_dump(),
        }
    }


# =========================
# Weekly Preferences API
# =========================

@app.get("/api/employee-week-preferences", tags=["Weekly Preferences"])
def get_employee_week_preferences(week_start_date: str):
    # Return weekly preferences for one week / Возвращаем недельные пожелания за одну неделю
    connection = get_connection()
    cursor = connection.cursor()

    cursor.execute("""
        SELECT
            ewp.id,
            ewp.employee_id,
            e.full_name AS employee_name,
            ewp.week_start_date,
            ewp.preference_date,
            ewp.preference_type
        FROM employee_week_preferences ewp
        JOIN employees e ON ewp.employee_id = e.id
        WHERE ewp.week_start_date = ?
        ORDER BY ewp.employee_id, ewp.preference_date
    """, (week_start_date,))

    preferences = [dict(row) for row in cursor.fetchall()]
    connection.close()

    return preferences


@app.post("/api/employee-week-preferences", tags=["Weekly Preferences"])
def save_employee_week_preference(preference: EmployeeWeekPreferenceCreate):
    # Save weekly employee preference / Сохраняем недельное пожелание сотрудника
    connection = get_connection()
    cursor = connection.cursor()

    # Check employee exists / Проверяем, что сотрудник существует
    cursor.execute("SELECT id FROM employees WHERE id = ?", (preference.employee_id,))
    employee = cursor.fetchone()

    if not employee:
        connection.close()
        raise HTTPException(status_code=404, detail="Employee not found")

    # Upsert preference / Обновляем или создаём пожелание
    cursor.execute("""
        SELECT id
        FROM employee_week_preferences
        WHERE employee_id = ? AND preference_date = ?
    """, (preference.employee_id, preference.preference_date))

    existing = cursor.fetchone()

    if existing:
        cursor.execute("""
            UPDATE employee_week_preferences
            SET
                week_start_date = ?,
                preference_type = ?
            WHERE employee_id = ? AND preference_date = ?
        """, (
            preference.week_start_date,
            preference.preference_type,
            preference.employee_id,
            preference.preference_date,
        ))
        saved_id = existing["id"]
    else:
        cursor.execute("""
            INSERT INTO employee_week_preferences (
                employee_id,
                week_start_date,
                preference_date,
                preference_type
            )
            VALUES (?, ?, ?, ?)
        """, (
            preference.employee_id,
            preference.week_start_date,
            preference.preference_date,
            preference.preference_type,
        ))
        saved_id = cursor.lastrowid

    connection.commit()
    connection.close()

    return {
        "message": "Weekly preference saved successfully",
        "preference": {
            "id": saved_id,
            **preference.model_dump(),
        }
    }


# =========================
# Requirements API
# =========================

@app.get("/api/shift-requirements", tags=["Requirements"])
def get_shift_requirements():
    # Return all shift requirements / Возвращаем все требования к сменам
    connection = get_connection()
    cursor = connection.cursor()

    cursor.execute("""
        SELECT
            sr.id,
            sr.position_id,
            p.name AS position_name,
            sr.shift_category,
            sr.required_total,
            sr.required_female_min
        FROM shift_requirements sr
        JOIN positions p ON sr.position_id = p.id
        ORDER BY sr.position_id, sr.shift_category
    """)

    requirements = [dict(row) for row in cursor.fetchall()]
    connection.close()

    return requirements


@app.post("/api/shift-requirements", tags=["Requirements"])
def add_shift_requirement(requirement: ShiftRequirementCreate):
    # Add shift requirement / Добавляем требование к смене
    connection = get_connection()
    cursor = connection.cursor()

    # Check position exists / Проверяем, что должность существует
    cursor.execute("SELECT id FROM positions WHERE id = ?", (requirement.position_id,))
    position = cursor.fetchone()

    if not position:
        connection.close()
        raise HTTPException(status_code=404, detail="Position not found")

    try:
        cursor.execute("""
            INSERT INTO shift_requirements (
                position_id,
                shift_category,
                required_total,
                required_female_min
            )
            VALUES (?, ?, ?, ?)
        """, (
            requirement.position_id,
            requirement.shift_category,
            requirement.required_total,
            requirement.required_female_min,
        ))

        connection.commit()
        new_id = cursor.lastrowid

    except sqlite3.IntegrityError:
        connection.close()
        raise HTTPException(
            status_code=400,
            detail="Requirement for this position and shift already exists"
        )

    connection.close()

    return {
        "message": "Shift requirement added successfully",
        "requirement": {
            "id": new_id,
            **requirement.model_dump(),
        }
    }


# =========================
# Schedule API
# =========================

@app.get("/api/schedule", tags=["Schedule"])
def get_schedule():
    # Return all schedule entries / Возвращаем все записи расписания
    connection = get_connection()
    cursor = connection.cursor()

    cursor.execute("""
        SELECT
            se.id,
            se.employee_id,
            se.position_id,
            se.date,
            se.shift_template_id,
            st.name AS shift_template_name,
            st.category AS shift_category,
            st.start_time,
            st.end_time,
            st.is_overnight
        FROM schedule_entries se
        JOIN shift_templates st ON se.shift_template_id = st.id
        ORDER BY se.date, se.employee_id, se.position_id, st.start_time
    """)

    schedule_entries = [dict(row) for row in cursor.fetchall()]
    connection.close()

    for entry in schedule_entries:
        entry["is_overnight"] = bool(entry["is_overnight"])

    return schedule_entries

# Build schedule export in Excel / Собираем экспорт расписания в Excel
@app.get("/api/schedule/export-excel", tags=["Schedule"])
# This endpoint generates an Excel file with the schedule for a given week and position. It includes employee names, their assigned shifts for each day of the week, and totals. The Excel file is styled for better readability.
def export_schedule_excel(week_start_date: str, position_id: int):
    connection = get_connection()
    cursor = connection.cursor()
    # Validate week_start_date format / Проверяем формат даты начала недели
    try:
        week_dates = build_week_dates(week_start_date)

        cursor.execute("SELECT * FROM positions WHERE id = ?", (position_id,))
        position_row = cursor.fetchone()
        if not position_row:
            raise HTTPException(status_code=404, detail="Position not found")

        position = row_to_position_dict(position_row)

        cursor.execute("""
            SELECT
                e.*,
                ep.is_primary,
                ep.priority_score,
                ep.is_fallback_only
            FROM employees e
            JOIN employee_positions ep ON ep.employee_id = e.id
            WHERE ep.position_id = ?
            ORDER BY
                ep.is_fallback_only ASC,
                ep.is_primary DESC,
                ep.priority_score DESC,
                e.full_name
        """, (position_id,))

        employees = []
        for row in cursor.fetchall():
            employee = row_to_employee_dict(row)
            employee["is_primary"] = bool(row["is_primary"])
            employee["priority_score"] = row["priority_score"]
            employee["is_fallback_only"] = bool(row["is_fallback_only"])
            employees.append(employee)
        # Get schedule entries for the position and week / Получаем записи расписания для должности и недели
        cursor.execute(f"""
            SELECT
                se.id,
                se.employee_id,
                se.position_id,
                se.date,
                se.shift_template_id,
                st.name AS shift_template_name,
                st.category AS shift_category,
                st.start_time,
                st.end_time,
                st.is_overnight
            FROM schedule_entries se
            JOIN shift_templates st ON se.shift_template_id = st.id
            WHERE se.position_id = ?
              AND se.date IN ({",".join(["?"] * len(week_dates))})
            ORDER BY se.date, se.employee_id, st.start_time, st.end_time
        """, [position_id, *week_dates])

        schedule_entries = [dict(row) for row in cursor.fetchall()]

        grouped_entries = {}
        for entry in schedule_entries:
            key = (entry["employee_id"], entry["date"])
            grouped_entries.setdefault(key, []).append(entry)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Schedule"
        worksheet.sheet_view.rightToLeft = True

        # Styles/ Стили
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        date_fill = PatternFill("solid", fgColor="F3F6F9")
        thin_side = Side(style="thin", color="CCCCCC")
        thin_border = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side
        )

        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        top_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Title / Заголовок
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
        worksheet["A1"] = f'Schedule export — {position["name"]} — week starting {week_start_date}'
        worksheet["A1"].font = Font(bold=True, size=14)
        worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

        # Headers/  Шапка
        worksheet["A3"] = "Employee"
        worksheet["B3"] = "Sunday"
        worksheet["C3"] = "Monday"
        worksheet["D3"] = "Tuesday"
        worksheet["E3"] = "Wednesday"
        worksheet["F3"] = "Thursday"
        worksheet["G3"] = "Friday"
        worksheet["H3"] = "Saturday"
        worksheet["I3"] = "Weekly total"

        for cell in worksheet[3]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border

        # Dates row/ Строка с датами
        worksheet["A4"] = ""
        for index, date in enumerate(week_dates, start=2):
            worksheet.cell(row=4, column=index, value=date)
        worksheet["I4"] = ""

        for cell in worksheet[4]:
            cell.fill = date_fill
            cell.alignment = center_alignment
            cell.border = thin_border

        # Data rows/ Данные
        start_row = 5

        for row_index, employee in enumerate(employees, start=start_row):
            worksheet.cell(row=row_index, column=1, value=employee["full_name"])
            worksheet.cell(row=row_index, column=1).font = Font(bold=True)
            worksheet.cell(row=row_index, column=1).alignment = top_alignment
            worksheet.cell(row=row_index, column=1).border = thin_border

            weekly_count = 0

            for day_offset, date in enumerate(week_dates, start=2):
                day_entries = grouped_entries.get((employee["id"], date), [])
                weekly_count += len(day_entries)

                cell_text = build_schedule_cell_text(day_entries)
                cell = worksheet.cell(row=row_index, column=day_offset, value=cell_text)
                cell.alignment = top_alignment
                cell.border = thin_border

            total_cell = worksheet.cell(row=row_index, column=9, value=weekly_count)
            total_cell.alignment = center_alignment
            total_cell.border = thin_border

        # Column widths/ Ширина колонок
        worksheet.column_dimensions["A"].width = 24
        worksheet.column_dimensions["B"].width = 22
        worksheet.column_dimensions["C"].width = 22
        worksheet.column_dimensions["D"].width = 22
        worksheet.column_dimensions["E"].width = 22
        worksheet.column_dimensions["F"].width = 22
        worksheet.column_dimensions["G"].width = 22
        worksheet.column_dimensions["H"].width = 22
        worksheet.column_dimensions["I"].width = 10

        # Freeze header/ Заморозить шапку
        worksheet.freeze_panes = "B5"

        # Row heights/ Высота строк
        for row_index, employee in enumerate(employees, start=start_row):
            max_lines = 1

            for day_offset, date in enumerate(week_dates, start=2):
                day_entries = grouped_entries.get((employee["id"], date), [])
                cell_text = build_schedule_cell_text(day_entries)

                line_count = cell_text.count("\n") + 1 if cell_text else 1
                if line_count > max_lines:
                    max_lines = line_count

            worksheet.row_dimensions[row_index].height = max(22, max_lines * 18)

                # Page setup/ Настройки страницы
        worksheet.page_setup.orientation = "landscape"

        # Pager size A4 / Размер бумаги A4
        worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4

        # Taken together, these settings will make the content fit on one page wide, and as many pages tall as needed/ Вместе эти настройки позволят поместить контент по ширине на одну страницу, а по высоте - на сколько нужно страниц
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 1

        # По высоте можно сколько нужно страниц
        worksheet.page_setup.fitToHeight = 0

        # Поля страницы
        worksheet.page_margins.left = 0.25
        worksheet.page_margins.right = 0.25
        worksheet.page_margins.top = 0.4
        worksheet.page_margins.bottom = 0.4
        worksheet.page_margins.header = 0.2
        worksheet.page_margins.footer = 0.2

        # Повторять шапку на каждой странице
        worksheet.print_title_rows = "1:4"

        # По центру по горизонтали
        worksheet.print_options.horizontalCentered = True

        # Сетка при печати выключена
        worksheet.print_options.gridLines = False
        
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        safe_position_name = position["name"].replace(" ", "_")
        filename = f'schedule_{safe_position_name}_{week_start_date}.xlsx'

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )

    finally:
        connection.close()


@app.post("/api/schedule", tags=["Schedule"])
def add_schedule_entry(entry: ScheduleEntryCreate):
    # Manual schedule entry / Ручное добавление смены без строгих ограничений
    connection = get_connection()
    cursor = connection.cursor()

    try:
        employee, position, shift_template = validate_schedule_entry_basic(connection, entry)

        cursor.execute("""
            INSERT INTO schedule_entries (
                employee_id,
                position_id,
                date,
                shift_template_id
            )
            VALUES (?, ?, ?, ?)
        """, (
            entry.employee_id,
            entry.position_id,
            entry.date,
            entry.shift_template_id
        ))

        connection.commit()

        return {
            "message": "Schedule entry added successfully",
            "schedule_entry": entry.model_dump(),
        }

    finally:
        connection.close()

def get_position_shift_count(connection, position_id: int, date: str, shift_category: str) -> int:
    cursor = connection.cursor()
    cursor.execute("""
        SELECT COUNT(*) AS cnt
        FROM schedule_entries se
        JOIN shift_templates st ON se.shift_template_id = st.id
        WHERE se.position_id = ?
          AND se.date = ?
          AND st.category = ?
    """, (position_id, date, shift_category))

    row = cursor.fetchone()
    return row["cnt"] if row else 0


def get_position_female_shift_count(connection, position_id: int, date: str, shift_category: str) -> int:
    cursor = connection.cursor()
    cursor.execute("""
        SELECT COUNT(*) AS cnt
        FROM schedule_entries se
        JOIN shift_templates st ON se.shift_template_id = st.id
        JOIN employees e ON se.employee_id = e.id
        WHERE se.position_id = ?
          AND se.date = ?
          AND st.category = ?
          AND e.sex = 'female'
    """, (position_id, date, shift_category))

    row = cursor.fetchone()
    return row["cnt"] if row else 0


def get_employee_week_category_count(connection, employee_id: int, week_start_date: str, category: str) -> int:
    cursor = connection.cursor()
    week_end_date = get_week_end_date(week_start_date)

    cursor.execute("""
        SELECT COUNT(*) AS cnt
        FROM schedule_entries se
        JOIN shift_templates st ON se.shift_template_id = st.id
        WHERE se.employee_id = ?
          AND se.date >= ?
          AND se.date <= ?
          AND st.category = ?
    """, (employee_id, week_start_date, week_end_date, category))

    row = cursor.fetchone()
    return row["cnt"] if row else 0


def get_employee_week_split_pair_count(connection, employee_id: int, week_start_date: str) -> int:
    count = 0
    for date in build_week_dates(week_start_date):
        if employee_has_split_pair_on_date(connection, employee_id, date):
            count += 1
    return count


def get_employee_week_heavy_count(connection, employee_id: int, week_start_date: str) -> int:
    return (
        get_employee_week_category_count(connection, employee_id, week_start_date, "night")
        + get_employee_week_split_pair_count(connection, employee_id, week_start_date)
    )


def get_team_min_split_count(connection, employees: list[dict], week_start_date: str) -> int:
    if not employees:
        return 0

    values = [
        get_employee_week_split_pair_count(connection, employee["id"], week_start_date)
        for employee in employees
    ]
    return min(values) if values else 0


def insert_schedule_entry_and_track(
    connection,
    cursor,
    entry: ScheduleEntryCreate,
    employee: dict,
    template: dict,
    created_entries: list[dict]
) -> None:
    cursor.execute("""
        INSERT INTO schedule_entries (
            employee_id,
            position_id,
            date,
            shift_template_id
        )
        VALUES (?, ?, ?, ?)
    """, (
        entry.employee_id,
        entry.position_id,
        entry.date,
        entry.shift_template_id
    ))

    created_entries.append({
        "employee_id": employee["id"],
        "employee_name": employee["full_name"],
        "date": entry.date,
        "shift_template_name": template["name"],
        "shift_category": template["category"]
    })


def get_best_regular_template_for_employee(
    connection,
    employee: dict,
    position_id: int,
    date: str,
    shift_category: str
):
    templates = get_sorted_templates_for_employee(
        connection,
        employee,
        date,
        shift_category
    )

    for template in templates:
        # Split-only не используем как обычную одиночную смену
        if template["is_split_only"]:
            continue

        entry = ScheduleEntryCreate(
            employee_id=employee["id"],
            position_id=position_id,
            date=date,
            shift_template_id=template["id"]
        )

        try:
            validate_schedule_entry_strict(connection, entry)
            return template
        except HTTPException:
            continue

    return None


def can_assign_split_pair_strictly(
    connection,
    employee: dict,
    position_id: int,
    date: str,
    morning_template: dict,
    evening_template: dict
) -> bool:
    """
    Проверяет split-пару строго для ОБЕИХ половин.
    Это нужно, чтобы split не обходил запрет "утро после ночи"
    и другие жёсткие ограничения.
    """
    cursor = connection.cursor()

    morning_entry = ScheduleEntryCreate(
        employee_id=employee["id"],
        position_id=position_id,
        date=date,
        shift_template_id=morning_template["id"]
    )
    evening_entry = ScheduleEntryCreate(
        employee_id=employee["id"],
        position_id=position_id,
        date=date,
        shift_template_id=evening_template["id"]
    )

    try:
        # Базово обе половины допустимы как split-части
        validate_schedule_entry_basic(
            connection,
            morning_entry,
            allow_split_without_pair=True
        )
        validate_schedule_entry_basic(
            connection,
            evening_entry,
            allow_split_without_pair=True
        )

        # Проверка утра В ПРИСУТСТВИИ вечера
        cursor.execute("""
            INSERT INTO schedule_entries (
                employee_id,
                position_id,
                date,
                shift_template_id
            )
            VALUES (?, ?, ?, ?)
        """, (
            evening_entry.employee_id,
            evening_entry.position_id,
            evening_entry.date,
            evening_entry.shift_template_id
        ))

        try:
            validate_schedule_entry_strict(connection, morning_entry)
        finally:
            cursor.execute("""
                DELETE FROM schedule_entries
                WHERE employee_id = ?
                  AND position_id = ?
                  AND date = ?
                  AND shift_template_id = ?
            """, (
                evening_entry.employee_id,
                evening_entry.position_id,
                evening_entry.date,
                evening_entry.shift_template_id
            ))

        # Проверка вечера В ПРИСУТСТВИИ утра
        cursor.execute("""
            INSERT INTO schedule_entries (
                employee_id,
                position_id,
                date,
                shift_template_id
            )
            VALUES (?, ?, ?, ?)
        """, (
            morning_entry.employee_id,
            morning_entry.position_id,
            morning_entry.date,
            morning_entry.shift_template_id
        ))

        try:
            validate_schedule_entry_strict(connection, evening_entry)
        finally:
            cursor.execute("""
                DELETE FROM schedule_entries
                WHERE employee_id = ?
                  AND position_id = ?
                  AND date = ?
                  AND shift_template_id = ?
            """, (
                morning_entry.employee_id,
                morning_entry.position_id,
                morning_entry.date,
                morning_entry.shift_template_id
            ))

        return True

    except HTTPException:
        return False
    

def get_valid_split_pair_templates_for_employee(connection, employee: dict, position_id: int, date: str):
    morning_templates = [
        template
        for template in get_shift_templates_by_category(connection, "morning")
        if template["is_split_only"]
    ]
    evening_templates = [
        template
        for template in get_shift_templates_by_category(connection, "evening")
        if template["is_split_only"]
    ]

    if not morning_templates or not evening_templates:
        return None, None

    for morning_template in morning_templates:
        for evening_template in evening_templates:
            morning_start, morning_end = build_shift_datetimes(date, morning_template)
            evening_start, evening_end = build_shift_datetimes(date, evening_template)

            break_minutes = get_break_minutes_between_shifts(
                morning_start,
                morning_end,
                evening_start,
                evening_end
            )

            if break_minutes < 60:
                continue

            if can_assign_split_pair_strictly(
                connection=connection,
                employee=employee,
                position_id=position_id,
                date=date,
                morning_template=morning_template,
                evening_template=evening_template
            ):
                return morning_template, evening_template

    return None, None


def assign_split_pair(
    connection,
    cursor,
    employee: dict,
    position_id: int,
    date: str,
    morning_template: dict,
    evening_template: dict,
    created_entries: list[dict]
) -> bool:
    morning_entry = ScheduleEntryCreate(
        employee_id=employee["id"],
        position_id=position_id,
        date=date,
        shift_template_id=morning_template["id"]
    )
    evening_entry = ScheduleEntryCreate(
        employee_id=employee["id"],
        position_id=position_id,
        date=date,
        shift_template_id=evening_template["id"]
    )

    # Сначала строго убеждаемся, что пара целиком валидна
    if not can_assign_split_pair_strictly(
        connection=connection,
        employee=employee,
        position_id=position_id,
        date=date,
        morning_template=morning_template,
        evening_template=evening_template
    ):
        return False

    try:
        cursor.execute("""
            INSERT INTO schedule_entries (
                employee_id,
                position_id,
                date,
                shift_template_id
            )
            VALUES (?, ?, ?, ?)
        """, (
            morning_entry.employee_id,
            morning_entry.position_id,
            morning_entry.date,
            morning_entry.shift_template_id
        ))

        cursor.execute("""
            INSERT INTO schedule_entries (
                employee_id,
                position_id,
                date,
                shift_template_id
            )
            VALUES (?, ?, ?, ?)
        """, (
            evening_entry.employee_id,
            evening_entry.position_id,
            evening_entry.date,
            evening_entry.shift_template_id
        ))

        created_entries.append({
            "employee_id": employee["id"],
            "employee_name": employee["full_name"],
            "date": date,
            "shift_template_name": morning_template["name"],
            "shift_category": morning_template["category"]
        })
        created_entries.append({
            "employee_id": employee["id"],
            "employee_name": employee["full_name"],
            "date": date,
            "shift_template_name": evening_template["name"],
            "shift_category": evening_template["category"]
        })

        return True

    except Exception:
        cursor.execute("""
            DELETE FROM schedule_entries
            WHERE employee_id = ?
              AND position_id = ?
              AND date = ?
              AND shift_template_id IN (?, ?)
        """, (
            employee["id"],
            position_id,
            date,
            morning_template["id"],
            evening_template["id"]
        ))
        return False


def get_employee_capacity_snapshot(connection, employee: dict, week_start_date: str, date: str) -> dict:
    week_count = get_employee_week_shift_count(connection, employee["id"], week_start_date)
    worked_days = get_employee_week_worked_days(connection, employee["id"], week_start_date)
    split_count = get_employee_week_split_pair_count(connection, employee["id"], week_start_date)
    night_count = get_employee_week_category_count(connection, employee["id"], week_start_date, "night")
    heavy_count = get_employee_week_heavy_count(connection, employee["id"], week_start_date)
    recent = get_employee_recent_heavy_stats(connection, employee["id"], date)

    already_has_shift_today = employee_has_shift_on_date(connection, employee["id"], date)
    projected_worked_days = worked_days if already_has_shift_today else worked_days + 1
    remaining_to_min = max(0, employee["min_shifts_per_week"] - week_count)
    remaining_to_target = max(0, employee["target_shifts_per_week"] - week_count)
    remaining_capacity = employee["max_shifts_per_week"] - week_count

    return {
        "week_count": week_count,
        "worked_days": worked_days,
        "split_count": split_count,
        "night_count": night_count,
        "heavy_count": heavy_count,
        "recent": recent,
        "already_has_shift_today": already_has_shift_today,
        "projected_worked_days": projected_worked_days,
        "remaining_to_min": remaining_to_min,
        "remaining_to_target": remaining_to_target,
        "remaining_capacity": remaining_capacity,
    }


def get_day_open_slots(connection, position_id: int, date: str, requirement_map: dict) -> dict:
    result = {}

    for category in ("morning", "evening", "night"):
        if category not in requirement_map:
            continue

        required_total = requirement_map[category]["required_total"]
        required_female = requirement_map[category]["required_female_min"]

        current_total = get_position_shift_count(connection, position_id, date, category)
        current_female = get_position_female_shift_count(connection, position_id, date, category)

        result[category] = {
            "missing_total": max(0, required_total - current_total),
            "missing_female": max(0, required_female - current_female),
        }

    return result


def get_regular_candidate_score(
    connection,
    employee: dict,
    week_start_date: str,
    date: str,
    shift_category: str,
    require_female: bool,
    strict_mode: bool = True
) -> tuple:
    snap = get_employee_capacity_snapshot(connection, employee, week_start_date, date)
    recent = snap["recent"]

    female_penalty = 0
    if require_female and employee["sex"] != "female":
        female_penalty = 100000

    # Кто ещё не добрал min — самый высокий приоритет
    if snap["remaining_to_min"] > 0:
        bucket = 0
    elif snap["remaining_to_target"] > 0:
        bucket = 1
    else:
        bucket = 2

    reserve_penalty = 0
    if snap["remaining_capacity"] <= 0:
        reserve_penalty += 100000
    elif snap["remaining_capacity"] == 1:
        reserve_penalty += 300
    elif snap["remaining_capacity"] == 2:
        reserve_penalty += 120

    if snap["projected_worked_days"] >= 6:
        reserve_penalty += 180

    heavy_penalty = 0
    if shift_category == "night":
        heavy_penalty += snap["night_count"] * 150
        if recent["prev1_night"]:
            heavy_penalty += 350 if strict_mode else 180
        if recent["prev1_split"]:
            heavy_penalty += 260 if strict_mode else 120
        if recent["heavy_streak"] >= 2:
            heavy_penalty += 100000 if strict_mode else 500
    else:
        if recent["prev1_night"]:
            heavy_penalty += 120 if strict_mode else 60
        if recent["prev1_split"]:
            heavy_penalty += 100 if strict_mode else 50
        if recent["heavy_streak"] >= 2:
            heavy_penalty += 700 if strict_mode else 200

    fallback_penalty = 250 if employee.get("is_fallback_only") else 0
    primary_bonus = -1 if employee.get("is_primary") else 0

    return (
        female_penalty,
        bucket,
        reserve_penalty,
        heavy_penalty,
        snap["week_count"],
        snap["worked_days"],
        snap["split_count"],
        snap["night_count"],
        fallback_penalty,
        primary_bonus,
        -int(employee.get("priority_score", 50)),
        employee["id"],
    )


def get_split_candidate_score(
    connection,
    employee: dict,
    employees: list[dict],
    week_start_date: str,
    date: str,
    strict_mode: bool = True
) -> tuple:
    snap = get_employee_capacity_snapshot(connection, employee, week_start_date, date)
    recent = snap["recent"]

    if snap["week_count"] + 2 > employee["max_shifts_per_week"]:
        return (1000000,)

    if snap["remaining_to_min"] > 0:
        bucket = 0
    elif snap["remaining_to_target"] > 0:
        bucket = 1
    else:
        bucket = 2

    team_min_split = get_team_min_split_count(connection, employees, week_start_date)
    split_excess = max(0, snap["split_count"] - team_min_split)

    reserve_penalty = 0
    if snap["remaining_capacity"] <= 1:
        reserve_penalty += 100000
    elif snap["remaining_capacity"] == 2:
        reserve_penalty += 220
    elif snap["remaining_capacity"] == 3:
        reserve_penalty += 90

    if snap["projected_worked_days"] >= 6:
        reserve_penalty += 220

    heavy_penalty = 0
    if recent["prev1_split"]:
        heavy_penalty += 360 if strict_mode else 180
    if recent["prev1_night"]:
        heavy_penalty += 320 if strict_mode else 160
    if recent["heavy_streak"] >= 2:
        heavy_penalty += 100000 if strict_mode else 600

    fallback_penalty = 300 if employee.get("is_fallback_only") else 0
    primary_bonus = -1 if employee.get("is_primary") else 0

    return (
        bucket,
        split_excess,
        reserve_penalty,
        heavy_penalty,
        snap["week_count"],
        snap["worked_days"],
        snap["split_count"],
        snap["night_count"],
        fallback_penalty,
        primary_bonus,
        -int(employee.get("priority_score", 50)),
        employee["id"],
    )


def get_employee_week_context(connection, employee: dict, week_start_date: str, date: str) -> dict:
    current_week_count = get_employee_week_shift_count(connection, employee["id"], week_start_date)
    current_worked_days = get_employee_week_worked_days(connection, employee["id"], week_start_date)
    current_split_count = get_employee_week_split_pair_count(connection, employee["id"], week_start_date)
    current_night_count = get_employee_week_category_count(connection, employee["id"], week_start_date, "night")
    current_heavy_count = get_employee_week_heavy_count(connection, employee["id"], week_start_date)

    already_has_shift_today = employee_has_shift_on_date(connection, employee["id"], date)
    projected_worked_days = current_worked_days if already_has_shift_today else current_worked_days + 1
    consecutive_days = get_employee_consecutive_work_days_around_date(connection, employee["id"], date)
    recent = get_employee_recent_heavy_stats(connection, employee["id"], date)

    return {
        "week_count": current_week_count,
        "worked_days": current_worked_days,
        "split_count": current_split_count,
        "night_count": current_night_count,
        "heavy_count": current_heavy_count,
        "already_has_shift_today": already_has_shift_today,
        "projected_worked_days": projected_worked_days,
        "consecutive_days": consecutive_days,
        "recent": recent,
        "remaining_to_min": max(0, employee["min_shifts_per_week"] - current_week_count),
        "remaining_to_target": max(0, employee["target_shifts_per_week"] - current_week_count),
        "remaining_capacity": employee["max_shifts_per_week"] - current_week_count,
    }


def get_regular_candidate_priority(
    connection,
    employee: dict,
    week_start_date: str,
    date: str,
    shift_category: str,
    require_female: bool = False
) -> tuple:
    ctx = get_employee_week_context(connection, employee, week_start_date, date)
    recent = ctx["recent"]

    female_penalty = 0
    if require_female and employee["sex"] != "female":
        female_penalty = 100000

    # 0 = надо срочно добирать до min
    # 1 = надо добирать до target
    # 2 = уже на/выше target
    if ctx["week_count"] < employee["min_shifts_per_week"]:
        bucket = 0
    elif ctx["week_count"] < employee["target_shifts_per_week"]:
        bucket = 1
    else:
        bucket = 2

    heavy_penalty = 0
    if shift_category == "night":
        heavy_penalty += ctx["night_count"] * 180
        if recent["prev1_night"]:
            heavy_penalty += 350
        if recent["prev1_split"]:
            heavy_penalty += 240
        if recent["heavy_streak"] >= 2:
            heavy_penalty += 100000
    else:
        if recent["prev1_night"]:
            heavy_penalty += 80
        if recent["prev1_split"]:
            heavy_penalty += 60
        if recent["heavy_streak"] >= 2:
            heavy_penalty += 600

    reserve_penalty = 0
    if ctx["remaining_capacity"] <= 0:
        reserve_penalty += 100000
    elif ctx["remaining_capacity"] == 1:
        reserve_penalty += 250
    elif ctx["remaining_capacity"] == 2:
        reserve_penalty += 90

    if ctx["projected_worked_days"] >= 6:
        reserve_penalty += 120

    if ctx["consecutive_days"] >= 6:
        reserve_penalty += 300
    elif ctx["consecutive_days"] == 5:
        reserve_penalty += 120
    elif ctx["consecutive_days"] == 4:
        reserve_penalty += 40

    fallback_penalty = 250 if employee.get("is_fallback_only") else 0
    primary_bonus_bucket = -1 if employee.get("is_primary") else 0

    return (
        female_penalty,
        bucket,
        reserve_penalty,
        heavy_penalty,
        ctx["week_count"],
        ctx["worked_days"],
        ctx["split_count"],
        fallback_penalty,
        primary_bonus_bucket,
        -int(employee.get("priority_score", 50)),
        employee["id"],
    )


def get_split_candidate_priority(
    connection,
    employee: dict,
    employees: list[dict],
    week_start_date: str,
    date: str
) -> tuple:
    ctx = get_employee_week_context(connection, employee, week_start_date, date)
    recent = ctx["recent"]

    if ctx["week_count"] + 2 > employee["max_shifts_per_week"]:
        return (1000000,)

    # 0 = ниже min, 1 = ниже target, 2 = уже на/выше target
    if ctx["week_count"] < employee["min_shifts_per_week"]:
        bucket = 0
    elif ctx["week_count"] < employee["target_shifts_per_week"]:
        bucket = 1
    else:
        bucket = 2

    team_min_split = get_team_min_split_count(connection, employees, week_start_date)
    split_excess = max(0, ctx["split_count"] - team_min_split)

    heavy_penalty = 0
    if recent["prev1_split"]:
        heavy_penalty += 320
    if recent["prev1_night"]:
        heavy_penalty += 280
    if recent["heavy_streak"] >= 2:
        heavy_penalty += 100000

    reserve_penalty = 0
    if ctx["remaining_capacity"] <= 1:
        reserve_penalty += 100000
    elif ctx["remaining_capacity"] == 2:
        reserve_penalty += 180
    elif ctx["remaining_capacity"] == 3:
        reserve_penalty += 60

    if ctx["projected_worked_days"] >= 6:
        reserve_penalty += 180

    fallback_penalty = 300 if employee.get("is_fallback_only") else 0
    primary_bonus_bucket = -1 if employee.get("is_primary") else 0

    return (
        bucket,
        split_excess,
        ctx["split_count"],
        ctx["heavy_count"],
        reserve_penalty,
        heavy_penalty,
        ctx["week_count"],
        ctx["worked_days"],
        ctx["night_count"],
        fallback_penalty,
        primary_bonus_bucket,
        -int(employee.get("priority_score", 50)),
        employee["id"],
    )


def find_best_regular_candidate(
    connection,
    employees: list[dict],
    position_id: int,
    week_start_date: str,
    date: str,
    shift_category: str,
    require_female: bool = False,
    strict_mode: bool = True
):
    candidates = []

    for employee in employees:
        if require_female and employee["sex"] != "female":
            continue

        snap = get_employee_capacity_snapshot(connection, employee, week_start_date, date)
        if snap["remaining_capacity"] <= 0:
            continue

        template = get_best_regular_template_for_employee(
            connection=connection,
            employee=employee,
            position_id=position_id,
            date=date,
            shift_category=shift_category
        )

        if template is None:
            continue

        score = get_regular_candidate_score(
            connection=connection,
            employee=employee,
            week_start_date=week_start_date,
            date=date,
            shift_category=shift_category,
            require_female=require_female,
            strict_mode=strict_mode
        )

        candidates.append((score, employee, template))

    if not candidates:
        return None, None, None

    candidates.sort(key=lambda item: item[0])
    score, employee, template = candidates[0]
    return employee, template, score


def find_best_split_candidate(
    connection,
    employees: list[dict],
    position_id: int,
    week_start_date: str,
    date: str,
    strict_mode: bool = True
):
    candidates = []

    for employee in employees:
        if not employee["can_work_mornings_and_evenings"]:
            continue

        morning_template, evening_template = get_valid_split_pair_templates_for_employee(
            connection=connection,
            employee=employee,
            position_id=position_id,
            date=date
        )

        if morning_template is None or evening_template is None:
            continue

        score = get_split_candidate_score(
            connection=connection,
            employee=employee,
            employees=employees,
            week_start_date=week_start_date,
            date=date,
            strict_mode=strict_mode
        )

        candidates.append((score, employee, morning_template, evening_template))

    if not candidates:
        return None, None, None, None

    candidates.sort(key=lambda item: item[0])
    score, employee, morning_template, evening_template = candidates[0]
    return employee, morning_template, evening_template, score



def try_fill_one_regular_slot(
    connection,
    cursor,
    employees: list[dict],
    position_id: int,
    week_start_date: str,
    date: str,
    shift_category: str,
    created_entries: list[dict],
    require_female: bool = False,
    strict_mode: bool = True
) -> bool:
    employee, template, _ = find_best_regular_candidate(
        connection=connection,
        employees=employees,
        position_id=position_id,
        week_start_date=week_start_date,
        date=date,
        shift_category=shift_category,
        require_female=require_female,
        strict_mode=strict_mode
    )

    if employee is None or template is None:
        return False

    entry = ScheduleEntryCreate(
        employee_id=employee["id"],
        position_id=position_id,
        date=date,
        shift_template_id=template["id"]
    )

    validate_schedule_entry_strict(connection, entry)
    insert_schedule_entry_and_track(connection, cursor, entry, employee, template, created_entries)
    return True


def try_fill_one_split_pair(
    connection,
    cursor,
    employees: list[dict],
    position_id: int,
    week_start_date: str,
    date: str,
    created_entries: list[dict],
    strict_mode: bool = True
) -> bool:
    employee, morning_template, evening_template, _ = find_best_split_candidate(
        connection=connection,
        employees=employees,
        position_id=position_id,
        week_start_date=week_start_date,
        date=date,
        strict_mode=strict_mode
    )

    if employee is None:
        return False

    return assign_split_pair(
        connection=connection,
        cursor=cursor,
        employee=employee,
        position_id=position_id,
        date=date,
        morning_template=morning_template,
        evening_template=evening_template,
        created_entries=created_entries
    )


def fill_day_with_relaxation(
    connection,
    cursor,
    employees: list[dict],
    position_id: int,
    week_start_date: str,
    date: str,
    requirement_map: dict,
    created_entries: list[dict],
    errors: list[str]
):
    # PASS 1 — strict
    for strict_mode in (True, False):
        # 1. Night
        if "night" in requirement_map:
            while get_position_shift_count(connection, position_id, date, "night") < requirement_map["night"]["required_total"]:
                assigned = try_fill_one_regular_slot(
                    connection=connection,
                    cursor=cursor,
                    employees=employees,
                    position_id=position_id,
                    week_start_date=week_start_date,
                    date=date,
                    shift_category="night",
                    created_entries=created_entries,
                    require_female=False,
                    strict_mode=strict_mode
                )
                if not assigned:
                    break

        # 2. Female minimum morning
        if "morning" in requirement_map:
            while get_position_female_shift_count(connection, position_id, date, "morning") < requirement_map["morning"]["required_female_min"]:
                assigned = try_fill_one_regular_slot(
                    connection=connection,
                    cursor=cursor,
                    employees=employees,
                    position_id=position_id,
                    week_start_date=week_start_date,
                    date=date,
                    shift_category="morning",
                    created_entries=created_entries,
                    require_female=True,
                    strict_mode=strict_mode
                )
                if not assigned:
                    break

        # 3. Female minimum evening
        if "evening" in requirement_map:
            while get_position_female_shift_count(connection, position_id, date, "evening") < requirement_map["evening"]["required_female_min"]:
                assigned = try_fill_one_regular_slot(
                    connection=connection,
                    cursor=cursor,
                    employees=employees,
                    position_id=position_id,
                    week_start_date=week_start_date,
                    date=date,
                    shift_category="evening",
                    created_entries=created_entries,
                    require_female=True,
                    strict_mode=strict_mode
                )
                if not assigned:
                    break

        # 4. Split when both morning and evening are missing
        if "morning" in requirement_map and "evening" in requirement_map:
            while True:
                open_slots = get_day_open_slots(connection, position_id, date, requirement_map)
                if open_slots["morning"]["missing_total"] <= 0 or open_slots["evening"]["missing_total"] <= 0:
                    break

                assigned = try_fill_one_split_pair(
                    connection=connection,
                    cursor=cursor,
                    employees=employees,
                    position_id=position_id,
                    week_start_date=week_start_date,
                    date=date,
                    created_entries=created_entries,
                    strict_mode=strict_mode
                )
                if not assigned:
                    break

        # 5. Regular morning
        if "morning" in requirement_map:
            while get_position_shift_count(connection, position_id, date, "morning") < requirement_map["morning"]["required_total"]:
                assigned = try_fill_one_regular_slot(
                    connection=connection,
                    cursor=cursor,
                    employees=employees,
                    position_id=position_id,
                    week_start_date=week_start_date,
                    date=date,
                    shift_category="morning",
                    created_entries=created_entries,
                    require_female=False,
                    strict_mode=strict_mode
                )
                if not assigned:
                    break

        # 6. Regular evening
        if "evening" in requirement_map:
            while get_position_shift_count(connection, position_id, date, "evening") < requirement_map["evening"]["required_total"]:
                assigned = try_fill_one_regular_slot(
                    connection=connection,
                    cursor=cursor,
                    employees=employees,
                    position_id=position_id,
                    week_start_date=week_start_date,
                    date=date,
                    shift_category="evening",
                    created_entries=created_entries,
                    require_female=False,
                    strict_mode=strict_mode
                )
                if not assigned:
                    break

    # Final warnings
    for category in ("morning", "evening", "night"):
        if category not in requirement_map:
            continue

        current_total = get_position_shift_count(connection, position_id, date, category)
        required_total = requirement_map[category]["required_total"]

        if current_total < required_total:
            errors.append(f"{category.capitalize()} still underfilled on {date}: {current_total}/{required_total}")


@app.post("/api/schedule/auto-generate", tags=["Schedule"])
def auto_generate_schedule(request_data: AutoGenerateScheduleRequest):
    connection = get_connection()
    cursor = connection.cursor()

    week_dates = build_week_dates(request_data.week_start_date)

    cleaned_orphans = cleanup_orphan_split_entries_for_position_week(
        connection,
        request_data.position_id,
        week_dates
    )

    cursor.execute("""
        SELECT
            e.*,
            ep.is_primary,
            ep.priority_score,
            ep.is_fallback_only
        FROM employees e
        JOIN employee_positions ep ON ep.employee_id = e.id
        WHERE ep.position_id = ?
        ORDER BY
            ep.is_fallback_only ASC,
            ep.is_primary DESC,
            ep.priority_score DESC,
            e.id
    """, (request_data.position_id,))

    employees = []
    for row in cursor.fetchall():
        employee = row_to_employee_dict(row)
        employee["is_primary"] = bool(row["is_primary"])
        employee["priority_score"] = row["priority_score"]
        employee["is_fallback_only"] = bool(row["is_fallback_only"])
        employees.append(employee)

    if not employees:
        connection.close()
        raise HTTPException(status_code=400, detail="No employees assigned to this position")

    cursor.execute("""
        SELECT *
        FROM shift_requirements
        WHERE position_id = ?
    """, (request_data.position_id,))
    requirements = [dict(row) for row in cursor.fetchall()]

    if not requirements:
        connection.close()
        raise HTTPException(status_code=400, detail="No shift requirements found for this position")

    requirement_map = {item["shift_category"]: item for item in requirements}

    created_entries = []
    errors = []

    try:
        for date in week_dates:
            fill_day_with_relaxation(
                connection=connection,
                cursor=cursor,
                employees=employees,
                position_id=request_data.position_id,
                week_start_date=request_data.week_start_date,
                date=date,
                requirement_map=requirement_map,
                created_entries=created_entries,
                errors=errors
            )

        # Post-clean orphan split entries one more time
        cleaned_orphans_after = cleanup_orphan_split_entries_for_position_week(
            connection,
            request_data.position_id,
            week_dates
        )

        if cleaned_orphans > 0:
            errors.insert(0, f"Removed {cleaned_orphans} orphan split-only entries before generation")

        if cleaned_orphans_after > 0:
            errors.append(f"Removed {cleaned_orphans_after} orphan split-only entries after generation")

        connection.commit()

        return {
            "message": "Auto-generation finished",
            "created_count": len(created_entries),
            "created_entries": created_entries,
            "errors": errors
        }

    finally:
        connection.close()


@app.delete("/api/schedule/{schedule_entry_id}", tags=["Schedule"])
def delete_schedule_entry(schedule_entry_id: int):
    # Manual delete / Ручное удаление смены.
    # Если удаляется половина split-пары, удаляем и вторую половину, чтобы не оставлять "сироту".
    connection = get_connection()
    cursor = connection.cursor()

    cursor.execute("""
        SELECT
            se.id,
            se.employee_id,
            se.position_id,
            se.date,
            se.shift_template_id,
            st.category,
            st.is_split_only
        FROM schedule_entries se
        JOIN shift_templates st ON se.shift_template_id = st.id
        WHERE se.id = ?
    """, (schedule_entry_id,))
    entry_row = cursor.fetchone()

    if not entry_row:
        connection.close()
        raise HTTPException(status_code=404, detail="Schedule entry not found")

    entry = dict(entry_row)
    deleted_ids = [schedule_entry_id]

    # Если это split-only смена, ищем вторую половину пары в тот же день
    if bool(entry["is_split_only"]) and entry["category"] in ("morning", "evening"):
        opposite_category = "evening" if entry["category"] == "morning" else "morning"

        cursor.execute("""
            SELECT
                se.id
            FROM schedule_entries se
            JOIN shift_templates st ON se.shift_template_id = st.id
            WHERE se.employee_id = ?
              AND se.position_id = ?
              AND se.date = ?
              AND st.category = ?
              AND st.is_split_only = 1
            LIMIT 1
        """, (
            entry["employee_id"],
            entry["position_id"],
            entry["date"],
            opposite_category
        ))

        pair_row = cursor.fetchone()
        if pair_row:
            deleted_ids.append(pair_row["id"])

    cursor.execute(f"""
        DELETE FROM schedule_entries
        WHERE id IN ({",".join(["?"] * len(deleted_ids))})
    """, deleted_ids)

    deleted_count = cursor.rowcount

    connection.commit()
    connection.close()

    return {
        "message": "Schedule entry deleted successfully",
        "deleted_count": deleted_count
    }

@app.post("/api/schedule/clear-week", tags=["Schedule"])
def clear_week_schedule(request_data: ClearWeekScheduleRequest):
    # Delete schedule entries for one position and one week / Удаляем расписание за неделю для одной должности
    connection = get_connection()
    cursor = connection.cursor()

    week_dates = build_week_dates(request_data.week_start_date)

    cursor.execute(f"""
        DELETE FROM schedule_entries
        WHERE position_id = ?
          AND date IN ({",".join(["?"] * len(week_dates))})
    """, [request_data.position_id, *week_dates])

    deleted_count = cursor.rowcount

    connection.commit()
    connection.close()

    return {
        "message": "Week schedule cleared successfully",
        "deleted_count": deleted_count
    }