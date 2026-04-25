from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


EXPORT_LABELS = {
    "en": {
        "schedule_title": "Schedule export",
        "week_starting": "week starting",
        "employee": "Employee",
        "row_number": "#",
        "weekly_total": "Total shifts / hours",
        "summary_title": "Coordinator summary",
        "position": "Position",
        "positions": "Positions",
        "week_start": "Week start",
        "assigned_employees": "Assigned employees",
        "schedule_entries": "Schedule entries",
        "worked_shifts": "Worked shifts",
        "no_show_entries": "No-show entries",
        "sick_days": "Sick days",
        "day_off_statuses": "Day-off statuses",
        "sick": "Sick",
        "day_off": "Day off",
        "no_show": "No-show",
        "hours": "h",
        "legend_title": "Legend",
        "position_colors": "Position colors",
        "status_colors": "Status colors",
        "cross_position_note": "A shift from another position is prefixed with that position name.",
        "row_number_note": "The # column is a separate employee row number for this export.",
        "weekdays": ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"],
    },
    "ru": {
        "schedule_title": "Экспорт расписания",
        "week_starting": "неделя с",
        "employee": "Сотрудник",
        "row_number": "№",
        "weekly_total": "Итого смен / часов",
        "summary_title": "Сводка координатора",
        "position": "Должность",
        "positions": "Должности",
        "week_start": "Начало недели",
        "assigned_employees": "Назначенные сотрудники",
        "schedule_entries": "Записи расписания",
        "worked_shifts": "Отработанные смены",
        "no_show_entries": "Неявки",
        "sick_days": "Больничные дни",
        "day_off_statuses": "Выходные",
        "sick": "Больничный",
        "day_off": "Выходной",
        "no_show": "Неявка",
        "hours": "ч",
        "legend_title": "Легенда",
        "position_colors": "Цвета должностей",
        "status_colors": "Цвета статусов",
        "cross_position_note": "Смена из другой должности подписывается названием этой должности.",
        "row_number_note": "Колонка № - отдельная нумерация строк сотрудников в этом экспорте.",
        "weekdays": ["Воскресенье", "Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"],
    },
    "he": {
        "schedule_title": "ייצוא סידור עבודה",
        "week_starting": "שבוע שמתחיל",
        "employee": "עובד",
        "row_number": "#",
        "weekly_total": "סה\"כ משמרות / שעות",
        "summary_title": "סיכום מתאם",
        "position": "תפקיד",
        "positions": "תפקידים",
        "week_start": "תחילת שבוע",
        "assigned_employees": "עובדים משויכים",
        "schedule_entries": "רשומות סידור",
        "worked_shifts": "משמרות עבודה",
        "no_show_entries": "אי הגעה",
        "sick_days": "ימי מחלה",
        "day_off_statuses": "ימי חופש",
        "sick": "מחלה",
        "day_off": "יום חופשי",
        "no_show": "אי הגעה",
        "hours": "ש'",
        "legend_title": "מקרא",
        "position_colors": "צבעי תפקידים",
        "status_colors": "צבעי סטטוסים",
        "cross_position_note": "משמרת מתפקיד אחר מסומנת בשם התפקיד בתחילת הטקסט.",
        "row_number_note": "עמודת # היא מספור שורות עובדים נפרד לייצוא הזה.",
        "weekdays": ["ראשון", "שני", "שלישי", "רביעי", "חמישי", "שישי", "שבת"],
    },
}


def get_export_label(key: str, lang: str) -> str:
    labels = EXPORT_LABELS.get(lang, EXPORT_LABELS["en"])
    return labels.get(key, key)


def get_weekday_labels(lang: str) -> list[str]:
    return EXPORT_LABELS.get(lang, EXPORT_LABELS["en"])["weekdays"]


def normalize_excel_color(value: str | None, fallback: str = "EFF6FF") -> str:
    raw = str(value or "").strip()
    if raw.startswith("#"):
        raw = raw[1:]
    if len(raw) == 6 and all(character in "0123456789abcdefABCDEF" for character in raw):
        raw = raw.upper()
    else:
        raw = fallback
    return f"FF{raw}" if len(raw) == 6 else raw


def time_to_minutes(value: str) -> int:
    hours, minutes = value.split(":")
    return int(hours) * 60 + int(minutes)


def entry_duration_minutes(entry: dict) -> int:
    start = time_to_minutes(entry["start_time"])
    end = time_to_minutes(entry["end_time"])
    if entry.get("is_overnight") or end <= start:
        end += 24 * 60
    return end - start


def format_hours(minutes: int, lang: str) -> str:
    if minutes % 60 == 0:
        value = str(minutes // 60)
    else:
        value = f"{minutes / 60:.1f}".rstrip("0").rstrip(".")
    return f"{value}{get_export_label('hours', lang)}"


def format_total(shift_count: int, minutes: int, lang: str) -> str:
    return f"{shift_count} / {format_hours(minutes, lang)}"


def build_entry_line(entry: dict, current_position_id: int, lang: str) -> tuple[str, str, bool]:
    shift_text = entry["shift_template_name"]
    position_color = None
    if entry["position_id"] != current_position_id:
        shift_text = f"{entry.get('position_name') or get_export_label('position', lang)}: {shift_text}"
        position_color = entry.get("position_color")
    return shift_text, position_color, bool(entry.get("no_show"))


def build_plain_text(parts: list[str]) -> str:
    return "\n".join(part for part in parts if part)


def build_schedule_cell_payload(
    entries: list[dict],
    current_position_id: int,
    day_status: dict | None = None,
    lang: str = "en",
) -> dict:
    sorted_entries = sorted(entries, key=lambda item: item["start_time"])
    status_type = day_status.get("status_type") if day_status else None

    if status_type == "day_off" and not sorted_entries:
        status_text = get_export_label("day_off", lang)
        return {
            "text": status_text,
            "lines": [{"text": status_text, "highlight_color": None}],
            "worked_count": 0,
            "worked_minutes": 0,
        }

    plain_lines = []
    formatted_lines = []
    for index, entry in enumerate(sorted_entries):
        entry_text, position_color, is_no_show = build_entry_line(entry, current_position_id, lang)
        line_parts = [entry_text]
        if is_no_show:
            status_text = get_export_label("no_show", lang)
            line_parts.append(status_text)
        line_text = " - ".join(line_parts)
        plain_lines.append(line_text)
        formatted_lines.append({"text": line_text, "highlight_color": position_color})

    if status_type in {"sick", "day_off"}:
        status_text = get_export_label(status_type, lang)
        plain_lines.append(status_text)
        formatted_lines.append({"text": status_text, "highlight_color": None})

    worked_entries = [
        entry
        for entry in sorted_entries
        if not entry.get("no_show") and status_type not in {"sick", "day_off"}
    ]
    plain_text = build_plain_text(plain_lines)
    return {
        "text": plain_text,
        "lines": formatted_lines,
        "worked_count": len(worked_entries),
        "worked_minutes": sum(entry_duration_minutes(entry) for entry in worked_entries),
    }


def choose_status_fill(day_status: dict | None = None) -> PatternFill | None:
    status_type = day_status.get("status_type") if day_status else None
    if status_type == "sick":
        return PatternFill("solid", fgColor=normalize_excel_color("FEE2E2"))
    if status_type == "day_off":
        return PatternFill("solid", fgColor=normalize_excel_color("F1F5F9"))
    return None


def build_unique_employees_by_priority_position(
    positions: list[dict],
    employees_by_position: dict[int, list[dict]],
) -> dict[int, list[dict]]:
    position_order = {position["id"]: index for index, position in enumerate(positions)}
    best_assignment_by_employee: dict[int, tuple[tuple[int, int, int, int], dict, int]] = {}

    for position in positions:
        position_id = position["id"]
        for employee in employees_by_position.get(position_id, []):
            assignment_rank = (
                0 if employee.get("is_fallback_only") else 1,
                1 if employee.get("is_primary") else 0,
                int(employee.get("priority_score") or 0),
                -position_order.get(position_id, 0),
            )
            current_best = best_assignment_by_employee.get(employee["id"])
            if current_best is None or assignment_rank > current_best[0]:
                best_assignment_by_employee[employee["id"]] = (assignment_rank, employee, position_id)

    unique_by_position = {position["id"]: [] for position in positions}
    for _rank, employee, position_id in best_assignment_by_employee.values():
        unique_by_position.setdefault(position_id, []).append(employee)

    for position_id, employees in unique_by_position.items():
        employees.sort(
            key=lambda employee: (
                1 if employee.get("is_fallback_only") else 0,
                -int(bool(employee.get("is_primary"))),
                -int(employee.get("priority_score") or 0),
                employee["id"],
            )
        )
    return unique_by_position


def build_all_schedule_export_workbook(
    positions: list[dict],
    week_start_date: str,
    week_dates: list[str],
    employees_by_position: dict[int, list[dict]],
    entries: list[dict],
    day_status_map: dict[tuple[int, str], dict],
    lang: str = "en",
) -> BytesIO:
    employees_by_priority_position = build_unique_employees_by_priority_position(positions, employees_by_position)

    grouped_entries: dict[tuple[int, str], list[dict]] = {}
    for entry in entries:
        grouped_entries.setdefault((entry["employee_id"], entry["date"]), []).append(entry)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = get_export_label("schedule_title", lang)
    worksheet.sheet_view.rightToLeft = lang == "he"

    header_fill = PatternFill("solid", fgColor=normalize_excel_color("D9EAF7"))
    date_fill = PatternFill("solid", fgColor=normalize_excel_color("F3F6F9"))
    position_fill = PatternFill("solid", fgColor=normalize_excel_color("E0F2FE"))
    legend_fill = PatternFill("solid", fgColor=normalize_excel_color("F8FAFC"))
    thin_side = Side(style="thin", color=normalize_excel_color("CCCCCC"))
    medium_side = Side(style="medium", color=normalize_excel_color("64748B"))
    thick_side = Side(style="thick", color=normalize_excel_color("334155"))
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    top_center = Alignment(horizontal="center", vertical="top", wrap_text=True)

    def employee_block_border(current_row: int, current_column: int, first_row: int, last_row: int) -> Border:
        return Border(
            left=medium_side if current_column in {1, 2} else thin_side,
            right=medium_side if current_column in {2, 11} else thin_side,
            top=medium_side if current_row == first_row else thin_side,
            bottom=medium_side if current_row == last_row else thin_side,
        )

    def number_column_border(current_row: int, first_row: int, last_row: int) -> Border:
        return Border(
            left=medium_side,
            right=medium_side,
            top=medium_side if current_row == first_row else thin_side,
            bottom=medium_side if current_row == last_row else thin_side,
        )

    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
    worksheet["A1"] = f"{get_export_label('schedule_title', lang)} - {get_export_label('week_starting', lang)} {week_start_date}"
    worksheet["A1"].font = Font(bold=True, size=14)
    worksheet["A1"].alignment = center

    headers = [
        get_export_label("position", lang),
        get_export_label("row_number", lang),
        get_export_label("employee", lang),
        *get_weekday_labels(lang),
        get_export_label("weekly_total", lang),
    ]
    for col_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=3, column=col_index, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    for index, current_date in enumerate(["", "", "", *week_dates, ""], start=1):
        cell = worksheet.cell(row=4, column=index, value=current_date)
        cell.fill = date_fill
        cell.alignment = center
        cell.border = thin_border

    total_worked_count = 0
    total_worked_minutes = 0
    total_employee_rows = 0
    row_number = 1
    row_index = 5
    has_rendered_position_block = False

    for position in positions:
        employees = employees_by_priority_position.get(position["id"], [])
        if not employees:
            continue

        if has_rendered_position_block:
            for separator_column in range(1, 12):
                separator_cell = worksheet.cell(row=row_index, column=separator_column, value="")
                separator_cell.fill = PatternFill("solid", fgColor=normalize_excel_color("000000"))
                separator_cell.border = Border(
                    left=thin_side,
                    right=thin_side,
                    top=thin_side,
                    bottom=thin_side,
                )
            worksheet.row_dimensions[row_index].height = 8
            row_index += 1

        position_first_row = row_index

        for employee in employees:
            day_payloads = []
            weekly_count = 0
            weekly_minutes = 0
            max_lines = 1
            for current_date in week_dates:
                day_entries = grouped_entries.get((employee["id"], current_date), [])
                day_status = day_status_map.get((employee["id"], current_date))
                payload = build_schedule_cell_payload(day_entries, position["id"], day_status, lang)
                day_payloads.append((payload, day_status))
                weekly_count += payload["worked_count"]
                weekly_minutes += payload["worked_minutes"]
                total_worked_count += payload["worked_count"]
                total_worked_minutes += payload["worked_minutes"]
                max_lines = max(max_lines, len(payload["lines"]) if payload["lines"] else 1)

            first_row = row_index
            last_row = row_index + max_lines - 1
            if max_lines > 1:
                worksheet.merge_cells(start_row=first_row, start_column=2, end_row=last_row, end_column=2)
                worksheet.merge_cells(start_row=first_row, start_column=3, end_row=last_row, end_column=3)
                worksheet.merge_cells(start_row=first_row, start_column=11, end_row=last_row, end_column=11)

            number_cell = worksheet.cell(row=first_row, column=2, value=row_number)
            number_cell.alignment = center
            number_cell.border = number_column_border(first_row, first_row, last_row)
            number_cell.font = Font(bold=True)
            number_cell.fill = legend_fill

            name_cell = worksheet.cell(row=first_row, column=3, value=employee["full_name"])
            name_cell.font = Font(bold=True)
            name_cell.alignment = center
            name_cell.border = thin_border
            for merge_row in range(first_row + 1, last_row + 1):
                worksheet.cell(row=merge_row, column=2).border = number_column_border(merge_row, first_row, last_row)
                worksheet.cell(row=merge_row, column=3).border = thin_border

            for day_offset, (payload, day_status) in enumerate(day_payloads, start=4):
                lines = payload["lines"] or [{"text": "", "highlight_color": None}]
                if max_lines > 1 and len(lines) <= 1:
                    worksheet.merge_cells(start_row=first_row, start_column=day_offset, end_row=last_row, end_column=day_offset)

                line_count = len(lines) if len(lines) > 1 else 1
                for line_offset in range(line_count):
                    line = lines[line_offset]
                    cell = worksheet.cell(row=first_row + line_offset, column=day_offset, value=line["text"])
                    cell.alignment = center
                    cell.border = thin_border
                    fill = None
                    if line.get("highlight_color"):
                        fill = PatternFill("solid", fgColor=normalize_excel_color(line["highlight_color"]))
                        cell.font = Font(color=normalize_excel_color("000000"), bold=True)
                    elif day_status and line["text"] == get_export_label(day_status.get("status_type"), lang):
                        fill = choose_status_fill(day_status)
                    if fill:
                        cell.fill = fill

            total_cell = worksheet.cell(row=first_row, column=11, value=format_total(weekly_count, weekly_minutes, lang))
            total_cell.alignment = center
            total_cell.border = thin_border
            total_cell.font = Font(bold=True)
            for merge_row in range(first_row + 1, last_row + 1):
                worksheet.cell(row=merge_row, column=11).border = thin_border

            for current_row in range(first_row, last_row + 1):
                worksheet.row_dimensions[current_row].height = 24
                for current_column in range(1, 12):
                    worksheet.cell(row=current_row, column=current_column).border = employee_block_border(
                        current_row,
                        current_column,
                        first_row,
                        last_row,
                    )
            row_index = last_row + 1
            row_number += 1
            total_employee_rows += 1

        position_last_row = row_index - 1
        if employees and position_last_row >= position_first_row:
            if position_last_row > position_first_row:
                worksheet.merge_cells(start_row=position_first_row, start_column=1, end_row=position_last_row, end_column=1)
            position_cell = worksheet.cell(row=position_first_row, column=1, value=position["name"])
            position_cell.font = Font(bold=True)
            position_cell.alignment = top_center
            position_cell.fill = PatternFill("solid", fgColor=normalize_excel_color(position.get("color"), "E0F2FE"))
            for current_row in range(position_first_row, position_last_row + 1):
                cell = worksheet.cell(row=current_row, column=1)
                cell.border = thin_border
                if not cell.fill or cell.fill.fill_type is None:
                    cell.fill = position_fill
            for current_column in range(1, 12):
                top_cell = worksheet.cell(row=position_first_row, column=current_column)
                top_cell.border = Border(
                    left=top_cell.border.left,
                    right=top_cell.border.right,
                    top=thick_side,
                    bottom=top_cell.border.bottom,
                )
                bottom_cell = worksheet.cell(row=position_last_row, column=current_column)
                bottom_cell.border = Border(
                    left=bottom_cell.border.left,
                    right=bottom_cell.border.right,
                    top=bottom_cell.border.top,
                    bottom=thick_side,
                )
            has_rendered_position_block = True

    legend_row = row_index + 2
    worksheet.merge_cells(start_row=legend_row, start_column=1, end_row=legend_row, end_column=11)
    legend_title_cell = worksheet.cell(row=legend_row, column=1, value=get_export_label("legend_title", lang))
    legend_title_cell.font = Font(bold=True, size=12)
    legend_title_cell.fill = header_fill
    legend_title_cell.alignment = center
    legend_title_cell.border = thin_border

    position_legend_row = legend_row + 2
    worksheet.cell(row=position_legend_row, column=1, value=get_export_label("position_colors", lang)).font = Font(bold=True)
    worksheet.cell(row=position_legend_row, column=1).border = thin_border
    current_legend_row = position_legend_row + 1
    for position in positions:
        color_cell = worksheet.cell(row=current_legend_row, column=1, value="")
        color_cell.fill = PatternFill("solid", fgColor=normalize_excel_color(position.get("color"), "E0F2FE"))
        color_cell.border = thin_border
        text_cell = worksheet.cell(row=current_legend_row, column=2, value=position["name"])
        text_cell.alignment = center
        text_cell.border = thin_border
        current_legend_row += 1

    status_legend_row = position_legend_row
    worksheet.cell(row=status_legend_row, column=4, value=get_export_label("status_colors", lang)).font = Font(bold=True)
    worksheet.cell(row=status_legend_row, column=4).border = thin_border
    status_items = [
        (get_export_label("sick", lang), PatternFill("solid", fgColor=normalize_excel_color("FEE2E2"))),
        (get_export_label("day_off", lang), PatternFill("solid", fgColor=normalize_excel_color("F1F5F9"))),
        (get_export_label("no_show", lang), None),
    ]
    for offset, (label, fill) in enumerate(status_items, start=1):
        color_cell = worksheet.cell(row=status_legend_row + offset, column=4, value="")
        color_cell.border = thin_border
        if fill:
            color_cell.fill = fill
        text_cell = worksheet.cell(row=status_legend_row + offset, column=5, value=label)
        text_cell.alignment = center
        text_cell.border = thin_border

    notes_row = max(current_legend_row, status_legend_row + len(status_items) + 1) + 1
    worksheet.merge_cells(start_row=notes_row, start_column=1, end_row=notes_row, end_column=11)
    notes_cell = worksheet.cell(
        row=notes_row,
        column=1,
        value=f"{get_export_label('row_number_note', lang)}\n{get_export_label('cross_position_note', lang)}",
    )
    notes_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    notes_cell.fill = legend_fill
    notes_cell.border = thin_border
    worksheet.row_dimensions[notes_row].height = 36

    for column in "ABCDEFGHIJK":
        if column == "A":
            worksheet.column_dimensions[column].width = 20
        elif column == "B":
            worksheet.column_dimensions[column].width = 8
        elif column == "K":
            worksheet.column_dimensions[column].width = 18
        else:
            worksheet.column_dimensions[column].width = 26
    worksheet.freeze_panes = "D5"
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.print_title_rows = "1:4"
    worksheet.print_options.horizontalCentered = True

    summary_sheet = workbook.create_sheet(get_export_label("summary_title", lang))
    summary_sheet.sheet_view.rightToLeft = lang == "he"
    summary_sheet["A1"] = get_export_label("summary_title", lang)
    summary_sheet["A1"].font = Font(bold=True, size=14)
    summary_rows = [
        ("positions", len(positions)),
        ("week_start", week_start_date),
        ("assigned_employees", total_employee_rows),
        ("schedule_entries", len(entries)),
        ("worked_shifts", format_total(total_worked_count, total_worked_minutes, lang)),
        ("no_show_entries", sum(1 for entry in entries if entry.get("no_show"))),
        ("sick_days", sum(1 for item in day_status_map.values() if item.get("status_type") == "sick")),
        ("day_off_statuses", sum(1 for item in day_status_map.values() if item.get("status_type") == "day_off")),
    ]
    for row_index, (key, value) in enumerate(summary_rows, start=3):
        summary_sheet.cell(row=row_index, column=1, value=get_export_label(key, lang))
        summary_sheet.cell(row=row_index, column=2, value=value)
    summary_sheet.column_dimensions["A"].width = 30
    summary_sheet.column_dimensions["B"].width = 28

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def build_schedule_export_workbook(
    position: dict,
    week_start_date: str,
    week_dates: list[str],
    employees: list[dict],
    entries: list[dict],
    day_status_map: dict[tuple[int, str], dict],
    lang: str = "en",
) -> BytesIO:
    grouped_entries: dict[tuple[int, str], list[dict]] = {}
    for entry in entries:
        grouped_entries.setdefault((entry["employee_id"], entry["date"]), []).append(entry)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = get_export_label("schedule_title", lang)
    worksheet.sheet_view.rightToLeft = lang == "he"

    header_fill = PatternFill("solid", fgColor=normalize_excel_color("D9EAF7"))
    date_fill = PatternFill("solid", fgColor=normalize_excel_color("F3F6F9"))
    thin_side = Side(style="thin", color=normalize_excel_color("CCCCCC"))
    medium_side = Side(style="medium", color=normalize_excel_color("64748B"))
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def employee_block_border(current_row: int, current_column: int, first_row: int, last_row: int) -> Border:
        return Border(
            left=medium_side if current_column == 1 else thin_side,
            right=medium_side if current_column == 9 else thin_side,
            top=medium_side if current_row == first_row else thin_side,
            bottom=medium_side if current_row == last_row else thin_side,
        )

    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    worksheet["A1"] = (
        f"{get_export_label('schedule_title', lang)} - {position['name']} - "
        f"{get_export_label('week_starting', lang)} {week_start_date}"
    )
    worksheet["A1"].font = Font(bold=True, size=14)
    worksheet["A1"].alignment = center

    headers = [get_export_label("employee", lang), *get_weekday_labels(lang), get_export_label("weekly_total", lang)]
    for col_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=3, column=col_index, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    for index, current_date in enumerate(["", *week_dates, ""], start=1):
        cell = worksheet.cell(row=4, column=index, value=current_date)
        cell.fill = date_fill
        cell.alignment = center
        cell.border = thin_border

    total_worked_count = 0
    total_worked_minutes = 0
    row_index = 5
    for employee in employees:
        day_payloads = []
        weekly_count = 0
        weekly_minutes = 0
        max_lines = 1
        for current_date in week_dates:
            day_entries = grouped_entries.get((employee["id"], current_date), [])
            day_status = day_status_map.get((employee["id"], current_date))
            payload = build_schedule_cell_payload(day_entries, position["id"], day_status, lang)
            day_payloads.append((payload, day_status))
            weekly_count += payload["worked_count"]
            weekly_minutes += payload["worked_minutes"]
            total_worked_count += payload["worked_count"]
            total_worked_minutes += payload["worked_minutes"]
            max_lines = max(max_lines, len(payload["lines"]) if payload["lines"] else 1)

        first_row = row_index
        last_row = row_index + max_lines - 1
        if max_lines > 1:
            worksheet.merge_cells(start_row=first_row, start_column=1, end_row=last_row, end_column=1)
            worksheet.merge_cells(start_row=first_row, start_column=9, end_row=last_row, end_column=9)

        name_cell = worksheet.cell(row=first_row, column=1, value=employee["full_name"])
        name_cell.font = Font(bold=True)
        name_cell.alignment = center
        name_cell.border = thin_border
        for merge_row in range(first_row + 1, last_row + 1):
            worksheet.cell(row=merge_row, column=1).border = thin_border

        for day_offset, (payload, day_status) in enumerate(day_payloads, start=2):
            lines = payload["lines"] or [{"text": "", "highlight_color": None}]
            if max_lines > 1 and len(lines) <= 1:
                worksheet.merge_cells(start_row=first_row, start_column=day_offset, end_row=last_row, end_column=day_offset)

            line_count = len(lines) if len(lines) > 1 else 1
            for line_offset in range(line_count):
                line = lines[line_offset]
                cell = worksheet.cell(row=first_row + line_offset, column=day_offset, value=line["text"])
                cell.alignment = center
                cell.border = thin_border
                fill = None
                if line.get("highlight_color"):
                    fill = PatternFill("solid", fgColor=normalize_excel_color(line["highlight_color"]))
                    cell.font = Font(color=normalize_excel_color("000000"), bold=True)
                elif day_status and line["text"] == get_export_label(day_status.get("status_type"), lang):
                    fill = choose_status_fill(day_status)
                if fill:
                    cell.fill = fill

        total_cell = worksheet.cell(row=first_row, column=9, value=format_total(weekly_count, weekly_minutes, lang))
        total_cell.alignment = center
        total_cell.border = thin_border
        total_cell.font = Font(bold=True)
        for merge_row in range(first_row + 1, last_row + 1):
            worksheet.cell(row=merge_row, column=9).border = thin_border
        for current_row in range(first_row, last_row + 1):
            worksheet.row_dimensions[current_row].height = 24
            for current_column in range(1, 10):
                worksheet.cell(row=current_row, column=current_column).border = employee_block_border(
                    current_row,
                    current_column,
                    first_row,
                    last_row,
                )
        row_index = last_row + 1

    for column in "ABCDEFGHI":
        worksheet.column_dimensions[column].width = 26 if column != "I" else 18
    worksheet.freeze_panes = "B5"
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.print_title_rows = "1:4"
    worksheet.print_options.horizontalCentered = True

    summary_sheet = workbook.create_sheet(get_export_label("summary_title", lang))
    summary_sheet.sheet_view.rightToLeft = lang == "he"
    summary_sheet["A1"] = get_export_label("summary_title", lang)
    summary_sheet["A1"].font = Font(bold=True, size=14)
    summary_rows = [
        ("position", position["name"]),
        ("week_start", week_start_date),
        ("assigned_employees", len(employees)),
        ("schedule_entries", len(entries)),
        ("worked_shifts", format_total(total_worked_count, total_worked_minutes, lang)),
        ("no_show_entries", sum(1 for entry in entries if entry.get("no_show"))),
        ("sick_days", sum(1 for item in day_status_map.values() if item.get("status_type") == "sick")),
        ("day_off_statuses", sum(1 for item in day_status_map.values() if item.get("status_type") == "day_off")),
    ]
    for row_index, (key, value) in enumerate(summary_rows, start=3):
        summary_sheet.cell(row=row_index, column=1, value=get_export_label(key, lang))
        summary_sheet.cell(row=row_index, column=2, value=value)
    summary_sheet.column_dimensions["A"].width = 30
    summary_sheet.column_dimensions["B"].width = 28

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
