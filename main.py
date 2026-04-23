import sqlite3
import sys
from dataclasses import dataclass
from datetime import date as Date, datetime, time, timedelta
from io import BytesIO
from pathlib import Path
from typing import Literal

from fastapi import FastAPI, HTTPException, Request
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pydantic import BaseModel, Field, model_validator

from database import get_connection, init_db


tags_metadata = [
    {"name": "Pages", "description": "Frontend pages / HTML страницы"},
    {"name": "Employees", "description": "Employee management / Сотрудники"},
    {"name": "Positions", "description": "Position management / Должности"},
    {"name": "Assignments", "description": "Employee-position assignments / Привязки"},
    {"name": "Shift Templates", "description": "Shift templates / Шаблоны смен"},
    {"name": "Preferences", "description": "Employee preferences / Пожелания"},
    {"name": "Weekly Preferences", "description": "Weekly preferences / Недельные пожелания"},
    {"name": "Requirements", "description": "Shift and coverage requirements / Требования"},
    {"name": "Schedule", "description": "Schedule management / Расписание"},
]

app = FastAPI(
    title="Schedule App - Nursing Staff Scheduling 0.11.3_alpha",
    description="Web application for nursing staff scheduling",
    version="0.11.3_alpha",
    openapi_tags=tags_metadata,
)


def get_base_path() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys._MEIPASS)
    return Path(__file__).resolve().parent


BASE_PATH = get_base_path()
init_db()

app.mount("/static", StaticFiles(directory=str(BASE_PATH / "static")), name="static")
templates = Jinja2Templates(directory=str(BASE_PATH / "templates"))

MAX_WORK_DAYS_PER_WEEK = 6
MAX_CONSECUTIVE_NIGHTS = 2
EMERGENCY_MAX_CONSECUTIVE_NIGHTS = 3
MAX_CONSECUTIVE_SPLIT_DAYS = 2
EMERGENCY_MAX_CONSECUTIVE_SPLIT_DAYS = 3
MIN_REST_MINUTES_AFTER_NIGHT_BEFORE_EVENING = 8 * 60
MIN_REST_MINUTES_BETWEEN_MORNING_AND_EVENING = 0
AFTER_NIGHT_EVENING_PENALTY = 1200


# =========================
# Models
# =========================


class EmployeeCreate(BaseModel):
    full_name: str = Field(min_length=2, max_length=100)
    sex: Literal["male", "female"]
    min_shifts_per_week: int = Field(ge=0, le=14)
    target_shifts_per_week: int = Field(ge=0, le=14)
    max_shifts_per_week: int = Field(ge=0, le=14)
    can_work_night: bool
    can_work_weekends: bool
    can_work_evenings_after_night: bool
    can_work_mornings_and_evenings: bool

    @model_validator(mode="after")
    def validate_shift_range(self):
        if self.min_shifts_per_week > self.max_shifts_per_week:
            raise ValueError("min_shifts_per_week cannot be greater than max_shifts_per_week")
        if self.target_shifts_per_week < self.min_shifts_per_week:
            raise ValueError("target_shifts_per_week cannot be less than min_shifts_per_week")
        if self.target_shifts_per_week > self.max_shifts_per_week:
            raise ValueError("target_shifts_per_week cannot be greater than max_shifts_per_week")
        return self


class PositionCreate(BaseModel):
    name: str = Field(min_length=2, max_length=100)
    requires_continuous_coverage: bool = False
    minimum_staff_presence: int = Field(ge=0, le=50, default=0)

    @model_validator(mode="after")
    def validate_presence(self):
        if not self.requires_continuous_coverage and self.minimum_staff_presence != 0:
            raise ValueError("minimum_staff_presence must be 0 if continuous coverage is disabled")
        return self


class EmployeePositionCreate(BaseModel):
    employee_id: int
    position_id: int
    is_primary: bool = False
    priority_score: int = Field(ge=0, le=100, default=50)
    is_fallback_only: bool = False


class ShiftTemplateCreate(BaseModel):
    name: str = Field(min_length=2, max_length=100)
    category: Literal["morning", "evening", "night"]
    start_time: str
    end_time: str
    is_overnight: bool = False
    is_active: bool = True
    is_split_only: bool = False


class ShiftRequirementCreate(BaseModel):
    position_id: int
    shift_category: Literal["morning", "evening", "night"]
    required_total: int = Field(ge=1, le=50)
    required_female_min: int = Field(ge=0, le=50)
    required_male_min: int = Field(ge=0, le=50, default=0)

    @model_validator(mode="after")
    def validate_gender_minimums(self):
        if self.required_female_min > self.required_total:
            raise ValueError("required_female_min cannot be greater than required_total")
        if self.required_male_min > self.required_total:
            raise ValueError("required_male_min cannot be greater than required_total")
        if self.required_female_min + self.required_male_min > self.required_total:
            raise ValueError("gender minimums cannot be greater than required_total")
        return self


class CoverageRequirementCreate(BaseModel):
    position_id: int
    start_time: str
    end_time: str
    required_total: int = Field(ge=0, le=50)
    required_female_min: int = Field(ge=0, le=50, default=0)
    required_male_min: int = Field(ge=0, le=50, default=0)
    is_overnight: bool = False

    @model_validator(mode="after")
    def validate_requirement(self):
        parse_time_string(self.start_time)
        parse_time_string(self.end_time)
        if self.required_female_min > self.required_total:
            raise ValueError("required_female_min cannot be greater than required_total")
        if self.required_male_min > self.required_total:
            raise ValueError("required_male_min cannot be greater than required_total")
        if self.required_female_min + self.required_male_min > self.required_total:
            raise ValueError("gender minimums cannot be greater than required_total")
        return self


class EmployeePreferenceCreate(BaseModel):
    employee_id: int
    allow_morning: bool
    allow_evening: bool
    allow_night: bool
    allow_morning_evening_combo: bool


class EmployeeWeekPreferenceCreate(BaseModel):
    employee_id: int
    week_start_date: str
    preference_date: str
    preference_type: Literal[
        "no_preference",
        "off_day",
        "vacation",
        "only_morning",
        "only_evening",
        "only_night",
        "not_morning",
        "not_evening",
        "not_night",
        "no_morning_evening_combo",
    ]


class EmployeeDayStatusCreate(BaseModel):
    employee_id: int
    date: str
    status_type: Literal["sick", "day_off"]


class ScheduleEntryCreate(BaseModel):
    employee_id: int
    position_id: int
    date: str
    shift_template_id: int


class ScheduleEntryStatusUpdate(BaseModel):
    no_show: bool


class AutoGenerateScheduleRequest(BaseModel):
    position_id: int
    week_start_date: str


class ClearWeekScheduleRequest(BaseModel):
    position_id: int
    week_start_date: str


class AppSettingsUpdate(BaseModel):
    min_rest_minutes_between_morning_and_evening: int | None = Field(
        default=None,
        ge=0,
        le=24 * 60,
    )
    min_rest_minutes_after_night_before_evening: int | None = Field(
        default=None,
        ge=0,
        le=24 * 60,
    )
    schedule_coverage_display_mode: Literal["category", "interval"] | None = None
    max_work_days_per_week: int | None = Field(default=None, ge=1, le=7)
    max_consecutive_nights: int | None = Field(default=None, ge=1, le=7)
    emergency_max_consecutive_nights: int | None = Field(default=None, ge=1, le=7)
    max_consecutive_split_days: int | None = Field(default=None, ge=1, le=7)
    emergency_max_consecutive_split_days: int | None = Field(default=None, ge=1, le=7)
    after_night_evening_penalty: int | None = Field(default=None, ge=0, le=10000)
    consecutive_night_penalty: int | None = Field(default=None, ge=0, le=10000)
    consecutive_split_penalty: int | None = Field(default=None, ge=0, le=10000)
    coverage_shortage_gain_weight: int | None = Field(default=None, ge=1, le=1000)
    coverage_overage_penalty_weight: int | None = Field(default=None, ge=0, le=1000)
    target_gender_bonus_weight: int | None = Field(default=None, ge=0, le=2000)
    wrong_gender_penalty_weight: int | None = Field(default=None, ge=0, le=2000)
    balance_missing_min_weight: int | None = Field(default=None, ge=0, le=10000)
    balance_target_distance_weight: int | None = Field(default=None, ge=0, le=10000)
    balance_over_target_weight: int | None = Field(default=None, ge=0, le=10000)
    balance_over_max_weight: int | None = Field(default=None, ge=0, le=50000)
    balance_worked_day_weight: int | None = Field(default=None, ge=0, le=10000)
    balance_night_weight: int | None = Field(default=None, ge=0, le=10000)
    balance_split_weight: int | None = Field(default=None, ge=0, le=10000)
    balance_consecutive_night_weight: int | None = Field(default=None, ge=0, le=10000)
    balance_consecutive_split_weight: int | None = Field(default=None, ge=0, le=10000)
    balance_excess_night_weight: int | None = Field(default=None, ge=0, le=50000)
    balance_excess_split_weight: int | None = Field(default=None, ge=0, le=50000)


@dataclass(frozen=True)
class Interval:
    start: int
    end: int

    def contains(self, start: int, end: int) -> bool:
        return self.start <= start and end <= self.end

    def overlaps(self, other: "Interval") -> bool:
        return self.start < other.end and other.start < self.end


# =========================
# Helpers
# =========================


def parse_time_string(value: str) -> time:
    return datetime.strptime(value, "%H:%M").time()


def parse_date_string(value: str) -> Date:
    return datetime.strptime(value, "%Y-%m-%d").date()


def time_to_minutes(value: str) -> int:
    parsed = parse_time_string(value)
    return parsed.hour * 60 + parsed.minute


def build_interval(start_time: str, end_time: str, is_overnight: bool = False) -> Interval:
    start = time_to_minutes(start_time)
    end = time_to_minutes(end_time)
    if is_overnight or end <= start:
        end += 24 * 60
    return Interval(start=start, end=end)


def build_week_dates(week_start_date: str) -> list[str]:
    start = parse_date_string(week_start_date)
    return [(start + timedelta(days=offset)).isoformat() for offset in range(7)]


def get_week_start_for_date(date_string: str) -> str:
    current = parse_date_string(date_string)
    days_since_sunday = (current.weekday() + 1) % 7
    return (current - timedelta(days=days_since_sunday)).isoformat()


def get_week_end_date(week_start_date: str) -> str:
    return build_week_dates(week_start_date)[-1]


def get_app_settings(connection) -> dict:
    cursor = connection.cursor()
    cursor.execute("SELECT key, value FROM app_settings")
    raw_settings = {row["key"]: row["value"] for row in cursor.fetchall()}

    def read_int(key: str, default: int) -> int:
        try:
            return int(raw_settings.get(key, default))
        except (TypeError, ValueError):
            return default

    return {
        "min_rest_minutes_between_morning_and_evening": read_int(
            "min_rest_minutes_between_morning_and_evening",
            MIN_REST_MINUTES_BETWEEN_MORNING_AND_EVENING,
        ),
        "min_rest_minutes_after_night_before_evening": read_int(
            "min_rest_minutes_after_night_before_evening",
            MIN_REST_MINUTES_AFTER_NIGHT_BEFORE_EVENING,
        ),
        "schedule_coverage_display_mode": (
            raw_settings.get("schedule_coverage_display_mode")
            if raw_settings.get("schedule_coverage_display_mode") in {"category", "interval"}
            else "interval"
        ),
        "max_work_days_per_week": read_int("max_work_days_per_week", MAX_WORK_DAYS_PER_WEEK),
        "max_consecutive_nights": read_int("max_consecutive_nights", MAX_CONSECUTIVE_NIGHTS),
        "emergency_max_consecutive_nights": read_int("emergency_max_consecutive_nights", EMERGENCY_MAX_CONSECUTIVE_NIGHTS),
        "max_consecutive_split_days": read_int("max_consecutive_split_days", MAX_CONSECUTIVE_SPLIT_DAYS),
        "emergency_max_consecutive_split_days": read_int("emergency_max_consecutive_split_days", EMERGENCY_MAX_CONSECUTIVE_SPLIT_DAYS),
        "after_night_evening_penalty": read_int("after_night_evening_penalty", AFTER_NIGHT_EVENING_PENALTY),
        "consecutive_night_penalty": read_int("consecutive_night_penalty", 500),
        "consecutive_split_penalty": read_int("consecutive_split_penalty", 450),
        "coverage_shortage_gain_weight": read_int("coverage_shortage_gain_weight", 100),
        "coverage_overage_penalty_weight": read_int("coverage_overage_penalty_weight", 25),
        "target_gender_bonus_weight": read_int("target_gender_bonus_weight", 250),
        "wrong_gender_penalty_weight": read_int("wrong_gender_penalty_weight", 120),
        "balance_missing_min_weight": read_int("balance_missing_min_weight", 300),
        "balance_target_distance_weight": read_int("balance_target_distance_weight", 70),
        "balance_over_target_weight": read_int("balance_over_target_weight", 80),
        "balance_over_max_weight": read_int("balance_over_max_weight", 10000),
        "balance_worked_day_weight": read_int("balance_worked_day_weight", 15),
        "balance_night_weight": read_int("balance_night_weight", 60),
        "balance_split_weight": read_int("balance_split_weight", 55),
        "balance_consecutive_night_weight": read_int("balance_consecutive_night_weight", 120),
        "balance_consecutive_split_weight": read_int("balance_consecutive_split_weight", 100),
        "balance_excess_night_weight": read_int("balance_excess_night_weight", 2000),
        "balance_excess_split_weight": read_int("balance_excess_split_weight", 1800),
    }


def save_app_settings(connection, settings: AppSettingsUpdate) -> None:
    cursor = connection.cursor()
    for key, value in settings.model_dump(exclude_none=True).items():
        cursor.execute(
            """
            INSERT INTO app_settings (key, value)
            VALUES (?, ?)
            ON CONFLICT(key)
            DO UPDATE SET value = excluded.value
            """,
            (key, str(value)),
        )


def is_weekend(date_string: str) -> bool:
    # Python: Monday=0. The app week starts on Sunday, but this still catches Friday/Saturday.
    weekday = parse_date_string(date_string).weekday()
    return weekday in (4, 5)


def row_to_employee_dict(row: sqlite3.Row) -> dict:
    return {
        "id": row["id"],
        "full_name": row["full_name"],
        "sex": row["sex"],
        "min_shifts_per_week": row["min_shifts_per_week"],
        "target_shifts_per_week": row["target_shifts_per_week"],
        "max_shifts_per_week": row["max_shifts_per_week"],
        "can_work_night": bool(row["can_work_night"]),
        "can_work_weekends": bool(row["can_work_weekends"]),
        "can_work_evenings_after_night": bool(row["can_work_evenings_after_night"]),
        "can_work_mornings_and_evenings": bool(row["can_work_mornings_and_evenings"]),
    }


def row_to_position_dict(row: sqlite3.Row) -> dict:
    return {
        "id": row["id"],
        "name": row["name"],
        "requires_continuous_coverage": bool(row["requires_continuous_coverage"]),
        "minimum_staff_presence": row["minimum_staff_presence"],
    }


def row_to_shift_template_dict(row: sqlite3.Row) -> dict:
    return {
        "id": row["id"],
        "name": row["name"],
        "category": row["category"],
        "start_time": row["start_time"],
        "end_time": row["end_time"],
        "is_overnight": bool(row["is_overnight"]),
        "is_active": bool(row["is_active"]),
        "is_split_only": bool(row["is_split_only"]),
    }


def row_to_coverage_requirement_dict(row: sqlite3.Row) -> dict:
    return {
        "id": row["id"],
        "position_id": row["position_id"],
        "position_name": row["position_name"] if "position_name" in row.keys() else None,
        "start_time": row["start_time"],
        "end_time": row["end_time"],
        "required_total": row["required_total"],
        "required_female_min": row["required_female_min"],
        "required_male_min": row["required_male_min"],
        "is_overnight": bool(row["is_overnight"]),
    }


def fetch_one_or_404(cursor, query: str, params: tuple, message: str):
    cursor.execute(query, params)
    row = cursor.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail=message)
    return row


def get_employee_week_shift_count(connection, employee_id: int, week_start_date: str) -> int:
    cursor = connection.cursor()
    cursor.execute(
        """
        SELECT COUNT(*) AS cnt
        FROM schedule_entries
        WHERE employee_id = ? AND date >= ? AND date <= ? AND no_show = 0
        """,
        (employee_id, week_start_date, get_week_end_date(week_start_date)),
    )
    return cursor.fetchone()["cnt"]


def get_employee_week_worked_dates(connection, employee_id: int, week_start_date: str) -> set[str]:
    cursor = connection.cursor()
    cursor.execute(
        """
        SELECT DISTINCT date
        FROM schedule_entries
        WHERE employee_id = ? AND date >= ? AND date <= ? AND no_show = 0
        """,
        (employee_id, week_start_date, get_week_end_date(week_start_date)),
    )
    return {row["date"] for row in cursor.fetchall()}


def entry_category(entry: dict) -> str:
    return entry.get("category") or entry.get("shift_category")


def get_employee_entries_for_date(connection, employee_id: int, date_string: str) -> list[dict]:
    cursor = connection.cursor()
    cursor.execute(
        """
        SELECT se.*, st.category, st.start_time, st.end_time, st.is_overnight, st.is_split_only
        FROM schedule_entries se
        JOIN shift_templates st ON st.id = se.shift_template_id
        WHERE se.employee_id = ? AND se.date = ? AND se.no_show = 0
        ORDER BY st.start_time
        """,
        (employee_id, date_string),
    )
    return [dict(row) for row in cursor.fetchall()]


def employee_has_night_on_date(connection, employee_id: int, date_string: str) -> bool:
    return any(entry["category"] == "night" for entry in get_employee_entries_for_date(connection, employee_id, date_string))


def get_previous_night_entries(connection, employee_id: int, date_string: str) -> list[dict]:
    previous_date = (parse_date_string(date_string) - timedelta(days=1)).isoformat()
    return [
        entry
        for entry in get_employee_entries_for_date(connection, employee_id, previous_date)
        if entry["category"] == "night"
    ]


def had_previous_night(connection, employee_id: int, date_string: str) -> bool:
    return bool(get_previous_night_entries(connection, employee_id, date_string))


def get_break_minutes_after_previous_night(connection, employee_id: int, date_string: str, template: dict) -> int | None:
    previous_nights = get_previous_night_entries(connection, employee_id, date_string)
    if not previous_nights:
        return None

    current_start = 24 * 60 + time_to_minutes(template["start_time"])
    shortest_break = None
    for entry in previous_nights:
        night_interval = build_interval(entry["start_time"], entry["end_time"], bool(entry["is_overnight"]))
        break_minutes = current_start - night_interval.end
        if shortest_break is None or break_minutes < shortest_break:
            shortest_break = break_minutes

    return shortest_break


def get_break_minutes_between_same_day_categories(
    connection,
    employee_id: int,
    date_string: str,
    template: dict,
    first_category: str,
    second_category: str,
    entries: list[dict] | None = None,
) -> int | None:
    template_interval = build_interval(template["start_time"], template["end_time"], template["is_overnight"])
    breaks = []

    for entry in entries if entries is not None else get_employee_entries_for_date(connection, employee_id, date_string):
        if entry_category(entry) == first_category and template["category"] == second_category:
            first_interval = build_interval(entry["start_time"], entry["end_time"], bool(entry["is_overnight"]))
            breaks.append(template_interval.start - first_interval.end)

        if entry_category(entry) == second_category and template["category"] == first_category:
            second_interval = build_interval(entry["start_time"], entry["end_time"], bool(entry["is_overnight"]))
            breaks.append(second_interval.start - template_interval.end)

    if not breaks:
        return None

    return min(breaks)


def employee_has_split_day(connection, employee_id: int, date_string: str) -> bool:
    entries = get_employee_entries_for_date(connection, employee_id, date_string)
    categories = {entry["category"] for entry in entries}
    return "morning" in categories and "evening" in categories


def would_have_night_day(connection, employee_id: int, date_string: str, template: dict) -> bool:
    return template["category"] == "night" or employee_has_night_on_date(connection, employee_id, date_string)


def would_have_split_day(connection, employee_id: int, date_string: str, template: dict) -> bool:
    entries = get_employee_entries_for_date(connection, employee_id, date_string)
    categories = {entry["category"] for entry in entries}
    categories.add(template["category"])
    return "morning" in categories and "evening" in categories


def count_consecutive_days_before(connection, employee_id: int, date_string: str, predicate) -> int:
    current_date = parse_date_string(date_string) - timedelta(days=1)
    count = 0

    while count < 7:
        if not predicate(connection, employee_id, current_date.isoformat()):
            break
        count += 1
        current_date -= timedelta(days=1)

    return count


def count_consecutive_days_after(connection, employee_id: int, date_string: str, predicate) -> int:
    current_date = parse_date_string(date_string) + timedelta(days=1)
    count = 0

    while count < 7:
        if not predicate(connection, employee_id, current_date.isoformat()):
            break
        count += 1
        current_date += timedelta(days=1)

    return count


def get_fatigue_penalty(connection, employee_id: int, date_string: str, template: dict) -> int:
    app_settings = get_app_settings(connection)
    penalty = 0

    if had_previous_night(connection, employee_id, date_string):
        if template["category"] == "evening":
            penalty += app_settings["after_night_evening_penalty"]
        elif template["category"] == "night":
            penalty += app_settings["after_night_evening_penalty"] // 2

    if would_have_night_day(connection, employee_id, date_string, template):
        previous_nights = count_consecutive_days_before(connection, employee_id, date_string, employee_has_night_on_date)
        next_nights = count_consecutive_days_after(connection, employee_id, date_string, employee_has_night_on_date)
        penalty += (previous_nights + next_nights) * app_settings["consecutive_night_penalty"]

    if would_have_split_day(connection, employee_id, date_string, template):
        previous_splits = count_consecutive_days_before(connection, employee_id, date_string, employee_has_split_day)
        next_splits = count_consecutive_days_after(connection, employee_id, date_string, employee_has_split_day)
        penalty += (previous_splits + next_splits) * app_settings["consecutive_split_penalty"]

    return penalty


def get_week_preference(connection, employee_id: int, date_string: str) -> str:
    cursor = connection.cursor()
    cursor.execute(
        """
        SELECT preference_type
        FROM employee_week_preferences
        WHERE employee_id = ? AND preference_date = ?
        """,
        (employee_id, date_string),
    )
    row = cursor.fetchone()
    return row["preference_type"] if row else "no_preference"


def get_employee_day_status(connection, employee_id: int, date_string: str):
    cursor = connection.cursor()
    cursor.execute(
        """
        SELECT status_type
        FROM employee_day_statuses
        WHERE employee_id = ? AND date = ?
        """,
        (employee_id, date_string),
    )
    return cursor.fetchone()


def get_general_preference(connection, employee_id: int) -> dict:
    cursor = connection.cursor()
    cursor.execute("SELECT * FROM employee_preferences WHERE employee_id = ?", (employee_id,))
    row = cursor.fetchone()
    if not row:
        return {
            "allow_morning": True,
            "allow_evening": True,
            "allow_night": True,
            "allow_morning_evening_combo": True,
        }
    return {
        "allow_morning": bool(row["allow_morning"]),
        "allow_evening": bool(row["allow_evening"]),
        "allow_night": bool(row["allow_night"]),
        "allow_morning_evening_combo": bool(row["allow_morning_evening_combo"]),
    }


def category_allowed_by_preferences(connection, employee: dict, date_string: str, category: str) -> bool:
    if category == "night" and not employee["can_work_night"]:
        return False
    if is_weekend(date_string) and not employee["can_work_weekends"]:
        return False

    general = get_general_preference(connection, employee["id"])
    if category == "morning" and not general["allow_morning"]:
        return False
    if category == "evening" and not general["allow_evening"]:
        return False
    if category == "night" and not general["allow_night"]:
        return False

    weekly = get_week_preference(connection, employee["id"], date_string)
    if weekly in ("off_day", "vacation"):
        return False
    if weekly.startswith("only_") and weekly != f"only_{category}":
        return False
    if weekly == f"not_{category}":
        return False
    return True


def can_employee_take_template(
    connection,
    employee: dict,
    position_id: int,
    date_string: str,
    template: dict,
    week_start_date: str,
    fatigue_relaxation: int = 0,
    staged_entries: list[dict] | None = None,
) -> bool:
    app_settings = get_app_settings(connection)
    staged_entries = staged_entries or []
    week_end_date = get_week_end_date(week_start_date)
    staged_week_entries = [
        entry
        for entry in staged_entries
        if entry["employee_id"] == employee["id"] and week_start_date <= entry["date"] <= week_end_date
    ]

    if get_employee_day_status(connection, employee["id"], date_string):
        return False
    if not category_allowed_by_preferences(connection, employee, date_string, template["category"]):
        return False
    if get_employee_week_shift_count(connection, employee["id"], week_start_date) + len(staged_week_entries) >= employee["max_shifts_per_week"]:
        return False

    worked_dates = get_employee_week_worked_dates(connection, employee["id"], week_start_date)
    worked_dates.update(entry["date"] for entry in staged_week_entries)
    projected_worked_dates = set(worked_dates)
    projected_worked_dates.add(date_string)
    if len(projected_worked_dates) > app_settings["max_work_days_per_week"]:
        return False

    cursor = connection.cursor()
    cursor.execute(
        """
        SELECT 1
        FROM employee_positions
        WHERE employee_id = ? AND position_id = ?
        """,
        (employee["id"], position_id),
    )
    if not cursor.fetchone():
        return False

    new_interval = build_interval(template["start_time"], template["end_time"], template["is_overnight"])
    existing_entries = [
        *get_employee_entries_for_date(connection, employee["id"], date_string),
        *[
            entry
            for entry in staged_entries
            if entry["employee_id"] == employee["id"] and entry["date"] == date_string
        ],
    ]
    existing_categories = {entry_category(entry) for entry in existing_entries}

    for entry in existing_entries:
        existing_interval = build_interval(entry["start_time"], entry["end_time"], bool(entry["is_overnight"]))
        if new_interval.overlaps(existing_interval):
            return False

    if existing_entries:
        allowed_pair = {"morning", "evening"}
        if not employee["can_work_mornings_and_evenings"]:
            return False
        if template["category"] not in allowed_pair or not existing_categories.issubset(allowed_pair):
            return False
        if not template["is_split_only"] or not all(bool(entry["is_split_only"]) for entry in existing_entries):
            return False
        if get_week_preference(connection, employee["id"], date_string) == "no_morning_evening_combo":
            return False
        if len(existing_entries) >= 2:
            return False

        morning_evening_break = get_break_minutes_between_same_day_categories(
            connection,
            employee["id"],
            date_string,
            template,
            "morning",
            "evening",
            entries=existing_entries,
        )
        if (
            morning_evening_break is not None
            and morning_evening_break < app_settings["min_rest_minutes_between_morning_and_evening"]
        ):
            return False

    if had_previous_night(connection, employee["id"], date_string):
        if template["category"] == "morning":
            return False

        if template["category"] == "evening":
            break_minutes = get_break_minutes_after_previous_night(
                connection,
                employee["id"],
                date_string,
                template,
            )
            if (
                break_minutes is None
                or break_minutes < app_settings["min_rest_minutes_after_night_before_evening"]
            ):
                return False

    if would_have_night_day(connection, employee["id"], date_string, template):
        previous_nights = count_consecutive_days_before(connection, employee["id"], date_string, employee_has_night_on_date)
        next_nights = count_consecutive_days_after(connection, employee["id"], date_string, employee_has_night_on_date)
        projected_nights = previous_nights + 1 + next_nights
        allowed_nights = (
            app_settings["max_consecutive_nights"]
            if fatigue_relaxation == 0
            else app_settings["emergency_max_consecutive_nights"]
        )
        if projected_nights > allowed_nights:
            return False

    projected_categories = set(existing_categories)
    projected_categories.add(template["category"])
    if "morning" in projected_categories and "evening" in projected_categories:
        previous_splits = count_consecutive_days_before(connection, employee["id"], date_string, employee_has_split_day)
        next_splits = count_consecutive_days_after(connection, employee["id"], date_string, employee_has_split_day)
        projected_splits = previous_splits + 1 + next_splits
        allowed_splits = (
            app_settings["max_consecutive_split_days"]
            if fatigue_relaxation == 0
            else app_settings["emergency_max_consecutive_split_days"]
        )
        if projected_splits > allowed_splits:
            return False

    return True


def explain_employee_template_rejection(
    connection,
    employee: dict,
    position_id: int,
    date_string: str,
    template: dict,
    week_start_date: str,
    fatigue_relaxation: int = 0,
    staged_entries: list[dict] | None = None,
) -> str | None:
    app_settings = get_app_settings(connection)
    staged_entries = staged_entries or []
    week_end_date = get_week_end_date(week_start_date)
    staged_week_entries = [
        entry
        for entry in staged_entries
        if entry["employee_id"] == employee["id"] and week_start_date <= entry["date"] <= week_end_date
    ]

    if get_employee_day_status(connection, employee["id"], date_string):
        return "day status blocks employee"
    if not category_allowed_by_preferences(connection, employee, date_string, template["category"]):
        return "employee preferences or permissions block this shift"
    if get_employee_week_shift_count(connection, employee["id"], week_start_date) + len(staged_week_entries) >= employee["max_shifts_per_week"]:
        return "employee reached weekly maximum shifts"

    worked_dates = get_employee_week_worked_dates(connection, employee["id"], week_start_date)
    worked_dates.update(entry["date"] for entry in staged_week_entries)
    projected_worked_dates = set(worked_dates)
    projected_worked_dates.add(date_string)
    if len(projected_worked_dates) > app_settings["max_work_days_per_week"]:
        return "mandatory weekly day off would be violated"

    cursor = connection.cursor()
    cursor.execute(
        """
        SELECT 1
        FROM employee_positions
        WHERE employee_id = ? AND position_id = ?
        """,
        (employee["id"], position_id),
    )
    if not cursor.fetchone():
        return "employee is not assigned to this position"

    new_interval = build_interval(template["start_time"], template["end_time"], template["is_overnight"])
    existing_entries = [
        *get_employee_entries_for_date(connection, employee["id"], date_string),
        *[
            entry
            for entry in staged_entries
            if entry["employee_id"] == employee["id"] and entry["date"] == date_string
        ],
    ]
    existing_categories = {entry_category(entry) for entry in existing_entries}

    for entry in existing_entries:
        existing_interval = build_interval(entry["start_time"], entry["end_time"], bool(entry["is_overnight"]))
        if new_interval.overlaps(existing_interval):
            return "employee already has an overlapping shift"

    if existing_entries:
        allowed_pair = {"morning", "evening"}
        if not employee["can_work_mornings_and_evenings"]:
            return "employee cannot work morning and evening on the same day"
        if template["category"] not in allowed_pair or not existing_categories.issubset(allowed_pair):
            return "employee already has another shift type that cannot be paired"
        if not template["is_split_only"] or not all(bool(entry["is_split_only"]) for entry in existing_entries):
            return "morning-evening combo requires split-only templates"
        if get_week_preference(connection, employee["id"], date_string) == "no_morning_evening_combo":
            return "weekly preference blocks morning-evening combo"
        if len(existing_entries) >= 2:
            return "employee already has two shifts that day"

        morning_evening_break = get_break_minutes_between_same_day_categories(
            connection,
            employee["id"],
            date_string,
            template,
            "morning",
            "evening",
            entries=existing_entries,
        )
        if (
            morning_evening_break is not None
            and morning_evening_break < app_settings["min_rest_minutes_between_morning_and_evening"]
        ):
            return "morning-evening rest gap is too short"

    if had_previous_night(connection, employee["id"], date_string):
        if template["category"] == "morning":
            return "morning after previous night is forbidden"

        if template["category"] == "evening":
            break_minutes = get_break_minutes_after_previous_night(
                connection,
                employee["id"],
                date_string,
                template,
            )
            if (
                break_minutes is None
                or break_minutes < app_settings["min_rest_minutes_after_night_before_evening"]
            ):
                return "night-evening rest gap is too short"

    if would_have_night_day(connection, employee["id"], date_string, template):
        previous_nights = count_consecutive_days_before(connection, employee["id"], date_string, employee_has_night_on_date)
        next_nights = count_consecutive_days_after(connection, employee["id"], date_string, employee_has_night_on_date)
        projected_nights = previous_nights + 1 + next_nights
        allowed_nights = (
            app_settings["max_consecutive_nights"]
            if fatigue_relaxation == 0
            else app_settings["emergency_max_consecutive_nights"]
        )
        if projected_nights > allowed_nights:
            return "too many consecutive night shifts"

    projected_categories = set(existing_categories)
    projected_categories.add(template["category"])
    if "morning" in projected_categories and "evening" in projected_categories:
        previous_splits = count_consecutive_days_before(connection, employee["id"], date_string, employee_has_split_day)
        next_splits = count_consecutive_days_after(connection, employee["id"], date_string, employee_has_split_day)
        projected_splits = previous_splits + 1 + next_splits
        allowed_splits = (
            app_settings["max_consecutive_split_days"]
            if fatigue_relaxation == 0
            else app_settings["emergency_max_consecutive_split_days"]
        )
        if projected_splits > allowed_splits:
            return "too many consecutive split shifts"

    return None


def validate_schedule_entry_basic(connection, entry: ScheduleEntryCreate):
    parse_date_string(entry.date)
    cursor = connection.cursor()
    employee_row = fetch_one_or_404(cursor, "SELECT * FROM employees WHERE id = ?", (entry.employee_id,), "Employee not found")
    fetch_one_or_404(cursor, "SELECT * FROM positions WHERE id = ?", (entry.position_id,), "Position not found")
    template_row = fetch_one_or_404(cursor, "SELECT * FROM shift_templates WHERE id = ?", (entry.shift_template_id,), "Shift template not found")
    employee = row_to_employee_dict(employee_row)
    template = row_to_shift_template_dict(template_row)
    week_start = get_week_start_for_date(entry.date)
    if not can_employee_take_template(connection, employee, entry.position_id, entry.date, template, week_start):
        raise HTTPException(status_code=400, detail="Employee cannot be assigned to this shift")
    return employee, template


def validate_manual_schedule_entry_basics(connection, entry: ScheduleEntryCreate):
    parse_date_string(entry.date)
    cursor = connection.cursor()
    fetch_one_or_404(cursor, "SELECT id FROM employees WHERE id = ?", (entry.employee_id,), "Employee not found")
    fetch_one_or_404(cursor, "SELECT id FROM positions WHERE id = ?", (entry.position_id,), "Position not found")
    fetch_one_or_404(cursor, "SELECT id FROM shift_templates WHERE id = ?", (entry.shift_template_id,), "Shift template not found")


def get_export_label(key: str, lang: str) -> str:
    labels = {
        "en": {
            "sick": "Sick",
            "day_off": "Day off",
            "no_show": "No-show",
        },
        "ru": {
            "sick": "Больничный",
            "day_off": "Выходной",
            "no_show": "Неявка",
        },
        "he": {
            "sick": "מחלה",
            "day_off": "יום חופשי",
            "no_show": "אי הגעה",
        },
    }
    return labels.get(lang, labels["en"]).get(key, key)


def build_schedule_cell_text(entries: list[dict], day_status: dict | None = None, lang: str = "en") -> str:
    if day_status and day_status.get("status_type") in {"sick", "day_off"}:
        return get_export_label(day_status["status_type"], lang)

    return "\n".join(
        get_export_label("no_show", lang) if entry.get("no_show") else entry["shift_template_name"]
        for entry in sorted(entries, key=lambda item: item["start_time"])
    )


# =========================
# Pages
# =========================


@app.get("/", tags=["Pages"])
def home_page(request: Request):
    return templates.TemplateResponse(request=request, name="index.html", context={})


@app.get("/employees", tags=["Pages"])
def employees_page(request: Request):
    return templates.TemplateResponse(request=request, name="employees.html", context={})


@app.get("/positions", tags=["Pages"])
def positions_page(request: Request):
    return templates.TemplateResponse(request=request, name="positions.html", context={})


@app.get("/employee-positions", tags=["Pages"])
def employee_positions_page(request: Request):
    return templates.TemplateResponse(request=request, name="employee_positions.html", context={})


@app.get("/shift-templates", tags=["Pages"])
def shift_templates_page(request: Request):
    return templates.TemplateResponse(request=request, name="shift_templates.html", context={})


@app.get("/coverage-requirements", tags=["Pages"])
def coverage_requirements_page(request: Request):
    return templates.TemplateResponse(request=request, name="coverage_requirements.html", context={})


@app.get("/weekly-preferences", tags=["Pages"])
def weekly_preferences_page(request: Request):
    return templates.TemplateResponse(request=request, name="weekly_preferences.html", context={})


@app.get("/schedule", tags=["Pages"])
def schedule_page(request: Request):
    return templates.TemplateResponse(request=request, name="schedule.html", context={})


@app.get("/settings", tags=["Pages"])
def settings_page(request: Request):
    return templates.TemplateResponse(request=request, name="settings.html", context={})


@app.get("/guide", tags=["Pages"])
def guide_page(request: Request):
    return templates.TemplateResponse(request=request, name="guide.html", context={})


# =========================
# CRUD APIs
# =========================


@app.get("/api/employees", tags=["Employees"])
def get_employees():
    connection = get_connection()
    try:
        cursor = connection.cursor()
        cursor.execute("SELECT * FROM employees ORDER BY id")
        return [row_to_employee_dict(row) for row in cursor.fetchall()]
    finally:
        connection.close()


@app.post("/api/employees", tags=["Employees"])
def add_employee(employee: EmployeeCreate):
    connection = get_connection()
    try:
        cursor = connection.cursor()
        cursor.execute(
            """
            INSERT INTO employees (
                full_name, sex, min_shifts_per_week, target_shifts_per_week, max_shifts_per_week,
                can_work_night, can_work_weekends, can_work_evenings_after_night, can_work_mornings_and_evenings
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                employee.full_name,
                employee.sex,
                employee.min_shifts_per_week,
                employee.target_shifts_per_week,
                employee.max_shifts_per_week,
                int(employee.can_work_night),
                int(employee.can_work_weekends),
                int(employee.can_work_evenings_after_night),
                int(employee.can_work_mornings_and_evenings),
            ),
        )
        connection.commit()
        return {"message": "Employee added successfully", "employee": {"id": cursor.lastrowid, **employee.model_dump()}}
    finally:
        connection.close()


@app.put("/api/employees/{employee_id}", tags=["Employees"])
def update_employee(employee_id: int, employee: EmployeeCreate):
    connection = get_connection()
    try:
        cursor = connection.cursor()
        fetch_one_or_404(cursor, "SELECT id FROM employees WHERE id = ?", (employee_id,), "Employee not found")
        cursor.execute(
            """
            UPDATE employees
            SET full_name = ?, sex = ?, min_shifts_per_week = ?, target_shifts_per_week = ?,
                max_shifts_per_week = ?, can_work_night = ?, can_work_weekends = ?,
                can_work_evenings_after_night = ?, can_work_mornings_and_evenings = ?
            WHERE id = ?
            """,
            (
                employee.full_name,
                employee.sex,
                employee.min_shifts_per_week,
                employee.target_shifts_per_week,
                employee.max_shifts_per_week,
                int(employee.can_work_night),
                int(employee.can_work_weekends),
                int(employee.can_work_evenings_after_night),
                int(employee.can_work_mornings_and_evenings),
                employee_id,
            ),
        )
        connection.commit()
        return {"message": "Employee updated successfully", "employee": {"id": employee_id, **employee.model_dump()}}
    finally:
        connection.close()


@app.delete("/api/employees/{employee_id}", tags=["Employees"])
def delete_employee(employee_id: int):
    connection = get_connection()
    try:
        cursor = connection.cursor()
        fetch_one_or_404(cursor, "SELECT id FROM employees WHERE id = ?", (employee_id,), "Employee not found")
        cursor.execute("DELETE FROM employees WHERE id = ?", (employee_id,))
        connection.commit()
        return {"message": "Employee deleted successfully"}
    finally:
        connection.close()


@app.get("/api/positions", tags=["Positions"])
def get_positions():
    connection = get_connection()
    try:
        cursor = connection.cursor()
        cursor.execute("SELECT * FROM positions ORDER BY id")
        return [row_to_position_dict(row) for row in cursor.fetchall()]
    finally:
        connection.close()


@app.post("/api/positions", tags=["Positions"])
def add_position(position: PositionCreate):
    connection = get_connection()
    try:
        cursor = connection.cursor()
        try:
            cursor.execute(
                """
                INSERT INTO positions (name, requires_continuous_coverage, minimum_staff_presence)
                VALUES (?, ?, ?)
                """,
                (position.name, int(position.requires_continuous_coverage), position.minimum_staff_presence),
            )
        except sqlite3.IntegrityError:
            raise HTTPException(status_code=400, detail="Position already exists")
        connection.commit()
        return {"message": "Position added successfully", "position": {"id": cursor.lastrowid, **position.model_dump()}}
    finally:
        connection.close()


@app.put("/api/positions/{position_id}", tags=["Positions"])
def update_position(position_id: int, position: PositionCreate):
    connection = get_connection()
    try:
        cursor = connection.cursor()
        fetch_one_or_404(cursor, "SELECT id FROM positions WHERE id = ?", (position_id,), "Position not found")
        try:
            cursor.execute(
                """
                UPDATE positions
                SET name = ?, requires_continuous_coverage = ?, minimum_staff_presence = ?
                WHERE id = ?
                """,
                (position.name, int(position.requires_continuous_coverage), position.minimum_staff_presence, position_id),
            )
        except sqlite3.IntegrityError:
            raise HTTPException(status_code=400, detail="Position already exists")
        connection.commit()
        return {"message": "Position updated successfully", "position": {"id": position_id, **position.model_dump()}}
    finally:
        connection.close()


@app.delete("/api/positions/{position_id}", tags=["Positions"])
def delete_position(position_id: int):
    connection = get_connection()
    try:
        cursor = connection.cursor()
        fetch_one_or_404(cursor, "SELECT id FROM positions WHERE id = ?", (position_id,), "Position not found")
        cursor.execute("DELETE FROM positions WHERE id = ?", (position_id,))
        connection.commit()
        return {"message": "Position deleted successfully"}
    finally:
        connection.close()


@app.get("/api/employee-positions", tags=["Assignments"])
def get_employee_positions():
    connection = get_connection()
    try:
        cursor = connection.cursor()
        cursor.execute(
            """
            SELECT ep.*, e.full_name AS employee_name, p.name AS position_name
            FROM employee_positions ep
            JOIN employees e ON e.id = ep.employee_id
            JOIN positions p ON p.id = ep.position_id
            ORDER BY ep.employee_id, ep.position_id
            """
        )
        items = [dict(row) for row in cursor.fetchall()]
        for item in items:
            item["is_primary"] = bool(item["is_primary"])
            item["is_fallback_only"] = bool(item["is_fallback_only"])
        return items
    finally:
        connection.close()


@app.post("/api/employee-positions", tags=["Assignments"])
def assign_employee_to_position(assignment: EmployeePositionCreate):
    connection = get_connection()
    try:
        cursor = connection.cursor()
        fetch_one_or_404(cursor, "SELECT id FROM employees WHERE id = ?", (assignment.employee_id,), "Employee not found")
        fetch_one_or_404(cursor, "SELECT id FROM positions WHERE id = ?", (assignment.position_id,), "Position not found")
        try:
            cursor.execute(
                """
                INSERT INTO employee_positions (employee_id, position_id, is_primary, priority_score, is_fallback_only)
                VALUES (?, ?, ?, ?, ?)
                """,
                (
                    assignment.employee_id,
                    assignment.position_id,
                    int(assignment.is_primary),
                    assignment.priority_score,
                    int(assignment.is_fallback_only),
                ),
            )
        except sqlite3.IntegrityError:
            raise HTTPException(status_code=400, detail="Employee is already assigned to this position")
        connection.commit()
        return {"message": "Employee assigned to position successfully", "assignment": assignment.model_dump()}
    finally:
        connection.close()


@app.put("/api/employee-positions", tags=["Assignments"])
def update_employee_position(assignment: EmployeePositionCreate):
    connection = get_connection()
    try:
        cursor = connection.cursor()
        fetch_one_or_404(cursor, "SELECT id FROM employees WHERE id = ?", (assignment.employee_id,), "Employee not found")
        fetch_one_or_404(cursor, "SELECT id FROM positions WHERE id = ?", (assignment.position_id,), "Position not found")
        cursor.execute(
            """
            UPDATE employee_positions
            SET is_primary = ?, priority_score = ?, is_fallback_only = ?
            WHERE employee_id = ? AND position_id = ?
            """,
            (
                int(assignment.is_primary),
                assignment.priority_score,
                int(assignment.is_fallback_only),
                assignment.employee_id,
                assignment.position_id,
            ),
        )
        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail="Assignment not found")
        connection.commit()
        return {"message": "Employee assignment updated successfully", "assignment": assignment.model_dump()}
    finally:
        connection.close()


@app.delete("/api/employee-positions", tags=["Assignments"])
def delete_employee_position(employee_id: int, position_id: int):
    connection = get_connection()
    try:
        cursor = connection.cursor()
        cursor.execute("DELETE FROM employee_positions WHERE employee_id = ? AND position_id = ?", (employee_id, position_id))
        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail="Assignment not found")
        connection.commit()
        return {"message": "Employee assignment deleted successfully"}
    finally:
        connection.close()


@app.get("/api/shift-templates", tags=["Shift Templates"])
def get_shift_templates(active_only: bool = False):
    connection = get_connection()
    try:
        cursor = connection.cursor()
        if active_only:
            cursor.execute("SELECT * FROM shift_templates WHERE is_active = 1 ORDER BY category, start_time, end_time")
        else:
            cursor.execute("SELECT * FROM shift_templates ORDER BY category, start_time, end_time")
        return [row_to_shift_template_dict(row) for row in cursor.fetchall()]
    finally:
        connection.close()


@app.post("/api/shift-templates", tags=["Shift Templates"])
def add_shift_template(template: ShiftTemplateCreate):
    parse_time_string(template.start_time)
    parse_time_string(template.end_time)
    connection = get_connection()
    try:
        cursor = connection.cursor()
        try:
            cursor.execute(
                """
                INSERT INTO shift_templates (name, category, start_time, end_time, is_overnight, is_active, is_split_only)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    template.name,
                    template.category,
                    template.start_time,
                    template.end_time,
                    int(template.is_overnight),
                    int(template.is_active),
                    int(template.is_split_only),
                ),
            )
        except sqlite3.IntegrityError:
            raise HTTPException(status_code=400, detail="Shift template already exists")
        connection.commit()
        return {"message": "Shift template added successfully", "shift_template": {"id": cursor.lastrowid, **template.model_dump()}}
    finally:
        connection.close()


@app.put("/api/shift-templates/{template_id}", tags=["Shift Templates"])
def update_shift_template(template_id: int, template: ShiftTemplateCreate):
    parse_time_string(template.start_time)
    parse_time_string(template.end_time)
    connection = get_connection()
    try:
        cursor = connection.cursor()
        fetch_one_or_404(cursor, "SELECT id FROM shift_templates WHERE id = ?", (template_id,), "Shift template not found")
        try:
            cursor.execute(
                """
                UPDATE shift_templates
                SET name = ?, category = ?, start_time = ?, end_time = ?, is_overnight = ?, is_active = ?, is_split_only = ?
                WHERE id = ?
                """,
                (
                    template.name,
                    template.category,
                    template.start_time,
                    template.end_time,
                    int(template.is_overnight),
                    int(template.is_active),
                    int(template.is_split_only),
                    template_id,
                ),
            )
        except sqlite3.IntegrityError:
            raise HTTPException(status_code=400, detail="Shift template with this name already exists")
        connection.commit()
        return {"message": "Shift template updated successfully", "shift_template": {"id": template_id, **template.model_dump()}}
    finally:
        connection.close()


@app.delete("/api/shift-templates/{template_id}", tags=["Shift Templates"])
def delete_shift_template(template_id: int):
    connection = get_connection()
    try:
        cursor = connection.cursor()
        cursor.execute("DELETE FROM shift_templates WHERE id = ?", (template_id,))
        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail="Shift template not found")
        connection.commit()
        return {"message": "Shift template deleted successfully"}
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=400, detail="Cannot delete shift template because it is used in schedule")
    finally:
        connection.close()


@app.get("/api/shift-requirements", tags=["Requirements"])
def get_shift_requirements():
    connection = get_connection()
    try:
        cursor = connection.cursor()
        cursor.execute(
            """
            SELECT sr.*, p.name AS position_name
            FROM shift_requirements sr
            JOIN positions p ON p.id = sr.position_id
            ORDER BY sr.position_id, sr.shift_category
            """
        )
        return [dict(row) for row in cursor.fetchall()]
    finally:
        connection.close()


@app.post("/api/shift-requirements", tags=["Requirements"])
def save_shift_requirement(requirement: ShiftRequirementCreate):
    connection = get_connection()
    try:
        cursor = connection.cursor()
        fetch_one_or_404(cursor, "SELECT id FROM positions WHERE id = ?", (requirement.position_id,), "Position not found")
        cursor.execute(
            """
            INSERT INTO shift_requirements (position_id, shift_category, required_total, required_female_min, required_male_min)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(position_id, shift_category)
            DO UPDATE SET required_total = excluded.required_total,
                          required_female_min = excluded.required_female_min,
                          required_male_min = excluded.required_male_min
            """,
            (
                requirement.position_id,
                requirement.shift_category,
                requirement.required_total,
                requirement.required_female_min,
                requirement.required_male_min,
            ),
        )
        connection.commit()
        return {"message": "Shift requirement saved successfully", "requirement": requirement.model_dump()}
    finally:
        connection.close()


@app.get("/api/coverage-requirements", tags=["Requirements"])
def get_coverage_requirements(position_id: int | None = None):
    connection = get_connection()
    try:
        cursor = connection.cursor()
        if position_id is None:
            cursor.execute(
                """
                SELECT cr.*, p.name AS position_name
                FROM coverage_requirements cr
                JOIN positions p ON p.id = cr.position_id
                ORDER BY cr.position_id, cr.start_time, cr.end_time
                """
            )
        else:
            cursor.execute(
                """
                SELECT cr.*, p.name AS position_name
                FROM coverage_requirements cr
                JOIN positions p ON p.id = cr.position_id
                WHERE cr.position_id = ?
                ORDER BY cr.start_time, cr.end_time
                """,
                (position_id,),
            )
        return [row_to_coverage_requirement_dict(row) for row in cursor.fetchall()]
    finally:
        connection.close()


@app.get("/api/app-settings", tags=["Requirements"])
def get_app_settings_api():
    connection = get_connection()
    try:
        return get_app_settings(connection)
    finally:
        connection.close()


@app.put("/api/app-settings", tags=["Requirements"])
def update_app_settings(settings: AppSettingsUpdate):
    connection = get_connection()
    try:
        save_app_settings(connection, settings)
        connection.commit()
        return {
            "message": "Application settings updated successfully",
            "settings": get_app_settings(connection),
        }
    finally:
        connection.close()


@app.post("/api/coverage-requirements", tags=["Requirements"])
def add_coverage_requirement(requirement: CoverageRequirementCreate):
    connection = get_connection()
    try:
        cursor = connection.cursor()
        fetch_one_or_404(cursor, "SELECT id FROM positions WHERE id = ?", (requirement.position_id,), "Position not found")
        cursor.execute(
            """
            INSERT INTO coverage_requirements
                (position_id, start_time, end_time, required_total, required_female_min, required_male_min, is_overnight)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                requirement.position_id,
                requirement.start_time,
                requirement.end_time,
                requirement.required_total,
                requirement.required_female_min,
                requirement.required_male_min,
                int(requirement.is_overnight),
            ),
        )
        connection.commit()
        return {"message": "Coverage requirement added successfully", "coverage_requirement": {"id": cursor.lastrowid, **requirement.model_dump()}}
    finally:
        connection.close()


@app.put("/api/coverage-requirements/{requirement_id}", tags=["Requirements"])
def update_coverage_requirement(requirement_id: int, requirement: CoverageRequirementCreate):
    connection = get_connection()
    try:
        cursor = connection.cursor()
        fetch_one_or_404(cursor, "SELECT id FROM coverage_requirements WHERE id = ?", (requirement_id,), "Coverage requirement not found")
        cursor.execute(
            """
            UPDATE coverage_requirements
            SET position_id = ?, start_time = ?, end_time = ?, required_total = ?, required_female_min = ?, required_male_min = ?, is_overnight = ?
            WHERE id = ?
            """,
            (
                requirement.position_id,
                requirement.start_time,
                requirement.end_time,
                requirement.required_total,
                requirement.required_female_min,
                requirement.required_male_min,
                int(requirement.is_overnight),
                requirement_id,
            ),
        )
        connection.commit()
        return {"message": "Coverage requirement updated successfully", "coverage_requirement": {"id": requirement_id, **requirement.model_dump()}}
    finally:
        connection.close()


@app.delete("/api/coverage-requirements/{requirement_id}", tags=["Requirements"])
def delete_coverage_requirement(requirement_id: int):
    connection = get_connection()
    try:
        cursor = connection.cursor()
        cursor.execute("DELETE FROM coverage_requirements WHERE id = ?", (requirement_id,))
        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail="Coverage requirement not found")
        connection.commit()
        return {"message": "Coverage requirement deleted successfully"}
    finally:
        connection.close()


@app.get("/api/employee-preferences", tags=["Preferences"])
def get_employee_preferences():
    connection = get_connection()
    try:
        cursor = connection.cursor()
        cursor.execute(
            """
            SELECT ep.*, e.full_name AS employee_name
            FROM employee_preferences ep
            JOIN employees e ON e.id = ep.employee_id
            ORDER BY ep.employee_id
            """
        )
        items = [dict(row) for row in cursor.fetchall()]
        for item in items:
            for key in ("allow_morning", "allow_evening", "allow_night", "allow_morning_evening_combo"):
                item[key] = bool(item[key])
        return items
    finally:
        connection.close()


@app.post("/api/employee-preferences", tags=["Preferences"])
def save_employee_preference(preference: EmployeePreferenceCreate):
    connection = get_connection()
    try:
        cursor = connection.cursor()
        fetch_one_or_404(cursor, "SELECT id FROM employees WHERE id = ?", (preference.employee_id,), "Employee not found")
        cursor.execute(
            """
            INSERT INTO employee_preferences
                (employee_id, allow_morning, allow_evening, allow_night, allow_morning_evening_combo)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(employee_id)
            DO UPDATE SET allow_morning = excluded.allow_morning,
                          allow_evening = excluded.allow_evening,
                          allow_night = excluded.allow_night,
                          allow_morning_evening_combo = excluded.allow_morning_evening_combo
            """,
            (
                preference.employee_id,
                int(preference.allow_morning),
                int(preference.allow_evening),
                int(preference.allow_night),
                int(preference.allow_morning_evening_combo),
            ),
        )
        connection.commit()
        return {"message": "Employee preference saved successfully", "preference": preference.model_dump()}
    finally:
        connection.close()


@app.get("/api/employee-week-preferences", tags=["Weekly Preferences"])
def get_employee_week_preferences(week_start_date: str):
    connection = get_connection()
    try:
        cursor = connection.cursor()
        cursor.execute(
            """
            SELECT ewp.*, e.full_name AS employee_name
            FROM employee_week_preferences ewp
            JOIN employees e ON e.id = ewp.employee_id
            WHERE ewp.week_start_date = ?
            ORDER BY ewp.employee_id, ewp.preference_date
            """,
            (week_start_date,),
        )
        return [dict(row) for row in cursor.fetchall()]
    finally:
        connection.close()


@app.post("/api/employee-week-preferences", tags=["Weekly Preferences"])
def save_employee_week_preference(preference: EmployeeWeekPreferenceCreate):
    connection = get_connection()
    try:
        cursor = connection.cursor()
        fetch_one_or_404(cursor, "SELECT id FROM employees WHERE id = ?", (preference.employee_id,), "Employee not found")
        cursor.execute(
            """
            INSERT INTO employee_week_preferences
                (employee_id, week_start_date, preference_date, preference_type)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(employee_id, preference_date)
            DO UPDATE SET week_start_date = excluded.week_start_date,
                          preference_type = excluded.preference_type
            """,
            (
                preference.employee_id,
                preference.week_start_date,
                preference.preference_date,
                preference.preference_type,
            ),
        )
        connection.commit()
        return {"message": "Weekly preference saved successfully", "preference": preference.model_dump()}
    finally:
        connection.close()


@app.delete("/api/employee-week-preferences", tags=["Weekly Preferences"])
def delete_employee_week_preference(employee_id: int, preference_date: str):
    connection = get_connection()
    try:
        cursor = connection.cursor()
        cursor.execute(
            """
            DELETE FROM employee_week_preferences
            WHERE employee_id = ? AND preference_date = ?
            """,
            (employee_id, preference_date),
        )
        connection.commit()
        return {"message": "Weekly preference deleted successfully", "deleted_count": cursor.rowcount}
    finally:
        connection.close()


@app.get("/api/employee-day-statuses", tags=["Schedule"])
def get_employee_day_statuses():
    connection = get_connection()
    try:
        cursor = connection.cursor()
        cursor.execute(
            """
            SELECT eds.*, e.full_name AS employee_name
            FROM employee_day_statuses eds
            JOIN employees e ON e.id = eds.employee_id
            ORDER BY eds.date, eds.employee_id
            """
        )
        return [dict(row) for row in cursor.fetchall()]
    finally:
        connection.close()


@app.post("/api/employee-day-statuses", tags=["Schedule"])
def save_employee_day_status(status: EmployeeDayStatusCreate):
    connection = get_connection()
    try:
        cursor = connection.cursor()
        fetch_one_or_404(cursor, "SELECT id FROM employees WHERE id = ?", (status.employee_id,), "Employee not found")
        cursor.execute(
            """
            INSERT INTO employee_day_statuses (employee_id, date, status_type)
            VALUES (?, ?, ?)
            ON CONFLICT(employee_id, date)
            DO UPDATE SET status_type = excluded.status_type
            """,
            (status.employee_id, status.date, status.status_type),
        )
        connection.commit()
        return {"message": "Employee day status saved successfully", "status": status.model_dump()}
    finally:
        connection.close()


@app.delete("/api/employee-day-statuses", tags=["Schedule"])
def delete_employee_day_status(employee_id: int, date: str):
    connection = get_connection()
    try:
        cursor = connection.cursor()
        cursor.execute("DELETE FROM employee_day_statuses WHERE employee_id = ? AND date = ?", (employee_id, date))
        connection.commit()
        return {"message": "Employee day status deleted successfully"}
    finally:
        connection.close()


def get_employee_day_status_map(connection, employee_ids: list[int], dates: list[str]) -> dict[tuple[int, str], dict]:
    if not employee_ids or not dates:
        return {}

    cursor = connection.cursor()
    employee_placeholders = ",".join(["?"] * len(employee_ids))
    date_placeholders = ",".join(["?"] * len(dates))
    cursor.execute(
        f"""
        SELECT *
        FROM employee_day_statuses
        WHERE employee_id IN ({employee_placeholders})
          AND date IN ({date_placeholders})
        """,
        [*employee_ids, *dates],
    )
    return {(row["employee_id"], row["date"]): dict(row) for row in cursor.fetchall()}


def sync_generated_day_off_statuses(
    connection,
    cursor,
    employees: list[dict],
    position_id: int,
    dates: list[str],
) -> int:
    employee_ids = [employee["id"] for employee in employees]
    if not employee_ids or not dates:
        return 0

    employee_placeholders = ",".join(["?"] * len(employee_ids))
    date_placeholders = ",".join(["?"] * len(dates))
    cursor.execute(
        f"""
        DELETE FROM employee_day_statuses
        WHERE status_type = 'day_off'
          AND employee_id IN ({employee_placeholders})
          AND date IN ({date_placeholders})
        """,
        [*employee_ids, *dates],
    )

    entries = get_schedule_entries(connection, dates=dates)
    employee_id_set = set(employee_ids)
    worked_days = {
        (entry["employee_id"], entry["date"])
        for entry in entries
        if entry["employee_id"] in employee_id_set and not entry.get("no_show")
    }
    day_status_map = get_employee_day_status_map(connection, employee_ids, dates)

    inserted_count = 0
    for employee_id in employee_ids:
        for date_string in dates:
            if (employee_id, date_string) in worked_days:
                continue
            if (employee_id, date_string) in day_status_map:
                continue
            cursor.execute(
                """
                INSERT INTO employee_day_statuses (employee_id, date, status_type)
                VALUES (?, ?, 'day_off')
                """,
                (employee_id, date_string),
            )
            inserted_count += 1

    return inserted_count


# =========================
# Schedule and generator
# =========================


def get_schedule_entries(connection, position_id: int | None = None, dates: list[str] | None = None) -> list[dict]:
    cursor = connection.cursor()
    clauses = []
    params: list = []
    if position_id is not None:
        clauses.append("se.position_id = ?")
        params.append(position_id)
    if dates:
        clauses.append(f"se.date IN ({','.join(['?'] * len(dates))})")
        params.extend(dates)
    where_sql = "WHERE " + " AND ".join(clauses) if clauses else ""
    cursor.execute(
        f"""
        SELECT
            se.id,
            se.employee_id,
            se.position_id,
            se.date,
            se.shift_template_id,
            se.no_show,
            st.name AS shift_template_name,
            st.category AS shift_category,
            st.start_time,
            st.end_time,
            st.is_overnight,
            st.is_split_only,
            e.full_name AS employee_name,
            e.sex AS employee_sex
        FROM schedule_entries se
        JOIN shift_templates st ON st.id = se.shift_template_id
        JOIN employees e ON e.id = se.employee_id
        {where_sql}
        ORDER BY se.date, se.employee_id, se.position_id, st.start_time
        """,
        params,
    )
    items = [dict(row) for row in cursor.fetchall()]
    for item in items:
        item["is_overnight"] = bool(item["is_overnight"])
        item["is_split_only"] = bool(item["is_split_only"])
        item["no_show"] = bool(item["no_show"])
    return items


@app.get("/api/schedule", tags=["Schedule"])
def get_schedule():
    connection = get_connection()
    try:
        return get_schedule_entries(connection)
    finally:
        connection.close()


@app.post("/api/schedule", tags=["Schedule"])
def add_schedule_entry(entry: ScheduleEntryCreate):
    connection = get_connection()
    try:
        validate_manual_schedule_entry_basics(connection, entry)
        cursor = connection.cursor()
        cursor.execute(
            """
            INSERT INTO schedule_entries (employee_id, position_id, date, shift_template_id)
            VALUES (?, ?, ?, ?)
            """,
            (entry.employee_id, entry.position_id, entry.date, entry.shift_template_id),
        )
        connection.commit()
        return {"message": "Schedule entry added successfully", "schedule_entry": {**entry.model_dump(), "id": cursor.lastrowid}}
    finally:
        connection.close()


@app.delete("/api/schedule/{schedule_entry_id}", tags=["Schedule"])
def delete_schedule_entry(schedule_entry_id: int):
    connection = get_connection()
    try:
        cursor = connection.cursor()
        cursor.execute("DELETE FROM schedule_entries WHERE id = ?", (schedule_entry_id,))
        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail="Schedule entry not found")
        connection.commit()
        return {"message": "Schedule entry deleted successfully", "deleted_count": cursor.rowcount}
    finally:
        connection.close()


@app.patch("/api/schedule/{schedule_entry_id}/status", tags=["Schedule"])
def update_schedule_entry_status(schedule_entry_id: int, status: ScheduleEntryStatusUpdate):
    connection = get_connection()
    try:
        cursor = connection.cursor()
        cursor.execute(
            "UPDATE schedule_entries SET no_show = ? WHERE id = ?",
            (int(status.no_show), schedule_entry_id),
        )
        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail="Schedule entry not found")
        connection.commit()
        return {"message": "Schedule entry status updated successfully", "id": schedule_entry_id, "no_show": status.no_show}
    finally:
        connection.close()


@app.post("/api/schedule/clear-week", tags=["Schedule"])
def clear_week_schedule(request_data: ClearWeekScheduleRequest):
    connection = get_connection()
    try:
        week_dates = build_week_dates(request_data.week_start_date)
        cursor = connection.cursor()
        cursor.execute(
            f"""
            DELETE FROM schedule_entries
            WHERE position_id = ? AND date IN ({','.join(['?'] * len(week_dates))})
            """,
            [request_data.position_id, *week_dates],
        )
        deleted_count = cursor.rowcount
        employees = load_position_employees(connection, request_data.position_id)
        employee_ids = [employee["id"] for employee in employees]
        if employee_ids:
            cursor.execute(
                f"""
                DELETE FROM employee_day_statuses
                WHERE status_type = 'day_off'
                  AND employee_id IN ({','.join(['?'] * len(employee_ids))})
                  AND date IN ({','.join(['?'] * len(week_dates))})
                """,
                [*employee_ids, *week_dates],
            )
        connection.commit()
        return {"message": "Week schedule cleared successfully", "deleted_count": deleted_count}
    finally:
        connection.close()


def load_position_employees(connection, position_id: int) -> list[dict]:
    cursor = connection.cursor()
    cursor.execute(
        """
        SELECT e.*, ep.is_primary, ep.priority_score, ep.is_fallback_only
        FROM employees e
        JOIN employee_positions ep ON ep.employee_id = e.id
        WHERE ep.position_id = ?
        ORDER BY ep.is_fallback_only ASC, ep.is_primary DESC, ep.priority_score DESC, e.id
        """,
        (position_id,),
    )
    employees = []
    for row in cursor.fetchall():
        employee = row_to_employee_dict(row)
        employee["is_primary"] = bool(row["is_primary"])
        employee["priority_score"] = row["priority_score"]
        employee["is_fallback_only"] = bool(row["is_fallback_only"])
        employees.append(employee)
    return employees


def load_active_templates(connection) -> list[dict]:
    cursor = connection.cursor()
    cursor.execute("SELECT * FROM shift_templates WHERE is_active = 1 ORDER BY start_time, end_time, category")
    return [row_to_shift_template_dict(row) for row in cursor.fetchall()]


def load_coverage_requirements_for_position(connection, position_id: int) -> list[dict]:
    cursor = connection.cursor()
    cursor.execute("SELECT * FROM coverage_requirements WHERE position_id = ? ORDER BY start_time", (position_id,))
    return [dict(row) for row in cursor.fetchall()]


def load_legacy_shift_requirements(connection, position_id: int) -> list[dict]:
    cursor = connection.cursor()
    cursor.execute("SELECT * FROM shift_requirements WHERE position_id = ?", (position_id,))
    return [dict(row) for row in cursor.fetchall()]


def build_atomic_slots(requirements: list[dict], templates: list[dict]) -> list[dict]:
    points = {0, 24 * 60}
    normalized_requirements = []
    for requirement in requirements:
        interval = build_interval(requirement["start_time"], requirement["end_time"], bool(requirement["is_overnight"]))
        normalized_requirements.append((interval, requirement))
        points.add(interval.start)
        points.add(interval.end)

    for template in templates:
        interval = build_interval(template["start_time"], template["end_time"], template["is_overnight"])
        points.add(interval.start)
        points.add(interval.end)

    sorted_points = sorted(points)
    slots = []
    for index in range(len(sorted_points) - 1):
        start = sorted_points[index]
        end = sorted_points[index + 1]
        if start == end:
            continue
        required_total = 0
        required_female = 0
        required_male = 0
        for interval, requirement in normalized_requirements:
            if interval.contains(start, end):
                required_total = max(required_total, requirement["required_total"])
                required_female = max(required_female, requirement["required_female_min"])
                required_male = max(required_male, requirement["required_male_min"])
        if required_total > 0 or required_female > 0 or required_male > 0:
            slots.append({
                "start": start,
                "end": end,
                "required_total": required_total,
                "required_female_min": required_female,
                "required_male_min": required_male,
            })
    return slots


def count_slot_coverage(entries: list[dict], slot: dict) -> tuple[int, int, int]:
    total = 0
    female = 0
    male = 0
    for entry in entries:
        if entry.get("no_show"):
            continue
        interval = build_interval(entry["start_time"], entry["end_time"], bool(entry["is_overnight"]))
        if interval.contains(slot["start"], slot["end"]):
            total += 1
            if entry.get("employee_sex") == "female":
                female += 1
            if entry.get("employee_sex") == "male":
                male += 1
    return total, female, male


def coverage_shortage(entries: list[dict], slots: list[dict]) -> int:
    shortage = 0
    for slot in slots:
        total, female, male = count_slot_coverage(entries, slot)
        shortage += max(0, slot["required_total"] - total) * 10
        shortage += max(0, slot["required_female_min"] - female) * 4
        shortage += max(0, slot["required_male_min"] - male) * 4
    return shortage


def coverage_overage(entries: list[dict], slots: list[dict]) -> int:
    overage = 0
    for slot in slots:
        total, _female, _male = count_slot_coverage(entries, slot)
        overage += max(0, total - slot["required_total"])
    return overage


def template_covers_slot(template: dict, slot: dict) -> bool:
    interval = build_interval(template["start_time"], template["end_time"], template["is_overnight"])
    return interval.contains(slot["start"], slot["end"])


def get_employee_week_category_count(connection, employee_id: int, week_start_date: str, category: str) -> int:
    cursor = connection.cursor()
    cursor.execute(
        """
        SELECT COUNT(*) AS cnt
        FROM schedule_entries se
        JOIN shift_templates st ON st.id = se.shift_template_id
        WHERE se.employee_id = ?
          AND se.date >= ?
          AND se.date <= ?
          AND st.category = ?
          AND se.no_show = 0
        """,
        (employee_id, week_start_date, get_week_end_date(week_start_date), category),
    )
    return cursor.fetchone()["cnt"]


def get_employee_week_split_day_count(connection, employee_id: int, week_start_date: str) -> int:
    return sum(
        1
        for date_string in build_week_dates(week_start_date)
        if employee_has_split_day(connection, employee_id, date_string)
    )


def candidate_priority(connection, employee: dict, date_string: str, week_start_date: str) -> tuple:
    week_count = get_employee_week_shift_count(connection, employee["id"], week_start_date)
    worked_dates = get_employee_week_worked_dates(connection, employee["id"], week_start_date)
    night_count = get_employee_week_category_count(connection, employee["id"], week_start_date, "night")
    split_count = get_employee_week_split_day_count(connection, employee["id"], week_start_date)
    if week_count < employee["min_shifts_per_week"]:
        bucket = 0
    elif week_count < employee["target_shifts_per_week"]:
        bucket = 1
    else:
        bucket = 2
    fallback_penalty = 1 if employee.get("is_fallback_only") else 0
    primary_bonus = 0 if employee.get("is_primary") else 1
    return (
        bucket,
        len(worked_dates),
        week_count,
        night_count,
        split_count,
        fallback_penalty,
        primary_bonus,
        -int(employee.get("priority_score", 50)),
        employee["id"],
    )


def create_entry_preview(employee: dict, position_id: int, date_string: str, template: dict) -> dict:
    return {
        "id": -1,
        "employee_id": employee["id"],
        "position_id": position_id,
        "date": date_string,
        "shift_template_id": template["id"],
        "shift_template_name": template["name"],
        "shift_category": template["category"],
        "start_time": template["start_time"],
        "end_time": template["end_time"],
        "is_overnight": template["is_overnight"],
        "is_split_only": template["is_split_only"],
        "employee_name": employee["full_name"],
        "employee_sex": employee["sex"],
    }


def template_start_minutes(template: dict) -> int:
    return time_to_minutes(template["start_time"])


def build_template_assignment_options(template: dict, templates: list[dict]) -> list[list[dict]]:
    if not template["is_split_only"]:
        return [[template]]

    if template["category"] not in {"morning", "evening"}:
        return []

    partner_category = "evening" if template["category"] == "morning" else "morning"
    partners = [
        candidate
        for candidate in templates
        if candidate["id"] != template["id"]
        and candidate["is_split_only"]
        and candidate["category"] == partner_category
    ]

    options = []
    for partner in partners:
        pair = sorted([template, partner], key=template_start_minutes)
        options.append(pair)
    return options


def build_valid_assignment_previews(
    connection,
    employee: dict,
    assignment_templates: list[dict],
    position_id: int,
    date_string: str,
    week_start_date: str,
    fatigue_relaxation: int = 0,
) -> list[dict] | None:
    staged_entries: list[dict] = []

    for template in assignment_templates:
        if not can_employee_take_template(
            connection,
            employee,
            position_id,
            date_string,
            template,
            week_start_date,
            fatigue_relaxation=fatigue_relaxation,
            staged_entries=staged_entries,
        ):
            return None
        staged_entries.append(create_entry_preview(employee, position_id, date_string, template))

    return staged_entries


def choose_best_interval_candidate(
    connection,
    employees: list[dict],
    templates: list[dict],
    position_id: int,
    date_string: str,
    week_start_date: str,
    current_entries: list[dict],
    slots: list[dict],
    fatigue_relaxation: int = 0,
    target_slot: dict | None = None,
):
    app_settings = get_app_settings(connection)
    baseline_shortage = coverage_shortage(current_entries, slots)
    baseline_overage = coverage_overage(current_entries, slots)
    target_needs_female = False
    target_needs_male = False
    target_needs_total = False
    if target_slot is not None:
        target_total, target_female, target_male = count_slot_coverage(current_entries, target_slot)
        target_needs_total = target_total < target_slot["required_total"]
        target_needs_female = target_female < target_slot["required_female_min"]
        target_needs_male = target_male < target_slot["required_male_min"]
    candidates = []

    for employee in employees:
        if target_slot is not None and not target_needs_total:
            if target_needs_female and employee["sex"] != "female":
                continue
            if target_needs_male and employee["sex"] != "male":
                continue
        for template in templates:
            if target_slot is not None and not template_covers_slot(template, target_slot):
                continue
            for assignment_templates in build_template_assignment_options(template, templates):
                previews = build_valid_assignment_previews(
                    connection,
                    employee,
                    assignment_templates,
                    position_id,
                    date_string,
                    week_start_date,
                    fatigue_relaxation=fatigue_relaxation,
                )
                if previews is None:
                    continue

                projected_entries = [*current_entries, *previews]
                shortage_gain = baseline_shortage - coverage_shortage(projected_entries, slots)
                overage_cost = coverage_overage(projected_entries, slots) - baseline_overage
                score = (
                    shortage_gain * app_settings["coverage_shortage_gain_weight"]
                    - overage_cost * app_settings["coverage_overage_penalty_weight"]
                )
                if target_slot is not None:
                    if target_needs_female and employee["sex"] == "female":
                        score += app_settings["target_gender_bonus_weight"]
                    if target_needs_male and employee["sex"] == "male":
                        score += app_settings["target_gender_bonus_weight"]
                    if (target_needs_female and employee["sex"] != "female") or (target_needs_male and employee["sex"] != "male"):
                        score -= app_settings["wrong_gender_penalty_weight"]
                if score <= 0:
                    continue

                fatigue_penalty = sum(
                    get_fatigue_penalty(connection, employee["id"], date_string, assignment_template)
                    for assignment_template in assignment_templates
                )
                night_balance = (
                    get_employee_week_category_count(connection, employee["id"], week_start_date, "night")
                    if any(assignment_template["category"] == "night" for assignment_template in assignment_templates)
                    else 0
                )
                split_balance = (
                    get_employee_week_split_day_count(connection, employee["id"], week_start_date)
                    if any(assignment_template["is_split_only"] for assignment_template in assignment_templates)
                    else 0
                )
                candidates.append((
                    (-score, fatigue_penalty, night_balance, split_balance, candidate_priority(connection, employee, date_string, week_start_date)),
                    employee,
                    assignment_templates,
                    score,
                ))

    if not candidates:
        return None

    candidates.sort(key=lambda item: item[0])
    return candidates[0][1], candidates[0][2], candidates[0][3]


def insert_generated_entry(cursor, employee: dict, position_id: int, date_string: str, template: dict, created_entries: list[dict]):
    cursor.execute(
        """
        INSERT INTO schedule_entries (employee_id, position_id, date, shift_template_id)
        VALUES (?, ?, ?, ?)
        """,
        (employee["id"], position_id, date_string, template["id"]),
    )
    created_entries.append(
        {
            "id": cursor.lastrowid,
            "employee_id": employee["id"],
            "employee_name": employee["full_name"],
            "date": date_string,
            "shift_template_name": template["name"],
            "shift_category": template["category"],
        }
    )


def insert_generated_assignment(cursor, employee: dict, position_id: int, date_string: str, templates: list[dict], created_entries: list[dict]):
    for template in templates:
        insert_generated_entry(cursor, employee, position_id, date_string, template, created_entries)


def count_candidate_options_for_slot(
    connection,
    employees: list[dict],
    templates: list[dict],
    position_id: int,
    date_string: str,
    week_start_date: str,
    slot: dict,
    fatigue_relaxation: int = 0,
) -> int:
    count = 0
    for employee in employees:
        for template in templates:
            if not template_covers_slot(template, slot):
                continue
            if any(
                build_valid_assignment_previews(
                    connection,
                    employee,
                    assignment_templates,
                    position_id,
                    date_string,
                    week_start_date,
                    fatigue_relaxation=fatigue_relaxation,
                )
                is not None
                for assignment_templates in build_template_assignment_options(template, templates)
            ):
                count += 1
    return count


def format_slot_time(slot: dict) -> str:
    return f"{slot['start'] // 60:02d}:{slot['start'] % 60:02d}-{slot['end'] // 60 % 24:02d}:{slot['end'] % 60:02d}"


def slot_to_report(slot: dict) -> dict:
    return {
        "start": f"{slot['start'] // 60:02d}:{slot['start'] % 60:02d}",
        "end": f"{slot['end'] // 60 % 24:02d}:{slot['end'] % 60:02d}",
        "required_total": slot["required_total"],
        "required_female_min": slot["required_female_min"],
        "required_male_min": slot["required_male_min"],
    }


def summarize_reasons(reason_counts: dict[str, int], limit: int = 4) -> str:
    if not reason_counts:
        return "no available employees found"
    ordered = sorted(reason_counts.items(), key=lambda item: (-item[1], item[0]))
    return "; ".join(f"{reason} ({count})" for reason, count in ordered[:limit])


def explain_unfilled_interval_slot(
    connection,
    employees: list[dict],
    templates: list[dict],
    position_id: int,
    date_string: str,
    week_start_date: str,
    slot: dict,
    require_female: bool = False,
    require_male: bool = False,
) -> str:
    covering_templates = [template for template in templates if template_covers_slot(template, slot)]
    if not covering_templates:
        return "no active shift template covers this time interval"

    reason_counts: dict[str, int] = {}
    for employee in employees:
        if require_female and employee["sex"] != "female":
            reason_counts["not enough female employees available"] = reason_counts.get("not enough female employees available", 0) + 1
            continue
        if require_male and employee["sex"] != "male":
            reason_counts["not enough male employees available"] = reason_counts.get("not enough male employees available", 0) + 1
            continue

        employee_had_relevant_template = False
        for template in covering_templates:
            options = build_template_assignment_options(template, templates)
            if not options:
                reason_counts["split-only template has no valid pair"] = reason_counts.get("split-only template has no valid pair", 0) + 1
                continue

            for assignment_templates in options:
                staged_entries: list[dict] = []
                option_reason = None
                for assignment_template in assignment_templates:
                    reason = explain_employee_template_rejection(
                        connection,
                        employee,
                        position_id,
                        date_string,
                        assignment_template,
                        week_start_date,
                        fatigue_relaxation=1,
                        staged_entries=staged_entries,
                    )
                    if reason:
                        option_reason = reason
                        break
                    staged_entries.append(create_entry_preview(employee, position_id, date_string, assignment_template))

                if option_reason is None:
                    employee_had_relevant_template = True
                    break
                reason_counts[option_reason] = reason_counts.get(option_reason, 0) + 1

            if employee_had_relevant_template:
                break

    if not reason_counts:
        return "candidates existed, but they did not improve coverage without overfilling other intervals"

    return summarize_reasons(reason_counts)


def build_interval_underfilled_message(
    connection,
    employees: list[dict],
    templates: list[dict],
    position_id: int,
    week_start_date: str,
    date_string: str,
    slot: dict,
    total: int,
    female: int,
    male: int,
) -> str:
    require_female = female < slot["required_female_min"]
    require_male = male < slot["required_male_min"]
    reasons = explain_unfilled_interval_slot(
        connection,
        employees,
        templates,
        position_id,
        date_string,
        week_start_date,
        slot,
        require_female=require_female,
        require_male=require_male,
    )
    return (
        f"{date_string} {format_slot_time(slot)} underfilled: "
        f"staff {total}/{slot['required_total']}, women {female}/{slot['required_female_min']}, "
        f"men {male}/{slot['required_male_min']}. "
        f"Reasons: {reasons}"
    )


def build_interval_underfilled_report(
    connection,
    employees: list[dict],
    templates: list[dict],
    position_id: int,
    week_start_date: str,
    date_string: str,
    slot: dict,
    total: int,
    female: int,
    male: int,
) -> dict:
    require_female = female < slot["required_female_min"]
    require_male = male < slot["required_male_min"]
    reasons = explain_unfilled_interval_slot(
        connection,
        employees,
        templates,
        position_id,
        date_string,
        week_start_date,
        slot,
        require_female=require_female,
        require_male=require_male,
    )
    return {
        "kind": "interval",
        "date": date_string,
        "slot": slot_to_report(slot),
        "actual": {
            "total": total,
            "female": female,
            "male": male,
        },
        "missing": {
            "total": max(0, slot["required_total"] - total),
            "female": max(0, slot["required_female_min"] - female),
            "male": max(0, slot["required_male_min"] - male),
        },
        "reasons": reasons,
    }


def build_legacy_underfilled_report(
    connection,
    employees: list[dict],
    templates: list[dict],
    position_id: int,
    week_start_date: str,
    date_string: str,
    requirement: dict,
    total: int,
    female: int,
    male: int,
) -> dict:
    require_female = female < requirement["required_female_min"]
    require_male = male < requirement["required_male_min"]
    reasons = explain_unfilled_legacy_category(
        connection,
        employees,
        templates,
        position_id,
        date_string,
        week_start_date,
        requirement["shift_category"],
        require_female=require_female,
        require_male=require_male,
    )
    return {
        "kind": "legacy_category",
        "date": date_string,
        "shift_category": requirement["shift_category"],
        "actual": {
            "total": total,
            "female": female,
            "male": male,
        },
        "missing": {
            "total": max(0, requirement["required_total"] - total),
            "female": max(0, requirement["required_female_min"] - female),
            "male": max(0, requirement["required_male_min"] - male),
        },
        "reasons": reasons,
    }


def explain_unfilled_legacy_category(
    connection,
    employees: list[dict],
    templates: list[dict],
    position_id: int,
    date_string: str,
    week_start_date: str,
    category: str,
    require_female: bool = False,
    require_male: bool = False,
) -> str:
    category_templates = [
        template
        for template in templates
        if template["category"] == category and not template["is_split_only"]
    ]
    if not category_templates:
        return f"no active non-split {category} templates"

    reason_counts: dict[str, int] = {}
    for employee in employees:
        if require_female and employee["sex"] != "female":
            reason_counts["not enough female employees available"] = reason_counts.get("not enough female employees available", 0) + 1
            continue
        if require_male and employee["sex"] != "male":
            reason_counts["not enough male employees available"] = reason_counts.get("not enough male employees available", 0) + 1
            continue

        for template in category_templates:
            reason = explain_employee_template_rejection(
                connection,
                employee,
                position_id,
                date_string,
                template,
                week_start_date,
                fatigue_relaxation=1,
            )
            if reason is None:
                break
            reason_counts[reason] = reason_counts.get(reason, 0) + 1

    return summarize_reasons(reason_counts)


def build_week_shortage_queue(
    connection,
    employees: list[dict],
    templates: list[dict],
    position_id: int,
    week_start_date: str,
    week_dates: list[str],
    slots: list[dict],
) -> list[dict]:
    queue = []
    for date_index, date_string in enumerate(week_dates):
        entries = get_schedule_entries(connection, position_id=position_id, dates=[date_string])
        for slot in slots:
            total, female, male = count_slot_coverage(entries, slot)
            missing_total = max(0, slot["required_total"] - total)
            missing_female = max(0, slot["required_female_min"] - female)
            missing_male = max(0, slot["required_male_min"] - male)
            if missing_total <= 0 and missing_female <= 0 and missing_male <= 0:
                continue

            strict_candidates = count_candidate_options_for_slot(
                connection,
                employees,
                templates,
                position_id,
                date_string,
                week_start_date,
                slot,
                fatigue_relaxation=0,
            )
            emergency_candidates = strict_candidates or count_candidate_options_for_slot(
                connection,
                employees,
                templates,
                position_id,
                date_string,
                week_start_date,
                slot,
                fatigue_relaxation=1,
            )
            scarcity = strict_candidates if strict_candidates > 0 else emergency_candidates + 1000
            queue.append(
                {
                    "date": date_string,
                    "date_index": date_index,
                    "slot": slot,
                    "is_night_slot": slot["start"] >= 23 * 60 or slot["end"] <= 7 * 60 or slot["end"] > 24 * 60,
                    "missing_total": missing_total,
                    "missing_female": missing_female,
                    "missing_male": missing_male,
                    "strict_candidates": strict_candidates,
                    "emergency_candidates": emergency_candidates,
                    "scarcity": scarcity,
                }
            )

    queue.sort(
        key=lambda item: (
            -int(item["is_night_slot"]),
            item["scarcity"],
            -item["date_index"],
            -item["missing_female"],
            -item["missing_male"],
            -item["missing_total"],
            item["slot"]["start"],
        )
    )
    return queue


def build_generation_feasibility_report(
    connection,
    employees: list[dict],
    templates: list[dict],
    coverage_requirements: list[dict],
    legacy_requirements: list[dict],
    position_id: int,
    week_start_date: str,
    week_dates: list[str],
) -> dict:
    issues: list[dict] = []
    employee_count = len(employees)
    female_count = sum(1 for employee in employees if employee["sex"] == "female")
    male_count = sum(1 for employee in employees if employee["sex"] == "male")

    if coverage_requirements:
        slots = build_atomic_slots(coverage_requirements, templates)
        if not slots:
            issues.append({
                "severity": "blocking",
                "kind": "coverage",
                "message": "No active coverage slots can be built from coverage requirements.",
            })

        for slot in slots:
            covering_templates = [template for template in templates if template_covers_slot(template, slot)]
            if not covering_templates:
                issues.append({
                    "severity": "blocking",
                    "kind": "template",
                    "slot": slot_to_report(slot),
                    "message": "No active shift template covers this required interval.",
                })
            if slot["required_total"] > employee_count:
                issues.append({
                    "severity": "blocking",
                    "kind": "staff",
                    "slot": slot_to_report(slot),
                    "message": "Required staff is greater than employees assigned to the position.",
                })
            if slot["required_female_min"] > female_count:
                issues.append({
                    "severity": "blocking",
                    "kind": "female_staff",
                    "slot": slot_to_report(slot),
                    "message": "Required female staff is greater than available female employees.",
                })
            if slot["required_male_min"] > male_count:
                issues.append({
                    "severity": "blocking",
                    "kind": "male_staff",
                    "slot": slot_to_report(slot),
                    "message": "Required male staff is greater than available male employees.",
                })

        for date_string in week_dates:
            for slot in slots:
                strict_candidates = count_candidate_options_for_slot(
                    connection,
                    employees,
                    templates,
                    position_id,
                    date_string,
                    week_start_date,
                    slot,
                    fatigue_relaxation=0,
                )
                emergency_candidates = strict_candidates or count_candidate_options_for_slot(
                    connection,
                    employees,
                    templates,
                    position_id,
                    date_string,
                    week_start_date,
                    slot,
                    fatigue_relaxation=1,
                )
                if emergency_candidates == 0:
                    issues.append({
                        "severity": "blocking",
                        "kind": "candidate",
                        "date": date_string,
                        "slot": slot_to_report(slot),
                        "message": "No eligible employee/template candidate can cover this interval.",
                    })
                elif strict_candidates == 0:
                    issues.append({
                        "severity": "warning",
                        "kind": "emergency_relaxation",
                        "date": date_string,
                        "slot": slot_to_report(slot),
                        "message": "This interval is only coverable with emergency fatigue relaxation.",
                    })
    else:
        for requirement in legacy_requirements:
            category_templates = [
                template
                for template in templates
                if template["category"] == requirement["shift_category"] and not template["is_split_only"]
            ]
            if not category_templates:
                issues.append({
                    "severity": "blocking",
                    "kind": "template",
                    "shift_category": requirement["shift_category"],
                    "message": "No active non-split template exists for this legacy shift requirement.",
                })
            if requirement["required_total"] > employee_count:
                issues.append({
                    "severity": "blocking",
                    "kind": "staff",
                    "shift_category": requirement["shift_category"],
                    "message": "Required staff is greater than employees assigned to the position.",
                })
            if requirement["required_female_min"] > female_count:
                issues.append({
                    "severity": "blocking",
                    "kind": "female_staff",
                    "shift_category": requirement["shift_category"],
                    "message": "Required female staff is greater than available female employees.",
                })
            if requirement["required_male_min"] > male_count:
                issues.append({
                    "severity": "blocking",
                    "kind": "male_staff",
                    "shift_category": requirement["shift_category"],
                    "message": "Required male staff is greater than available male employees.",
                })

    for issue in issues:
        issue["constraint_type"] = "hard" if issue["severity"] == "blocking" else "soft"

    hard_constraints = [issue for issue in issues if issue["constraint_type"] == "hard"]
    soft_constraints = [issue for issue in issues if issue["constraint_type"] == "soft"]

    return {
        "status": "blocking" if hard_constraints else "ok",
        "issues": issues,
        "hard_constraints": hard_constraints,
        "soft_constraints": soft_constraints,
    }


def format_feasibility_issue(issue: dict) -> str:
    parts = []
    if issue.get("date"):
        parts.append(str(issue["date"]))
    if issue.get("slot"):
        slot = issue["slot"]
        parts.append(f"{slot['start']}-{slot['end']}")
    if issue.get("shift_category"):
        parts.append(str(issue["shift_category"]))

    prefix = " ".join(parts)
    if prefix:
        return f"Pre-check {issue['severity']}: {prefix}: {issue['message']}"
    return f"Pre-check {issue['severity']}: {issue['message']}"


def fill_week_by_interval_coverage(
    connection,
    cursor,
    employees: list[dict],
    templates: list[dict],
    requirements: list[dict],
    position_id: int,
    week_start_date: str,
    week_dates: list[str],
    created_entries: list[dict],
    errors: list[str],
    unfilled_reports: list[dict],
):
    slots = build_atomic_slots(requirements, templates)
    if not slots:
        errors.append("No active coverage slots for this week")
        return

    guard = 0
    while guard < 1500:
        guard += 1
        queue = build_week_shortage_queue(
            connection,
            employees,
            templates,
            position_id,
            week_start_date,
            week_dates,
            slots,
        )
        if not queue:
            break

        assigned = False
        for shortage in queue:
            current_entries = get_schedule_entries(connection, position_id=position_id, dates=[shortage["date"]])
            for fatigue_relaxation in (0, 1):
                candidate = choose_best_interval_candidate(
                    connection=connection,
                    employees=employees,
                    templates=templates,
                    position_id=position_id,
                    date_string=shortage["date"],
                    week_start_date=week_start_date,
                    current_entries=current_entries,
                    slots=slots,
                    fatigue_relaxation=fatigue_relaxation,
                    target_slot=shortage["slot"],
                )
                if candidate is None:
                    continue

                employee, assignment_templates, _score = candidate
                insert_generated_assignment(cursor, employee, position_id, shortage["date"], assignment_templates, created_entries)
                if fatigue_relaxation == 1:
                    errors.append(f"{shortage['date']}: emergency fatigue relaxation was used to cover a slot")
                assigned = True
                break

            if assigned:
                break

        if not assigned:
            break

    for date_string in week_dates:
        final_entries = get_schedule_entries(connection, position_id=position_id, dates=[date_string])
        for slot in slots:
            total, female, male = count_slot_coverage(final_entries, slot)
            if total < slot["required_total"] or female < slot["required_female_min"] or male < slot["required_male_min"]:
                unfilled_reports.append(build_interval_underfilled_report(
                    connection,
                    employees,
                    templates,
                    position_id,
                    week_start_date,
                    date_string,
                    slot,
                    total,
                    female,
                    male,
                ))
                errors.append(build_interval_underfilled_message(
                    connection,
                    employees,
                    templates,
                    position_id,
                    week_start_date,
                    date_string,
                    slot,
                    total,
                    female,
                    male,
                ))


def fill_day_by_interval_coverage(
    connection,
    cursor,
    employees: list[dict],
    templates: list[dict],
    requirements: list[dict],
    position_id: int,
    week_start_date: str,
    date_string: str,
    created_entries: list[dict],
    errors: list[str],
    unfilled_reports: list[dict],
):
    slots = build_atomic_slots(requirements, templates)
    if not slots:
        errors.append(f"No active coverage slots for {date_string}")
        return

    guard = 0
    while guard < 200:
        guard += 1
        current_entries = get_schedule_entries(connection, position_id=position_id, dates=[date_string])
        if coverage_shortage(current_entries, slots) == 0:
            break
        candidate = None
        for fatigue_relaxation in (0, 1):
            candidate = choose_best_interval_candidate(
                connection=connection,
                employees=employees,
                templates=templates,
                position_id=position_id,
                date_string=date_string,
                week_start_date=week_start_date,
                current_entries=current_entries,
                slots=slots,
                fatigue_relaxation=fatigue_relaxation,
            )
            if candidate is not None:
                if fatigue_relaxation == 1:
                    errors.append(f"{date_string}: emergency fatigue relaxation was used to cover a slot")
                break
        if candidate is None:
            break
        employee, assignment_templates, _score = candidate
        insert_generated_assignment(cursor, employee, position_id, date_string, assignment_templates, created_entries)

    final_entries = get_schedule_entries(connection, position_id=position_id, dates=[date_string])
    remaining = coverage_shortage(final_entries, slots)
    if remaining > 0:
        for slot in slots:
            total, female, male = count_slot_coverage(final_entries, slot)
            if total < slot["required_total"] or female < slot["required_female_min"] or male < slot["required_male_min"]:
                unfilled_reports.append(build_interval_underfilled_report(
                    connection,
                    employees,
                    templates,
                    position_id,
                    week_start_date,
                    date_string,
                    slot,
                    total,
                    female,
                    male,
                ))
                errors.append(build_interval_underfilled_message(
                    connection,
                    employees,
                    templates,
                    position_id,
                    week_start_date,
                    date_string,
                    slot,
                    total,
                    female,
                    male,
                ))


def fill_day_by_legacy_categories(
    connection,
    cursor,
    employees: list[dict],
    templates: list[dict],
    requirements: list[dict],
    position_id: int,
    week_start_date: str,
    date_string: str,
    created_entries: list[dict],
    errors: list[str],
    unfilled_reports: list[dict],
):
    for requirement in requirements:
        category = requirement["shift_category"]
        category_templates = [template for template in templates if template["category"] == category and not template["is_split_only"]]
        while True:
            entries = [
                entry for entry in get_schedule_entries(connection, position_id=position_id, dates=[date_string])
                if entry["shift_category"] == category
            ]
            female_count = sum(1 for entry in entries if entry["employee_sex"] == "female")
            male_count = sum(1 for entry in entries if entry["employee_sex"] == "male")
            if (
                len(entries) >= requirement["required_total"]
                and female_count >= requirement["required_female_min"]
                and male_count >= requirement["required_male_min"]
            ):
                break

            require_female = female_count < requirement["required_female_min"]
            require_male = male_count < requirement["required_male_min"]
            candidates = []
            for fatigue_relaxation in (0, 1):
                for employee in employees:
                    if require_female and employee["sex"] != "female":
                        continue
                    if require_male and employee["sex"] != "male":
                        continue
                    for template in category_templates:
                        if can_employee_take_template(
                            connection,
                            employee,
                            position_id,
                            date_string,
                            template,
                            week_start_date,
                            fatigue_relaxation=fatigue_relaxation,
                        ):
                            fatigue_penalty = get_fatigue_penalty(connection, employee["id"], date_string, template)
                            candidates.append((fatigue_penalty, candidate_priority(connection, employee, date_string, week_start_date), employee, template))
                if candidates:
                    if fatigue_relaxation == 1:
                        errors.append(f"{date_string} {category}: emergency fatigue relaxation was used to cover a slot")
                    break
            if not candidates:
                report = build_legacy_underfilled_report(
                    connection,
                    employees,
                    templates,
                    position_id,
                    week_start_date,
                    date_string,
                    requirement,
                    len(entries),
                    female_count,
                    male_count,
                )
                unfilled_reports.append(report)
                errors.append(
                    f"{date_string} {category} underfilled: staff {len(entries)}/{requirement['required_total']}, "
                    f"women {female_count}/{requirement['required_female_min']}, "
                    f"men {male_count}/{requirement['required_male_min']}. "
                    f"Reasons: {report['reasons']}"
                )
                break
            candidates.sort()
            _, _, employee, template = candidates[0]
            insert_generated_entry(cursor, employee, position_id, date_string, template, created_entries)


def max_consecutive_in_week(connection, employee_id: int, week_dates: list[str], predicate) -> int:
    best = 0
    current = 0

    for date_string in week_dates:
        if predicate(connection, employee_id, date_string):
            current += 1
            best = max(best, current)
        else:
            current = 0

    return best


def append_fatigue_summary_warnings(connection, employees: list[dict], week_dates: list[str], week_start_date: str, errors: list[str]) -> None:
    app_settings = get_app_settings(connection)
    for employee in employees:
        worked_days = len(get_employee_week_worked_dates(connection, employee["id"], week_start_date))
        if worked_days > app_settings["max_work_days_per_week"]:
            errors.append(
                f"{employee['full_name']} has {worked_days} worked days in the week; mandatory weekly day off is violated"
            )

        max_nights = max_consecutive_in_week(connection, employee["id"], week_dates, employee_has_night_on_date)
        if max_nights > app_settings["max_consecutive_nights"]:
            errors.append(
                f"{employee['full_name']} has {max_nights} consecutive night days; normal limit is {app_settings['max_consecutive_nights']}"
            )

        max_splits = max_consecutive_in_week(connection, employee["id"], week_dates, employee_has_split_day)
        if max_splits > app_settings["max_consecutive_split_days"]:
            errors.append(
                f"{employee['full_name']} has {max_splits} consecutive split days; normal limit is {app_settings['max_consecutive_split_days']}"
            )


def get_employee_week_entries(
    connection,
    employee_id: int,
    week_start_date: str,
    exclude_entry_ids: set[int] | None = None,
    staged_entries: list[dict] | None = None,
) -> list[dict]:
    exclude_entry_ids = exclude_entry_ids or set()
    staged_entries = staged_entries or []

    cursor = connection.cursor()
    cursor.execute(
        """
        SELECT
            se.id,
            se.employee_id,
            se.position_id,
            se.date,
            se.shift_template_id,
            se.no_show,
            st.name AS shift_template_name,
            st.category,
            st.start_time,
            st.end_time,
            st.is_overnight,
            st.is_split_only
        FROM schedule_entries se
        JOIN shift_templates st ON st.id = se.shift_template_id
        WHERE se.employee_id = ?
          AND se.date >= ?
          AND se.date <= ?
          AND se.no_show = 0
        """,
        (employee_id, week_start_date, get_week_end_date(week_start_date)),
    )

    entries = [dict(row) for row in cursor.fetchall() if row["id"] not in exclude_entry_ids]
    entries.extend(staged_entries)
    return entries


def max_consecutive_projected(entries: list[dict], week_dates: list[str], predicate) -> int:
    entries_by_date: dict[str, list[dict]] = {}
    for entry in entries:
        entries_by_date.setdefault(entry["date"], []).append(entry)

    best = 0
    current = 0
    for date_string in week_dates:
        if predicate(entries_by_date.get(date_string, [])):
            current += 1
            best = max(best, current)
        else:
            current = 0
    return best


def projected_employee_score(employee: dict, entries: list[dict], week_dates: list[str], app_settings: dict) -> int:
    week_count = len(entries)
    worked_dates = {entry["date"] for entry in entries}
    night_count = sum(1 for entry in entries if entry_category(entry) == "night")
    split_day_count = sum(
        1
        for date_string in week_dates
        if {"morning", "evening"}.issubset({entry_category(entry) for entry in entries if entry["date"] == date_string})
    )
    max_nights = max_consecutive_projected(
        entries,
        week_dates,
        lambda day_entries: any(entry_category(entry) == "night" for entry in day_entries),
    )
    max_splits = max_consecutive_projected(
        entries,
        week_dates,
        lambda day_entries: {"morning", "evening"}.issubset({entry_category(entry) for entry in day_entries}),
    )

    score = 0
    score += max(0, employee["min_shifts_per_week"] - week_count) * app_settings["balance_missing_min_weight"]
    score += abs(employee["target_shifts_per_week"] - week_count) * app_settings["balance_target_distance_weight"]
    score += max(0, week_count - employee["target_shifts_per_week"]) * app_settings["balance_over_target_weight"]
    score += max(0, week_count - employee["max_shifts_per_week"]) * app_settings["balance_over_max_weight"]
    score += len(worked_dates) * app_settings["balance_worked_day_weight"]
    score += max(0, len(worked_dates) - app_settings["max_work_days_per_week"]) * app_settings["balance_over_max_weight"]
    score += night_count * app_settings["balance_night_weight"]
    score += split_day_count * app_settings["balance_split_weight"]
    score += max_nights * app_settings["balance_consecutive_night_weight"]
    score += max_splits * app_settings["balance_consecutive_split_weight"]
    score += max(0, max_nights - app_settings["max_consecutive_nights"]) * app_settings["balance_excess_night_weight"]
    score += max(0, max_splits - app_settings["max_consecutive_split_days"]) * app_settings["balance_excess_split_weight"]
    return score


def row_to_template_for_assignment(row: dict) -> dict:
    return {
        "id": row["shift_template_id"],
        "name": row["shift_template_name"],
        "category": entry_category(row),
        "start_time": row["start_time"],
        "end_time": row["end_time"],
        "is_overnight": bool(row["is_overnight"]),
        "is_active": True,
        "is_split_only": bool(row["is_split_only"]),
    }


def build_generated_assignment_groups(connection, created_entries: list[dict], position_id: int) -> list[list[dict]]:
    created_ids = [entry["id"] for entry in created_entries if entry.get("id")]
    if not created_ids:
        return []

    placeholders = ",".join(["?"] * len(created_ids))
    cursor = connection.cursor()
    cursor.execute(
        f"""
        SELECT
            se.id,
            se.employee_id,
            se.position_id,
            se.date,
            se.shift_template_id,
            st.name AS shift_template_name,
            st.category,
            st.start_time,
            st.end_time,
            st.is_overnight,
            st.is_split_only
        FROM schedule_entries se
        JOIN shift_templates st ON st.id = se.shift_template_id
        WHERE se.id IN ({placeholders})
          AND se.position_id = ?
        ORDER BY se.date, se.employee_id, st.start_time
        """,
        [*created_ids, position_id],
    )
    rows = [dict(row) for row in cursor.fetchall()]
    rows_by_id = {row["id"]: row for row in rows}
    visited: set[int] = set()
    groups: list[list[dict]] = []

    for row in rows:
        if row["id"] in visited:
            continue

        if row["is_split_only"] and row["category"] in {"morning", "evening"}:
            group = [
                candidate
                for candidate in rows
                if candidate["employee_id"] == row["employee_id"]
                and candidate["position_id"] == row["position_id"]
                and candidate["date"] == row["date"]
                and candidate["is_split_only"]
                and candidate["category"] in {"morning", "evening"}
            ]
        else:
            group = [row]

        group = [item for item in group if item["id"] in rows_by_id and item["id"] not in visited]
        if not group:
            continue
        for item in group:
            visited.add(item["id"])
        groups.append(sorted(group, key=lambda item: time_to_minutes(item["start_time"])))

    return groups


def post_optimize_generated_schedule(
    connection,
    cursor,
    employees: list[dict],
    position_id: int,
    week_start_date: str,
    week_dates: list[str],
    created_entries: list[dict],
    errors: list[str],
) -> int:
    app_settings = get_app_settings(connection)
    employee_by_id = {employee["id"]: employee for employee in employees}
    groups = build_generated_assignment_groups(connection, created_entries, position_id)
    moved_count = 0

    for group in groups:
        if not group:
            continue

        original_employee = employee_by_id.get(group[0]["employee_id"])
        if original_employee is None:
            continue

        date_string = group[0]["date"]
        entry_ids = {entry["id"] for entry in group}
        assignment_templates = [row_to_template_for_assignment(entry) for entry in group]
        original_entries_before = get_employee_week_entries(connection, original_employee["id"], week_start_date)
        original_entries_after = get_employee_week_entries(
            connection,
            original_employee["id"],
            week_start_date,
            exclude_entry_ids=entry_ids,
        )
        original_score_before = projected_employee_score(original_employee, original_entries_before, week_dates, app_settings)
        original_score_after = projected_employee_score(original_employee, original_entries_after, week_dates, app_settings)

        best_employee = None
        best_improvement = 0
        for candidate in employees:
            if candidate["id"] == original_employee["id"]:
                continue

            previews = build_valid_assignment_previews(
                connection,
                candidate,
                assignment_templates,
                position_id,
                date_string,
                week_start_date,
                fatigue_relaxation=0,
            )
            if previews is None:
                continue

            candidate_entries_before = get_employee_week_entries(connection, candidate["id"], week_start_date)
            candidate_entries_after = get_employee_week_entries(
                connection,
                candidate["id"],
                week_start_date,
                staged_entries=previews,
            )
            before_score = original_score_before + projected_employee_score(
                candidate,
                candidate_entries_before,
                week_dates,
                app_settings,
            )
            after_score = original_score_after + projected_employee_score(
                candidate,
                candidate_entries_after,
                week_dates,
                app_settings,
            )
            improvement = before_score - after_score

            if improvement > best_improvement:
                best_improvement = improvement
                best_employee = candidate

        if best_employee is None or best_improvement < 80:
            continue

        cursor.execute(
            f"""
            UPDATE schedule_entries
            SET employee_id = ?
            WHERE id IN ({','.join(['?'] * len(entry_ids))})
            """,
            [best_employee["id"], *entry_ids],
        )
        for created_entry in created_entries:
            if created_entry.get("id") in entry_ids:
                created_entry["employee_id"] = best_employee["id"]
                created_entry["employee_name"] = best_employee["full_name"]
        moved_count += len(entry_ids)

    return moved_count


@app.post("/api/schedule/auto-generate", tags=["Schedule"])
def auto_generate_schedule(request_data: AutoGenerateScheduleRequest):
    connection = get_connection()
    try:
        cursor = connection.cursor()
        fetch_one_or_404(cursor, "SELECT id FROM positions WHERE id = ?", (request_data.position_id,), "Position not found")
        week_dates = build_week_dates(request_data.week_start_date)
        employees = load_position_employees(connection, request_data.position_id)
        if not employees:
            raise HTTPException(status_code=400, detail="No employees assigned to this position")

        templates_list = load_active_templates(connection)
        if not templates_list:
            raise HTTPException(status_code=400, detail="No active shift templates found")

        coverage_requirements = load_coverage_requirements_for_position(connection, request_data.position_id)
        legacy_requirements = load_legacy_shift_requirements(connection, request_data.position_id)
        if not coverage_requirements and not legacy_requirements:
            raise HTTPException(status_code=400, detail="No coverage or shift requirements found for this position")

        created_entries: list[dict] = []
        errors: list[str] = []
        unfilled_reports: list[dict] = []
        feasibility_report = build_generation_feasibility_report(
            connection,
            employees,
            templates_list,
            coverage_requirements,
            legacy_requirements,
            request_data.position_id,
            request_data.week_start_date,
            week_dates,
        )
        for issue in feasibility_report["issues"]:
            errors.append(format_feasibility_issue(issue))

        if coverage_requirements:
            fill_week_by_interval_coverage(
                connection,
                cursor,
                employees,
                templates_list,
                coverage_requirements,
                request_data.position_id,
                request_data.week_start_date,
                week_dates,
                created_entries,
                errors,
                unfilled_reports,
            )
        else:
            for date_string in week_dates:
                fill_day_by_legacy_categories(
                    connection,
                    cursor,
                    employees,
                    templates_list,
                    legacy_requirements,
                    request_data.position_id,
                    request_data.week_start_date,
                    date_string,
                    created_entries,
                    errors,
                    unfilled_reports,
                )

        optimization_moved_count = post_optimize_generated_schedule(
            connection,
            cursor,
            employees,
            request_data.position_id,
            request_data.week_start_date,
            week_dates,
            created_entries,
            errors,
        )

        append_fatigue_summary_warnings(connection, employees, week_dates, request_data.week_start_date, errors)
        day_off_count = sync_generated_day_off_statuses(
            connection,
            cursor,
            employees,
            request_data.position_id,
            week_dates,
        )

        connection.commit()
        return {
            "message": "Auto-generation finished",
            "created_count": len(created_entries),
            "created_entries": created_entries,
            "day_off_count": day_off_count,
            "optimization_moved_count": optimization_moved_count,
            "feasibility_report": feasibility_report,
            "unfilled_reports": unfilled_reports,
            "errors": errors,
        }
    finally:
        connection.close()


@app.get("/api/schedule/export-excel", tags=["Schedule"])
def export_schedule_excel(week_start_date: str, position_id: int, lang: str = "en"):
    connection = get_connection()
    try:
        if lang not in {"en", "ru", "he"}:
            lang = "en"
        week_dates = build_week_dates(week_start_date)
        cursor = connection.cursor()
        position_row = fetch_one_or_404(cursor, "SELECT * FROM positions WHERE id = ?", (position_id,), "Position not found")
        position = row_to_position_dict(position_row)
        employees = load_position_employees(connection, position_id)
        entries = get_schedule_entries(connection, position_id=position_id, dates=week_dates)
        day_status_map = get_employee_day_status_map(connection, [employee["id"] for employee in employees], week_dates)
        grouped_entries: dict[tuple[int, str], list[dict]] = {}
        for entry in entries:
            grouped_entries.setdefault((entry["employee_id"], entry["date"]), []).append(entry)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Schedule"
        worksheet.sheet_view.rightToLeft = True

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        date_fill = PatternFill("solid", fgColor="F3F6F9")
        thin_side = Side(style="thin", color="CCCCCC")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
        worksheet["A1"] = f"Schedule export - {position['name']} - week starting {week_start_date}"
        worksheet["A1"].font = Font(bold=True, size=14)
        worksheet["A1"].alignment = center

        headers = ["Employee", "Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Weekly total"]
        for col_index, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=3, column=col_index, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border

        for index, current_date in enumerate(["", *week_dates, ""], start=1):
            cell = worksheet.cell(row=4, column=index, value=current_date)
            cell.fill = date_fill
            cell.alignment = center
            cell.border = thin_border

        for row_index, employee in enumerate(employees, start=5):
            name_cell = worksheet.cell(row=row_index, column=1, value=employee["full_name"])
            name_cell.font = Font(bold=True)
            name_cell.alignment = center
            name_cell.border = thin_border
            weekly_count = 0
            max_lines = 1
            for day_offset, current_date in enumerate(week_dates, start=2):
                day_entries = grouped_entries.get((employee["id"], current_date), [])
                day_status = day_status_map.get((employee["id"], current_date))
                if not (day_status and day_status.get("status_type") in {"sick", "day_off"}):
                    weekly_count += sum(1 for entry in day_entries if not entry.get("no_show"))
                text = build_schedule_cell_text(day_entries, day_status, lang)
                max_lines = max(max_lines, text.count("\n") + 1 if text else 1)
                cell = worksheet.cell(row=row_index, column=day_offset, value=text)
                cell.alignment = center
                cell.border = thin_border
            total_cell = worksheet.cell(row=row_index, column=9, value=weekly_count)
            total_cell.alignment = center
            total_cell.border = thin_border
            worksheet.row_dimensions[row_index].height = max(22, max_lines * 18)

        for column in "ABCDEFGHI":
            worksheet.column_dimensions[column].width = 24 if column != "I" else 12
        worksheet.freeze_panes = "B5"
        worksheet.page_setup.orientation = "landscape"
        worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.print_title_rows = "1:4"
        worksheet.print_options.horizontalCentered = True

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        safe_position_name = position["name"].replace(" ", "_")
        filename = f"schedule_{safe_position_name}_{week_start_date}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    finally:
        connection.close()
